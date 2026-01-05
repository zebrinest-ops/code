# -*- coding: utf-8 -*-
"""
Created on Fri Mar  7 10:57:07 2025

@author: Administrator
"""

import numpy as np
import pandas as pd
import math
import os,gc
from glob import glob
import xarray as xr
from datetime import datetime
import warnings
from pathlib import Path
from scipy.interpolate import interp1d
from scipy.interpolate import RegularGridInterpolator
from scipy import stats
import statsmodels.api as sm
import metpy.calc
from metpy.units import units
from metpy.calc import relative_humidity_from_dewpoint
import metpy.calc as mpcalc
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.lines as mlines
from matplotlib.gridspec import GridSpec
from matplotlib.ticker import MultipleLocator, AutoMinorLocator
with warnings.catch_warnings():
    warnings.filterwarnings("ignore", category=FutureWarning)
    # 此处代码中的FutureWarning将被忽略
# 铁塔海拔
Altitude = 49
output_path = "E:\\Beijing2024\\数据处理结果2nd\\"
try:
    os.makedirs(output_path, exist_ok=True)
except:
    print(f"权限不足，无法创建目录：{output_path}")
    exit(1)
def goff_gratch_formula_dataframe(df):#冰面的公式另外需要列出，计算夏天此处省略
    def goff_gratch_formula(T):
        T_ref = 373.15  # Reference temperature in Kelvin (100°C)
        log10_es = (
            -7.90298 * (T_ref / T - 1) +
            5.02808 * math.log10(T_ref / T) -
            1.3816e-7 * (10**(11.344 * (1 - T / T_ref)) - 1) +
            8.1328e-3 * (10**(-3.49149 * (T_ref / T - 1)) - 1) +
            math.log10(1013.246)
        )
        es = 10**log10_es
        return es
    
    # Apply the Goff-Gratch formula to each element in the DataFrame
    es_df = df.map(goff_gratch_formula)
    return es_df
############处理铁塔数据###################
# Set the path to your target directory
path = r'E:\Beijing2024\铁塔\2024铁塔平均场6月6至6月98月5日至8月10日10天数据(1)\0805-0810\\'

# Get a list of all files in the folder
files = [f for f in os.listdir(path) if f.endswith('.txt')]  # Assuming all files are .txt files

# Initialize an empty list to collect data frames
all_data = []

# Loop through each file in the folder
for file in files:
    file_path = os.path.join(path, file)
    
    # Read data from the file
    df = pd.read_table(file_path, skiprows=5, header=None, sep=r'\s+')
    print(file)
    # Iterate over the data and create timestamps
    for i in range(1, 4321):  # i 从 1 到 4320
        # 获取当前时间点的各个部分
        if (18*i - 2) >= len(df):
            break
        year = "20" + str(df.iloc[18*i-2, 1])  # 获取年份（假设年份在第 1 列）
        day = df.iloc[18*i-2, 3]  # 获取日期（假设日期在第 3 列）
        month = df.iloc[18*i-2, 5]  # 获取月份（假设月份在第 5 列）
        hour = df.iloc[18*i-1, 1]  # 获取小时（假设小时在第 1 列）
        minute = df.iloc[18*i-1, 3]  # 获取分钟（假设分钟在第 3 列）
        second = df.iloc[18*i-1, 5]  # 获取秒（假设秒在第 5 列）

        # 创建时间戳
        timestamps = pd.to_datetime(f"{year}-{month}-{day} {hour}:{minute}:{second}")

        # 将时间戳添加到 DataFrame 中
        df.loc[18*i-17:18*i-3, 'timestamps'] = timestamps
    
    # Append the processed DataFrame to the list
    all_data.append(df)

# Concatenate all DataFrames from all files into one large DataFrame
final_df = pd.concat(all_data, ignore_index=True)

tieta_filtered = final_df[~final_df.iloc[:, 0].str.contains('=|-', na=False)]
tieta_filtered.columns = tieta_filtered.iloc[0]
tieta_filtered.columns.values[6] = 'timestamps'
tieta_filtered = tieta_filtered[1:].reset_index(drop=True)  # 删除第一行并重置索引
tieta_filtered.set_index('timestamps', inplace=True)
tieta_filtered.replace('999.0', np.nan, inplace=True)
#Vnw和Vse是由同一条线上两个传感器测到的，可以任取其中一个或者作平均
tieta_filtered['ws_mean(m/s)'] = tieta_filtered.iloc[:, 1:3].apply(pd.to_numeric, errors='coerce').mean(axis=1)
tieta_filtered['T(c)'] = tieta_filtered['T(c)'].apply(lambda x: pd.to_numeric(x,  errors='coerce'))
tieta_filtered['T(c)'] = tieta_filtered['T(c)'].astype(float).round(3)
tieta_filtered['T(K)'] = tieta_filtered['T(c)']+273.15
tieta_filtered['RH(%)'] = tieta_filtered['RH(%)'].apply(lambda x: pd.to_numeric(x,  errors='coerce'))
tieta_filtered['RH(%)'] = tieta_filtered['RH(%)'].astype(float).round(1)
tieta_filtered['D(deg)'] = tieta_filtered['D(deg)'].apply(lambda x: pd.to_numeric(x,  errors='coerce'))
tieta_filtered['D(deg)'] = tieta_filtered['D(deg)'].astype(float).round(1)
tieta_filtered['H(m)'] = tieta_filtered['H(m)'].apply(lambda x: pd.to_numeric(x,  errors='coerce'))
tieta_filtered['H(m)'] = tieta_filtered['H(m)'].astype(float).round(0)

tieta_filtered['u'] = tieta_filtered['ws_mean(m/s)'] *np.sin(np.deg2rad(tieta_filtered['D(deg)']))
tieta_filtered['v'] = tieta_filtered['ws_mean(m/s)'] *np.cos(np.deg2rad(tieta_filtered['D(deg)']))

tieta_filtered = tieta_filtered.drop(columns=['Vnw(m/s)', 'Vse(m/s)', 'T(c)'])
tieta_filtered.columns = ['H(m)', 'wd_tieta', 'RH_tieta', 'ws_tieta', 'T(K)_tieta','u','v']

tieta_1min = (tieta_filtered.reset_index().groupby(['H(m)', pd.Grouper(key='timestamps', freq='1min')]).mean()).reset_index()
tieta_5min = (tieta_filtered.reset_index().groupby(['H(m)', pd.Grouper(key='timestamps', freq='5min')]).mean()).reset_index()
tieta_60min = (tieta_filtered.reset_index().groupby(['H(m)', pd.Grouper(key='timestamps', freq='60min')]).mean()).reset_index()



tieta_filtered = tieta_filtered.reset_index()
tieta_filtered['timestamps'] = tieta_filtered['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
tieta_1min['timestamps'] = tieta_1min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
tieta_5min['timestamps'] = tieta_5min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
tieta_60min['timestamps'] = tieta_60min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')

# 调整列顺序
new_columns = ['timestamps', 'H(m)'] + [col for col in tieta_filtered.columns if col not in ['timestamps', 'H(m)']]
tieta_filtered = tieta_filtered[new_columns]
tieta_1min = tieta_1min[new_columns]
tieta_5min = tieta_5min[new_columns]
tieta_60min = tieta_60min[new_columns]

# 按时间和高度排序
tieta_filtered = tieta_filtered.sort_values(by=['timestamps', 'H(m)'])
tieta_1min = tieta_1min.sort_values(by=['timestamps', 'H(m)'])
tieta_5min = tieta_5min.sort_values(by=['timestamps', 'H(m)'])
tieta_60min = tieta_60min.sort_values(by=['timestamps', 'H(m)'])

tieta_1min['height_plus_alt'] = tieta_1min['H(m)'] + Altitude
tieta_1min['Level (hPa)'] = pd.DataFrame(metpy.calc.height_to_pressure_std(tieta_1min['height_plus_alt'].to_list() * units('m')).m)
tieta_1min = tieta_1min.set_index(['timestamps', 'Level (hPa)'])

#tieta_filtered.to_csv(os.path.join(path, 'tieta_origin_merged.csv'), index=False)
# with pd.ExcelWriter(output_path + 'tieta0805-0810_1-5-60.xlsx',  engine='xlsxwriter', date_format='yyyy/mm/dd hh:mm:ss') as writer:
#     # 写入数据到不同sheet
#     tieta_filtered.to_excel(writer,  sheet_name='tieta0805-0810_1s', index=False)
#     tieta_1min.reset_index().to_excel(writer,  sheet_name='tieta0805-0810_1min', index=False)
#     tieta_5min.to_excel(writer,  sheet_name='tieta0805-0810_5min', index=False)
#     tieta_60min.to_excel(writer,  sheet_name='tieta0805-0810_60min', index=False)
    
#     # 设置列宽（仅适用于xlsxwriter引擎）
#     for sheet_name in writer.sheets: 
#         worksheet = writer.sheets[sheet_name] 
#         worksheet.set_column('A:A',  20)  # 注意参数格式为'A:A'而非'A'


#######超声风数据处理#######
#注释掉的代码为处理原始数据的代码，运行一次保存合并后数据，后续读取时直接读取合并后原始文件
# main_dir = r'E:\Beijing2024\超声风\2024-08-05---2024-08-10'

# # 获取所有需要处理的文件路径及对应高度（包含所有扩展名）
# file_list = []
# for raw_folder in glob(os.path.join(main_dir, '*', 'raw')):
#     height = os.path.basename(os.path.dirname(raw_folder))
#     for data_file in glob(os.path.join(raw_folder, '*.*')):  # 匹配所有文件类型
#         file_list.append((data_file, height))

# required_columns = [
#     'CO2 (mmol/m^3)', 'CO2 (mg/m^3)',
#     'H2O (mmol/m^3)', 'H2O (g/m^3)',
#     'Temperature (C)', 'Pressure (kPa)',
#     'CO2 (umol/mol)', 'H2O (mmol/mol)',
#     'Dew Point (C)', 'U (m/s)', 'V (m/s)',
#     'W (m/s)', 'T (C)'
# ]

# # 关键时间列检查清单
# time_columns = ['Date', 'Time']

# dfs = []

# for file_path, height in file_list:
#     try:
#         # 读取数据（自动检测分隔符）
#         df = pd.read_csv(file_path, sep=None, engine='python', skiprows=7, header=0)
#         print(file_path)
#         # 检查必要时间列是否存在
#         missing_cols = [col for col in time_columns if col not in df.columns]
#         if missing_cols:
#             raise KeyError(f"缺少关键列 {missing_cols}，无法生成时间戳")
#         df['Time'] = df['Time'].str.replace(r':(\d{3})$', r'.\1', regex=True)  
        
#         df['timestamps'] = pd.to_datetime(df['Date'] + ' ' + df['Time'], format='%Y-%m-%d %H:%M:%S.%f')

#         # 添加高度列并选择所需列
#         df['Height'] = height
#         selected_cols = ['timestamps', 'Height'] + [col for col in required_columns if col in df.columns]
#         df_selected = df[selected_cols]
        
#         dfs.append(df_selected)
        
#     except Exception as e:
#         print(f"跳过文件 {os.path.basename(file_path)}，原因：{str(e)}")
#         continue

# # 合并处理后的数据
# if dfs:
#     combined_df = pd.concat(dfs, ignore_index=True)
    
#     # 处理压力列
#     if 'Pressure (kPa)' in combined_df.columns:
#         combined_df['Pressure (hPa)'] = combined_df['Pressure (kPa)'] * 10
#         combined_df.drop('Pressure (kPa)', axis=1, inplace=True)
#     else:
#         print("警告：未找到 Pressure (kPa) 列")
    
#     print("合并完成，有效记录数:", len(combined_df))
#     print("最终列顺序:", combined_df.columns.tolist())
# else:
#     combined_df = pd.DataFrame()
#     print("未合并任何有效数据")

###############处理超声风数据##############################
combined_df = pd.read_parquet('E:\\Beijing2024\\超声风\\turb.parquet', engine='pyarrow')
#combined_df = pd.read_csv('E:\\Beijing2024\\超声风\\2024-08-05---2024-08-10\\combined_ultrasonic_wind.csv')
combined_df = combined_df.rename(columns={'Timestamp':  'timestamps'})       
combined_df['timestamps'] = pd.to_datetime(combined_df['timestamps'])
combined_df = combined_df.replace(-9999, np.nan)

# 结果示例
print("\n示例数据：")
print(combined_df.head(2) if not combined_df.empty else "空数据集")
ultrasonic_turb_1min = (combined_df.groupby(['Height', pd.Grouper(key='timestamps', freq='1min')]).mean()).reset_index()
ultrasonic_turb_1min['hws'] = np.sqrt(np.square(ultrasonic_turb_1min['U (m/s)']) + np.square(ultrasonic_turb_1min['V (m/s)']))
ultrasonic_turb_1min['wd'] =  (90 - np.degrees(np.arctan2(ultrasonic_turb_1min['V (m/s)'], -ultrasonic_turb_1min['U (m/s)']))) % 360

ultrasonic_turb_1s = (combined_df.groupby(['Height', pd.Grouper(key='timestamps', freq='1s')]).mean()).reset_index()
ultrasonic_turb_1s['hws'] = np.sqrt(np.square(ultrasonic_turb_1s['U (m/s)']) + np.square(ultrasonic_turb_1s['V (m/s)']))
ultrasonic_turb_1s['wd'] =  (90 - np.degrees(np.arctan2(ultrasonic_turb_1s['V (m/s)'], -ultrasonic_turb_1s['U (m/s)']))) % 360


#U：风从南往北为+值    风从北往南为-值  V：风从东往西为+值    风从西往东为-值

# 定义各高度层的校正信息（包括U/V校正因子和风向夹角θ）
correction_info = {
    '8m': {'factor': 1/np.cos(math.radians(6)), 'theta': 6},
    '16m': {'factor': 1/np.cos(math.radians(8)), 'theta': 8},
    '47m': {'factor': 1/np.cos(math.radians(4)), 'theta': 4},
    '80m': {'factor': 1/np.cos(math.radians(2)), 'theta': 2},
    '140m': {'factor': 1/np.cos(math.radians(2)), 'theta': 2},
    '200m': {'factor': 1/np.cos(math.radians(3)), 'theta': 3},
    '280m': {'factor': 1/np.cos(math.radians(1)), 'theta': 1}
}
def correct_wind_data(df, correction_info):
    """
    校正风速和风向，结果保存到新列，并隔离原始数据
    
    参数:
    df (pd.DataFrame): 必须包含列 ['Height', 'U (m/s)', 'V (m/s)', 'wd']
    correction_info (dict): 校正参数，键为高度字符串（如 '8m'）
    
    返回:
    pd.DataFrame: 包含校正列的新数据框，原始数据框完全不被修改
    """
    # 创建数据框的拷贝以隔离原始数据
    df = df.copy()
    
    # 检查必要列是否存在
    required_columns = ['Height', 'U (m/s)', 'V (m/s)', 'wd']
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        raise ValueError(f"输入数据框缺少必要列: {missing_cols}")
    
    # 确保height列是字符串类型
    df['Height'] = df['Height'].astype(str).str.strip()  # 同时去除首尾空格
    
    # 动态映射参数
    df['factor'] = df['Height'].map(lambda x: correction_info.get(x, {}).get('factor', 1.0))
    df['theta'] = df['Height'].map(lambda x: correction_info.get(x, {}).get('theta', 0.0))
    
    # 校正并保存到新列
    df['U_corrected (m/s)'] = df['U (m/s)'] * df['factor']
    df['V_corrected (m/s)'] = df['V (m/s)'] * df['factor']
    df['wd_corrected'] = (df['wd'] + df['theta']) % 360
    
    # 删除临时列
    df.drop(['factor', 'theta'], axis=1, inplace=True)
    
    return df

turb_calibrated_1s = correct_wind_data(ultrasonic_turb_1s,correction_info)
turb_calibrated_1s['timestamps'] = pd.to_datetime(turb_calibrated_1s['timestamps']) 
turb_calibrated_1s['timestamps'] = turb_calibrated_1s['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
turb_calibrated_1s['timestamps'] = pd.to_datetime(turb_calibrated_1s['timestamps'], format='%Y/%m/%d %H:%M:%S')

turb_calibrated_1min = correct_wind_data(ultrasonic_turb_1min,correction_info)
turb_calibrated_1min['timestamps'] = pd.to_datetime(turb_calibrated_1min['timestamps']) 
turb_calibrated_1min['timestamps'] = turb_calibrated_1min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
turb_calibrated_1min = (turb_calibrated_1s.groupby(['Height', pd.Grouper(key='timestamps', freq='1min')]).mean()).reset_index()
turb_calibrated_1min['hws_corrected'] = np.sqrt(np.square(turb_calibrated_1min['U_corrected (m/s)']) + np.square(turb_calibrated_1min['V_corrected (m/s)']))
turb_calibrated_1min['wd_corrected'] =  (90 - np.degrees(np.arctan2(turb_calibrated_1min['V_corrected (m/s)'], -turb_calibrated_1min['U_corrected (m/s)']))) % 360
turb_calibrated_1min['Height'] = list(map(lambda x: float(x.strip('m')), turb_calibrated_1min['Height']))
turb_calibrated_1min['Height_plus_alt'] = turb_calibrated_1min['Height'] + Altitude
turb_calibrated_1min['Level (hPa)'] = pd.DataFrame(metpy.calc.height_to_pressure_std(turb_calibrated_1min['Height_plus_alt'].to_list() * units('m')).m)
turb_calibrated_1min['T (K)'] = turb_calibrated_1min['Temperature (C)'] +273.15
turb_calibrated_1min['rh'] = 100*goff_gratch_formula_dataframe(turb_calibrated_1min['Dew Point (C)']+273.15)/goff_gratch_formula_dataframe(turb_calibrated_1min['Temperature (C)']+273.15)
turb_calibrated_1min['q*10^-2'] = 100*0.622*goff_gratch_formula_dataframe(turb_calibrated_1min['Dew Point (C)']+273.15)/(turb_calibrated_1min['Pressure (hPa)']
                                                                                                                        -0.378*goff_gratch_formula_dataframe(turb_calibrated_1min['Dew Point (C)']+273.15))
turb_calibrated_1min = turb_calibrated_1min.sort_values(by=['timestamps', 'Height'], ascending=[True, True])
turb_calibrated_1min['rh'] = relative_humidity_from_dewpoint(turb_calibrated_1min['Temperature (C)'].values * units.degC, turb_calibrated_1min['Dew Point (C)'].values * units.degC).to('percent').magnitude


turb_calibrated_5min = (turb_calibrated_1s.groupby(['Height', pd.Grouper(key='timestamps', freq='5min')]).mean()).reset_index()
turb_calibrated_5min['hws_corrected'] = np.sqrt(np.square(turb_calibrated_5min['U_corrected (m/s)']) + np.square(turb_calibrated_5min['V_corrected (m/s)']))
turb_calibrated_5min['wd_corrected'] =  (90 - np.degrees(np.arctan2(turb_calibrated_5min['V_corrected (m/s)'], -turb_calibrated_5min['U_corrected (m/s)']))) % 360
turb_calibrated_5min['Height'] = list(map(lambda x: float(x.strip('m')), turb_calibrated_5min['Height']))
turb_calibrated_5min['Height_plus_alt'] = turb_calibrated_5min['Height'] + Altitude
turb_calibrated_5min['Level (hPa)'] = pd.DataFrame(metpy.calc.height_to_pressure_std(turb_calibrated_5min['Height_plus_alt'].to_list() * units('m')).m)
turb_calibrated_5min['rh'] = 100*goff_gratch_formula_dataframe(turb_calibrated_5min['Dew Point (C)']+273.15)/goff_gratch_formula_dataframe(turb_calibrated_5min['Temperature (C)']+273.15)
turb_calibrated_5min['q*10^-2'] = 100*0.622*goff_gratch_formula_dataframe(turb_calibrated_5min['Dew Point (C)']+273.15)/(turb_calibrated_5min['Pressure (hPa)'] 
                                                                                                                         - 0.378*goff_gratch_formula_dataframe(turb_calibrated_5min['Dew Point (C)']+273.15))
turb_calibrated_5min = turb_calibrated_5min.sort_values(by=['timestamps', 'Height'], ascending=[True, True])
turb_calibrated_5min['T (K)'] = turb_calibrated_5min['Temperature (C)'] +273.15
turb_calibrated_5min['rh'] = relative_humidity_from_dewpoint(turb_calibrated_5min['Temperature (C)'].values * units.degC, turb_calibrated_5min['Dew Point (C)'].values * units.degC).to('percent').magnitude


turb_calibrated_60min = (turb_calibrated_1s.groupby(['Height', pd.Grouper(key='timestamps', freq='60min')]).mean()).reset_index()
#turb_calibrated_60min = turb_calibrated_60min.set_index(['timestamps','Height'])
#turb_calibrated_60min = turb_calibrated_60min[['T (C)','rh','U (m/s)', 'V (m/s)','W (m/s)','q*10^-2']]
#turb_calibrated_60min['W (m/s)'] = 100 * turb_calibrated_60min['W (m/s)']
#turb_calibrated_60min = turb_calibrated_60min.rename(columns={'W (m/s)':  'W (10⁻² {m/s})'})
turb_calibrated_60min['Height'] = list(map(lambda x: float(x.strip('m')), turb_calibrated_60min['Height']))
#turb_calibrated_60min = turb_calibrated_60min.reset_index()
turb_calibrated_60min['hws_corrected'] = np.sqrt(np.square(turb_calibrated_60min['U_corrected (m/s)']) + np.square(turb_calibrated_60min['V_corrected (m/s)']))
turb_calibrated_60min['wd_corrected'] =  (90 - np.degrees(np.arctan2(turb_calibrated_60min['V_corrected (m/s)'], -turb_calibrated_60min['U_corrected (m/s)']))) % 360
turb_calibrated_60min['T (K)'] = turb_calibrated_60min['Temperature (C)'] +273.15
turb_calibrated_60min['rh'] = relative_humidity_from_dewpoint(turb_calibrated_60min['Temperature (C)'].values * units.degC, turb_calibrated_60min['Dew Point (C)'].values * units.degC).to('percent').magnitude

def process_turb_data(df_input, full_time):
    results = []
    for height in df_input['Height'].unique():
        # 提取当前高度层数据
        df_height = df_input[df_input['Height'] == height].copy()
        
        # 处理时间索引并重新对齐
        df_height = (df_height
                     .set_index(pd.to_datetime(df_height['timestamps']))  # 转换并设索引
                     .reindex(full_time))                                # 重新索引
        
        # 填充高度并重置索引
        df_height = (df_height
                     .assign(Height=height)                             # 填充高度
                     .reset_index()
                     .rename(columns={'index': 'timestampss'}))          # 重命名时间列
        
        results.append(df_height)
    
    return pd.concat(results, ignore_index=True)

# 生成公共时间范围
full_time = pd.date_range(start='2024-06-05 00:00:00', end='2024-08-10 23:59:59', freq='1min')
turb_calibrated_1min = process_turb_data(turb_calibrated_1min, full_time)
# full_time = pd.date_range(start='2024-06-05 00:00:00', end='2024-08-10 23:59:59', freq='5min')
# turb_calibrated_5min = process_turb_data(turb_calibrated_5min, full_time)
full_time = pd.date_range(start='2024-06-05 00:00:00', end='2024-08-10 23:59:59', freq='60min')
turb_calibrated_60min = process_turb_data(turb_calibrated_60min, full_time)

turb_calibrated_1min = turb_calibrated_1min.drop(turb_calibrated_1min.columns[2], axis=1)
turb_calibrated_1min = turb_calibrated_1min.rename(columns={'timestampss': 'timestamps'})
# turb_calibrated_5min = turb_calibrated_5min.drop(turb_calibrated_5min.columns[2], axis=1)
# turb_calibrated_5min = turb_calibrated_5min.rename(columns={'timestampss': 'timestamps'})
turb_calibrated_60min = turb_calibrated_60min.drop(turb_calibrated_60min.columns[2], axis=1)
turb_calibrated_60min = turb_calibrated_60min.rename(columns={'timestampss': 'timestamps'})

turb_calibrated_1min['timestamps'] = turb_calibrated_1min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
#turb_calibrated_5min['timestamps'] = turb_calibrated_5min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
turb_calibrated_60min['timestamps'] = turb_calibrated_60min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')

turb_calibrated_1min = turb_calibrated_1min.sort_values(
    by=['Height','timestamps'],  # 按时间→气压层级双排序
    ascending=[True,True]                # 默认升序（时间从早到晚，层级从低到高）
).dropna()
# turb_calibrated_5min = turb_calibrated_5min.sort_values(
#     by=['Height','timestamps'],  # 按时间→气压层级双排序
#     ascending=[True,True]                # 默认升序（时间从早到晚，层级从低到高）
# )
turb_calibrated_60min = turb_calibrated_60min.sort_values(
    by=['Height','timestamps'],  # 按时间→气压层级双排序
    ascending=[True,True]                # 默认升序（时间从早到晚，层级从低到高）
).dropna()

# turb_calibrated_1s.to_csv(output_path + 'turb_1s_corrected.csv',index=False)

# with pd.ExcelWriter(output_path + 'ultrasonic_1_5_60.xlsx', date_format='yyyy/mm/dd hh:mm:ss') as writer:

#     turb_calibrated_1min.to_excel(writer, sheet_name='turb_1min_corrected', index=False)
        
#     turb_calibrated_5min.to_excel(writer, sheet_name='turb_5min_corrected', index=False)
    
#     turb_calibrated_60min.to_excel(writer, sheet_name='turb_60min_corrected', index=False)
    
#     workbook = writer.book  
#     for sheet_name in writer.sheets: 
#         worksheet = writer.sheets[sheet_name] 
#         worksheet.set_column('A:Z',  17.5)  # 注意参数格式为'A:A'而非'A'



#####处理微波辐射计数据#####
folder_path = 'E:\\Beijing2024\\MWR\\ahrh\\04-07\\'  # 指定文件夹路径
ah = pd.read_excel(folder_path+'ah&rh_merged.xlsx',sheet_name="ah_data")
ah = ah.rename(columns={'timestamp':  'timestamps'})       
ah['timestamps'] = pd.to_datetime(ah['timestamps'])
ah.set_index('timestamps', inplace=True)
rh = pd.read_excel(folder_path+'ah&rh_merged.xlsx',sheet_name="rh_data")
rh = rh.rename(columns={'timestamp':  'timestamps'})
rh['timestamps'] = pd.to_datetime(rh['timestamps'])
rh.set_index('timestamps', inplace=True)

folder_path = 'E:\\Beijing2024\\MWR\\temp\\0702-0810\\'
temp = pd.read_excel(folder_path+'temp merged.xlsx',sheet_name="0702-0810")
temp = temp.rename(columns={'timestamp':  'timestamps'}) 
temp['timestamps'] = pd.to_datetime(temp['timestamps'], format='%Y/%m/%d %H:%M:%S')
temp.set_index('timestamps', inplace=True)

#####处理微波辐射计数据#####
folder_path2 = 'E:\\Beijing2024\\MWR\\ahrh\\04-06\\'  # 指定文件夹路径
ah2 = pd.read_excel(folder_path2+'ah&rh_merged.xlsx',sheet_name="ah_data")
ah2 = ah2.rename(columns={'timestamp':  'timestamps'})       
ah2['timestamps'] = pd.to_datetime(ah2['timestamps'])
ah2.set_index('timestamps', inplace=True)
rh2 = pd.read_excel(folder_path2+'ah&rh_merged.xlsx',sheet_name="rh_data")
rh2 = rh2.rename(columns={'timestamp':  'timestamps'})
rh2['timestamps'] = pd.to_datetime(rh2['timestamps'])
rh2.set_index('timestamps', inplace=True)

folder_path2 = 'E:\\Beijing2024\\MWR\\temp\\0605-0624北京\\'
temp2 = pd.read_excel(folder_path2+'temp merged.xlsx',sheet_name="0605-0624")
temp2 = temp2.rename(columns={'timestamp':  'timestamps'}) 
temp2['timestamps'] = pd.to_datetime(temp2['timestamps'], format='%Y/%m/%d %H:%M:%S')
temp2.set_index('timestamps', inplace=True)

temp = pd.concat([temp,temp2], axis=0)
rh = pd.concat([rh,rh2], axis=0)


###############################################################
folder_path = 'E:\\Beijing2024\\MWR\\met\\'  # 指定路径
all_data = []  # 创建一个空列表以存储数据帧

# 遍历文件夹中的所有文件
for filename in os.listdir(folder_path):
    file_path = os.path.join(folder_path, filename)  # 获取文件完整路径
    # 使用pandas读取文件，跳过前8行
    data = pd.read_table(file_path, skiprows=19,delimiter=',', header=None, encoding='ANSI')
    all_data.append(data)  # 将数据帧添加到列表中
merged_data = pd.concat(all_data, ignore_index=True)

met=merged_data.iloc[:,:10]

met.columns = ['Year', 'Month', 'Day', 'Hour', 'Minute', 'Second','rain_flag','pressure','temperature','RH']
met['Year'] = met['Year'].apply(lambda x: 2000 + x)

# 将两位数年份转换为四位数年份
# 创建一个新的列来存储时间戳
met['timestamp'] = pd.to_datetime(met[['Year', 'Month', 'Day', 'Hour', 'Minute', 'Second']])
met['timestamp'] = met['timestamp'] + pd.Timedelta(hours=8)

df1_met = pd.DataFrame(met)
df1_met['timestamp'] = pd.to_datetime(df1_met['timestamp'])
df1_met.set_index('timestamp', inplace=True)
met_1min=df1_met.resample('1min').mean().interpolate()
met_1min = met_1min[['pressure', 'temperature', 'RH']]
###################################################################

start_time = '2024-06-05 10:35:00'#temp.index.min()
end_time = '2024-08-10 11:58:59'#temp.index.max()

temp_1min = temp[(temp.index >= start_time)&(temp.index <= end_time)]
temp_1min_index = pd.date_range(start=start_time,end=end_time,freq='1min')  # '1T' 表示1分钟
combined_index = temp_1min.index.union(temp_1min_index).drop_duplicates().sort_values()
temp_1min_reindexed = temp_1min.reindex(combined_index) 
temp_1min_reindexed_interpolated = temp_1min_reindexed.interpolate(method='time') 
temp_1min_reindexed_interpolated.update(temp_1min) 
temp_1min_reindexed_interpolated_final = temp_1min_reindexed_interpolated.reset_index().rename(columns={'index':  'timestamps'})
temp_1min_reindexed_interpolated_final = temp_1min_reindexed_interpolated_final[temp_1min_reindexed_interpolated_final['timestamps'].dt.second  == 0].reset_index(drop=True)

rh_1min = rh[(rh.index >= start_time)&(rh.index <= end_time)]
rh_1min_index = pd.date_range(start=start_time,end=end_time,freq='1min')  # '1T' 表示1分钟
combined_index = rh_1min.index.union(rh_1min_index).drop_duplicates().sort_values()
rh_1min_reindexed = rh_1min.reindex(combined_index) 
rh_1min_reindexed_interpolated = rh_1min_reindexed.interpolate(method='time') 
rh_1min_reindexed_interpolated.update(rh_1min) 
rh_1min_reindexed_interpolated_final = rh_1min_reindexed_interpolated.reset_index().rename(columns={'index':  'timestamps'})
rh_1min_reindexed_interpolated_final = rh_1min_reindexed_interpolated_final[rh_1min_reindexed_interpolated_final['timestamps'].dt.second  == 0].reset_index(drop=True)

ah_1min = ah[(ah.index >= start_time)&(ah.index <= end_time)]
ah_1min_index = pd.date_range(start=start_time,end=end_time,freq='1min')  # '1T' 表示1分钟
combined_index = ah_1min.index.union(ah_1min_index).drop_duplicates().sort_values()
ah_1min_reindexed = ah_1min.reindex(combined_index) 
ah_1min_reindexed_interpolated = ah_1min_reindexed.interpolate(method='time') 
ah_1min_reindexed_interpolated.update(ah_1min) 
ah_1min_reindexed_interpolated_final = ah_1min_reindexed_interpolated.reset_index().rename(columns={'index':  'timestamps'})
ah_1min_reindexed_interpolated_final = ah_1min_reindexed_interpolated_final[ah_1min_reindexed_interpolated_final['timestamps'].dt.second  == 0].reset_index(drop=True)

met = met.set_index("timestamp")
met_1min = met[(met.index >= start_time)&(met.index <= end_time)]
met_1min_index = pd.date_range(start=start_time,end=end_time,freq='1min')  # '1T' 表示1分钟

temp_1min_reindexed_interpolated_final.set_index('timestamps', inplace=True)
temp_60min = temp_1min_reindexed_interpolated_final.resample('h').mean() 

es_1min = goff_gratch_formula_dataframe(temp_1min_reindexed_interpolated_final)
es_60min = goff_gratch_formula_dataframe(temp_60min)

rh_1min_reindexed_interpolated_final['timestamps'] = pd.to_datetime(rh_1min_reindexed_interpolated_final['timestamps'], format='%Y-%m-%d %H:%M:%S')
rh_1min_reindexed_interpolated_final.set_index('timestamps', inplace=True)
rh_60min = rh_1min_reindexed_interpolated_final.resample('h').mean()
met_60min = met_1min.resample('h').mean()

L = 0.0065  # 温度梯度 (K/m)
g = 9.8017104   # 重力加速度 (m/s^2)
M = 0.0289644  # 空气平均摩尔质量 (kg/mol)
R = 8.3144598  # 气体常数 (J/(mol·K))
def calculate_pressure_series(P0_series, T0_series, heights):
    
    pressures = np.zeros((len(P0_series), len(heights)))
    
    for i in range(len(P0_series)):
        P0 = P0_series[i]
        T0 = T0_series[i]
        pressures[i, :] = P0 * (1 - (L * heights) / T0) ** (g * M / (R * L))        
    return pressures

height_temp=[col for col in temp_1min.columns if isinstance(col, int)]
heights = np.array(height_temp)  # 高度数组 (m)

p_60min = pd.DataFrame(calculate_pressure_series(met_60min['pressure'].to_numpy(), met_60min['temperature'].values, heights), 
                       index=met_60min.index, columns=heights)

p_1min = pd.DataFrame(calculate_pressure_series(met_1min['pressure'].to_numpy(), met_1min['temperature'].values, heights), 
                       index=met_1min.index, columns=heights)

common_timestamps_1min = met_1min.index.intersection(temp_1min_reindexed_interpolated_final.index)
common_timestamps_60min = met_1min.index.intersection(temp_60min.index)

e_1min=np.multiply(es_1min.loc[common_timestamps_1min] , rh_1min_reindexed_interpolated_final.loc[common_timestamps_1min]/100)
e_60min=np.multiply(es_60min.loc[common_timestamps_60min] , rh_60min.loc[common_timestamps_60min[0:119]]/100)




q_1min=0.622*e_1min/(p_1min.loc[common_timestamps_1min]-0.378*e_1min.loc[common_timestamps_1min])
q_60min=0.622*e_60min/(p_60min-0.378*e_60min)

θ_1min=temp_1min_reindexed_interpolated_final.loc[common_timestamps_1min]*(1000/p_1min.loc[common_timestamps_1min]) ** 0.28571428571428564
θ_60min=temp_60min.loc[common_timestamps_60min[0:119]]*(1000/p_60min) ** 0.28571428571428564

θv_1min=θ_1min*(1+0.608*q_1min)
θv_60min=temp_60min.loc[common_timestamps_60min[0:119]]*(1+0.608*q_60min)*(1000/p_60min) ** 0.28571428571428564

t_mwr_long = temp_1min_reindexed_interpolated_final.reset_index().melt(id_vars='timestamps',  var_name='height', value_name='T(K)') 
rh_mwr_long = rh_1min_reindexed_interpolated_final.reset_index().melt(id_vars='timestamps',  var_name='height', value_name='RH(%)') 
#ah_mwr_long = ah_1min_reindexed_interpolated_final.reset_index().melt(id_vars='timestamps',  var_name='height', value_name='ah({g/m-3})') 
q_mwr_long = q_1min.reset_index().melt(id_vars='index',  var_name='height', value_name='q(kg/kg)') 
t_mwr_long = t_mwr_long.set_index(['timestamps', 'height'])
q_mwr_long = q_mwr_long.rename(columns={'index':  'timestamps'})
q_mwr_long = q_mwr_long.set_index(['timestamps', 'height'])
rh_mwr_long = rh_mwr_long.set_index(['timestamps', 'height'])


# 只保留时间和高度两列 
df_mwr = pd.concat([t_mwr_long,  rh_mwr_long, q_mwr_long], axis=1)
df_mwr = df_mwr.sort_index(level=[0, 1], ascending=[True, True])
df_mwr = df_mwr.loc[pd.IndexSlice["2024-06-05":"2024-08-11", :], :]
df_mwr = df_mwr.reset_index()
df_mwr['height_plus_alt'] = df_mwr['height'] + Altitude + 2
#df_mwr = df_mwr.reset_index().dropna()
df_mwr.columns = ['timestamps','height','T(K)','RH(%)','q(kg/kg)','height_plus_alt']
df_mwr['Level (hPa)'] = pd.DataFrame(metpy.calc.height_to_pressure_std(df_mwr['height_plus_alt'].to_list() * units('m')).m)
df_mwr['timestamps'] = pd.to_datetime(df_mwr['timestamps'])
df_mwr['BJT'] = df_mwr['timestamps'] + pd.Timedelta(hours=8)
df_mwr['BJT'] =df_mwr['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
df_mwr['BJT'] = pd.to_datetime(df_mwr['BJT'])
df_mwr_5min = (df_mwr.groupby(['Level (hPa)', pd.Grouper(key='BJT', freq='5min')]).mean()).reset_index()

df_mwr_60min = (df_mwr.groupby(['Level (hPa)', pd.Grouper(key='BJT', freq='60min')]).mean()).reset_index()
df_mwr = df_mwr.set_index(['BJT', 'Level (hPa)'])
df_mwr = df_mwr.sort_index(level=[0,1])  # 先按时间排序，再按高度排序
df_mwr_60min = df_mwr_60min.set_index(['BJT', 'Level (hPa)'])
df_mwr_60min = df_mwr_60min.sort_index(level=[0,1])  # 先按时间排序，再按高度排序

temp_1min_reindexed_interpolated_final = temp_1min_reindexed_interpolated_final.reset_index()
temp_1min_reindexed_interpolated_final['timestamps'] = temp_1min_reindexed_interpolated_final['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
rh_1min_reindexed_interpolated_final = rh_1min_reindexed_interpolated_final.reset_index()
rh_1min_reindexed_interpolated_final['timestamps'] = rh_1min_reindexed_interpolated_final['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
ah_1min_reindexed_interpolated_final['timestamps'] = ah_1min_reindexed_interpolated_final['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
q_1min = q_1min.reset_index()
q_1min = q_1min.rename(columns={'index':  'timestamps'})
q_1min['timestamps'] = q_1min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')
θv_1min = θv_1min.reset_index()
θv_1min = θv_1min.rename(columns={'index':  'timestamps'})
θv_1min['timestamps'] = θv_1min['timestamps'].dt.strftime('%Y/%m/%d %H:%M:%S')

#turb_calibrated_1s.to_csv(output_path + 'turb_1s_corrected.csv',index=False)

# with pd.ExcelWriter(output_path + 'MWR Raw data-1min.xlsx', engine='openpyxl', date_format='yyyy/mm/dd hh:mm:ss') as writer:

#     temp_1min_reindexed_interpolated_final.to_excel(writer, sheet_name='T_1min', index=False)
    
#     rh_1min_reindexed_interpolated_final.to_excel(writer, sheet_name='rh_1min', index=False)
    
#     ah_1min_reindexed_interpolated_final.to_excel(writer, sheet_name='ah_1min', index=False)
    
#     q_1min.to_excel(writer, sheet_name='q_1min', index=False)
    
#     θv_1min.to_excel(writer, sheet_name='θv_1min', index=False)

#     for sheet_name in writer.sheets:
#         worksheet = writer.sheets[sheet_name]
#             # 设置A到F列
#         for col in ['A']:
#             worksheet.column_dimensions[col].width = 20




###MWR插值到再分析###
target_heights = np.arange(275, 1001, 25)
var_series = df_mwr_60min[['T(K)','RH(%)','q(kg/kg)']]

def interpolate_group_multi(group, target_heights):
    # 提取当前时间组的高度值
    heights = group.index.get_level_values('Level (hPa)').values
    
    # 初始化结果容器
    interpolated_data = {}
    
    # 遍历每个变量进行插值
    for col in group.columns:
        values = group[col].values
        
        # 数据对齐性检查
        if len(heights) != len(values):
            print(f"变量 {col} 数据长度不匹配: heights={len(heights)}, values={len(values)}")
            interpolated_data[col] = np.full(len(target_heights), np.nan)
            continue
        
        # 处理多维数组（如降维）
        if values.ndim > 1:
            values = values.squeeze()
        
        # 创建插值函数（线性插值）
        try:
            f = interp1d(heights, values, kind='linear', fill_value='extrapolate')
            interpolated_values = f(target_heights)
        except Exception as e:
            print(f"变量 {col} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan)
        
        interpolated_data[col] = interpolated_values
    
    # 转换为DataFrame并设置目标高度为索引
    return pd.DataFrame(interpolated_data, index=target_heights)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('BJT').apply(
    interpolate_group_multi, 
    target_heights=target_heights
)
# 展开多层索引
interpolated_df = interpolated_df.unstack(level=1).stack(level=0,future_stack=True)

# 重置索引并命名列
interpolated_df = interpolated_df.reset_index()

# 转换核心步骤
long_df = (
    interpolated_df.melt(
        id_vars=['BJT', 'level_1'],
        var_name='pressure_level',
        value_name='value'
    )
    .query("level_1 in ['T(K)', 'RH(%)', 'q(kg/kg)']")  # 过滤有效参数
    .assign(pressure_level=lambda x: x.pressure_level.astype(int))  # 转换压力层级为整型
    .pivot_table(
        index=['BJT', 'pressure_level'],
        columns='level_1',
        values='value'
    )
    .reset_index()
    .rename_axis(None, axis=1)  # 清理列名
    .sort_values('BJT')
)
sorted_df = long_df.sort_values(
    by=['BJT', 'pressure_level'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)

sorted_df = sorted_df.rename(columns={'pressure_level':  'Level (hPa)'})

# 设置双重索引
sorted_df = sorted_df.set_index(['BJT', 'Level (hPa)'])
mwr_interp_to_reanalysis = sorted_df



####MWR插值到铁塔#####
target_heights = tieta_1min['H(m)'].unique()
filtered_df_mwr = df_mwr.query("height <= 341")

starttime = pd.to_datetime('2024-06-05')
endtime = filtered_df_mwr.reset_index()['timestamps'].unique().max()
mwr_tobe_interp = filtered_df_mwr.loc[pd.IndexSlice[starttime:endtime, :], :]
    
starttime = pd.to_datetime('2024/06/05')
endtime = pd.to_datetime(filtered_df_mwr.reset_index()['timestamps'].unique().max(),format='%Y/%m/%d %H:%M:%S')
new_index = pd.to_datetime(tieta_1min.index.levels[0])
tieta_1min.index = tieta_1min.index.set_levels(new_index, level=0)
tieta_tobe_interp = tieta_1min.loc[pd.IndexSlice[starttime:endtime, :], :]
tieta_tobe_interp = tieta_tobe_interp.reset_index()
var_series = mwr_tobe_interp[['T(K)', 'RH(%)']]

def interpolate_group_multi(group, target_heights, mwr_tobe_interp):
    current_time = group.name
    try:
        # 直接通过索引获取对应时间的高度数据
        heights_df = mwr_tobe_interp.loc[[current_time]].sort_values('Level (hPa)')
    except KeyError:
        print(f"时间 {current_time}: 在tieta_tobe_interp中未找到对应高度数据")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    heights = heights_df['height'].values
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    # 插值逻辑（保持不变）
    interpolated_data = {}
    for col in group.columns:
        values = group[col].values
        if np.isnan(values).all():
            interpolated_data[col] = np.full(len(target_heights), np.nan)
            continue
        try:
            f = interp1d(heights, values, kind='linear', bounds_error=False, fill_value='extrapolate')
            interpolated_values = f(target_heights)
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan)
        interpolated_data[col] = interpolated_values
    
    return pd.DataFrame(interpolated_data, index=target_heights).infer_objects(copy=False)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('BJT').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    mwr_tobe_interp=mwr_tobe_interp
)
# 展开多层索引
interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'height_tieta'})

sorted_df = interpolated_df.sort_values(
    by=['BJT', 'height_tieta'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
# 设置双重索引
sorted_df = sorted_df.set_index(['BJT', 'height_tieta'])
mwr_interp_to_tieta = sorted_df



####MWR插值到超声#####
target_heights = turb_calibrated_1min.reset_index()['Height'].unique()
#turb_calibrated_1min = turb_calibrated_1min.set_index(['timestamps','Height'])
filtered_df_mwr = df_mwr.query("height <= 280")

starttime = pd.to_datetime('2024-08-04')
endtime = filtered_df_mwr.reset_index()['timestamps'].unique().max()
mwr_tobe_interp = filtered_df_mwr.loc[pd.IndexSlice[starttime:endtime, :], :]

starttime = pd.to_datetime('2024/08/04')
endtime = pd.to_datetime(filtered_df_mwr.reset_index()['timestamps'].unique().max(),format='%Y/%m/%d %H:%M:%S')

# new_index = pd.to_datetime(turb_calibrated_1min.index.levels[0])
# turb_calibrated_1min.index = turb_calibrated_1min.index.set_levels(new_index, level=0)
# ultra_tobe_interp = turb_calibrated_1min.loc[pd.IndexSlice[starttime:endtime, :], :]
#tieta_tobe_interp = tieta_tobe_interp.reset_index()
var_series = mwr_tobe_interp[['T(K)', 'RH(%)','q(kg/kg)']]

def interpolate_group_multi(group, target_heights, mwr_tobe_interp):
    current_time = group.name
    try:
        # 直接通过索引获取对应时间的高度数据
        heights_df = mwr_tobe_interp.loc[[current_time]].sort_values('Level (hPa)')
    except KeyError:
        print(f"时间 {current_time}: 在tieta_tobe_interp中未找到对应高度数据")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    heights = heights_df['height'].values
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    # 插值逻辑（保持不变）
    interpolated_data = {}
    for col in group.columns:
        values = group[col].values
        if np.isnan(values).all():
            interpolated_data[col] = np.full(len(target_heights), np.nan)
            continue
        try:
            f = interp1d(heights, values, kind='linear', bounds_error=False, fill_value='extrapolate')
            interpolated_values = f(target_heights)
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan)
        interpolated_data[col] = interpolated_values
    
    return pd.DataFrame(interpolated_data, index=target_heights).infer_objects(copy=False)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('BJT').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    mwr_tobe_interp=mwr_tobe_interp
)
# 展开多层索引

interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'height_ultra'})

sorted_df = interpolated_df.sort_values(
    by=['BJT', 'height_ultra'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
# 设置双重索引
sorted_df = sorted_df.set_index(['BJT', 'height_ultra'])
mwr_interp_to_ultra = sorted_df


mwr_interp_to_ultra = mwr_interp_to_ultra.reset_index()
mwr_interp_to_ultra['BJT'] = pd.to_datetime(mwr_interp_to_ultra['BJT'])

mwr_interp_to_ultra['BJT'] = mwr_interp_to_ultra['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
mwr_interp_to_ultra = mwr_interp_to_ultra.sort_values(
    by=['height_ultra','BJT'],  # 按时间→气压层级双排序
    ascending=[False,True]                # 默认升序（时间从早到晚，层级从低到高）
)

mwr_interp_to_tieta = mwr_interp_to_tieta.reset_index()
mwr_interp_to_tieta['BJT'] = pd.to_datetime(mwr_interp_to_tieta['BJT'])

mwr_interp_to_tieta['BJT'] = mwr_interp_to_tieta['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
mwr_interp_to_tieta = mwr_interp_to_tieta.sort_values(
    by=['height_tieta','BJT'],  # 按时间→气压层级双排序
    ascending=[False,True]                # 默认升序（时间从早到晚，层级从低到高）
)

mwr_interp_to_reanalysis = mwr_interp_to_reanalysis.reset_index()
mwr_interp_to_reanalysis['BJT'] = pd.to_datetime(mwr_interp_to_reanalysis['BJT'])
mwr_interp_to_reanalysis['BJT'] = mwr_interp_to_reanalysis['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
mwr_interp_to_reanalysis = mwr_interp_to_reanalysis.sort_values(
    by=['Level (hPa)','BJT'],  # 按时间→气压层级双排序
    ascending=[False,True]                # 默认升序（时间从早到晚，层级从低到高）
)


# with pd.ExcelWriter(output_path + 'MWR Interpolation Results.xlsx', engine='openpyxl', date_format='yyyy/mm/dd hh:mm:ss') as writer:

#     mwr_interp_to_reanalysis.to_excel(writer, sheet_name='MWR插值再分析', index=False)
    
#     mwr_interp_to_tieta.to_excel(writer, sheet_name='MWR插值铁塔', index=False)
    
#     mwr_interp_to_ultra.to_excel(writer, sheet_name='MWR插值超声', index=False)

# # 调整列宽（openpyxl语法）
#     for sheet_name in writer.sheets:
#         worksheet = writer.sheets[sheet_name]
#         worksheet.column_dimensions['A'].width = 20  # 设置A列宽度为20
#         worksheet.column_dimensions['B'].width = 20
#         worksheet.column_dimensions['F'].width = 20


###################################################处理CMA-RA数据集######################################
# ===================== 配置区 =====================
PARENT_DIR = r"E:\Beijing2024\CMA-RA\Reanalysis"
OUTPUT_DIR = r"E:\Beijing2024\CMA-RA"
TARGET_LON = 116.3705
TARGET_LAT = 39.9745
EXCEL_NAME = f"CRA_Interp_tieta_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

# ==================== 核心处理逻辑 ====================
data = {
    '常规变量': {},
    '特殊变量': {}
}

variable_dirs = [d for d in os.listdir(PARENT_DIR) 
                if os.path.isdir(os.path.join(PARENT_DIR, d))]

for variable_dir in variable_dirs:
    input_dir = os.path.join(PARENT_DIR, variable_dir)
    
    # 精确筛选.grib2文件（排除.grib2.xxx文件）
    files = sorted([
        f for f in glob(os.path.join(input_dir, "*"))
        if os.path.isfile(f) and os.path.splitext(f)[1].lower() == ".grib2"
    ])
    
    if not files:
        print(f"⚠️ 无有效GRIB2文件跳过目录: {variable_dir}")
        continue

    try:
        # 禁用索引文件并忽略警告
        os.environ["CFGRIB_DISABLE_INDEX"] = "True"
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            
            # 获取网格信息
            with xr.open_dataset(files[0], engine='cfgrib',decode_timedelta=True,
                               backend_kwargs={'filter_by_keys': {'typeOfLevel': 'isobaricInhPa'},
                                               'indexpath': False}
                               ) as ds:
                var_name = list(ds.data_vars)[0]
                print(f"✔️ 正在处理变量: {variable_dir} | 识别到变量名: {var_name}")
                
                unit = ds[var_name].attrs.get('units', '')
                
                # 处理坐标
                lons = np.unique(ds.longitude.values)
                lats = np.unique(ds.latitude.values)
                lons.sort()
                lats.sort()

                # 边界检查
                if not (lons[0] <= TARGET_LON <= lons[-1] and lats[0] <= TARGET_LAT <= lats[-1]):
                    print(f"🚫 坐标越界跳过: {variable_dir}")
                    continue

                # 计算网格索引
                i_lon = max(0, np.searchsorted(lons, TARGET_LON, side='right')-1)
                i_lat = np.searchsorted(lats, TARGET_LAT, side='left')

    except Exception as e:
        print(f"❌ 初始化失败: {variable_dir} | {str(e)}")
        continue

    # 确定缩放因子
    SCALE_FACTOR = 1e2 if variable_dir in [
        'Cloud Ice', 'Specific Humidity', 
        'Absolute Vorticity', 'Cloud Mixing Ratio'
    ] else 1
    print(f"🔄 应用缩放因子: {SCALE_FACTOR} 于变量: {var_name}")

    # 单位处理
    if var_name == 'r':  # 相对湿度特殊处理
        scaled_unit = "%"
        col_name = f"rh ({scaled_unit})"
        #SCALE_FACTOR = 100  # 假设原始数据是小数形式（0-1）
    elif var_name == 't':
        scaled_unit = "K"
        col_name = f"Temp ({scaled_unit})"
    else:
        scaled_unit = f"10⁻2 {unit}" if SCALE_FACTOR == 1e2 else unit
        col_name = f"{var_name} ({scaled_unit})" if unit else var_name


    # 处理所有文件
    for file in files:
        try:
            filename = os.path.basename(file)
            time_str = filename[27:37]  # 根据实际文件名调整
            dt = datetime.strptime(time_str, "%Y%m%d%H")
            
            with xr.open_dataset(file, engine='cfgrib',decode_timedelta=True,
                               backend_kwargs={'filter_by_keys': {'typeOfLevel': 'isobaricInhPa'}}) as ds:
                for level in ds.isobaricInhPa.values:
                    grid_data = ds[var_name].sel(isobaricInhPa=level).values
                    
                    # 应用缩放因子
                    if SCALE_FACTOR != 1:
                        grid_data = grid_data * SCALE_FACTOR

                    # 执行插值
                    interpolator = RegularGridInterpolator(
                        (lats, lons), grid_data, bounds_error=False)
                    value = interpolator([[TARGET_LAT, TARGET_LON]])[0]
                    #value = np.nan_to_num(value,  nan=0)
                    
                    # 确定工作表分组
                    sheet_group = '特殊变量' if variable_dir in [
                        #'Vertical velocity',   # 对应w变量
                        'Cloud Ice',           # 对应cice变量 
                        'Cloud Mixing Ratio'   # 对应clwmr变量
                        ] else '常规变量'

                    timestamp = dt.strftime("%Y-%m-%d %H:00")
                    level_key = (timestamp, int(level))
                    
                    # 更新数据结构
                    if level_key not in data[sheet_group]:
                        data[sheet_group][level_key] = {
                            'timestamps': timestamp,
                            'Level (hPa)': int(level)
                        }
                    data[sheet_group][level_key][col_name] = value

        except Exception as e:
            print(f"❌ 处理失败: {filename} | {str(e)}")
            continue

    # 打印网格信息
    print(f"\n🔍 完成目录: {variable_dir}")
    print(f"网格坐标索引: 经度[{i_lon}:{i_lon+1}] 纬度[{i_lat-1}:{i_lat}]")
    print("实际坐标点:")
    print(f"  NW: {lats[i_lat]:.4f}°N, {lons[i_lon]:.4f}°E")
    print(f"  NE: {lats[i_lat]:.4f}°N, {lons[i_lon+1]:.4f}°E")
    print(f"  SW: {lats[i_lat-1]:.4f}°N, {lons[i_lon]:.4f}°E")
    print(f"  SE: {lats[i_lat-1]:.4f}°N, {lons[i_lon+1]:.4f}°E\n")

# # 生成Excel文件
# with pd.ExcelWriter(os.path.join(output_path, EXCEL_NAME), engine='openpyxl', date_format='yyyy/mm/dd hh:mm:ss') as writer:
#     for sheet_name in ['常规变量', '特殊变量']:
#         if not data[sheet_name]:
#             continue
            
#         df = pd.DataFrame(list(data[sheet_name].values())).fillna(0)
#         df = df.sort_values(by=['timestamps', 'Level (hPa)'], 
#                           ascending=[True, False])
        
#         if sheet_name == '常规变量':  
            
#             #df['height(m)'] = pd.DataFrame(metpy.calc.geopotential_to_height(units.Quantity(df['gh (gpm)'].to_list(), 'm^2/s^2')).m)
#             df['vws(10⁻² {m/s})'] = 100*pd.DataFrame((metpy.calc.vertical_velocity(df['w (Pa s**-1)'].to_list() * units('Pa/s'), 
#                                                                         df['Level (hPa)'].to_list() * units.hPa, 
#                                                                         df['Temp (K)'].to_list() * units.K,
#                                                                         (df['q (10⁻2 kg kg**-1)']*10**-2).to_list() * units('kg/kg'))).m)
#             df['hws(m/s)'] = np.sqrt(np.square(df['u (m s**-1)']) + np.square(df['v (m s**-1)']))
#             df['hwd(deg)'] = np.mod(180+ np.degrees(np.arctan2(-df['u (m s**-1)'],-df['u (m s**-1)'])),360)

#             omega = df[['timestamps','Level (hPa)','w (Pa s**-1)']]
#             omega = omega.set_index(['timestamps','Level (hPa)'])
#             df = df.drop(columns=['w (Pa s**-1)'])
#             df['timestamps'] = pd.to_datetime(df['timestamps'])
#             df['BJT'] = df['timestamps'] + pd.Timedelta(hours=8)
#             df['BJT'] = df['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
            
#             cols = ['timestamps','BJT', 'Level (hPa)'] + sorted(
#                 [c for c in df.columns if c not in ['timestamps','BJT', 'Level (hPa)']],
#                 key=lambda x: x.lower()
#             )
#             df = df[cols]
#             sheet1 = df
#         else:
            
#             df = df.set_index(['timestamps','Level (hPa)'])            
#             df = pd.concat([df,  omega], axis=1)
#             df = df.reset_index()
#             df['w (Pa s**-1)'] = df['w (Pa s**-1)'] *100
#             df.rename(columns={'w (Pa s**-1)': 'omega(10⁻² {Pa/s})'}, inplace=True)   
#             df['timestamps'] = pd.to_datetime(df['timestamps'])
#             df['BJT'] = df['timestamps'] + pd.Timedelta(hours=8)
#             df['BJT'] = df['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
#             cols = ['timestamps','BJT', 'Level (hPa)'] + sorted(
#                 [c for c in df.columns if c not in ['timestamps','BJT', 'Level (hPa)']],
#                 key=lambda x: x.lower()
#             )
#             df = df[cols]
            
#             sheet2 = df
#         # 智能列排序
#         cols = ['timestamps', 'Level (hPa)'] + sorted(
#             [c for c in df.columns if c not in ['timestamps', 'Level (hPa)']],
#             key=lambda x: x.replace('rh  (%)', 'A').lower()
#         )
#         df = df[cols]
        
#         df.to_excel(writer, sheet_name=sheet_name, index=False)
#         ws = writer.sheets[sheet_name]
#         ws.freeze_panes = 'A2'  # 核心代码
#         # 设置样式
#         #header_font = Font(color="FFFFFF", bold=True)
#         #header_fill = PatternFill("solid", fgColor="4F81BD")
        
#         # # 自适应列宽
#         # # ==== 修改后的列宽设置 ====
#         # for col_idx, col_name in enumerate(df.columns, 1):
#         #     col_letter = get_column_letter(col_idx)
            
#         #     if col_name == 'Level (hPa)':
#         #         ws.column_dimensions[col_letter].width = 10  # 固定宽度
#         #     else:
#         #         # 计算该列最大内容长度（跳过标题）
#         #         max_length = max(
#         #             len(str(cell)) for cell in df[col_name].astype(str)
#         #         )
#         #         adjusted_width = (max_length + 2) * 1.2
#         #         ws.column_dimensions[col_letter].width = adjusted_width
            
#         # for cell in ws[1]:
#         #     cell.font = header_font
#         #     cell.fill = header_fill

# print(f"\n✅ 处理完成! 结果文件已保存至:\n{os.path.join(OUTPUT_DIR, EXCEL_NAME)}")

##############CRAv1.5插值到47和280米对应1000/975hPa##############

# 初始化结果变量,生成两个表存储处理结果
cra_regular = None
cra_special = None
omega = None

for sheet_name in ['常规变量', '特殊变量']:
    if not data.get(sheet_name):
        continue  # 如果数据为空则跳过
        
    # 创建DataFrame并填充缺失值
    df = pd.DataFrame(list(data[sheet_name].values())).fillna(0)
    # 按时间戳和层级排序
    df = df.sort_values(by=['timestamps', 'Level (hPa)'], ascending=[True, False])
    
    if sheet_name == '常规变量':
        # --------------- 物理量计算 ---------------
        # # 位势高度转几何高度
        # df['height(m)'] = pd.DataFrame(
        #     metpy.calc.geopotential_to_height(
        #         units.Quantity(df['gh (gpm)'].to_list(), 'm^2/s^2')
        #     ).m
        # )
        
        # 垂直速度转换 (需要温度、气压、比湿数据)
        df['vws(10⁻² {m/s})'] = 100 * pd.DataFrame(
            metpy.calc.vertical_velocity(
                df['w (Pa s**-1)'].to_list() * units('Pa/s'),
                df['Level (hPa)'].to_list() * units.hPa,
                df['Temp (K)'].to_list() * units.K,
                (df['q (10⁻2 kg kg**-1)'] * 1e-2).to_list() * units('kg/kg')
            ).m
        )
        
        # 水平风速和风向
        df['hws(m/s)'] = np.sqrt(df['u (m s**-1)']**2 + df['v (m s**-1)']**2)
        df['hwd(deg)'] = np.mod(180+ np.degrees(np.arctan2(-df['u (m s**-1)'],-df['u (m s**-1)'])),360)
        
        # 保存omega变量供特殊变量表使用
        omega = df[['timestamps', 'Level (hPa)', 'w (Pa s**-1)']]
        omega = omega.set_index(['timestamps', 'Level (hPa)'])
        df = df.drop(columns=['w (Pa s**-1)'])
        
        # 智能列排序
        cols = ['timestamps', 'Level (hPa)'] + sorted(
            [c for c in df.columns if c not in ['timestamps', 'Level (hPa)']],
            key=lambda x: x.replace('rh  (%)', 'A').lower()  # 保持rh列靠前
        )
        cra_regular = df[cols]
        
    else:  # 处理特殊变量
        # 合并omega变量
        df = df.set_index(['timestamps', 'Level (hPa)'])
        df = pd.concat([df, omega], axis=1).reset_index()
        
        # 单位转换和重命名
        df['w (Pa s**-1)'] *= 100
        df.rename(columns={'w (Pa s**-1)': 'omega(10⁻² {Pa/s})'}, inplace=True)
        
        # 列排序（逻辑同常规变量）
        cols = ['timestamps', 'Level (hPa)'] + sorted(
            [c for c in df.columns if c not in ['timestamps', 'Level (hPa)']],
            key=lambda x: x.replace('rh  (%)', 'A').lower()
        )
        cra_special = df[cols]

target_heights = np.array([65,80,103,120,140,160,180,200,240,280,320])
cra_regular_tobe_interp = cra_regular.copy()
cra_regular_tobe_interp = cra_regular_tobe_interp.rename(columns={'Level (hPa)':  'pressure_level'})
cra_regular_tobe_interp = cra_regular_tobe_interp.query("pressure_level >= 950")
cra_regular_tobe_interp = cra_regular_tobe_interp.set_index(['timestamps','pressure_level'])
cra_regular_tobe_interp = cra_regular_tobe_interp[['Temp (K)','q (10⁻2 kg kg**-1)','rh (%)','u (m s**-1)','v (m s**-1)', 'vws(10⁻² {m/s})']]
cra_regular_tobe_interp = cra_regular_tobe_interp.reset_index()
cra_regular_tobe_interp['pressure_level'] =pd.to_numeric(cra_regular_tobe_interp['pressure_level'], errors='coerce', downcast='float')
cra_regular_tobe_interp['height_plus_alt'] = 1000*pd.DataFrame(metpy.calc.pressure_to_height_std(cra_regular_tobe_interp['pressure_level'].to_list() * units('hPa')).m) - Altitude



starttime = pd.to_datetime('2024-06-05')
endtime = cra_regular_tobe_interp.reset_index()['timestamps'].unique().max()
var_series = cra_regular_tobe_interp.reset_index()
var_series = var_series.sort_values(
    by=['timestamps', 'height_plus_alt'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
def interpolate_group_multi(group, target_heights, cra_tobe_interp):
    current_time = group.name
    try:
        # 直接通过索引获取对应时间的高度数据
        heights_df = cra_tobe_interp[cra_tobe_interp['timestamps'] == current_time].sort_values('height_plus_alt') 
    except KeyError:
        print(f"时间 {current_time}: 在era5_tobe_interp中未找到对应高度数据")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    heights = heights_df['height_plus_alt'].values
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    # 插值逻辑（保持不变）
    interpolated_data = {}
    for col in group.columns:
        values = group[col].values
        if np.isnan(values).all():
            interpolated_data[col] = np.full(len(target_heights), np.nan)
            continue
        try:
            f = interp1d(heights, values, kind='linear', bounds_error=False, fill_value='extrapolate')
            interpolated_values = f(target_heights)
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan)
        interpolated_data[col] = interpolated_values
    
    return pd.DataFrame(interpolated_data, index=target_heights).infer_objects(copy=False)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('timestamps').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    cra_tobe_interp=var_series,
    include_groups=False
)

interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'cra_tobe_interp'})
interpolated_df['hws(m/s)'] = np.sqrt(np.square(interpolated_df['u (m s**-1)']) + np.square(interpolated_df['v (m s**-1)']))
interpolated_df['hwd(deg)'] = np.degrees(np.arctan2(-interpolated_df['u (m s**-1)'], -interpolated_df['v (m s**-1)'])) % 360

sorted_df = interpolated_df.sort_values(
    by=['timestamps', 'height_plus_alt'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)

# 设置双重索引
sorted_df = sorted_df.set_index(['timestamps', 'height_plus_alt'])
CRA_interp_to_65_80_280 = sorted_df[['Temp (K)','q (10⁻2 kg kg**-1)','rh (%)','hws(m/s)','hwd(deg)', 'vws(10⁻² {m/s})']]
CRA_interp_to_65_80_280 = CRA_interp_to_65_80_280.reset_index()
CRA_interp_to_65_80_280['timestamps'] =  pd.to_datetime(CRA_interp_to_65_80_280['timestamps'])
CRA_interp_to_65_80_280['BJT'] = CRA_interp_to_65_80_280['timestamps'] + pd.Timedelta(hours=8)
CRA_interp_to_65_80_280['BJT'] =  pd.to_datetime(CRA_interp_to_65_80_280['BJT'])
CRA_interp_to_65_80_280['BJT'] = CRA_interp_to_65_80_280['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')


target_heights = temp_1min_reindexed_interpolated_final.set_index("timestamps").columns
cra_regular_tobe_interp = cra_regular.copy()
cra_regular_tobe_interp = cra_regular_tobe_interp.rename(columns={'Level (hPa)':  'pressure_level'})
cra_regular_tobe_interp = cra_regular_tobe_interp.query("pressure_level >= 250")
cra_regular_tobe_interp = cra_regular_tobe_interp.set_index(['timestamps','pressure_level'])
cra_regular_tobe_interp = cra_regular_tobe_interp[['Temp (K)','q (10⁻2 kg kg**-1)','rh (%)','u (m s**-1)','v (m s**-1)', 'vws(10⁻² {m/s})']]
cra_regular_tobe_interp = cra_regular_tobe_interp.reset_index()
cra_regular_tobe_interp['pressure_level'] =pd.to_numeric(cra_regular_tobe_interp['pressure_level'], errors='coerce', downcast='float')
cra_regular_tobe_interp['height_plus_alt'] = 1000*pd.DataFrame(metpy.calc.pressure_to_height_std(cra_regular_tobe_interp['pressure_level'].to_list() * units('hPa')).m) - Altitude

starttime = pd.to_datetime('2024-08-05')
endtime = cra_regular_tobe_interp.reset_index()['timestamps'].unique().max()
var_series = cra_regular_tobe_interp
var_series = var_series.sort_values(
    by=['timestamps', 'height_plus_alt'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
def interpolate_group_multi(group, target_heights, cra_tobe_interp):
    current_time = group.name
    try:
        # 添加明确的类型转换
        heights_df = cra_tobe_interp[cra_tobe_interp['timestamps'] == current_time].sort_values('height_plus_alt')
        heights_df = heights_df.astype({'height_plus_alt': np.float64})  # 确保数值类型
    except KeyError:
        print(f"时间 {current_time}: 在era5_tobe_interp中未找到对应高度数据")
        # 明确指定数据类型为float
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    heights = heights_df['height_plus_alt'].values.astype(np.float64)  # 强制类型转换
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    interpolated_data = {}
    for col in group.columns:
        # 确保输入数据为浮点类型
        values = group[col].astype(np.float64).values
        if np.isnan(values).all():
            # 创建数组时指定dtype
            interpolated_data[col] = np.full(len(target_heights), np.nan, dtype=np.float64)
            continue
        try:
            f = interp1d(heights, values, kind='linear', bounds_error=False, fill_value=np.nan)
            interpolated_values = f(target_heights.astype(np.float64))  # 确保目标高度为数值类型
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan, dtype=np.float64)
        interpolated_data[col] = interpolated_values
    
    # 创建DataFrame时强制类型为float
    return pd.DataFrame(interpolated_data, index=target_heights.astype(np.float64)).astype(np.float64)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('timestamps').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    cra_tobe_interp=var_series,
    include_groups=False
)

interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'cra_tobe_interp'})
interpolated_df['hws(m/s)'] = np.sqrt(np.square(interpolated_df['u (m s**-1)']) + np.square(interpolated_df['v (m s**-1)']))
interpolated_df['hwd(deg)'] = np.degrees(np.arctan2(-interpolated_df['u (m s**-1)'], -interpolated_df['v (m s**-1)'])) % 360

sorted_df = interpolated_df.sort_values(
    by=['timestamps', 'height_plus_alt'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)

# 设置双重索引
sorted_df = sorted_df.set_index(['timestamps', 'height_plus_alt'])
CRA_interp_to_MWR = sorted_df[['pressure_level', 'Temp (K)', 'q (10⁻2 kg kg**-1)',
       'rh (%)', 'u (m s**-1)', 'v (m s**-1)', 'vws(10⁻² {m/s})', 'hws(m/s)',
       'hwd(deg)']]
CRA_interp_to_MWR = CRA_interp_to_MWR.reset_index()
CRA_interp_to_MWR['timestamps'] =  pd.to_datetime(CRA_interp_to_MWR['timestamps'])
CRA_interp_to_MWR['BJT'] = CRA_interp_to_MWR['timestamps'] + pd.Timedelta(hours=8)
CRA_interp_to_MWR['BJT'] =  pd.to_datetime(CRA_interp_to_MWR['BJT'])
CRA_interp_to_MWR['BJT'] = CRA_interp_to_MWR['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')


target_heights = np.arange(50, 2776, 25)
cra_regular_tobe_interp = cra_regular.copy()
cra_regular_tobe_interp = cra_regular_tobe_interp.rename(columns={'Level (hPa)':  'pressure_level'})
cra_regular_tobe_interp = cra_regular_tobe_interp.query("pressure_level >= 250")
cra_regular_tobe_interp = cra_regular_tobe_interp.set_index(['timestamps','pressure_level'])
cra_regular_tobe_interp = cra_regular_tobe_interp[['Temp (K)','q (10⁻2 kg kg**-1)','rh (%)','u (m s**-1)','v (m s**-1)', 'vws(10⁻² {m/s})']]
cra_regular_tobe_interp = cra_regular_tobe_interp.reset_index()
cra_regular_tobe_interp['pressure_level'] =pd.to_numeric(cra_regular_tobe_interp['pressure_level'], errors='coerce', downcast='float')
cra_regular_tobe_interp['height_plus_alt'] = 1000*pd.DataFrame(metpy.calc.pressure_to_height_std(cra_regular_tobe_interp['pressure_level'].to_list() * units('hPa')).m) - Altitude

starttime = pd.to_datetime('2024-08-05')
endtime = cra_regular_tobe_interp.reset_index()['timestamps'].unique().max()
var_series = cra_regular_tobe_interp
var_series = var_series.sort_values(
    by=['timestamps', 'height_plus_alt'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
def interpolate_group_multi(group, target_heights, cra_tobe_interp):
    current_time = group.name
    try:
        # 添加明确的类型转换
        heights_df = cra_tobe_interp[cra_tobe_interp['timestamps'] == current_time].sort_values('height_plus_alt')
        heights_df = heights_df.astype({'height_plus_alt': np.float64})  # 确保数值类型
    except KeyError:
        print(f"时间 {current_time}: 在era5_tobe_interp中未找到对应高度数据")
        # 明确指定数据类型为float
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    heights = heights_df['height_plus_alt'].values.astype(np.float64)  # 强制类型转换
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    interpolated_data = {}
    for col in group.columns:
        # 确保输入数据为浮点类型
        values = group[col].astype(np.float64).values
        if np.isnan(values).all():
            # 创建数组时指定dtype
            interpolated_data[col] = np.full(len(target_heights), np.nan, dtype=np.float64)
            continue
        try:
            f = interp1d(heights, values, kind='linear', bounds_error=False, fill_value=np.nan)
            interpolated_values = f(target_heights.astype(np.float64))  # 确保目标高度为数值类型
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan, dtype=np.float64)
        interpolated_data[col] = interpolated_values
    
    # 创建DataFrame时强制类型为float
    return pd.DataFrame(interpolated_data, index=target_heights.astype(np.float64)).astype(np.float64)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('timestamps').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    cra_tobe_interp=var_series,
    include_groups=False
)

interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'cra_tobe_interp'})
interpolated_df['hws(m/s)'] = np.sqrt(np.square(interpolated_df['u (m s**-1)']) + np.square(interpolated_df['v (m s**-1)']))
interpolated_df['hwd(deg)'] = np.degrees(np.arctan2(-interpolated_df['u (m s**-1)'], -interpolated_df['v (m s**-1)'])) % 360

sorted_df = interpolated_df.sort_values(
    by=['timestamps', 'height_plus_alt'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)

# 设置双重索引
sorted_df = sorted_df.set_index(['timestamps', 'height_plus_alt'])
CRA_interp_to_WL = sorted_df[['pressure_level', 'Temp (K)', 'q (10⁻2 kg kg**-1)',
       'rh (%)', 'u (m s**-1)', 'v (m s**-1)', 'vws(10⁻² {m/s})', 'hws(m/s)',
       'hwd(deg)']]
CRA_interp_to_WL = CRA_interp_to_WL.reset_index()
CRA_interp_to_WL['timestamps'] =  pd.to_datetime(CRA_interp_to_WL['timestamps'])
CRA_interp_to_WL['BJT'] = CRA_interp_to_WL['timestamps'] + pd.Timedelta(hours=8)
CRA_interp_to_WL['BJT'] =  pd.to_datetime(CRA_interp_to_WL['BJT'])
CRA_interp_to_WL['BJT'] = CRA_interp_to_WL['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
# with pd.ExcelWriter(output_path + 'CRA插到铁塔、MWR和WL.xlsx', engine='openpyxl', date_format='yyyy/mm/dd hh:mm:ss') as writer:

#     CRA_interp_to_65_80_280.to_excel(writer, sheet_name='CRA_to_65_80_280', index=False)
#     # for sheet_name in writer.sheets:
#     #     worksheet = writer.sheets[sheet_name]
#     #         # 设置A到F列
#     #     for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
#     #         worksheet.column_dimensions[col].width = 20
    
#     CRA_interp_to_WL.dropna().to_excel(writer, sheet_name='CRA_to_WL', index=False)
#     CRA_interp_to_MWR.dropna().to_excel(writer, sheet_name='CRA_to_MWR', index=False)
#     #CRA_interp_to_ultra.dropna().to_excel(writer, sheet_name='CRA_to_ultra', index=False)
#     for sheet_name in writer.sheets:
#         worksheet = writer.sheets[sheet_name]
#             # 设置A到F列
#         for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
#             worksheet.column_dimensions[col].width = 20

Altitude = 49
# ###############################ERA5提取并插值新版#######################################
# # 1. 加载数据并筛选变量
# ds = xr.open_dataset(r'E:\Beijing2024\ERA5\0804-0811 hourly pressure level.nc')
# target_vars = ['q', 'u', 'v', 'w', 't', 'r']
# variables = [var for var in ds.data_vars if var in target_vars]

# # 2. 目标坐标
# target_lon = 116.3705
# target_lat = 39.9745

# # 3. 获取原始网格信息
# raw_lats = ds.latitude.values  # ERA5纬度通常为降序
# raw_lons = ds.longitude.values

# # 4. 自动定位包围网格点 ------------------------------------------------------
# # 寻找最近的4个格点索引
# lat_idx = np.searchsorted(raw_lats[::-1], target_lat)  # 处理降序纬度
# lon_idx = np.searchsorted(raw_lons, target_lon)

# # 确定包围格点的索引
# if lat_idx == 0 or lat_idx == len(raw_lats):
#     raise ValueError("目标纬度超出范围")
# if lon_idx == 0 or lon_idx == len(raw_lons):
#     raise ValueError("目标经度超出范围")

# # 获取实际参与插值的四个格点索引（原始数据索引）
# surrounding_points = [
#     (lat_idx-1, lon_idx-1),  # 西南
#     (lat_idx-1, lon_idx),    # 东南
#     (lat_idx, lon_idx-1),    # 西北
#     (lat_idx, lon_idx)       # 东北
# ]

# # 转换为原始纬度索引（因纬度降序排列的特殊处理）
# adjusted_points = [
#     (len(raw_lats)-1 - i, j) for i, j in surrounding_points
# ]

# # 5. 输出网格点信息 ----------------------------------------------------------
# print("插值使用的网格点信息：")
# print(f"目标坐标：{target_lat}°N, {target_lon}°E")
# print("包围点原始索引（lat_index, lon_index）及坐标：")

# for i, (lat_i, lon_i) in enumerate(adjusted_points, 1):
#     lat_val = raw_lats[lat_i].item()
#     lon_val = raw_lons[lon_i].item()
#     print(f"点{i}：({lat_i}, {lon_i}) -> {lat_val}°N, {lon_val}°E")

# # 6. 数据预处理 --------------------------------------------------------------
# # 反转纬度数据以适应RegularGridInterpolator要求
# lats = raw_lats[::-1]  # 现在变为升序
# lons = raw_lons

# # 创建时空索引
# times = ds.valid_time.values
# levels = ds.pressure_level.values
# index = pd.MultiIndex.from_product([times, levels], 
#                                   names=["valid_time", "pressure_level"])

# # 初始化结果DataFrame
# final_df = pd.DataFrame({
#     "valid_time": index.get_level_values("valid_time"),
#     "pressure_level": index.get_level_values("pressure_level")
# })

# # 7. 批量插值核心逻辑 --------------------------------------------------------
# for var in variables:
#     # 转置维度并反转纬度顺序
#     data = ds[var].transpose("latitude", "longitude", "valid_time", "pressure_level").values[::-1]
    
#     # 创建插值器
#     interpolator = RegularGridInterpolator(
#         (lats, lons),
#         data,
#         method="linear",
#         bounds_error=False,
#         fill_value=np.nan
#     )
    
#     # 执行插值
#     result = interpolator([[target_lat, target_lon]]).squeeze(0)
#     final_df[var] = result.ravel()

# final_df['vws(10⁻² {m/s})'] = pd.DataFrame((metpy.calc.vertical_velocity(final_df['w'].to_list() * units('Pa/s'), 
#                                                             final_df['pressure_level'].to_list() * units.hPa, 
#                                                             final_df['t'].to_list() * units.K,mixing_ratio = 0
#                                                             )).m)#(final_df['q']).to_list() * units('kg/kg')
# final_df['hws(m/s)'] = np.sqrt(np.square(final_df['u']) + np.square(final_df['v']))
# final_df['hwd(deg)'] = np.mod(180+ np.degrees(np.arctan2(final_df['u'],final_df['v'])),360)
# final_df['height'] = 1000*pd.DataFrame(metpy.calc.pressure_to_height_std(final_df['pressure_level'].to_list() * units('hPa')).m) - Altitude
# #对应指定点的above ground level
# # 8. 输出验证 ---------------------------------------------------------------
# print("\n插值结果示例：")
# print(final_df.head())
# print("\n数据结构验证：")
# print(f"总行数：{len(final_df)} (时间点{len(times)} × 气压层{len(levels)})")
# print(f"保留变量：{variables}")
# ERA5_interp_BJ = final_df
# ERA5_interp_BJ['valid_time'] =  pd.to_datetime(ERA5_interp_BJ['valid_time'])
# ERA5_interp_BJ['BJT'] = ERA5_interp_BJ['valid_time'] + pd.Timedelta(hours=8)
# ERA5_interp_BJ['BJT'] =  pd.to_datetime(ERA5_interp_BJ['BJT'])
# ERA5_interp_BJ['BJT'] = ERA5_interp_BJ['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')

# cols = ['BJT', 'valid_time','pressure_level'] + sorted(
#     [c for c in ERA5_interp_BJ.columns if c not in ['BJT','valid_time', 'pressure_level']],
#     key=lambda x: x.lower()
# )
# ERA5_interp_BJ = ERA5_interp_BJ[cols]


import os
import xarray as xr
import numpy as np
import pandas as pd
from scipy.interpolate import RegularGridInterpolator
import metpy.calc as mpcalc
from metpy.units import units

# ================== 基本设置 ==================
data_dir = r'E:\Beijing2024\ERA5'

# 只遍历当前目录下的 .nc 文件，不进入子文件夹
files = [
    f for f in os.listdir(data_dir)
    if os.path.isfile(os.path.join(data_dir, f)) and f.endswith('.nc')
]

target_lon = 116.3705
target_lat = 39.9745

target_vars = ['q', 'u', 'v', 'w', 't', 'r']  # 按你原来设定
Altitude = 49.0  # 举例：地面高度（m），你按真实高度改

all_dfs = []  # 存储每个文件的结果，最后再 concat

for file in files:
    file_path = os.path.join(data_dir, file)
    print(f"正在处理文件：{file_path}")

    # 1. 打开数据集
    ds = xr.open_dataset(file_path)
    variables = [var for var in ds.data_vars if var in target_vars]

    # 2. 原始经纬度
    raw_lats = ds.latitude.values  # 通常为降序
    raw_lons = ds.longitude.values

    # 3. 定位包围目标点的网格索引
    lat_idx = np.searchsorted(raw_lats[::-1], target_lat)
    lon_idx = np.searchsorted(raw_lons, target_lon)

    if lat_idx == 0 or lat_idx == len(raw_lats):
        raise ValueError("目标纬度超出范围")
    if lon_idx == 0 or lon_idx == len(raw_lons):
        raise ValueError("目标经度超出范围")

    surrounding_points = [
        (lat_idx-1, lon_idx-1),
        (lat_idx-1, lon_idx),
        (lat_idx,   lon_idx-1),
        (lat_idx,   lon_idx),
    ]
    adjusted_points = [
        (len(raw_lats)-1 - i, j) for i, j in surrounding_points
    ]

    # 4. 为插值准备“升序纬度”
    lats = raw_lats[::-1]
    lons = raw_lons

    # 5. 时间与层次
    times = ds.valid_time.values
    levels = ds.pressure_level.values

    index = pd.MultiIndex.from_product(
        [times, levels],
        names=['valid_time', 'pressure_level']
    )

    # ⚠️ 关键：每个文件先建一个骨架 DataFrame
    file_df = pd.DataFrame({
        'valid_time': index.get_level_values('valid_time'),
        'pressure_level': index.get_level_values('pressure_level')
    })

    # 6. 变量循环：在 file_df 中一列一列填数据
    for var in variables:
        print(f"  插值变量：{var}")

        # 维度调整：lat, lon, time, level 再反转纬度
        data = ds[var].transpose(
            'latitude', 'longitude', 'valid_time', 'pressure_level'
        ).values[::-1]

        # 设置插值器
        interpolator = RegularGridInterpolator(
            (lats, lons),
            data,
            method='linear',
            bounds_error=False,
            fill_value=np.nan
        )

        # 对单点插值，结果形状：(time, level)
        result = interpolator([[target_lat, target_lon]]).squeeze(0)

        # 展平成一维，与 index 对齐
        file_df[var] = result.ravel()

    # 7. 这个文件的结果暂存
    all_dfs.append(file_df)

# 8. 所有文件拼接
final_df = pd.concat(all_dfs, ignore_index=True)

# 9. 计算派生量（要求已有 u, v, w, t, pressure_level）
if {'w', 't', 'pressure_level'}.issubset(final_df.columns):
    final_df['vws(10⁻² {m/s})'] = (
        mpcalc.vertical_velocity(
            final_df['w'].to_numpy() * units('Pa/s'),
            final_df['pressure_level'].to_numpy() * units.hPa,
            final_df['t'].to_numpy() * units.K,
            mixing_ratio=0 * units('kg/kg')
        ).m
    )

if {'u', 'v'}.issubset(final_df.columns):
    final_df['hws(m/s)'] = np.sqrt(
        np.square(final_df['u']) + np.square(final_df['v'])
    )
    final_df['hwd(deg)'] = np.mod(180+ np.degrees(np.arctan2(final_df['u'],final_df['v'])),360)

# 标准大气高度（相对海平面），再减去站点海拔得到 AGL
final_df['height'] = (
    1000.0
    * mpcalc.pressure_to_height_std(
        final_df['pressure_level'].to_numpy() * units.hPa
    ).m
    - Altitude
)

# 10. 时间处理与列排序
final_df['valid_time'] = pd.to_datetime(final_df['valid_time'])
final_df['BJT'] = (final_df['valid_time'] + pd.Timedelta(hours=8)).dt.strftime(
    '%Y/%m/%d %H:%M:%S'
)

cols = ['BJT', 'valid_time', 'pressure_level'] + sorted(
    [c for c in final_df.columns if c not in ['BJT', 'valid_time', 'pressure_level']],
    key=lambda x: x.lower()
)
final_df = final_df[cols]

print("\n插值结果示例：")
print(final_df.head())
print(f"\n总行数：{len(final_df)}")
print(f"列：{final_df.columns.tolist()}")

# with pd.ExcelWriter(output_path + 'ERA5_interp_BJ.xlsx', engine='openpyxl', date_format='yyyy/mm/dd hh:mm:ss') as writer:

#     ERA5_interp_BJ.to_excel(writer, sheet_name='ERA5_interp_BJ', index=False)

#     for sheet_name in writer.sheets:
#         worksheet = writer.sheets[sheet_name]
#             # 设置A到F列
#         for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I','L']:
#             worksheet.column_dimensions[col].width = 20


##############ERA5插值到65和280米对应1000/975hPa##############
filtered_final_df = final_df.query("pressure_level >= 950")
filtered_final_df = filtered_final_df.rename(columns={'valid_time':  'timestamps'})
filtered_final_df = filtered_final_df.set_index(['timestamps','height'])

target_heights = np.array([65,80,103,120,140,160,180,200,240,280,320])#65m and 280m
starttime = pd.to_datetime('2024-06-05')
endtime = filtered_final_df.reset_index()['timestamps'].unique().max()
era5_tobe_interp = filtered_final_df[['t','r','u','v','q','vws(10⁻² {m/s})']]
var_series = era5_tobe_interp.reset_index()

def interpolate_group_multi(group, target_heights, era5_tobe_interp):
    current_time = group.name
    try:
        # 直接通过索引获取对应时间的高度数据
        heights_df = era5_tobe_interp[era5_tobe_interp['timestamps'] == current_time].sort_values('height') 
    except KeyError:
        print(f"时间 {current_time}: 在era5_tobe_interp中未找到对应高度数据")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    heights = heights_df['height'].values
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns)
    
    # 插值逻辑（保持不变）
    interpolated_data = {}
    for col in group.columns:
        values = group[col].values
        if np.isnan(values).all():
            interpolated_data[col] = np.full(len(target_heights), np.nan)
            continue
        try:
            f = interp1d(heights, values, kind='linear', bounds_error=False, fill_value='extrapolate')
            interpolated_values = f(target_heights)
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan)
        interpolated_data[col] = interpolated_values
    
    return pd.DataFrame(interpolated_data, index=target_heights).infer_objects(copy=False)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('timestamps').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    era5_tobe_interp=var_series,
    include_groups=False
)
# 展开多层索引

interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'era5_tobe_interp'})
#interpolated_df['height'] = interpolated_df['height']
interpolated_df['hws(m/s)'] = np.sqrt(np.square(interpolated_df['u']) + np.square(interpolated_df['v']))
interpolated_df['hwd(deg)'] = np.degrees(np.arctan2(-interpolated_df['u'], -interpolated_df['v'])) % 360

sorted_df = interpolated_df.sort_values(
    by=['timestamps', 'height'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
# 设置双重索引
#sorted_df = sorted_df.set_index(['timestamps', 'height'])
ERA5_interp_to_65_80_280 = sorted_df#.reset_index()[['timestamps','tieta_tobe_interp','height','t', 'r','q','hws(m/s)', 'hwd(deg)','vws(10⁻² {m/s})']]
ERA5_interp_to_65_80_280.loc[:, 'q'] = ERA5_interp_to_65_80_280['q'] * 100
ERA5_interp_to_65_80_280 = ERA5_interp_to_65_80_280.rename(columns={'q':  'q*10^-2'})
ERA5_interp_to_65_80_280['timestamps'] = pd.to_datetime(ERA5_interp_to_65_80_280['timestamps'])
ERA5_interp_to_65_80_280['BJT'] = ERA5_interp_to_65_80_280['timestamps'] + pd.Timedelta(hours=8)
ERA5_interp_to_65_80_280['BJT'] = ERA5_interp_to_65_80_280['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
cols = ['BJT', 'timestamps','era5_tobe_interp','height'] + sorted(
    [c for c in ERA5_interp_to_65_80_280.columns if c not in ['BJT', 'timestamps','era5_tobe_interp','height']],
    key=lambda x: x.lower()
)
ERA5_interp_to_65_80_280 = ERA5_interp_to_65_80_280[cols]



##############ERA5插值到MWR##############

filtered_final_df = final_df.query("pressure_level >= 250")
filtered_final_df = filtered_final_df.rename(columns={'valid_time':  'timestamps'})
filtered_final_df = filtered_final_df.set_index(['timestamps','height'])

target_heights = temp_1min_reindexed_interpolated_final.set_index("timestamps").columns
starttime = pd.to_datetime('2024-08-05')
endtime = filtered_final_df.reset_index()['timestamps'].unique().max()
era5_tobe_interp = filtered_final_df[['t','r','u','v','q','vws(10⁻² {m/s})']]
var_series = era5_tobe_interp.reset_index()

def interpolate_group_multi(group, target_heights, era5_tobe_interp):
    current_time = group.name
    
    # 强制转换目标高度为浮点型
    target_heights = np.asarray(target_heights, dtype=np.float64)
    
    try:
        # 确保height列是数值型
        heights_df = era5_tobe_interp[era5_tobe_interp['timestamps'] == current_time].copy()
        heights_df['height'] = pd.to_numeric(heights_df['height'], errors='coerce')  # 安全转换
        heights_df = heights_df.sort_values('height').dropna(subset=['height'])
    except KeyError as e:
        print(f"时间 {current_time}: 关键列缺失 - {str(e)}")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 有效高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    # 确保高度数据是浮点数组
    heights = heights_df['height'].values.astype(np.float64)
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    interpolated_data = {}
    for col in group.columns:
        # 安全数据类型转换
        try:
            values = pd.to_numeric(group[col], errors='coerce').values.astype(np.float64)
        except Exception as e:
            print(f"列 {col} 包含非数值数据: {str(e)}")
            values = np.full(len(group), np.nan, dtype=np.float64)
        
        if np.isnan(values).all():
            interpolated_data[col] = np.full(len(target_heights), np.nan, dtype=np.float64)
            continue
            
        try:
            # 验证插值输入
            valid_mask = ~np.isnan(heights) & ~np.isnan(values)
            f = interp1d(heights[valid_mask], 
                         values[valid_mask], 
                         kind='linear', 
                         bounds_error=False, 
                         fill_value=np.nan)
            interpolated_values = f(target_heights)
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan, dtype=np.float64)
        
        interpolated_data[col] = interpolated_values
    
    # 强制输出类型为float64
    return pd.DataFrame(interpolated_data, index=target_heights).astype(np.float64)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('timestamps').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    era5_tobe_interp=var_series,
    include_groups=False
)
# 展开多层索引

interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'era5_tobe_interp'})
interpolated_df['height'] = interpolated_df['height'] - Altitude
interpolated_df['hws(m/s)'] = np.sqrt(np.square(interpolated_df['u']) + np.square(interpolated_df['v']))
interpolated_df['hwd(deg)'] = np.degrees(np.arctan2(-interpolated_df['u'], -interpolated_df['v'])) % 360

sorted_df = interpolated_df.sort_values(
    by=['timestamps', 'height'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
# 设置双重索引
#sorted_df = sorted_df.set_index(['timestamps', 'height'])
ERA5_interp_to_MWR = sorted_df#.reset_index()[['timestamps','tieta_tobe_interp','height','t', 'r','q','hws(m/s)', 'hwd(deg)','vws(10⁻² {m/s})']]
ERA5_interp_to_MWR.loc[:, 'q'] = ERA5_interp_to_MWR['q'] * 100
ERA5_interp_to_MWR = ERA5_interp_to_MWR.rename(columns={'q':  'q*10^-2'})
ERA5_interp_to_MWR['timestamps'] = pd.to_datetime(ERA5_interp_to_MWR['timestamps'])
ERA5_interp_to_MWR['BJT'] = ERA5_interp_to_MWR['timestamps'] + pd.Timedelta(hours=8)
ERA5_interp_to_MWR['BJT'] = ERA5_interp_to_MWR['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
cols = ['BJT', 'timestamps','era5_tobe_interp','height'] + sorted(
    [c for c in ERA5_interp_to_MWR.columns if c not in ['BJT', 'timestamps','era5_tobe_interp','height']],
    key=lambda x: x.lower()
)
ERA5_interp_to_MWR = ERA5_interp_to_MWR[cols]



##############ERA5插值到WL#############
filtered_final_df = final_df.query("pressure_level >= 250")
filtered_final_df = filtered_final_df.rename(columns={'valid_time':  'timestamps'})
filtered_final_df = filtered_final_df.set_index(['timestamps','height'])
target_heights = np.arange(50, 2776, 25)
starttime = pd.to_datetime('2024-08-05')
endtime = filtered_final_df.reset_index()['timestamps'].unique().max()
era5_tobe_interp = filtered_final_df[['t','r','u','v','q','vws(10⁻² {m/s})']]
var_series = era5_tobe_interp.reset_index()

def interpolate_group_multi(group, target_heights, era5_tobe_interp):
    current_time = group.name
    
    # 强制转换目标高度为浮点型
    target_heights = np.asarray(target_heights, dtype=np.float64)
    
    try:
        # 确保height列是数值型
        heights_df = era5_tobe_interp[era5_tobe_interp['timestamps'] == current_time].copy()
        heights_df['height'] = pd.to_numeric(heights_df['height'], errors='coerce')  # 安全转换
        heights_df = heights_df.sort_values('height').dropna(subset=['height'])
    except KeyError as e:
        print(f"时间 {current_time}: 关键列缺失 - {str(e)}")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    if len(heights_df) == 0:
        print(f"时间 {current_time}: 有效高度数据为空")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    # 确保高度数据是浮点数组
    heights = heights_df['height'].values.astype(np.float64)
    if len(heights) != len(group):
        print(f"时间 {current_time}: 高度数据长度({len(heights)})与变量数据({len(group)})不匹配")
        return pd.DataFrame(np.nan, index=target_heights, columns=group.columns).astype(np.float64)
    
    interpolated_data = {}
    for col in group.columns:
        # 安全数据类型转换
        try:
            values = pd.to_numeric(group[col], errors='coerce').values.astype(np.float64)
        except Exception as e:
            print(f"列 {col} 包含非数值数据: {str(e)}")
            values = np.full(len(group), np.nan, dtype=np.float64)
        
        if np.isnan(values).all():
            interpolated_data[col] = np.full(len(target_heights), np.nan, dtype=np.float64)
            continue
            
        try:
            # 验证插值输入
            valid_mask = ~np.isnan(heights) & ~np.isnan(values)
            f = interp1d(heights[valid_mask], 
                         values[valid_mask], 
                         kind='linear', 
                         bounds_error=False, 
                         fill_value=np.nan)
            interpolated_values = f(target_heights)
        except Exception as e:
            print(f"变量 {col} 在时间 {current_time} 插值失败: {str(e)}")
            interpolated_values = np.full(len(target_heights), np.nan, dtype=np.float64)
        
        interpolated_data[col] = interpolated_values
    
    # 强制输出类型为float64
    return pd.DataFrame(interpolated_data, index=target_heights).astype(np.float64)
# 按时间分组并应用插值
interpolated_df = var_series.groupby('timestamps').apply(
    interpolate_group_multi, 
    target_heights=target_heights,
    era5_tobe_interp=var_series,
    include_groups=False
)
# 展开多层索引

interpolated_df = interpolated_df.reset_index().rename(columns={'level_1':  'era5_tobe_interp'})
interpolated_df['height'] = interpolated_df['height'] - Altitude
interpolated_df['hws(m/s)'] = np.sqrt(np.square(interpolated_df['u']) + np.square(interpolated_df['v']))
interpolated_df['hwd(deg)'] = np.degrees(np.arctan2(-interpolated_df['u'], -interpolated_df['v'])) % 360

sorted_df = interpolated_df.sort_values(
    by=['timestamps', 'height'],  # 按时间→气压层级双排序
    ascending=[True, True]                # 默认升序（时间从早到晚，层级从低到高）
)
# 设置双重索引
#sorted_df = sorted_df.set_index(['timestamps', 'height'])
ERA5_interp_to_WL = sorted_df#.reset_index()[['timestamps','tieta_tobe_interp','height','t', 'r','q','hws(m/s)', 'hwd(deg)','vws(10⁻² {m/s})']]
ERA5_interp_to_WL.loc[:, 'q'] = ERA5_interp_to_WL['q'] * 100
ERA5_interp_to_WL = ERA5_interp_to_WL.rename(columns={'q':  'q*10^-2'})
ERA5_interp_to_WL['timestamps'] = pd.to_datetime(ERA5_interp_to_WL['timestamps'])
ERA5_interp_to_WL['BJT'] = ERA5_interp_to_WL['timestamps'] + pd.Timedelta(hours=8)
ERA5_interp_to_WL['BJT'] = ERA5_interp_to_WL['BJT'].dt.strftime('%Y/%m/%d %H:%M:%S')
cols = ['BJT', 'timestamps','era5_tobe_interp','height'] + sorted(
    [c for c in ERA5_interp_to_WL.columns if c not in ['BJT', 'timestamps','era5_tobe_interp','height']],
    key=lambda x: x.lower()
)
ERA5_interp_to_WL = ERA5_interp_to_WL[cols]

# with pd.ExcelWriter(output_path + 'ERA5插到铁塔、MWR和WL.xlsx', engine='openpyxl', date_format='yyyy/mm/dd hh:mm:ss') as writer:

#     ERA5_interp_to_65_80_280.to_excel(writer, sheet_name='ERA5_interp_to_65_80_280', index=False)

#     for sheet_name in writer.sheets:
#         worksheet = writer.sheets[sheet_name]
#             # 设置A到F列
#         for col in ['A','B','C','L']:
#             worksheet.column_dimensions[col].width = 20
            
#     CRA_interp_to_WL.dropna().to_excel(writer, sheet_name='ERA5_to_WL', index=False)
#     CRA_interp_to_MWR.dropna().to_excel(writer, sheet_name='ERA5_to_MWR', index=False)
#     for sheet_name in writer.sheets:
#         worksheet = writer.sheets[sheet_name]
#             # 设置A到F列
#         for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
#             worksheet.column_dimensions[col].width = 20            

######################################################################################################################################################


# hws_cra_wide = CRA_interp_to_WL.pivot(
#     index='timestamps',       # 保持时间戳为行索引
#     columns='height_plus_alt',    # 将气压层级展开为列
#     values='hws(m/s)'        # 指定需要转换的数值列
# ).dropna(axis=1)

# vws_cra_wide = CRA_interp_to_WL.pivot(
#     index='timestamps',       # 保持时间戳为行索引
#     columns='height_plus_alt',    # 将气压层级展开为列
#     values='vws(10⁻² {m/s})'        # 指定需要转换的数值列
# ).dropna(axis=1)

# hws_era5_wide = ERA5_interp_to_WL.pivot(
#     index='timestamps',       # 保持时间戳为行索引
#     columns='era5_tobe_interp',    # 将气压层级展开为列
#     values='hws(m/s)'        # 指定需要转换的数值列
# ).dropna(axis=1)

# vws_era5_wide = ERA5_interp_to_WL.pivot(
#     index='timestamps',       # 保持时间戳为行索引
#     columns='era5_tobe_interp',    # 将气压层级展开为列
#     values='vws(10⁻² {m/s})'        # 指定需要转换的数值列
# ).dropna(axis=1)

    
# hws_wl_wide = wind_60min_hws_joined.reset_index().query("height % 25 == 0").pivot(
#     index='BJT',       # 保持时间戳为行索引
#     columns='height',    # 将气压层级展开为列
#     values='hws'        # 指定需要转换的数值列
# )
    
# vws_wl_wide = wind_60min_vws_joined.reset_index().query("height % 25 == 0").pivot(
#     index='BJT',       # 保持时间戳为行索引
#     columns='height',    # 将气压层级展开为列
#     values='vws'        # 指定需要转换的数值列
# )
    

# wl_hws_minus_cra = hws_wl_wide.sub(hws_cra_wide)
# wl_vws_minus_cra = (vws_wl_wide).sub(-vws_cra_wide/100)#再分析资料与风雷达的符号相反

# wl_hws_minus_era5 = hws_wl_wide.sub(hws_era5_wide)
# wl_vws_minus_era5 = (vws_wl_wide).sub(-vws_era5_wide/100)

# with pd.ExcelWriter(output_path + '风雷达减再分析差值.xlsx', engine='openpyxl', date_format='yyyy/mm/dd hh:mm:ss') as writer:
    
#     wl_hws_minus_cra.reset_index().to_excel(writer, sheet_name='WL-CRA_hws', index=False)
    
#     wl_vws_minus_cra.reset_index().to_excel(writer, sheet_name='WL-CRA_vws', index=False)
    
#     wl_hws_minus_era5.reset_index().to_excel(writer, sheet_name='WL-ERA5_hws', index=False)
    
#     wl_vws_minus_era5.reset_index().to_excel(writer, sheet_name='WL-ERA5_vws', index=False)
    
#     writer.book.create_sheet("我是分隔符")
    
#     hws_wl_wide.reset_index().to_excel(writer, sheet_name='hws_wl_wide', index=False)
    
#     vws_wl_wide.reset_index().to_excel(writer, sheet_name='vws_wl_wide', index=False)
    
#     hws_cra_wide.reset_index().to_excel(writer, sheet_name='hws_cra_wide', index=False)
    
#     vws_cra_wide.reset_index().to_excel(writer, sheet_name='vws_cra_wide', index=False)
    
#     hws_era5_wide.reset_index().to_excel(writer, sheet_name='hws_era5_wide', index=False)
    
#     vws_era5_wide.reset_index().to_excel(writer, sheet_name='vws_era5_wide', index=False)
    
#     for sheet_name in writer.sheets:
#         worksheet = writer.sheets[sheet_name]
#             # 设置A到F列
#         for col in ['A']:
#             worksheet.column_dimensions[col].width = 20
    
######################################################################################################################################################

# cra_temp = cra_regular.pivot(
#     index='timestamps',       # 保持时间戳为行索引
#     columns='Level (hPa)',    # 将气压层级展开为列
#     values='Temp (K)'        # 指定需要转换的数值列
# )

# cra_hws = cra_regular.pivot(
#     index='timestamps',       # 保持时间戳为行索引
#     columns='Level (hPa)',    # 将气压层级展开为列
#     values='hws(m/s)'        # 指定需要转换的数值列
# )
# cra_hws = cra_regular.pivot(
#     index='timestamps',       # 保持时间戳为行索引
#     columns='Level (hPa)',    # 将气压层级展开为列
#     values='vws(10⁻² {m/s})'        # 指定需要转换的数值列
# )

#################################################下面是画单个图的代码

# # === 读取探空数据 (Excel) ===
# # 假设原始 Excel 文件名为 "original combined2.xlsx"，目标工作表名为 '2024080508'
# snd = pd.read_excel("E:\\Beijing2024\\探空数据-54511\\original combined2.xlsx", sheet_name="2024080508")  # :contentReference[oaicite:3]{index=3}

# # === 过滤其他数据源在2024-08-05 08:00的剖面 ===
# target_time = pd.to_datetime('2024-08-05 08:00')
# cra_regular['timestamps'] = pd.to_datetime(cra_regular['timestamps'])
# # 假定 ERA5_interp_BJ、cra_regular、df_mwr_60min 已经以 DataFrame 形式加载
# era5_profile = ERA5_interp_BJ[ERA5_interp_BJ['valid_time'] == target_time].copy()
# cra_profile = cra_regular[cra_regular['timestamps'] == target_time].copy()
# mwr_profile = df_mwr_60min.reset_index().loc[df_mwr_60min.reset_index()['BJT'] == target_time].copy()

# # === 单位转换与计算高度 ===
# # ERA5 与 CRA 数据：压力 (hPa) -> 高度 (m)
# era5_profile['height'] = mpcalc.pressure_to_height_std(era5_profile['pressure_level'].to_list() * units('hPa')).to('m').magnitude  
# cra_profile['height'] = mpcalc.pressure_to_height_std(cra_profile['Level (hPa)'].to_list() * units.hPa).to('m').magnitude  

# # 探空数据：温度摄氏度转为开尔文，并将压力 -> 高度
# snd['temperature_K'] = snd['temperature_C'] + 273.15
# snd['height'] = mpcalc.pressure_to_height_std(snd['pressure_hPa'].to_list() * units.hPa).to('m').magnitude

# # MWR 数据：假设已含高度 (m) 和温度 (K)
# # （无需转换高度，只需排序）
# mwr_profile = mwr_profile.sort_values('height')

# # 对各剖面按高度排序（便于绘图）
# era5_profile = era5_profile.sort_values('height')
# cra_profile = cra_profile.sort_values('height')[0:37]
# snd = snd.sort_values('height')

# # === 绘图 ===
# plt.rcParams.update({'font.size': 30,
#                      'axes.linewidth':3,
#                      'axes.labelsize': 30,
#                      'xtick.labelsize': 30,
#                      'ytick.labelsize': 30})

# fig, ax = plt.subplots(figsize=(10, 13))

# # 绘制温度-高度曲线，各数据源不同颜色线条
# ax.plot(era5_profile['t'], era5_profile['height'], color='blue',   linewidth=3)
# ax.plot(cra_profile['Temp (K)'], cra_profile['height'],     color='red',    linewidth=3)
# ax.plot(mwr_profile['T(K)'],    mwr_profile['height'],     color='magenta',linewidth=3)
# ax.plot(snd['temperature_K'],   snd['height'],              color='black',  linewidth=3)

# # 设置左侧坐标轴：高度 (红色)，并启用左右双侧刻度线

# ax.spines['left'].set_color('red') 
# ax.spines['right'].set_color('blue') 
# ax.spines['bottom'].set_color('black') 
# ax.spines['top'].set_color('black') 

# ax.set_xlabel('Temperature(K)')
# ax.set_ylabel('Height(m a.g.l.)',color='red')

# ax.tick_params(axis='y', colors='red', length = 12,width=2)

# # 添加右侧辅助坐标轴：气压 (蓝色)，使用MetPy转换函数&#8203;:contentReference[oaicite:7]{index=7}
# secax = ax.secondary_yaxis('right',
#     functions=(lambda h: mpcalc.height_to_pressure_std(h * units.m).to('hPa').magnitude,
#                lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude))
# secax.set_ylabel('Pressure(hPa)', color='blue',rotation = 270, labelpad = 20)
# secax.yaxis.label.set_color('blue')
# secax.tick_params(axis='y', colors='blue',length=12,width=2)

# # 绘制水平虚线：示例标示 2000 m (红) 和对应 500 hPa (蓝)
# site_elev = 49
# alt_ticks = np.array([0, 500, 2000, 
#                       3000, 5000, 8000, 10000, 12500, 15000])        # AGL
# p_ticks   = np.array([1000, 925, 850,
#                       700, 500, 275])                  # hPa
# p2alt = lambda p: mpcalc.pressure_to_height_std(
#         p*units.hPa).to('m').magnitude

# for y in alt_ticks:
#     ax.axhline(y, color='red', lw=2, zorder=0, linestyle=(0,(5,10))) 
# for p in p_ticks:
#     ax.axhline(p2alt(p), color='blue', lw=2, zorder=0, linestyle=(0,(5,10)))
# ax.set_yticks(alt_ticks)  # 设置左侧主刻度
# secax.set_yticks(p_ticks)   # 设置右侧主刻度
# secax.set_ylim(1000, 275)   # 气压轴范围倒置（符合大气科学惯例）
# ax.set_ylim(0, 10000) 
# ax.xaxis.set_major_locator(MultipleLocator(10))
# ax.xaxis.set_minor_locator(MultipleLocator(5))
# ax.tick_params(axis='x',which='minor',length=6,width=2,color='black')
# plt.tick_params(axis='x', length=12, width=2)
# ax.set_xlim(240, 310)
# plt.title(f"{target_time}", loc='center', pad=20, y=1.01,fontsize = 25)
# plt.tight_layout()
# plt.show()

# -*- coding: utf-8 -*-

############铁塔数据散点对比####################
# ────────────────────────────── 基本路径与文件 ──────────────────────────────
#ultra_path  = r"E:\Beijing2024\数据处理结果2nd\ultrasonic_1_5_60.xlsx"
#tieta_path  = r"E:\Beijing2024\数据处理结果2nd\tieta0805-0810_1-5-60.xlsx"
#output_dir  = r"E:\Beijing2024\出图TIF"
#os.makedirs(output_dir, exist_ok=True)

# ───────────────────────────★ 新增：时间窗口参数 ───────────────────────────
TIME_START1 = pd.Timestamp("2024-08-09 12:46")   # ← 根据需要修改
TIME_END1   = pd.Timestamp("2024-08-10 00:03")
# 第二个时间窗口（新增）
TIME_START2 = pd.Timestamp("2024-08-10 16:20")   # ← 新增时间窗口开始
TIME_END2   = pd.Timestamp("2024-08-10 17:30")   # ← 新增时间窗口结束

# 第二个时间窗口（新增）
TIME_START3 = pd.Timestamp("2024-06-07 09:15")   # ← 新增时间窗口开始
TIME_END3   = pd.Timestamp("2024-06-07 15:00")   # ← 新增时间窗口结束

# ───────────────────────────── 读取 Excel ─────────────────────────────
# 超声（取前26列）
ultra = turb_calibrated_1min   #pd.read_excel(ultra_path,  sheet_name="turb_1min_corrected", usecols=range(26))
# 铁塔（取前8列）
tieta = tieta_1min.reset_index()   #pd.read_excel(tieta_path, sheet_name="tieta0805-0810_1min",     usecols=range(8))
# ---- 确保时间列是 datetime 类型 ----
ultra["timestamps"] = pd.to_datetime(ultra["timestamps"], errors="coerce")
tieta["timestamps"] = pd.to_datetime(tieta["timestamps"], errors="coerce")

# ───────────────────── 筛选目标高度层 & 预处理 ─────────────────────
heights = [280, 200, 140, 80]
ultra = ultra[ultra["Height"].isin(heights)]
ultra["rh"] = ultra["rh"].clip(upper=100)           # RH 上限 100 %
tieta = tieta[tieta["H(m)"].isin(heights)]

# ─────────────────────────── 合并（时间 & 高度） ───────────────────────────
merged = (
    pd.merge(
        ultra[["timestamps", "Height", "rh", "hws_corrected", "T (K)", "wd_corrected"]],
        tieta[["timestamps", "H(m)",  "RH_tieta", "ws_tieta", "T(K)_tieta", "wd_tieta"]],
        left_on=["timestamps", "Height"],
        right_on=["timestamps", "H(m)"],
        suffixes=("_ultra", "_tieta")
    )
    .dropna()
)

# ───────────────────────────── 变量映射 ─────────────────────────────
variables = {
    "temp": ("T (K)",       "T(K)_tieta"),
    "hws" : ("hws_corrected","ws_tieta"),
    "rh"  : ("rh",          "RH_tieta"),
}

# ─────────────────── 预计算统计量 & 保存时间戳 ★新增 ───────────────────
results = {}
for h in heights:
    subset = merged[merged["Height"] == h]
    results[h] = {}
    for var, (ultra_col, tieta_col) in variables.items():
        x = subset[ultra_col]
        y = subset[tieta_col]

        # 基本统计
        pearson_r, _ = stats.pearsonr(x, y)
        X = sm.add_constant(x)
        model = sm.OLS(y, X).fit()
        r2 = model.rsquared
        adj_r2 = 1 - (1 - r2) * (len(x) - 1) / (len(x) - 2)

        # 置信区间
        pred = model.get_prediction(X)
        ci = pred.conf_int(alpha=0.05)

        results[h][var] = {
            "x": x,
            "y": y,
            "t": subset["timestamps"],   # ★新增：保存对应时间戳
            "pearson_r": pearson_r,
            "r2": r2,
            "adj_r2": adj_r2,
            "ci": ci,
            "model": model,              # 回归模型备用
        }
# ────────────────────────── 画图配置 ──────────────────────────
fig = plt.figure(figsize=(30, 40), dpi=100)
gs  = GridSpec(nrows=4, ncols=6, figure=fig,
               width_ratios=[1, 0.3]*3, height_ratios=[1]*4)
plt.rcParams.update(plt.rcParamsDefault)

color_map = {"temp": "#ff0000", "hws": "#0000ff", "rh": "#00aa00"}
palette   = {"X": "#4C72B0", "Y": "#DD8452"}
sns.set_theme(style="whitegrid", font_scale=1.2)
# ★时间窗口布尔掩码：三个时间段一起高亮


# ───────────────────────────── 循环绘制 ─────────────────────────────
for row, height in enumerate(heights):
    for col, var in enumerate(variables.keys()):
        data          = results[height][var]
        x, y, t       = data["x"], data["y"], data["t"]
        current_color = color_map[var]

        # 两个时间窗口合并为一个掩码
        mask_1 = (t >= TIME_START1) & (t <= TIME_END1)
        mask_2 = (t >= TIME_START2) & (t <= TIME_END2)
        mask_3 = (t >= TIME_START3) & (t <= TIME_END3)

        mask_time  = mask_1 | mask_2 | mask_3   # 三个时间窗口合并
        mask_other = ~mask_time                 # 其他时间点

        # ───── 散点 & 回归线 ─────
        ax_scatter = fig.add_subplot(gs[row, 2*col])

        # 回归线（不绘制默认散点）- 交换x和y
        sns.regplot(
            x=y, y=x, ax=ax_scatter, scatter=False,
            line_kws={"color": "#949494", "lw": 4},
            ci=95,scatter_kws={"alpha":0},
            
        )

        # # 所有点（主数据） - 圆点 - 交换x和y
        ax_scatter.scatter(
            y[~mask_time], x[~mask_time],
            marker="o", color=current_color, alpha=1, s=25, edgecolors="none"
        )
        # ★第一个时间窗口点 - 三角标记 - 交换x和y
        ax_scatter.scatter(
            y[mask_time], x[mask_time],
            marker="o", color='orange', alpha=1,
            s=50, edgecolors="none", linewidths=0.5, 
        )

        # 1:1 参考虚线 - 交换x和y
        ax_scatter.plot([y.min(), y.max()], [y.min(), y.max()], "k--", lw=4)

        # 统计文字
        ax_scatter.text(
            0.025, 0.90,
            f"Pearson R = {data['pearson_r']:.2f}\nAdjusted R² = {data['adj_r2']:.2f}",
            transform=ax_scatter.transAxes, fontsize=32, fontname="Arial",
            bbox={'facecolor': 'none', 'edgecolor': 'none'}
        )

        # 统一坐标轴、刻度 & 样式（保持原逻辑，略）
        ax_scatter.xaxis.set_major_locator(MultipleLocator(5))
        ax_scatter.xaxis.set_minor_locator(AutoMinorLocator(2))
        ax_scatter.yaxis.set_minor_locator(AutoMinorLocator(2))
        ax_scatter.tick_params(axis='both', labelsize=36, direction='out',
                               length=20, width=3, color='black',
                               bottom=True, left=True)
        ax_scatter.tick_params(which='minor', axis='both',
                               length=10, width=3, direction='out',
                               bottom=True, left=True)
        ax_scatter.grid(False)
        ax_scatter.spines[['top', 'right']].set_visible(False)
        ax_scatter.spines[['left', 'bottom']].set_linewidth(3)
        ax_scatter.spines[['left', 'bottom']].set_color('black')

        # 变量专属坐标范围 - 交换x和y
        if var == 'temp':
            ax_scatter.set_xlim(287, 310); ax_scatter.set_ylim(287, 310)
            ax_scatter.set_xticks(np.arange(290, 311, 5))
            ax_scatter.set_yticks(np.arange(290, 311, 5))
        elif var == 'hws':
            ax_scatter.set_xlim(-0.2, 16); ax_scatter.set_ylim(-0.2, 16)
            ax_scatter.set_xticks(np.arange(0, 17, 4))
            ax_scatter.set_yticks(np.arange(0, 17, 4))
        elif var == 'rh':
            ax_scatter.set_xlim(30, 105); ax_scatter.set_ylim(30, 105)
            ax_scatter.set_xticks(np.arange(30, 105, 10))
            ax_scatter.set_yticks(np.arange(30, 105, 10))

        # 轴标题 & 标签 - 交换x和y标签
        if row == 0:
            title_map = {'temp': "Temperature (K)",
                         'hws' : "Horizontal Wind Speed (m/s)",
                         'rh'  : "RH (%)"}
            ax_scatter.set_title(title_map[var], fontsize=36, pad=25, fontweight='bold')
        if row == len(heights)-1 and col == 1:
            ax_scatter.set_xlabel(
                "Automatic Weather Station (AWS)",  # 交换为AWS
                fontsize=40, fontweight='bold', labelpad=15
            )
        else:
            ax_scatter.set_xlabel('')
        ax_scatter.set_ylabel('')

        # ───── 半小提琴图（保持原逻辑，仅样式微调） ─────
        ax_violin = fig.add_subplot(gs[row, 2*col+1])
        # 交换x和y - 现在x是AWS，y是ECM
        df_melt = pd.DataFrame({
            "value": pd.concat([y, x]),  # 交换x和y
            "source": ["X"]*len(y) + ["Y"]*len(x)  # 交换x和y
        })
        sns.violinplot(
            x="source", y="value", data=df_melt, ax=ax_violin,
            hue="source", palette=palette, linewidth=3, linecolor='black',
            inner="box", legend=False, split=True,
            inner_kws=dict(box_width=20, whis_width=3, color="0.75")
        )

        # 小提琴坐标与散点保持一致 - 交换x和y
        if var == 'temp':
            ax_violin.set_ylim(287, 310); ax_violin.set_yticks(np.arange(287, 311, 5))
        elif var == 'hws':
            ax_violin.set_ylim(-0.2, 16); ax_violin.set_yticks(np.arange(0, 16.01, 4))
        elif var == 'rh':
            ax_violin.set_ylim(10, 106.1); ax_violin.set_yticks(np.arange(10, 106.1, 10))

        ax_violin.tick_params(axis='both', which='major', labelsize=36,
                              direction='out', length=20, width=3, color='black',
                              bottom=True, left=True)
        ax_violin.yaxis.set_minor_locator(AutoMinorLocator(2))
        ax_violin.tick_params(which='minor', axis='both', length=10, width=3,
                              direction='out', bottom=False, left=True, color='black')
        ax_violin.grid(False)
        ax_violin.spines[['top', 'right']].set_visible(False)
        ax_violin.spines[['left', 'bottom']].set_linewidth(3)
        ax_violin.spines[['left', 'bottom']].set_color('black')
        ax_violin.set_facecolor('none')
        ax_violin.set_xlabel(''); ax_violin.set_ylabel('')

# ─────────────────────────── 全局 Y 轴标签 ───────────────────────────
fig.text(-0.015, 0.5,
         'Eddy Covariance Measurement (ECM)',  # 交换为ECM
         rotation='vertical', va='center', fontsize=40, fontweight='bold')

#───────────────────────────── 保存 ─────────────────────────────
plt.tight_layout()
plt.savefig(
    os.path.join("E:\\Beijing2024\\出图TIF\\tieta_vs_ultra_selected2.tif"),
    dpi=600, bbox_inches="tight", pad_inches=0.1, format="tif",
    facecolor='none', pil_kwargs={"compression": "tiff_adobe_deflate"}
)
plt.close()
gc.collect()


# """
# -------------------------------------------------------------------
# 重新做风雷达质量控制
# -------------------------------------------------------------------
# """
# import os
# import numpy as np
# import pandas as pd
# import matplotlib.pyplot as plt
# import metpy.calc as mpcalc
# from metpy.units import units

# # ---------- 0. 全局常量 ----------
# RAW_PATH1 = r"E:\Beijing2024\Wind Lidar\北京202310-202408\北京202310-202408.csv"
# RAW_PATH2 = r"E:\Beijing2024\Wind Lidar\0805-0810\merged_sorted_data.csv"
# OUT_DIR   = r"E:\Beijing2024\出图TIF\lidar qc"
# os.makedirs(OUT_DIR, exist_ok=True)

# START_BJT = "2024-08-04 00:00:00"
# END_BJT   = "2024-08-11 00:00:00"
# start_dt  = pd.to_datetime(START_BJT)
# end_dt    = pd.to_datetime(END_BJT)


# WINDOW_TXT = f"{START_BJT} – {END_BJT}"
# TIME_TAG   = f"{start_dt:%Y%m%d}_{end_dt:%Y%m%d}"
# # 额外增加一个时间窗（例如给 0606-0609）
# EXTRA_START_BJT = "2024-06-06 00:00:00"
# EXTRA_END_BJT   = "2024-06-09 23:59:59"
# extra_start_dt  = pd.to_datetime(EXTRA_START_BJT)
# extra_end_dt    = pd.to_datetime(EXTRA_END_BJT)

# # ---- CNR 阈值（水平 / 垂直） ----
# CNR_LIM_H = -22     # Horizontal dB
# CNR_LIM_V = -25     # Vertical dB

# # ---- 全局风速阈值（非对称） ----
# HWS_MIN, HWS_MAX =  0, 60   # m s-¹
# VWS_MIN, VWS_MAX = -8,6   # m s-¹

# # ---- 高度上限（区分水平 / 垂直） ----
# HEIGHT_LIM_H = 2700   # m
# HEIGHT_LIM_V = 2500   # m   ← 举例，按需调整

# # ---- 分层风速上限 ----
# HWS_LAYER_LIMITS = {300: 18, 3000: 60}
# # 上升气流（vws > 0）：各层最大允许值
# VWS_LAYER_LIMITS_UP = {
#     (0,   650):  5,
#     (650, 1800): 10,
#     (1800, 3000): 2
# }
# VWS_LAYER_LIMITS_DN = {
#     (0,   500):  -9,
#     (500, 1000): -7,
#     (1000,1500): -6,
#     (1500,2300): -3.5,
#     (2300,2500): -2
# }

# USE_LAYER_LIMITS_H = True
# USE_LAYER_LIMITS_V = True

# # ---------- 1. 绘图风格 ----------
# gc.collect()
# plt.rcParams.update({
#     "font.size": 16,
#     "axes.titlesize": 16,
#     "axes.labelsize": 16
# })
# # --------------------------------------------------
# # 2. 读取与预处理
# # --------------------------------------------------
# print("[INFO] 读取原始文件 …")

# RAW_PATH1 = r"E:\Beijing2024\Wind Lidar\北京202310-202408\北京202310-202408.csv"
# RAW_PATH2 = r"E:\Beijing2024\Wind Lidar\0805-0810\merged_sorted_data.csv"
# w1 = pd.read_csv(RAW_PATH1).drop(columns=["Unnamed: 0"], errors="ignore")
# w2 = pd.read_csv(RAW_PATH2)
# # 新增：0606-0609 的水平 / 垂直文件
# RAW_PATH3_H = r"E:\Beijing2024\Wind Lidar\0606-0609\wl_0606_h.xlsx"  # 水平风速
# RAW_PATH3_V = r"E:\Beijing2024\Wind Lidar\0606-0609\wl_0606_v.xlsx"  # 垂直风速

# # 新增：0606-0609 水平 / 垂直
# w3_h = pd.read_excel(RAW_PATH3_H)   # 水平风速文件 wl_0606_h.xlsx
# w3_v = pd.read_excel(RAW_PATH3_V)   # 垂直风速文件 wl_0606_v.xlsx


# def preprocess_horizontal(df: pd.DataFrame, label: str,
#                           time_col_guess=("Timestamp", "timestamp")) -> pd.DataFrame:
#     """处理带水平风(必有 hws/wd)的文件，生成: BJT, height, hws, wd, u, v, vws(如有), CNR, Level(hPa)."""
#     print(f"[INFO] 预处理水平风 DataFrame {label} ({len(df):,} rows)…")
#     df = df.copy()

#     # ------- 时间列统一为 BJT -------
#     for cand in time_col_guess:
#         if cand in df.columns:
#             df.rename(columns={cand: "timestamps"}, inplace=True)
#             break
#     if "timestamps" not in df.columns:
#         raise ValueError(f"{label} 中没有时间列 (Timestamp/timestamp)")

#     df["timestamps"] = pd.to_datetime(df["timestamps"])
#     df["BJT"] = df["timestamps"] + pd.Timedelta(hours=8)
#     df.drop(columns="timestamps", inplace=True)

#     # ------- 高度 -------
#     df.rename(columns={"Altitude [m]": "height"}, inplace=True, errors="ignore")

#     # ------- 水平风速 + 风向 -------
#     df.rename(columns={
#         "Horizontal Wind Speed [m/s]": "hws",
#         "Horizontal Wind Direction [°]": "wd"
#     }, inplace=True, errors="ignore")

#     if not {"hws", "wd"}.issubset(df.columns):
#         raise ValueError(f"{label} 缺少 hws / wd 列")

#     # 计算 u, v
#     df["u"] = -df["hws"] * np.sin(np.radians(df["wd"]))
#     df["v"] = -df["hws"] * np.cos(np.radians(df["wd"]))

#     # ------- 垂直风速(如果本文件有的话) -------
#     df.rename(columns={"Vertical Wind Speed [m/s]": "vws"}, inplace=True, errors="ignore")
#     if "vws" not in df.columns:
#         df["vws"] = np.nan   # 对于纯水平风文件先占位，后面用垂直文件补

#     # ------- CNR -------
#     df.rename(columns={"CNR [dB]": "CNR"}, inplace=True, errors="ignore")
#     if "CNR" not in df.columns:
#         df["CNR"] = np.nan

#     # ------- 标准大气压高度 -> 压力 -------
#     if "height" in df.columns:
#         df["Level (hPa)"] = mpcalc.height_to_pressure_std(df["height"].values * units.m).m
#     else:
#         df["Level (hPa)"] = np.nan

#     df.sort_values(["BJT", "height"], inplace=True)
#     return df.reset_index(drop=True)


# def preprocess_vertical(df: pd.DataFrame, label: str,
#                         time_col_guess=("Timestamp", "timestamp")) -> pd.DataFrame:
#     """处理垂直风文件，只关心: BJT, height, vws（和可选 CNR）。"""
#     print(f"[INFO] 预处理垂直风 DataFrame {label} ({len(df):,} rows)…")
#     df = df.copy()

#     # ------- 时间列 -> BJT -------
#     for cand in time_col_guess:
#         if cand in df.columns:
#             df.rename(columns={cand: "timestamps"}, inplace=True)
#             break
#     if "timestamps" not in df.columns:
#         raise ValueError(f"{label} 中没有时间列 (Timestamp/timestamp)")

#     df["timestamps"] = pd.to_datetime(df["timestamps"])
#     df["BJT"] = df["timestamps"] + pd.Timedelta(hours=8)
#     df.drop(columns="timestamps", inplace=True)

#     # ------- 高度 -------
#     df.rename(columns={"Altitude [m]": "height"}, inplace=True, errors="ignore")

#     # 垂直风速
#     df.rename(columns={"Vertical Wind Speed [m/s]": "vws"}, inplace=True, errors="ignore")
#     if "vws" not in df.columns:
#         raise ValueError(f"{label} 中没有 Vertical Wind Speed [m/s] 列？")

#     # CNR（如果有，可以一起带上）
#     df.rename(columns={"CNR [dB]": "CNR_v"}, inplace=True, errors="ignore")

#     df.sort_values(["BJT", "height"], inplace=True)
#     return df.reset_index(drop=True)


# # ===== 2.1 先分别处理 w1, w2（原有逻辑不变） =====
# w1_proc = preprocess_horizontal(w1, "#1_202310-202408", time_col_guess=("Timestamp",))
# w2_proc = preprocess_horizontal(w2, "#2_0805-0810",      time_col_guess=("timestamp",))


# # ===== 2.2 单独处理 0606-0609 的水平 / 垂直，然后合并成同时有 hws+vws 的 DataFrame =====
# w3_h_proc = preprocess_horizontal(w3_h, "#3_0606-0609_h", time_col_guess=("Timestamp", "timestamp"))
# w3_v_proc = preprocess_vertical(w3_v, "#3_0606-0609_v",   time_col_guess=("Timestamp", "timestamp"))

# # 为了用垂直文件的 vws 替换掉水平文件中占位的 NaN vws：
# w3_h_proc = w3_h_proc.drop(columns=["vws"], errors="ignore")

# # 在 (BJT, height) 上合并，把 vws 加进去
# w3_merged = pd.merge(
#     w3_h_proc,
#     w3_v_proc[["BJT", "height", "vws"]],   # 这里只把 vws merge 进来
#     on=["BJT", "height"],
#     how="left"
# )

# print(f"[INFO] 0606-0609 合并后行数：{len(w3_merged):,}")


# # ===== 2.3 最终 raw_all：包括水平和垂直速度 =====
# all_df = [w1_proc, w2_proc, w3_merged]
# raw_all = pd.concat(all_df, ignore_index=True)
# print(f"[INFO] 合并后行数：{len(raw_all):,}")

# # 按时间窗裁剪
# raw_all = raw_all[(raw_all["BJT"] >= start_dt) & (raw_all["BJT"] <= end_dt)].reset_index(drop=True)
# print(f"[INFO] 时间窗内行数：{len(raw_all):,}")

# # --------------------------------------------------
# # 2. 读取与预处理
# # --------------------------------------------------
# print("[INFO] 读取 CSV …")
# w1 = pd.read_csv(RAW_PATH1).drop(columns=["Unnamed: 0"], errors="ignore")
# w2 = pd.read_csv(RAW_PATH2)

# w1.rename(columns={"Timestamp": "timestamps"}, inplace=True)
# w2.rename(columns={"timestamp": "timestamps"}, inplace=True)

# all_df = []
# for idx, df in enumerate([w1, w2], 1):
#     print(f"[INFO] 预处理 DataFrame #{idx} ({len(df):,} rows)…")
#     df["timestamps"] = pd.to_datetime(df["timestamps"])
#     df["BJT"] = df["timestamps"] + pd.Timedelta(hours=8)
#     df.drop(columns="timestamps", inplace=True)

#     df.rename(columns={"Altitude [m]": "height"}, inplace=True, errors="ignore")
#     if {"Horizontal Wind Speed [m/s]", "Horizontal Wind Direction [°]"}.issubset(df.columns):
#         df.rename(columns={
#             "Horizontal Wind Speed [m/s]": "hws",
#             "Horizontal Wind Direction [°]": "wd"
#         }, inplace=True)
#     if not {"hws", "wd"}.issubset(df.columns):
#         raise ValueError("缺少 hws / wd 列")

#     df["u"] = -df["hws"] * np.sin(np.radians(df["wd"]))
#     df["v"] = -df["hws"] * np.cos(np.radians(df["wd"]))

#     df.rename(columns={"Vertical Wind Speed [m/s]": "vws"}, inplace=True, errors="ignore")
#     if "vws" not in df.columns:
#         df["vws"] = np.nan

#     df.rename(columns={"CNR [dB]": "CNR"}, inplace=True, errors="ignore")
#     if "CNR" not in df.columns:
#         df["CNR"] = np.nan

#     df["Level (hPa)"] = mpcalc.height_to_pressure_std(df["height"].values*units.m).m
#     df.sort_values(["BJT", "height"], inplace=True)
#     all_df.append(df.reset_index(drop=True))

# raw_all = pd.concat(all_df, ignore_index=True)
# print(f"[INFO] 合并后行数：{len(raw_all):,}")

# # 主时间窗 + 额外时间窗，一起保留
# mask_main  = (raw_all["BJT"] >= start_dt)      & (raw_all["BJT"] <= end_dt)
# mask_extra = (raw_all["BJT"] >= extra_start_dt) & (raw_all["BJT"] <= extra_end_dt)

# raw_all = raw_all[mask_main | mask_extra].reset_index(drop=True)
# print(
#     f"[INFO] 主时间窗({START_BJT}–{END_BJT}) + "
#     f"额外时间窗({EXTRA_START_BJT}–{EXTRA_END_BJT}) 过滤后行数：{len(raw_all):,}"
# )

# # vertical_all=raw_all.dropna(subset=['vws'])
# # vertical_all=vertical_all.set_index("BJT")
# # horizontal_all = raw_all.dropna(subset=['hws'])
# # horizontal_all=horizontal_all.set_index("BJT")
# # horizontal_all.reset_index().to_csv("E:\\20240805-20240810_h.csv",index=False)
# # vertical_all.reset_index().to_csv("E:\\20240805-20240810_v.csv",index=False)
# # --------------------------------------------------
# # 3. 原始统计（±k σ）
# # --------------------------------------------------
# def bands(mu, sigma, ks=(1,1.5,2,2.5,3)):
#     return {f"±{k}σ": (mu - k*sigma, mu + k*sigma) for k in ks}

# for col, tag in [("hws", "Horizontal"), ("vws", "Vertical")]:
#     mu = raw_all[col].mean(skipna=True)
#     sd = raw_all[col].std(skipna=True)
#     print(f"[STAT] {tag} Speed  mean={mu:.3f}, σ={sd:.3f}")
#     for k, (lo, hi) in bands(mu, sd).items():
#         print(f"       {k}: ({lo:.3f}, {hi:.3f}) m/s")

# # --------------------------------------------------
# # 4. QC 掩码
# # --------------------------------------------------
# def build_qc_mask(df: pd.DataFrame, speed_col: str) -> pd.Series:
#     """speed_col='hws' / 'vws' → 生成 QC 布尔掩码"""
#     if speed_col == "hws":
#         mask = (df["CNR"] >= CNR_LIM_H) & df["hws"].between(HWS_MIN, HWS_MAX)
#         if HEIGHT_LIM_H is not None:
#             mask &= df["height"] <= HEIGHT_LIM_H
#         if USE_LAYER_LIMITS_H:
#             for z_top, lim in HWS_LAYER_LIMITS.items():
#                 mask &= ~((df["height"] <= z_top) & (df["hws"] > lim))
#     elif speed_col == "vws":
#         # ① CNR + 全局非对称区间
#         mask = (df["CNR"] >= CNR_LIM_V) & df["vws"].between(VWS_MIN, VWS_MAX)
    
#         # ② 高度上限
#         if HEIGHT_LIM_V is not None:
#             mask &= df["height"] <= HEIGHT_LIM_V
    
#         # ③ 层内上限（区分正 / 负）
#         if USE_LAYER_LIMITS_V:
#             # -- 上升：vws > 0 --
#             for (z_lo, z_hi), max_up in VWS_LAYER_LIMITS_UP.items():
#                 cond = (df["height"] > z_lo) & (df["height"] <= z_hi) & (df["vws"] > max_up)
#                 mask &= ~cond
    
#             # -- 下沉：vws < 0 --
#             for (z_lo, z_hi), min_dn in VWS_LAYER_LIMITS_DN.items():
#                 cond = (
#                     (df["height"] >  z_lo) &        # 位于该层下界之上
#                     (df["height"] <= z_hi) &        # 且位于该层上界之内
#                     (df["vws"]      <  min_dn)      # 且下沉幅度超过阈值（更负）
#                 )
#                 mask &= ~cond       # 满足 cond → 异常，取反后 & 进总掩码

#     else:
#         raise ValueError("speed_col 必须 'hws' 或 'vws'")
#     return mask

# mask_h = build_qc_mask(raw_all, "hws")
# mask_v = build_qc_mask(raw_all, "vws")
# cleaned_h = raw_all[mask_h].reset_index(drop=True)
# cleaned_v = raw_all[mask_v].reset_index(drop=True)

# print(f"[INFO] QC 后水平行数：{len(cleaned_h):,}")
# print(f"[INFO] QC 后垂直行数：{len(cleaned_v):,}")

# import pandas as pd
# import dask.dataframe as dd
# from dask.diagnostics import ProgressBar


# # ----------------- 1. 定义优化的数据类型 -----------------
# dtype_spec = {
#     # ID 类列，使用 int32 节省空间
#     'Settings ID': 'Int32',
#     'Resolution ID': 'Int32',
#     'Scan ID': 'Int32',
#     'LOS ID': 'Int32',
#     'Sequence ID': 'Int32',

#     # 所有测量值，使用 float32 精度足够且省空间
#     'Azimuth [°]': 'float32',
#     'Elevation [°]': 'float32',
#     'height': 'float32',
#     'hws': 'float32',
#     'vws': 'float32',
#     'wd': 'float32',
#     'Radial Wind Speed [m/s]': 'float32',
#     'CNR': 'float32',
#     'Confidence Index [%]': 'float32',
#     'u': 'float32',
#     'v': 'float32',
#     'Level (hPa)': 'float32',
#     'ws_ci': 'float32',
#     'rws_ci': 'float32',
#     'rws': 'float32',
#     'doppler_spectrum_Mean_Error': 'float32',
#     'doppler_spectrum_width': 'float32',
#     'range_to_measurement_volume': 'float32',

#     # 状态类列，使用 category 类型，节省空间效果最佳
#     'Confidence Index Status': 'category',
#     'ws_status': 'category',
#     'rws_status': 'category',
# }

# # ----------------- 2. 应用数据类型转换 -----------------
# # 先转换除了时间列之外的所有列
# cleaned_v_optimized = cleaned_v.astype(dtype_spec)

# # 单独处理时间列 'BJT'，转换为标准 datetime 格式
# # errors='coerce' 会将无法转换的日期设为 NaT (Not a Time)，避免报错
# #cleaned_h_optimized['BJT'] = pd.to_datetime(cleaned_h_optimized['BJT'], errors='coerce')

# # ----------------- 3. 写入 Parquet -----------------
# output_parquet = r'E:/Beijing2024/Wind Lidar/cleaned_v_data.parquet'

# # 将优化后的 Pandas DataFrame 转为 Dask DataFrame 以便分块写入
# ddf = dd.from_pandas(cleaned_v_optimized, npartitions=1) # npartitions可根据CPU核心数调整

# with ProgressBar():
#     ddf.to_parquet(
#         output_parquet,
#         engine='pyarrow',
#         compression='zstd'  # 或 'zstd'
#     )

# print(f"数据已成功写入 {output_parquet}")







# # --------------------------------------------------
# # 5. 绘图工具
# # --------------------------------------------------
# def _ensure_raw_not_exist(fname: str, tag: str) -> bool:
#     if tag.lower() == "raw" and os.path.exists(fname):
#         print(f"[SKIP] Raw 图已存在：{os.path.basename(fname)}")
#         return False
#     return True

# def scatter_cnr_speed(df, speed_col, label_tag,
#                       cnr_lim, s_min, s_max,
#                       height_lim, tag, show_qc_info):
#     """Height–Wind-Speed 散点"""
#     speed_sig = f"{speed_col}{s_min}to{s_max}ms"
#     fname = (os.path.join(
#         OUT_DIR,
#         f"{label_tag}_{tag}_{TIME_TAG}.png"
#     ) if not show_qc_info else
#         os.path.join(
#             OUT_DIR,
#             f"{label_tag}_{tag}_CNR{cnr_lim}dB_{speed_sig}_z{height_lim}m_{TIME_TAG}.png"
#         ))
#     if not _ensure_raw_not_exist(fname, tag):
#         return
#     if df.empty:
#         print(f"[WARN] {label_tag} 数据为空，跳过")
#         return

#     fig, ax = plt.subplots(figsize=(8,10))
#     ax.scatter(df[speed_col], df["height"], color="dodgerblue", s=12, linewidths=0)
#     ax.set_xlabel(f"{label_tag} Wind Speed (m/s)")
#     ax.set_ylabel("Height (m)")

#     if show_qc_info:
#         thres = f"CNR ≥ {cnr_lim} dB, {s_min} ≤ {speed_col} ≤ {s_max} m/s, z ≤ {height_lim} m"
#         title = f"{tag} {label_tag} Wind Speed vs Height\n({thres})\n{WINDOW_TXT}"
#     else:
#         title = f"{tag} {label_tag} Wind Speed vs Height\n{WINDOW_TXT}"
#     ax.set_title(title, fontsize=16)
#     plt.tight_layout()
#     fig.savefig(fname, dpi=150)
#     plt.close(fig)
#     print(f"[SAVED] {fname}")

# def scatter_speed_vs_cnr(df, speed_col, label_tag,
#                          cnr_lim, s_min, s_max,
#                          height_lim, tag, show_qc_info):
#     """CNR–Wind-Speed 散点"""
#     speed_sig = f"{speed_col}{s_min}to{s_max}ms"
#     fname = (os.path.join(
#         OUT_DIR,
#         f"CNR_vs_{label_tag}_{tag}_{TIME_TAG}.png"
#     ) if not show_qc_info else
#         os.path.join(
#             OUT_DIR,
#             f"CNR_vs_{label_tag}_{tag}_CNR{cnr_lim}dB_{speed_sig}_z{height_lim}m_{TIME_TAG}.png"
#         ))
#     if not _ensure_raw_not_exist(fname, tag):
#         return
#     if df.empty:
#         print(f"[WARN] {label_tag} 数据为空，跳过")
#         return

#     fig, ax = plt.subplots(figsize=(8,8))
#     ax.scatter(df[speed_col], df["CNR"], color="dodgerblue", s=12, linewidths=0)
#     ax.set_xlabel(f"{label_tag} Wind Speed (m/s)")
#     ax.set_ylabel("CNR (dB)")

#     if show_qc_info:
#         thres = f"CNR ≥ {cnr_lim} dB, {s_min} ≤ {speed_col} ≤ {s_max} m/s, z ≤ {height_lim} m"
#         title = f"{tag} CNR vs {label_tag} Wind Speed\n({thres})\n{WINDOW_TXT}"
#     else:
#         title = f"{tag} CNR vs {label_tag} Wind Speed\n{WINDOW_TXT}"
#     ax.set_title(title, fontsize=16)
#     plt.tight_layout()
#     fig.savefig(fname, dpi=150)
#     plt.close(fig)
#     print(f"[SAVED] {fname}")

# # --------------------------------------------------
# # 6. 调用绘图
# # --------------------------------------------------
# # --- Raw ---
# scatter_cnr_speed(raw_all, "hws", "Horizontal",
#                   CNR_LIM_H, HWS_MIN, HWS_MAX, HEIGHT_LIM_H, "Raw", False)
# scatter_cnr_speed(raw_all, "vws", "Vertical",
#                   CNR_LIM_V, VWS_MIN, VWS_MAX, HEIGHT_LIM_V, "Raw", False)

# scatter_speed_vs_cnr(raw_all, "hws", "Horizontal",
#                      CNR_LIM_H, HWS_MIN, HWS_MAX, HEIGHT_LIM_H, "Raw", False)
# scatter_speed_vs_cnr(raw_all, "vws", "Vertical",
#                      CNR_LIM_V, VWS_MIN, VWS_MAX, HEIGHT_LIM_V, "Raw", False)

# # --- QC ---
# scatter_cnr_speed(cleaned_h, "hws", "Horizontal",
#                   CNR_LIM_H, HWS_MIN, HWS_MAX, HEIGHT_LIM_H, "QC", True)
# scatter_cnr_speed(cleaned_v, "vws", "Vertical",
#                   CNR_LIM_V, VWS_MIN, VWS_MAX, HEIGHT_LIM_V, "QC", True)

# scatter_speed_vs_cnr(cleaned_h, "hws", "Horizontal",
#                      CNR_LIM_H, HWS_MIN, HWS_MAX, HEIGHT_LIM_H, "QC", True)
# scatter_speed_vs_cnr(cleaned_v, "vws", "Vertical",
#                      CNR_LIM_V, VWS_MIN, VWS_MAX, HEIGHT_LIM_V, "QC", True)

# print(f"[DONE] 所有图已输出至 {OUT_DIR}")

# def scatter_compare_in_out(df_good, df_raw, mask,  # good=通过QC
#                            x_col, y_col,           # 轴变量
#                            x_lab, y_lab,           # 轴标签
#                            title_tag, fname_tag):  # 图标题 & 文件签名
#     """
#     双层散点：QC 通过点(蓝) + Outlier(红)
#     df_good : 通过 QC 的行
#     df_raw  : 原始全部行
#     mask    : 布尔掩码 (True=good, False=outlier)
#     x_col/y_col : 列名
#     """
#     fname = os.path.join(
#         OUT_DIR,
#         f"{fname_tag}_Compare_{TIME_TAG}.png"
#     )
#     # 若想避免重复生成可加跳过逻辑
#     fig, ax = plt.subplots(figsize=(8, 8 if 'CNR' in [x_lab, y_lab] else 10))
#     # ---- ② Outlier（红）----
#     df_out = df_raw[~mask]
#     ax.scatter(df_out[x_col], df_out[y_col],
#                s=12, linewidths=0, color="red", label="outlier")

#     # ---- ① QC 通过（蓝）----
#     ax.scatter(df_good[x_col], df_good[y_col],
#                s=12, linewidths=0, color="dodgerblue", label="QC passed")


#     ax.set_xlabel(x_lab)
#     ax.set_ylabel(y_lab)
#     ax.set_title(f"{title_tag}\n{WINDOW_TXT}", fontsize=16)

#     # legend 右上
#     ax.legend(loc="upper right", markerscale=1.2, fontsize=12, frameon=True)

#     plt.tight_layout()
#     fig.savefig(fname, dpi=150)
#     plt.close(fig)
#     print(f"[SAVED] {fname}")

# # —— 额外：outlier 行 —— #
# outlier_h = raw_all[~mask_h].reset_index(drop=True)
# outlier_v = raw_all[~mask_v].reset_index(drop=True)

# # ---------- 6-B. Compare：QC vs Outlier ----------

# # 1) CNR vs HWS
# scatter_compare_in_out(cleaned_h, raw_all, mask_h,
#                        x_col="hws", y_col="CNR",
#                        x_lab="Horizontal Wind Speed (m/s)",
#                        y_lab="CNR (dB)",
#                        title_tag="CNR vs Horizontal Wind Speed (QC vs Outlier)",
#                        fname_tag="CNR_vs_HWS")

# # 2) CNR vs VWS
# scatter_compare_in_out(cleaned_v, raw_all, mask_v,
#                        x_col="vws", y_col="CNR",
#                        x_lab="Vertical Wind Speed (m/s)",
#                        y_lab="CNR (dB)",
#                        title_tag="CNR vs Vertical Wind Speed (QC vs Outlier)",
#                        fname_tag="CNR_vs_VWS")

# # 3) Height vs HWS
# scatter_compare_in_out(cleaned_h, raw_all, mask_h,
#                        x_col="hws", y_col="height",
#                        x_lab="Horizontal Wind Speed (m/s)",
#                        y_lab="Height (m)",
#                        title_tag="Height vs Horizontal Wind Speed (QC vs Outlier)",
#                        fname_tag="Height_vs_HWS")

# # 4) Height vs VWS
# scatter_compare_in_out(cleaned_v, raw_all, mask_v,
#                        x_col="vws", y_col="height",
#                        x_lab="Vertical Wind Speed (m/s)",
#                        y_lab="Height (m)",
#                        title_tag="Height vs Vertical Wind Speed (QC vs Outlier)",
#                        fname_tag="Height_vs_VWS")

cleaned_v = pd.read_parquet("E:\\Beijing2024\\Wind Lidar\\cleaned_v_data.parquet\\cleaned_v_data.parquet", engine='pyarrow')
cleaned_h = pd.read_parquet("E:\\Beijing2024\\Wind Lidar\\cleaned_h_data.parquet\\cleaned_h_data.parquet", engine='pyarrow')

# --------------------------------------------------
# 7. 先时间平均 (用 u/v) → 再垂直插值
# --------------------------------------------------

TARGET_Z = np.array([80, 140, 200, 280])   # m
COLS     = [f"{z}m" for z in TARGET_Z]

def interp_profile(group, val_col):
    """单条时间廓线 -> 4 个插值值；禁止外推"""
    z   = group["height"].values
    val = group[val_col].values
    msk = ~np.isnan(val)
    z, val = z[msk], val[msk]

    if len(z) < 4:              # 点太少，全部 NaN
        return np.full_like(TARGET_Z, np.nan, dtype=float)

    idx = np.argsort(z)         # 保证高度单调
    z, val = z[idx], val[idx]

    f = interp1d(z, val, kind="linear",
                 bounds_error=False, fill_value=np.nan)
    return f(TARGET_Z)

def interpolate_df(df, val_col):
    """
    按时间戳 groupby 后对 val_col 做垂直插值，
    输出：index = BJT，columns = [80m, 140m, 200m, 280m]
    """
    arr = df.groupby("BJT")[["height", val_col]].apply(
        lambda g: interp_profile(g, val_col)
    ).to_numpy()
    out = pd.DataFrame(
        np.vstack(arr), index=df["BJT"].unique(), columns=COLS
    ).sort_index()
    return out


# --------- 7.1 水平风：用已有 u/v，先时间平均再插值 ---------
def resample_uv_then_interp(df_h,
                            freqs=("1min",  "60min")):#"5min", "30min",
    """
    df_h: 必须至少包含 ['BJT', 'height', 'u', 'v'] 列
    返回：
      hws_avg: {freq: DataFrame (BJT × 4 高度)}
      wd_avg:  {freq: DataFrame (BJT × 4 高度)}
    """
    # 只取需要的列，丢掉 u/v 为 NaN 的行
    df = df_h[["BJT", "height", "u", "v"]].copy()
    df = df.dropna(subset=["u", "v"])

    # 以 BJT 为时间索引，方便 resample
    df = df.set_index("BJT")

    hws_out = {}
    wd_out  = {}

    for freq in freqs:
        # 按高度分组，再对时间做重采样平均
        # 结果列：["height", "BJT", "u", "v"]
        uv_avg = (
            df
            .groupby("height")[["u", "v"]]
            .resample(freq)
            .mean()
            .reset_index()
        )

        # --- 用平均后的 u/v 重算风速和风向 ---
        u = uv_avg["u"].to_numpy()
        v = uv_avg["v"].to_numpy()

        # 风速
        hws = np.sqrt(u * u + v * v)

        # 风向（气象风向，来自的方向，单位：度）
        # atan2(y, x) = atan2(v, u)
        wd_rad = np.arctan2(v, u)
        wd_deg = (270.0 - np.degrees(wd_rad)) % 360.0

        uv_avg["hws"] = hws
        uv_avg["wd"]  = wd_deg

        # --- 对该时间分辨率下的 hws / wd 做垂直插值 ---
        hws_i = interpolate_df(uv_avg[["BJT", "height", "hws"]], "hws")
        wd_i  = interpolate_df(uv_avg[["BJT", "height", "wd"]],  "wd")

        hws_out[freq] = hws_i
        wd_out[freq]  = wd_i

    return hws_out, wd_out


# --------- 7.2 垂直风：标量先时间平均，再插值 ---------
def resample_scalar_then_interp(df, val_col,
                                freqs=("1min", "60min")):#"5min", "30min",
    """
    对标量（如 vws）先按高度+时间平均，再插值到 TARGET_Z
    返回：{freq: DataFrame (BJT × 4 高度)}
    """
    df2 = df.set_index("BJT")
    out = {}

    for freq in freqs:
        avg = (
            df2
            .groupby("height")[[val_col]]
            .resample(freq)
            .mean()
            .reset_index()
        )
        out[freq] = interpolate_df(avg[["BJT", "height", val_col]], val_col)

    return out


from scipy.interpolate import interp1d
import numpy as np
import pandas as pd

def time_interp_60min_small_gaps(df_60, windows):
    """
    对 60min 平均风速沿时间轴做线性插值，规则：
      - 只在给定的若干时间窗 windows 内插值；
      - 只填“连续 NaN 长度 <= 2”的小洞；
      - 这一段 NaN 前后必须都有非 NaN 数据；
      - 禁止外推：只在全局有效数据时间范围内插值。

    参数
    ----
    df_60 : DataFrame
        index 为 DatetimeIndex（60min），列为各高度（如 '80m','140m','200m','280m'）
    windows : list[tuple[pd.Timestamp, pd.Timestamp]]
        例如：
        [
            (pd.Timestamp("2024-06-06 00:00:00"), pd.Timestamp("2024-06-10 23:59:59")),
            (pd.Timestamp("2024-08-05 00:00:00"), pd.Timestamp("2024-08-10 23:59:59")),
        ]
    """
    df = df_60.copy()

    # 时间 → 数值（秒）作为自变量
    t = df.index.view("int64") / 1e9   # nanosecond -> second

    # 窗口转成秒
    win_secs = []
    for start, end in windows:
        start = pd.to_datetime(start)
        end   = pd.to_datetime(end)
        s_sec = start.value / 1e9
        e_sec = end.value   / 1e9
        win_secs.append((s_sec, e_sec))

    for col in df.columns:
        y = df[col].values.astype(float)
        mask_valid = ~np.isnan(y)

        # 有效点太少，不做任何插值
        if mask_valid.sum() < 2:
            continue

        t_valid = t[mask_valid]
        y_valid = y[mask_valid]

        t_min = t_valid.min()
        t_max = t_valid.max()

        # 构造线性插值函数（只用于 t_min~t_max 之间）
        f = interp1d(
            t_valid, y_valid,
            kind="linear"
        )

        y_new = y.copy()

        # ---- 找连续 NaN 区段 ----
        nan_mask = np.isnan(y)
        n = len(y)
        i = 0
        while i < n:
            if nan_mask[i]:
                # 起点
                start = i
                # 找这一段连续 NaN 的结束位置
                while i + 1 < n and nan_mask[i + 1]:
                    i += 1
                end = i
                length = end - start + 1

                # 条件 1：长度 <= 2
                # 条件 2：两侧有非 NaN（不是首尾，而且左右邻接点都是 valid）
                if (length <= 3 and
                    start > 0 and end < n - 1 and
                    (not nan_mask[start - 1]) and
                    (not nan_mask[end + 1])):

                    # 这一小段 NaN 逐点判断是否在窗口内 & 在有效数据时间范围内
                    for j in range(start, end + 1):
                        tj = t[j]

                        # 在任意一个窗口内？
                        in_window = any((tj >= ws) and (tj <= we)
                                        for (ws, we) in win_secs)
                        if (not in_window):
                            continue

                        # 禁止外推：必须在 [t_min, t_max] 内
                        if (tj < t_min) or (tj > t_max):
                            continue

                        # 符合所有条件，才插值
                        y_new[j] = float(f(tj))

                # 移动到下一段
                i += 1
            else:
                i += 1

        df[col] = y_new

    return df

# --------- 7.3 调用 ---------
# 水平风：得到每个时间分辨率下插值后的 hws、wd
hws_avg, wd_avg = resample_uv_then_interp(cleaned_h)

# 垂直风：得到每个时间分辨率下插值后的 vws
vws_avg = resample_scalar_then_interp(cleaned_v, "vws")

# 定义两个时间窗（示例，按你自己的时间改）
WIN1 = (pd.Timestamp("2024-06-06 00:00:00"), pd.Timestamp("2024-06-10 23:59:59"))
WIN2 = (pd.Timestamp("2024-08-04 00:00:00"), pd.Timestamp("2024-08-10 23:59:59"))

windows = [WIN1, WIN2]

# 对 60min 平均风速，在两个指定时间窗内做时间插值（禁止外推）
hws_avg["60min"] = time_interp_60min_small_gaps(hws_avg["60min"], windows)
hws_avg["1min"] = time_interp_60min_small_gaps(hws_avg["1min"], windows)

# —— 示例：访问水平风速 5 min 均值 DataFrame ——
# hws_avg["5min"]
# vws_avg["60min"]

# 如需保存：
# hws_avg["5min"].to_csv(os.path.join(OUT_DIR, "hws_5min_mean.csv"))
# vws_avg["60min"].to_csv(os.path.join(OUT_DIR, "vws_60min_mean.csv"))



#80-75m    140-150m   200-200m    280-275m

# import seaborn as sns
# import matplotlib.pyplot as plt
# import matplotlib.ticker as ticker
# from mpl_toolkits.axes_grid1.inset_locator import inset_axes
# import matplotlib.patches as mpatches   # 用来生成自定义图例手柄

# plt.figure(figsize=(6, 6))

# # ---------- 1. 主直方图 + KDE ----------
# ax = sns.histplot(
#     raw_all['hws'],
#     bins=50,
#     kde=True,
#     color='blue',
#     line_kws={'color': 'black', 'lw': 1.5},
#     ax=None,              # 让 seaborn 自动创建轴
#     label='HWS Histogram'     # 给直方图一个名字
# )

# # 单独给 KDE 线条改标签
# ax.lines[0].set_label('Gaussian KDE')

# # ---------- 2. 下方嵌入箱线图 ----------
# # 在主坐标系内部再开一条“横条形”子坐标轴
# ax_box = inset_axes(
#     ax,
#     width="100%", height="10%",      # 100% 宽、10% 高
#     loc='lower center',
#     bbox_to_anchor=(0, -0.28, 1, 1), # 向下偏移一些，不遮住直方图
#     bbox_transform=ax.transAxes,
#     borderpad=0
# )

# sns.boxplot(
#     x=raw_all['hws'],
#     ax=ax_box,
#     width=0.6,
#     color='gray',
#     fliersize=0            # 不显示异常值点
# )

# # 箱线图子轴美化（去掉 y 轴 & 字注）
# ax_box.set(yticks=[], ylabel='', xlabel='')
# for spine in ax_box.spines.values():
#     spine.set_color('black')
#     spine.set_linewidth(2)

# # ---------- 3. 轴刻度 & 文字 ----------
# ax.xaxis.set_major_locator(ticker.MultipleLocator(20))
# ax.yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=False))
# ax.yaxis.offsetText.set_visible(False)   # 隐藏科学计数 offset
# ax.set_ylabel(r"Frequency ($\times10^{6}$)", fontsize=14)
# ax.set_xlabel("Horizontal Wind Speed (m/s)", fontsize=14)
# ax.set_title("Distribution of Horizontal Wind Speed", pad=0, fontsize=14)

# # 坐标轴边框
# for spine in ['left', 'right', 'top', 'bottom']:
#     ax.spines[spine].set_color('black')
#     ax.spines[spine].set_linewidth(2)

# # 主/副刻度
# ax.tick_params(axis='both', which='major', direction='out',
#                length=12, width=2, bottom=True, left=True)
# ax.tick_params(axis='both', which='minor', direction='out',
#                length=6, width=2, bottom=True, left=True)
# ax.grid(False)

# # ---------- 4. 图例 ----------
# # 自定义箱线图手柄（灰色矩形）
# box_handle = mpatches.Patch(facecolor='gray', edgecolor='black', label='Box plot')
# handles, labels = ax.get_legend_handles_labels()
# handles.append(box_handle); labels.append('HWS Box plot')

# ax.legend(handles, labels, loc='upper right', frameon=False)

# plt.tight_layout()
# plt.show()


"""
先处理微波辐射计数据
-------------------------------------------------------------------
"""

# ---------- 0. 选择并标准化列 ---------- #
columns_needed = ['timestamps', 'height_plus_alt', 'T(K)', 'RH(%)']
mwr_selected   = df_mwr[columns_needed].copy()
mwr_selected['timestamps'] = (
    pd.to_datetime(mwr_selected['timestamps']) + pd.Timedelta(hours=8)
)
mwr_selected = mwr_selected.rename(columns={'height_plus_alt': 'height_m'})  # 观测高度 (m)

# ---------- 1. 透视为“行=时间、列=高度” ---------- #
temperature_wide = (
    mwr_selected
    .pivot_table(index='timestamps', columns='height_m', values='T(K)')
    .sort_index(axis=1)
)

humidity_wide = (
    mwr_selected
    .pivot_table(index='timestamps', columns='height_m', values='RH(%)')
    .sort_index(axis=1)
)

# ---------- 2. 垂直线性插值（禁止外推） ---------- #
target_heights_m = [80, 140, 200, 280]

def interpolate_to_target_levels(wide_dataframe, target_levels):
    """在列方向做线性插值，仅保留目标高度；若缺层则整行剔除。"""
    # 先把目标高度并入列，再插值
    dataframe_with_targets = wide_dataframe.reindex(
        columns=sorted(set(wide_dataframe.columns).union(target_levels))
    )
    interpolated_dataframe = dataframe_with_targets.interpolate(axis=1, method='values')
    # 仅保留目标高度列，并去掉仍存在 NaN 的行（代表外推会产生缺值）
    return interpolated_dataframe[target_levels].dropna(how='any')

temperature_interpolated = interpolate_to_target_levels(temperature_wide, target_heights_m)
humidity_interpolated    = interpolate_to_target_levels(humidity_wide,    target_heights_m)

# ---------- 3. 合并温度与湿度剖面 ---------- #
# 得到多层列索引：第一层变量名，第二层高度
combined_interpolated_profiles = pd.concat(
    {'T(K)': temperature_interpolated, 'RH(%)': humidity_interpolated},
    axis=1
).sort_index(axis=1, level=0)

# ---------- 4. 时间平均 ---------- #
mean_1min_profiles  = combined_interpolated_profiles.resample('1min').mean().dropna(how='all')
mean_5min_profiles  = combined_interpolated_profiles.resample('5min').mean().dropna(how='all')
mean_30min_profiles  = combined_interpolated_profiles.resample('30min').mean().dropna(how='all')
mean_60min_profiles = combined_interpolated_profiles.resample('60min').mean().dropna(how='all')

# # ---------- 5. 保存 ---------- #
# mean_5min_profiles.to_parquet('MWR_mean_5min_80-280m.parquet')
# mean_60min_profiles.to_parquet('MWR_mean_60min_80-280m.parquet')
# 也可以用 Excel：
# mean_5min_profiles.to_excel('MWR_mean_5min_80-280m.xlsx',   merge_cells=False)
# mean_60min_profiles.to_excel('MWR_mean_60min_80-280m.xlsx', merge_cells=False)
"""
Tower multi-level time–series (lines) + inter‑dataset scatter matrix

1min
-------------------------------------------------------------------
"""
#####
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.gridspec import GridSpec
from matplotlib.ticker import MultipleLocator, AutoMinorLocator, FixedLocator
from matplotlib.ticker import FuncFormatter

# --------------------------------------------------
# 0. Constants & rcParams
# --------------------------------------------------
HEIGHTS = [280, 200, 140, 80]
VAR_GROUPS = ["Temp", "RH", "HWS", "VWS"]
COLMAP = dict(MWR="red", 
              LiDAR="darkorange",
              TowerAvg="black", 
              TowerTurb="slategray", 
              CRA="magenta", 
              ERA5="blue", 
              )

DATA_START = pd.Timestamp("2024-06-05 00:00")
DATA_END   = pd.Timestamp("2024-08-11 00:00")
# 数据筛选窗口（只把这段时间的数据读进来、做 QC）
PLOT_START = pd.Timestamp("2024-08-05 01:30")
PLOT_END   = pd.Timestamp("2024-08-11 00:00")
# ==== 新增：两个不连续的绘图时间窗口（自己改成想要的时间段） ====
SEG1_START = pd.Timestamp("2024-06-06 00:00")
SEG1_END   = pd.Timestamp("2024-06-10 00:00")
SEG2_START = pd.Timestamp("2024-08-05 01:30")
SEG2_END   = pd.Timestamp("2024-08-11 00:00")

ALPHA_MAP = {
    'MWR'      : 0.4,
    'TowerAvg' : 0.4,
    'TowerTurb': 0.4,
    'CRA'      : 0.6,
    'ERA5'     : 0.6,
    'LiDAR'    : 0.4,
}
ALPHA_MAP2 = {
    'MWR'      : 1,
    'TowerAvg' : 1,
    'TowerTurb': 0.5,
    'CRA'      : 1,
    'ERA5'     : 1,
    'LiDAR'    : 0.5,
}

LABEL_MAP = {
    'MWR'      : 'MWR',
    'LiDAR'    : 'DWL',
    'TowerAvg' : 'AWS',
    'TowerTurb': 'ECM',#eddy covariance flux measurement system
    'CRA'      : 'CRA',
    'ERA5'     : 'ERA5',
}
# axis limits for each variable
VAR_RANGE = {
    'Temp': (288, 310),
    'RH'  : (10, 103),
    'HWS' : (-0.99, 18),
    'VWS' : (-7, 4)
}
title_map = {
    'Temp': 'Temperature (T)',
    'RH'  : 'Relative Humidity (RH)',
    'HWS' : 'Horizontal Wind Speed (HWS)',
    'VWS' : 'Vertical Wind Speed (VWS)'
}
OUT_DIR = r"E:\Beijing2024\出图TIF"; os.makedirs(OUT_DIR, exist_ok=True)
OUT_FILE = os.path.join(OUT_DIR, "Tower_MultiLevel_TimeScatter2-1min3.tif")
plt.rcParams.update({    
    'font.size': 50,
    'font.family': 'sans-serif',                # 先指定家族
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     # 把 Arial 设为首选
    'axes.titlepad': 8,
    'legend.fontsize': 50,
    'axes.labelsize': 50,
    'xtick.labelsize': 50,
    'ytick.labelsize': 50,
    'xtick.major.size': 30,
    'xtick.minor.size': 15,
    'ytick.major.size': 30,
    'ytick.minor.size': 15,
    'xtick.direction': 'out',
    'ytick.direction': 'out',
    'xtick.major.width': 6,
    'xtick.minor.width': 6,
    'ytick.major.width': 6,
    'ytick.minor.width': 6,
})
# --------------------------------------------------
# 1. Helpers
# --------------------------------------------------

def ensure_dt(ser: pd.Series) -> pd.Series:
    ser = ser.copy(); ser.index = pd.to_datetime(ser.index); return ser.sort_index()

def pivot_height(df: pd.DataFrame, *, idx: str, col: str, val: str, h: int) -> pd.Series:
    tmp = df.reset_index(); tmp[idx] = pd.to_datetime(tmp[idx])
    return ensure_dt(tmp.pivot(index=idx, columns=col, values=val).loc[:, h])

def align_x_to_ref(ref: pd.Series, other: pd.Series) -> pd.DataFrame:
    return pd.concat([ref, other], axis=1, join='inner').dropna(how='any')

# --------------------------------------------------
# 2.  Build Series dictionary
# --------------------------------------------------
turb_calibrated_1min['rh'] = turb_calibrated_1min['rh'].clip(upper=100)
series_dict = {v: {h: {} for h in HEIGHTS} for v in VAR_GROUPS}

# ---- Temperature -----------------------------------------------------------
for h in HEIGHTS:
    series_dict['Temp'][h]['TowerAvg']  = pivot_height(tieta_1min, idx='timestamps', col='H(m)', val='T(K)_tieta', h=h)
    series_dict['Temp'][h]['TowerTurb'] = pivot_height(turb_calibrated_1min, idx='timestamps', col='Height', val='T (K)', h=h)
    series_dict['Temp'][h]['MWR']  = ensure_dt(mean_1min_profiles.xs('T(K)', level=0, axis=1)[h])
    series_dict['Temp'][h]['CRA']       = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['Temp (K)'])
    series_dict['Temp'][h]['ERA5']      = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['t'])

# ---- Relative Humidity -----------------------------------------------------
for h in HEIGHTS:
    series_dict['RH'][h]['TowerAvg']  = pivot_height(tieta_1min, idx='timestamps', col='H(m)', val='RH_tieta', h=h)
    series_dict['RH'][h]['TowerTurb'] = pivot_height(turb_calibrated_1min, idx='timestamps', col='Height', val='rh', h=h)
    series_dict['RH'][h]['MWR']  = ensure_dt(mean_1min_profiles.xs('RH(%)', level=0, axis=1)[h])
    series_dict['RH'][h]['CRA']       = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['rh (%)'])
    series_dict['RH'][h]['ERA5']      = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['r'])

# ---- Horizontal Wind Speed -------------------------------------------------
hws_qc_1min = (
    hws_avg["1min"]
    .rename(columns=lambda x: x.replace('m', ''))  # 先处理列名（移除'm'）
    .stack()
    .reset_index()
    .rename(columns={'level_1': 'height', 0: 'hws'})
    .astype({'height': int})  # 确保height为整数
)

for h in HEIGHTS:
    series_dict['HWS'][h]['LiDAR']    = ensure_dt(
        hws_qc_1min.query('height == @h').set_index('level_0')['hws'])
    series_dict['HWS'][h]['TowerAvg'] = pivot_height(tieta_1min, idx='timestamps', col='H(m)', val='ws_tieta', h=h)
    series_dict['HWS'][h]['TowerTurb']= pivot_height(turb_calibrated_1min, idx='timestamps', col='Height', val='hws_corrected', h=h)
    series_dict['HWS'][h]['CRA']      = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['hws(m/s)'])
    series_dict['HWS'][h]['ERA5']     = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['hws(m/s)'])

# ---- Vertical Wind Speed ---------------------------------------------------
vws_qc_1min = (
    vws_avg["1min"]
    .rename(columns=lambda x: x.replace('m', ''))  # 先处理列名（移除'm'）
    .stack()
    .reset_index()
    .rename(columns={'level_1': 'height', 0: 'vws'})
    .astype({'height': int})  # 确保height为整数
)

for h in HEIGHTS:
    series_dict['VWS'][h]['LiDAR']    = ensure_dt(
        vws_qc_1min.reset_index().query('height == @h').set_index('level_0')['vws'])
    series_dict['VWS'][h]['TowerTurb'] = pivot_height(turb_calibrated_1min, idx='timestamps', col='Height', val='W (m/s)', h=h)
    series_dict['VWS'][h]['CRA']       = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['vws(10⁻² {m/s})'] * -1)
    series_dict['VWS'][h]['ERA5']      = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['vws(10⁻² {m/s})'] * -100)

for v in VAR_GROUPS:
    for h in HEIGHTS:
        for src, ser in series_dict[v][h].items():
            series_dict[v][h][src] = ser.loc[DATA_START:DATA_END]

series_dict_1min = series_dict
# ==== 定义两个时间窗口 ====
WIN1_START = pd.Timestamp("2024-06-06 00:00")   # 新增窗口：6月6–10日
WIN1_END   = pd.Timestamp("2024-06-10 00:00")

WIN2_START = PLOT_START                         # 原来的 8 月窗口
WIN2_END   = PLOT_END

def make_month_first_formatter(first_date):
    """
    第一个 tick 显示 'Jun 06'，后面的 tick 只显示 '07', '08' ...
    first_date: 这个时间段的起始日期（比如 WIN1_START 或 WIN2_START）
    """
    first_date = pd.to_datetime(first_date).normalize()

    def _fmt(x, pos=None):
        dt = mdates.num2date(x)          # matplotlib 内部浮点 -> datetime
        d  = pd.Timestamp(dt).normalize()

        if d == first_date:
            # 第一个刻度：英文月份缩写 + 日
            return d.strftime('%b %d')   # 例如 'Jun 06'
        else:
            # 其它刻度：只显示日
            return d.strftime('%d')      # 例如 '07'
    return FuncFormatter(_fmt)

# --------------------------------------------------
# 3. Plot with broken x-axis
# --------------------------------------------------
fig = plt.figure(figsize=(74, 60))  # , constrained_layout=True
outer = GridSpec(2, 2, figure=fig, hspace=0.22, wspace=0.125)

for g_idx, var in enumerate(VAR_GROUPS):
    row, col = divmod(g_idx, 2)
    gspec = outer[row, col].subgridspec(
        len(HEIGHTS), 2,
        width_ratios=[3, 1.2],
        wspace=0.18,
        hspace=0.55
    )

    # column header
    ax_head = fig.add_subplot(gspec[0, 0])
    ax_head.axis('off')
    if var == 'VWS':
        ax_head.set_title(
            "Vertical Wind Speed (VWS)\n"           # 第一行普通文字，自动是正体
            r"$(VWS_{80/140/200/280}^{\mathrm{ERA5,CRA}}\times 35)$",  # 第二行数学公式，保留上下标
            pad=35, x=0.5, fontweight='bold', fontsize=60  # 整体加粗，第一行肯定是粗体
        )
    else:
        ax_head.set_title(title_map.get(var, var),
                          pad=35, fontweight='bold', fontsize=60)

    vmin, vmax = VAR_RANGE[var]
    
    for r_idx, h in enumerate(HEIGHTS):
        # ---------------- Line with broken x-axis ------------------
        # 左列再拆成两个子轴：左=6 月，右=8 月
        line_gs = gspec[r_idx, 0].subgridspec(1, 2, width_ratios=[1, 1],
                                              wspace=0.05)
        ax_left  = fig.add_subplot(line_gs[0, 0])
        ax_right = fig.add_subplot(line_gs[0, 1], sharey=ax_left)
    
        # 画曲线：两个时间段
        for src, ser in series_dict[var][h].items():
            if var == 'VWS' and src == 'TowerAvg':
                continue
            if ser is None or ser.empty:
                continue
    
            color = COLMAP.get(src, 'k')
            alpha = ALPHA_MAP2.get(src, 0.6)
            lw    = 4
    
            s1 = ser.loc[WIN1_START:WIN1_END]
            s2 = ser.loc[WIN2_START:WIN2_END]
    
            if not s1.empty:
                ax_left.plot(s1.index, s1.values,
                             color=color, alpha=alpha, lw=lw)
            if not s2.empty:
                ax_right.plot(s2.index, s2.values,
                              color=color, alpha=alpha, lw=lw)
    
        # x 范围
        ax_left.set_xlim(WIN1_START, WIN1_END)
        ax_right.set_xlim(WIN2_START, WIN2_END)
    
        # y 范围 & 刻度（只设置在左轴上）
        ax_left.set_ylim(vmin, vmax)
        if var == 'Temp':
            ax_left.yaxis.set_major_locator(MultipleLocator(4))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
        elif var == 'RH':
            ax_left.yaxis.set_major_locator(MultipleLocator(20))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
        elif var == 'HWS':
            ax_left.yaxis.set_major_locator(MultipleLocator(4))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
        else:  # VWS
            for ax in (ax_left, ax_right):
                ax.axhline(
                    0,
                    color='black',
                    linewidth=6,
                    zorder=1
                )
            ax_left.yaxis.set_major_locator(MultipleLocator(2))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
    
        # 右轴隐藏 y 轴刻度和标签（只保留一个 y 轴）
        ax_right.tick_params(
            axis='y', which='both',
            left=False, right=False,
            labelleft=False, labelright=False
        )
    
        # 高度标签（写在左轴上）
        ax_left.text(
            0.1, 1.05, f"{h} m",
            transform=ax_left.transAxes,
            ha='center', va='bottom',
            fontsize=56, fontweight='bold'
        )
    
        # y-label：第三行写单位
        if r_idx == 2:
            if var == 'Temp':
                ylab = "T (K)"
            elif var == 'RH':
                ylab = "RH (%)"
            elif var == 'HWS':
                ylab = "HWS (m/s)"
            elif var == 'VWS':
                ylab = "VWS (m/s)"
            else:
                ylab = ""
    
            ax_left.text(
                -0.2, 1.5, ylab,
                transform=ax_left.transAxes,
                fontsize=50, ha='right', va='center',
                rotation=90, fontweight="bold"
            )
    
        # 边框 & x 轴刻度格式
        for ax in (ax_left, ax_right):
            for sp in ['top', 'right']:
                ax.spines[sp].set_visible(False)
            for sp in ['bottom', 'left']:
                ax.spines[sp].set_linewidth(6)
    
            # ---- x 轴刻度：DayLocator + 自定义 formatter ----
            # 主刻度：每天一个
            ax.xaxis.set_major_locator(mdates.DayLocator())
            # 副刻度：每天中午 12 点
            ax.xaxis.set_minor_locator(mdates.HourLocator(byhour=[12]))
    
        # 左边（第一个时间窗）：第一个刻度显示 'Jun 06'，后面 '07' '08' ...
        ax_left.xaxis.set_major_formatter(
            make_month_first_formatter(WIN1_START)
        )
    
        # 右边（第二个时间窗）：第一个刻度显示 'Aug 10'，后面 '11' ...
        ax_right.xaxis.set_major_formatter(
            make_month_first_formatter(WIN2_START)
        )
    
        # broken axis：只在底部画两个斜杠
        ax_left.spines['right'].set_visible(False)
        ax_right.spines['left'].set_visible(False)
    
        d = .015
        kwargs = dict(color='k', clip_on=False, linewidth=4)
    
        # 左轴右侧（底部）斜杠
        ax_left.plot(
            (1 - d, 1 + d), (-d, +d),
            transform=ax_left.transAxes, **kwargs
        )
        # 右轴左侧（底部）斜杠
        ax_right.plot(
            (-d, +d), (-d, +d),
            transform=ax_right.transAxes, **kwargs
        )
    
        # 最后一行显示 x 轴标签，其余行隐藏
        if r_idx == len(HEIGHTS) - 1:
            # ====== 新增部分开始 ======
            # 在每个x轴下方添加月份标签
            month1 = WIN1_START.strftime('%b') # 获取月份缩写，如 'Jun'
            month2 = WIN2_START.strftime('%b') # 获取月份缩写，如 'Aug'
    
            ax_left.text(
                0.04, -0.27, month1,
                transform=ax_left.transAxes,
                ha='center', va='top',
                fontsize=50, fontweight='bold'
            )
            ax_right.text(
                0.085, -0.27, month2,
                transform=ax_right.transAxes,
                ha='center', va='top',
                fontsize=50, fontweight='bold'
            )    
            ax_left.set_xlabel("Datetime (BJT)", fontsize=50, x=1, labelpad = 45)
            ax_left.tick_params(labelbottom=True)
            ax_right.tick_params(labelbottom=True)
        else:
            ax_left.tick_params(labelbottom=False)
            ax_right.tick_params(labelbottom=False)

        # ---------------- Scatter 右侧 --------------
        ax_scat = fig.add_subplot(gspec[r_idx, 1])

        # 确定散点参考 ref
        ref = series_dict[var][h].get('TowerAvg')
        if ref is not None and not ref.empty:
            ref_source = 'AWS'
        else:
            ref = series_dict[var][h].get('TowerTurb')
            if ref is not None and not ref.empty:
                ref_source = 'ECM'
            else:
                ref = None
                ref_source = 'Other Source'
                for _s in series_dict[var][h].values():
                    if _s is not None and not _s.empty:
                        ref = _s
                        break

        # 只有最底行写 Ref 文本
        if r_idx == len(HEIGHTS) - 1:
            ax_scat.text(
                0.5, -0.35, f'Ref.: {ref_source}',
                transform=ax_scat.transAxes,
                fontsize=50, ha='center', va='top'
            )
        else:
            ax_scat.tick_params(labelbottom=False)

        # 1:1 线 & 基本范围
        ax_scat.plot(
            [vmin, vmax], [vmin, vmax],
            ls=(0, (5, 10)), lw=3, color='black',
            zorder=0, alpha=0.5
        )
        ax_scat.set_xlim(vmin, vmax)
        ax_scat.set_ylim(vmin, vmax)

        # 画散点
        for src, ser in series_dict[var][h].items():
            if src == 'TowerAvg':
                continue
            if var == 'VWS' and src == 'TowerTurb':
                continue
            if ref is None or ser is None or ser.empty:
                continue

            aligned = align_x_to_ref(ref, ser)
            if aligned.empty:
                continue

            ax_scat.scatter(
                aligned.iloc[:, 0], aligned.iloc[:, 1],
                s=200, linewidths=0, marker='o',
                color=COLMAP.get(src, 'k'),
                alpha=ALPHA_MAP.get(src, 0.5)
            )

        # 根据变量调整散点轴刻度
        if var == 'HWS':
            ax_scat.set_xlim(0, 16)
            ax_scat.xaxis.set_major_locator(MultipleLocator(4))
            ax_scat.xaxis.set_minor_locator(MultipleLocator(2))
            ax_scat.yaxis.set_major_locator(MultipleLocator(4))
            ax_scat.yaxis.set_minor_locator(MultipleLocator(2))

        elif var == 'VWS':
            ax_scat.set_xlim(-3, 3)
            ax_scat.xaxis.set_major_locator(MultipleLocator(2))
            ax_scat.xaxis.set_minor_locator(MultipleLocator(1))
            ax_scat.yaxis.set_major_locator(MultipleLocator(2))
            ax_scat.yaxis.set_minor_locator(MultipleLocator(1))

        elif var == 'RH':
            ax_scat.xaxis.set_major_locator(FixedLocator([30, 65, 100]))
            ax_scat.xaxis.set_minor_locator(AutoMinorLocator(2))
            ax_scat.yaxis.set_major_locator(FixedLocator([30, 65, 100]))
            ax_scat.yaxis.set_minor_locator(AutoMinorLocator(2))

        elif var == 'Temp':
            ax_scat.xaxis.set_major_locator(MultipleLocator(4))
            ax_scat.xaxis.set_minor_locator(AutoMinorLocator(2))
            ax_scat.yaxis.set_major_locator(MultipleLocator(4))
            ax_scat.yaxis.set_minor_locator(AutoMinorLocator(2))

        # 散点图边框
        for sp in ['top', 'right']:
            ax_scat.spines[sp].set_visible(False)
        for sp in ['bottom', 'left']:
            ax_scat.spines[sp].set_linewidth(6)

# --------------------------------------------------
# 4. Legend & save
# --------------------------------------------------
handles = [
    plt.Line2D([], [],
               color=COLMAP[k],
               lw=15,
               label=LABEL_MAP[k])
    for k in COLMAP
]
fig.legend(
    handles, [h.get_label() for h in handles],
    loc='upper center', ncol=3, frameon=False,
    title_fontsize=80, bbox_to_anchor=(0.5, 0.935)
)

fig.savefig(
    OUT_FILE, dpi=300, format='tif',
    transparent=True, bbox_inches='tight',
    pil_kwargs={'compression': 'tiff_adobe_deflate'}
)
plt.close(fig)
print(f"Plot saved → {OUT_FILE}")
gc.collect()


# # 定义输出路径
# excel_path_1min = os.path.join(OUT_DIR, "Tower_MultiLevel_TimeScatter2-1min.xlsx")

# # 创建Excel写入器
# with pd.ExcelWriter(excel_path_1min, engine='xlsxwriter') as writer:
#     # 遍历所有变量组（Temp/RH/HWS/VWS）
#     for var_group in VAR_GROUPS:
#         # 创建空列表存储所有数据
#         all_data = []
        
#         # 遍历所有高度和来源
#         for height in HEIGHTS:
#             for source, series in series_dict_1min[var_group][height].items():
#                 # 创建临时DataFrame
#                 temp_df = pd.DataFrame({
#                     'DateTime': series.index,
#                     'Value': series.values,
#                     'Height': height,
#                     'Source': source,
#                     'Variable': var_group
#                 })
#                 all_data.append(temp_df)
        
#         # 合并所有数据
#         combined_df = pd.concat(all_data, ignore_index=True)
        
#         # 将数据透视为宽格式（每个来源一列）
#         pivot_df = combined_df.pivot_table(
#             index=['DateTime', 'Height', 'Variable'],
#             columns='Source',
#             values='Value'
#         ).reset_index()
        
#         # 重命名列名和索引
#         pivot_df.columns.name = None
#         pivot_df = pivot_df.rename(columns={'DateTime': '时间 (UTC)'})
        
#         # 按时间排序
#         pivot_df.sort_values('时间 (UTC)', inplace=True)
        
#         # 保存到Excel工作表
#         sheet_name = f"{var_group}_1min"
#         pivot_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
#         # 获取工作簿和工作表对象用于格式设置
#         workbook = writer.book
#         worksheet = writer.sheets[sheet_name]
        
#         # 设置列宽
#         for i, col in enumerate(pivot_df.columns):
#             # 时间列设置更宽
#             if '时间' in col:
#                 worksheet.set_column(i, i, 25)
#             else:
#                 worksheet.set_column(i, i, 15)
                
#         # 添加筛选器
#         worksheet.autofilter(0, 0, 0, len(pivot_df.columns)-1)

# print(f"1分钟数据已成功导出到: {excel_path_1min}")


"""
Tower multi-level time–series (lines) + inter‑dataset scatter matrix

60min
-------------------------------------------------------------------
"""
#####
# 重置所有参数为默认值
plt.rcParams.update(plt.rcParamsDefault)


import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.gridspec import GridSpec
from matplotlib.ticker import MultipleLocator, AutoMinorLocator, FixedLocator

# --------------------------------------------------
# 0. Constants & rcParams
# --------------------------------------------------
HEIGHTS = [280, 200, 140, 80]
VAR_GROUPS = ["Temp", "RH", "HWS", "VWS"]
COLMAP = dict(MWR="red", 
              LiDAR="darkorange",
              TowerAvg="black", 
              TowerTurb="slategray", 
              CRA="magenta", 
              ERA5="blue", 
              )

DATA_START = pd.Timestamp("2024-06-05 00:00")
DATA_END   = pd.Timestamp("2024-08-11 00:00")
# 数据筛选窗口（只把这段时间的数据读进来、做 QC）
PLOT_START = pd.Timestamp("2024-08-05 01:30")
PLOT_END   = pd.Timestamp("2024-08-11 00:00")
# ==== 新增：两个不连续的绘图时间窗口（自己改成想要的时间段） ====
SEG1_START = pd.Timestamp("2024-06-06 00:00")
SEG1_END   = pd.Timestamp("2024-06-10 00:00")
SEG2_START = pd.Timestamp("2024-08-05 01:30")
SEG2_END   = pd.Timestamp("2024-08-11 00:00")

ALPHA_MAP = {
    'MWR'      : 0.4,
    'TowerAvg' : 0.4,
    'TowerTurb': 0.4,
    'CRA'      : 0.6,
    'ERA5'     : 0.6,
    'LiDAR'    : 0.4,
}
ALPHA_MAP2 = {
    'MWR'      : 1,
    'TowerAvg' : 1,
    'TowerTurb': 0.5,
    'CRA'      : 1,
    'ERA5'     : 1,
    'LiDAR'    : 0.5,
}

LABEL_MAP = {
    'MWR'      : 'MWR',
    'LiDAR'    : 'DWL',
    'TowerAvg' : 'AWS',
    'TowerTurb': 'ECM',#eddy covariance flux measurement system
    'CRA'      : 'CRA',
    'ERA5'     : 'ERA5',
}
# axis limits for each variable
VAR_RANGE = {
    'Temp': (288, 310),
    'RH'  : (10, 103),
    'HWS' : (-0.99, 14),
    'VWS' : (-4, 4)
}
title_map = {
    'Temp': 'Temperature (T)',
    'RH'  : 'Relative Humidity (RH)',
    'HWS' : 'Horizontal Wind Speed (HWS)',
    'VWS' : 'Vertical Wind Speed (VWS)'
}
OUT_DIR = r"E:\Beijing2024\出图TIF"; os.makedirs(OUT_DIR, exist_ok=True)
OUT_FILE = os.path.join(OUT_DIR, "Tower_MultiLevel_TimeScatter2-60min3.tif")
plt.rcParams.update({    
    'font.size': 50,
    'font.family': 'sans-serif',                # 先指定家族
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     # 把 Arial 设为首选
    'axes.titlepad': 8,
    'legend.fontsize': 50,
    'axes.labelsize': 50,
    'xtick.labelsize': 50,
    'ytick.labelsize': 50,
    'xtick.major.size': 30,
    'xtick.minor.size': 15,
    'ytick.major.size': 30,
    'ytick.minor.size': 15,
    'xtick.direction': 'out',
    'ytick.direction': 'out',
    'xtick.major.width': 6,
    'xtick.minor.width': 6,
    'ytick.major.width': 6,
    'ytick.minor.width': 6,
})
# --------------------------------------------------
# 1. Helpers
# --------------------------------------------------

def ensure_dt(ser: pd.Series) -> pd.Series:
    ser = ser.copy(); ser.index = pd.to_datetime(ser.index); return ser.sort_index()

def pivot_height(df: pd.DataFrame, *, idx: str, col: str, val: str, h: int) -> pd.Series:
    tmp = df.reset_index(); tmp[idx] = pd.to_datetime(tmp[idx])
    return ensure_dt(tmp.pivot(index=idx, columns=col, values=val).loc[:, h])

def align_x_to_ref(ref: pd.Series, other: pd.Series) -> pd.DataFrame:
    return pd.concat([ref, other], axis=1, join='inner').dropna(how='any')

# --------------------------------------------------
# 2.  Build Series dictionary
# --------------------------------------------------
turb_calibrated_60min['rh'] = turb_calibrated_60min['rh'].clip(upper=100)
series_dict = {v: {h: {} for h in HEIGHTS} for v in VAR_GROUPS}

# ---- Temperature -----------------------------------------------------------
for h in HEIGHTS:
    series_dict['Temp'][h]['TowerAvg']  = pivot_height(tieta_60min, idx='timestamps', col='H(m)', val='T(K)_tieta', h=h)
    series_dict['Temp'][h]['TowerTurb'] = pivot_height(turb_calibrated_60min, idx='timestamps', col='Height', val='T (K)', h=h)
    series_dict['Temp'][h]['MWR'] = ensure_dt(mean_60min_profiles.xs('T(K)', level=0, axis=1)[h])
    series_dict['Temp'][h]['CRA']       = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['Temp (K)'])
    series_dict['Temp'][h]['ERA5']      = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['t'])

# ---- Relative Humidity -----------------------------------------------------
for h in HEIGHTS:
    series_dict['RH'][h]['TowerAvg']  = pivot_height(tieta_60min, idx='timestamps', col='H(m)', val='RH_tieta', h=h)
    series_dict['RH'][h]['TowerTurb'] = pivot_height(turb_calibrated_60min, idx='timestamps', col='Height', val='rh', h=h)
    series_dict['RH'][h]['MWR']       = ensure_dt(mean_60min_profiles.xs('RH(%)', level=0, axis=1)[h])
    series_dict['RH'][h]['CRA']       = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['rh (%)'])
    series_dict['RH'][h]['ERA5']      = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['r'])

# ---- Horizontal Wind Speed -------------------------------------------------
hws_qc_60min = (
    hws_avg["60min"]
    .rename(columns=lambda x: x.replace('m', ''))  # 先处理列名（移除'm'）
    .stack()
    .reset_index()
    .rename(columns={'level_1': 'height', 0: 'hws'})
    .astype({'height': int})  # 确保height为整数
)

for h in HEIGHTS:
    series_dict['HWS'][h]['LiDAR']    = ensure_dt(
        hws_qc_60min.query('height == @h').set_index('level_0')['hws'])
    series_dict['HWS'][h]['TowerAvg'] = pivot_height(tieta_60min, idx='timestamps', col='H(m)', val='ws_tieta', h=h)
    series_dict['HWS'][h]['TowerTurb']= pivot_height(turb_calibrated_60min, idx='timestamps', col='Height', val='hws_corrected', h=h)
    series_dict['HWS'][h]['CRA']      = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['hws(m/s)'])
    series_dict['HWS'][h]['ERA5']     = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['hws(m/s)'])

# ---- Vertical Wind Speed ---------------------------------------------------
vws_qc_60min = (
    vws_avg["60min"]
    .rename(columns=lambda x: x.replace('m', ''))  # 先处理列名（移除'm'）
    .stack()
    .reset_index()
    .rename(columns={'level_1': 'height', 0: 'vws'})
    .astype({'height': int})  # 确保height为整数
)

for h in HEIGHTS:
    series_dict['VWS'][h]['LiDAR']    = ensure_dt(
        vws_qc_60min.reset_index().query('height == @h').set_index('level_0')['vws'])
    series_dict['VWS'][h]['TowerTurb'] = pivot_height(turb_calibrated_60min, idx='timestamps', col='Height', val='W (m/s)', h=h)
    series_dict['VWS'][h]['CRA']       = ensure_dt(
        CRA_interp_to_65_80_280.query('height_plus_alt == @h').set_index('BJT')['vws(10⁻² {m/s})'] * -0.35)
    series_dict['VWS'][h]['ERA5']      = ensure_dt(
        ERA5_interp_to_65_80_280.query('height == @h').set_index('BJT')['vws(10⁻² {m/s})'] * -35)

for v in VAR_GROUPS:
    for h in HEIGHTS:
        for src, ser in series_dict[v][h].items():
            series_dict[v][h][src] = ser.loc[DATA_START:DATA_END]

series_dict_60min = series_dict

# ==== 定义两个时间窗口 ====
WIN1_START = pd.Timestamp("2024-06-06 00:00")   # 新增窗口：6月6–10日
WIN1_END   = pd.Timestamp("2024-06-10 00:00")

WIN2_START = PLOT_START                         # 原来的 8 月窗口
WIN2_END   = PLOT_END

def make_month_first_formatter(first_date):
    """
    第一个 tick 显示 'Jun 06'，后面的 tick 只显示 '07', '08' ...
    first_date: 这个时间段的起始日期（比如 WIN1_START 或 WIN2_START）
    """
    first_date = pd.to_datetime(first_date).normalize()

    def _fmt(x, pos=None):
        dt = mdates.num2date(x)          # matplotlib 内部浮点 -> datetime
        d  = pd.Timestamp(dt).normalize()

        if d == first_date:
            # 第一个刻度：英文月份缩写 + 日
            return d.strftime('%b %d')   # 例如 'Jun 06'
        else:
            # 其它刻度：只显示日
            return d.strftime('%d')      # 例如 '07'
    return FuncFormatter(_fmt)

# # 定义输出路径
# excel_path = os.path.join(OUT_DIR, "Tower_MultiLevel_TimeScatter2-60min.xlsx")

# # 创建Excel写入器
# with pd.ExcelWriter(excel_path, engine='xlsxwriter') as writer:
#     # 遍历所有变量组（Temp/RH/HWS/VWS）
#     for var_group in VAR_GROUPS:
#         # 创建空列表存储所有数据
#         all_data = []
        
#         # 遍历所有高度和来源
#         for height in HEIGHTS:
#             for source, series in series_dict_60min[var_group][height].items():
#                 # 创建临时DataFrame
#                 temp_df = pd.DataFrame({
#                     'DateTime': series.index,
#                     'Value': series.values,
#                     'Height': height,
#                     'Source': source,
#                     'Variable': var_group
#                 })
#                 all_data.append(temp_df)
        
#         # 合并所有数据
#         combined_df = pd.concat(all_data, ignore_index=True)
        
#         # 将数据透视为宽格式（每个来源一列）
#         pivot_df = combined_df.pivot_table(
#             index=['DateTime', 'Height', 'Variable'],
#             columns='Source',
#             values='Value'
#         ).reset_index()
        
#         # 重命名列名和索引
#         pivot_df.columns.name = None
#         pivot_df = pivot_df.rename(columns={'DateTime': '时间 (UTC)'})
        
#         # 按时间排序
#         pivot_df.sort_values('时间 (UTC)', inplace=True)
        
#         # 保存到Excel工作表
#         sheet_name = f"{var_group}_60min"
#         pivot_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
#         # 获取工作簿和工作表对象用于格式设置
#         workbook = writer.book
#         worksheet = writer.sheets[sheet_name]
        
#         # 设置列宽
#         for i, col in enumerate(pivot_df.columns):
#             # 时间列设置更宽
#             if '时间' in col:
#                 worksheet.set_column(i, i, 25)
#             else:
#                 worksheet.set_column(i, i, 15)
                
#         # 添加筛选器
#         worksheet.autofilter(0, 0, 0, len(pivot_df.columns)-1)

# print(f"数据已成功导出到: {excel_path}")

# --------------------------------------------------
# 3. Plot with broken x-axis
# --------------------------------------------------
fig = plt.figure(figsize=(74, 60))  # , constrained_layout=True
outer = GridSpec(2, 2, figure=fig, hspace=0.22, wspace=0.125)

for g_idx, var in enumerate(VAR_GROUPS):
    row, col = divmod(g_idx, 2)
    gspec = outer[row, col].subgridspec(
        len(HEIGHTS), 2,
        width_ratios=[3, 1.2],
        wspace=0.18,
        hspace=0.55
    )

    # column header
    ax_head = fig.add_subplot(gspec[0, 0])
    ax_head.axis('off')
    if var == 'VWS':
        ax_head.set_title(
            "Vertical Wind Speed (VWS)\n"           # 第一行普通文字，自动是正体
            r"$(VWS_{80/140/200/280}^{\mathrm{ERA5,CRA}}\times 35)$",  # 第二行数学公式，保留上下标
            pad=35, x=0.5, fontweight='bold', fontsize=60  # 整体加粗，第一行肯定是粗体
        )
    else:
        ax_head.set_title(title_map.get(var, var),
                          pad=35, fontweight='bold', fontsize=60)

    vmin, vmax = VAR_RANGE[var]
    
    for r_idx, h in enumerate(HEIGHTS):
        # ---------------- Line with broken x-axis ------------------
        # 左列再拆成两个子轴：左=6 月，右=8 月
        line_gs = gspec[r_idx, 0].subgridspec(1, 2, width_ratios=[1, 1],
                                              wspace=0.05)
        ax_left  = fig.add_subplot(line_gs[0, 0])
        ax_right = fig.add_subplot(line_gs[0, 1], sharey=ax_left)
    
        # 画曲线：两个时间段
        for src, ser in series_dict[var][h].items():
            if var == 'VWS' and src == 'TowerAvg':
                continue
            if ser is None or ser.empty:
                continue
    
            color = COLMAP.get(src, 'k')
            alpha = ALPHA_MAP2.get(src, 0.6)
            lw    = 4
    
            s1 = ser.loc[WIN1_START:WIN1_END]
            s2 = ser.loc[WIN2_START:WIN2_END]
    
            if not s1.empty:
                ax_left.plot(s1.index, s1.values,
                             color=color, alpha=alpha, lw=lw)
            if not s2.empty:
                ax_right.plot(s2.index, s2.values,
                              color=color, alpha=alpha, lw=lw)
    
        # x 范围
        ax_left.set_xlim(WIN1_START, WIN1_END)
        ax_right.set_xlim(WIN2_START, WIN2_END)
    
        # y 范围 & 刻度（只设置在左轴上）
        ax_left.set_ylim(vmin, vmax)
        if var == 'Temp':
            ax_left.yaxis.set_major_locator(MultipleLocator(4))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
        elif var == 'RH':
            ax_left.yaxis.set_major_locator(MultipleLocator(20))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
        elif var == 'HWS':
            ax_left.yaxis.set_major_locator(MultipleLocator(4))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
        else:  # VWS
            for ax in (ax_left, ax_right):
                ax.axhline(
                    0,
                    color='black',
                    linewidth=6,
                    zorder=1
                )
            ax_left.yaxis.set_major_locator(MultipleLocator(2))
            ax_left.yaxis.set_minor_locator(AutoMinorLocator(2))
    
        # 右轴隐藏 y 轴刻度和标签（只保留一个 y 轴）
        ax_right.tick_params(
            axis='y', which='both',
            left=False, right=False,
            labelleft=False, labelright=False
        )
    
        # 高度标签（写在左轴上）
        ax_left.text(
            0.1, 1.05, f"{h} m",
            transform=ax_left.transAxes,
            ha='center', va='bottom',
            fontsize=56, fontweight='bold'
        )
    
        # y-label：第三行写单位
        if r_idx == 2:
            if var == 'Temp':
                ylab = "T (K)"
            elif var == 'RH':
                ylab = "RH (%)"
            elif var == 'HWS':
                ylab = "HWS (m/s)"
            elif var == 'VWS':
                ylab = "VWS (m/s)"
            else:
                ylab = ""
    
            ax_left.text(
                -0.2, 1.5, ylab,
                transform=ax_left.transAxes,
                fontsize=50, ha='right', va='center',
                rotation=90, fontweight="bold"
            )
    
        # 边框 & x 轴刻度格式
        for ax in (ax_left, ax_right):
            for sp in ['top', 'right']:
                ax.spines[sp].set_visible(False)
            for sp in ['bottom', 'left']:
                ax.spines[sp].set_linewidth(6)
    
            # ---- x 轴刻度：DayLocator + 自定义 formatter ----
            # 主刻度：每天一个
            ax.xaxis.set_major_locator(mdates.DayLocator())
            # 副刻度：每天中午 12 点
            ax.xaxis.set_minor_locator(mdates.HourLocator(byhour=[12]))
    
        # 左边（第一个时间窗）：第一个刻度显示 'Jun 06'，后面 '07' '08' ...
        ax_left.xaxis.set_major_formatter(
            make_month_first_formatter(WIN1_START)
        )
    
        # 右边（第二个时间窗）：第一个刻度显示 'Aug 10'，后面 '11' ...
        ax_right.xaxis.set_major_formatter(
            make_month_first_formatter(WIN2_START)
        )
    
        # broken axis：只在底部画两个斜杠
        ax_left.spines['right'].set_visible(False)
        ax_right.spines['left'].set_visible(False)
    
        d = .015
        kwargs = dict(color='k', clip_on=False, linewidth=4)
    
        # 左轴右侧（底部）斜杠
        ax_left.plot(
            (1 - d, 1 + d), (-d, +d),
            transform=ax_left.transAxes, **kwargs
        )
        # 右轴左侧（底部）斜杠
        ax_right.plot(
            (-d, +d), (-d, +d),
            transform=ax_right.transAxes, **kwargs
        )
    
        # 最后一行显示 x 轴标签，其余行隐藏
        if r_idx == len(HEIGHTS) - 1:
            # ====== 新增部分开始 ======
            # 在每个x轴下方添加月份标签
            month1 = WIN1_START.strftime('%b') # 获取月份缩写，如 'Jun'
            month2 = WIN2_START.strftime('%b') # 获取月份缩写，如 'Aug'
    
            ax_left.text(
                0.04, -0.27, month1,
                transform=ax_left.transAxes,
                ha='center', va='top',
                fontsize=50, fontweight='bold'
            )
            ax_right.text(
                0.085, -0.27, month2,
                transform=ax_right.transAxes,
                ha='center', va='top',
                fontsize=50, fontweight='bold'
            )    
            ax_left.set_xlabel("Datetime (BJT)", fontsize=50, x=1, labelpad = 45)
            ax_left.tick_params(labelbottom=True)
            ax_right.tick_params(labelbottom=True)
        else:
            ax_left.tick_params(labelbottom=False)
            ax_right.tick_params(labelbottom=False)

        # ---------------- Scatter 右侧 --------------
        ax_scat = fig.add_subplot(gspec[r_idx, 1])

        # 确定散点参考 ref
        ref = series_dict[var][h].get('TowerAvg')
        if ref is not None and not ref.empty:
            ref_source = 'AWS'
        else:
            ref = series_dict[var][h].get('TowerTurb')
            if ref is not None and not ref.empty:
                ref_source = 'AWS'
            else:
                ref = None
                ref_source = 'Other Source'
                for _s in series_dict[var][h].values():
                    if _s is not None and not _s.empty:
                        ref = _s
                        break

        # 只有最底行写 Ref 文本
        if r_idx == len(HEIGHTS) - 1:
            ax_scat.text(
                0.5, -0.35, f'Ref.: {ref_source}',
                transform=ax_scat.transAxes,
                fontsize=50, ha='center', va='top'
            )
        else:
            ax_scat.tick_params(labelbottom=False)

        # 1:1 线 & 基本范围
        ax_scat.plot(
            [vmin, vmax], [vmin, vmax],
            ls=(0, (5, 10)), lw=3, color='black',
            zorder=0, alpha=0.5
        )
        ax_scat.set_xlim(vmin, vmax)
        ax_scat.set_ylim(vmin, vmax)

        # 画散点
        for src, ser in series_dict[var][h].items():
            if src == 'TowerAvg':
                continue
            if var == 'VWS' and src == 'TowerTurb':
                continue
            if ref is None or ser is None or ser.empty:
                continue

            aligned = align_x_to_ref(ref, ser)
            if aligned.empty:
                continue

            ax_scat.scatter(
                aligned.iloc[:, 0], aligned.iloc[:, 1],
                s=200, linewidths=0, marker='o',
                color=COLMAP.get(src, 'k'),
                alpha=ALPHA_MAP.get(src, 0.5)
            )

        # 根据变量调整散点轴刻度
        if var == 'HWS':
            ax_scat.set_xlim(0, 16)
            ax_scat.xaxis.set_major_locator(MultipleLocator(4))
            ax_scat.xaxis.set_minor_locator(MultipleLocator(2))
            ax_scat.yaxis.set_major_locator(MultipleLocator(4))
            ax_scat.yaxis.set_minor_locator(MultipleLocator(2))

        elif var == 'VWS':
            ax_scat.set_xlim(-3, 3)
            ax_scat.xaxis.set_major_locator(MultipleLocator(2))
            ax_scat.xaxis.set_minor_locator(MultipleLocator(1))
            ax_scat.yaxis.set_major_locator(MultipleLocator(2))
            ax_scat.yaxis.set_minor_locator(MultipleLocator(1))

        elif var == 'RH':
            ax_scat.xaxis.set_major_locator(FixedLocator([30, 65, 100]))
            ax_scat.xaxis.set_minor_locator(AutoMinorLocator(2))
            ax_scat.yaxis.set_major_locator(FixedLocator([30, 65, 100]))
            ax_scat.yaxis.set_minor_locator(AutoMinorLocator(2))

        elif var == 'Temp':
            ax_scat.xaxis.set_major_locator(MultipleLocator(4))
            ax_scat.xaxis.set_minor_locator(AutoMinorLocator(2))
            ax_scat.yaxis.set_major_locator(MultipleLocator(4))
            ax_scat.yaxis.set_minor_locator(AutoMinorLocator(2))

        # 散点图边框
        for sp in ['top', 'right']:
            ax_scat.spines[sp].set_visible(False)
        for sp in ['bottom', 'left']:
            ax_scat.spines[sp].set_linewidth(6)

# --------------------------------------------------
# 4. Legend & save
# --------------------------------------------------
handles = [
    plt.Line2D([], [],
               color=COLMAP[k],
               lw=15,
               label=LABEL_MAP[k])
    for k in COLMAP
]
fig.legend(
    handles, [h.get_label() for h in handles],
    loc='upper center', ncol=3, frameon=False,
    title_fontsize=80, bbox_to_anchor=(0.5, 0.935)
)

fig.savefig(
    OUT_FILE, dpi=300, format='tif',
    transparent=True, bbox_inches='tight',
    pil_kwargs={'compression': 'tiff_adobe_deflate'}
)
plt.close(fig)
print(f"Plot saved → {OUT_FILE}")
gc.collect()

"""
Generate **side‑by‑side horizontal box + jitter** plots for error statistics of
four variables (rows) across three temporal resolutions (columns).
"""


# -----------------------------------------------------------------------------
# 0.  Configuration
# -----------------------------------------------------------------------------
#0607:0916-1500

plt.rcParams.update(plt.rcParamsDefault)
OUT_DIR = Path(r"E:\Beijing2024\出图TIF")
OUT_DIR.mkdir(parents=True, exist_ok=True)
FIGSIZE = (16, 18)  # inches (width, height)

# ─ Time-window for filtering data ─
EXCLUDE_WINDOWS = [
    (pd.Timestamp("2024-08-09 12:46"), pd.Timestamp("2024-08-10 00:03")),
    (pd.Timestamp("2024-06-07 09:16"), pd.Timestamp("2024-06-07 15:00")),
    (pd.Timestamp("2024-08-10 16:20"), pd.Timestamp("2024-08-10 17:30"))
]

plt.rcParams.update({
    'font.family': 'sans-serif',                # 先指定家族
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     # 把 Arial 设为首选
    "font.size": 22,
    "axes.titlesize": 22,
    "axes.labelsize": 22,
    "xtick.labelsize": 22,
    "ytick.labelsize": 22,
    "legend.fontsize": 22,
    'xtick.major.width': 2,
    'xtick.minor.width': 2,
    'ytick.major.width': 2,
    'ytick.minor.width': 2,
})

COLMAP = {
    "MWR": "red",
    "TowerTurb": "slategray",
    "CRA": "magenta",
    "ERA5": "blue",
    "LiDAR": "darkorange",
}
DISPLAY_NAMES = {
    "TowerTurb": "ECFMS",   
    "LiDAR":"DWL"    
}
VAR_GROUPS = ["Temp", "RH", "HWS", "VWS"]
ROW_LABELS = {
    "Temp": "Temperature (K)\nRef. : AWS",
    "RH": "Relative Humidity (%)\nRef. : AWS",
    "HWS": "Horizontal Wind Speed (m/s)\nRef. : AWS",
    "VWS": "Vertical Wind Speed (m/s)\nRef. : ECM",
}
COL_LABELS = ["1min",  "60min"]#"30min",

# -----------------------------------------------------------------------------
# Expected data in memory
# -----------------------------------------------------------------------------
series_dict_1min  = globals().get("series_dict_1min")
#series_dict_30min = globals().get("series_dict_30min")
series_dict_60min = globals().get("series_dict_60min")
RESOLUTION_DICT = {
    "1min": series_dict_1min,
    #"30min": series_dict_30min,
    "60min": series_dict_60min,
}

REF_MAP = {"Temp": "TowerAvg", "RH": "TowerAvg", "HWS": "TowerAvg", "VWS": "TowerTurb"}
SRC_MAP = {
    "Temp": ["TowerTurb", "MWR", "CRA", "ERA5"],
    "RH":   ["TowerTurb", "MWR", "CRA", "ERA5"],
    "HWS":  ["LiDAR", "TowerTurb", "ERA5", "CRA"],
    "VWS":  ["LiDAR", "ERA5", "CRA"],
}

# -----------------------------------------------------------------------------
# Helper: build long‑form dataframe of errors
# -----------------------------------------------------------------------------
def exclude_time_windows(series, windows):
    """逐个时间段排除数据"""
    for start, end in windows:
        series = series[(series.index < start) | (series.index > end)]
    return series

def align_diff(ref: pd.Series, tgt: pd.Series, *, inside: bool) -> pd.Series:
    """Return tgt−ref for either inside (True) or outside (False) time window."""
    # slice either inside or outside window
    if inside:
        ref = ref.loc[TIME_START:TIME_END]
        tgt = tgt.loc[TIME_START:TIME_END]
    else:
        ref = ref.loc[(ref.index < TIME_START) | (ref.index > TIME_END)]
        tgt = tgt.loc[(tgt.index < TIME_START) | (tgt.index > TIME_END)]
    aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
    return aligned.iloc[:, 1] - aligned.iloc[:, 0]
# ---------------------------------------------------------------------------
# 专门给 1 min 用的 source 映射   ★ CHG ★
# ---------------------------------------------------------------------------
SRC_MAP_1MIN = {
    "Temp": ["TowerTurb", "MWR"],
    "RH":   ["TowerTurb", "MWR"],
    "HWS":  ["LiDAR", "TowerTurb"],
    "VWS":  ["LiDAR"],
}

XLIMS_OUTSIDE = {"Temp": (-7, 7), "RH": (-35, 50),
                  "HWS": (-10, 7.5),     "VWS": (-7.5, 2.7)}
STEP_OUTSIDE  = {"Temp": 2, "RH": 20, "HWS": 5, "VWS": 2.5}
# ---------------------------------------------------------------------------
# Helper: build long-form dataframe of errors   ★ NEW ★
# ---------------------------------------------------------------------------
def build_error_df() -> pd.DataFrame:
    """
    统计 **全时段** 误差。  
    - 对温度(T)的 MWR，在 TIME_START–TIME_END 内的数据会被剔除。
    """
    recs = []
    for res, sdict in RESOLUTION_DICT.items():
        if sdict is None:
            continue

        for var in VAR_GROUPS:
            ref_key  = REF_MAP[var]
            src_list = SRC_MAP_1MIN[var] if res == "1min" else SRC_MAP[var]

            for h, series_map in sdict[var].items():
                if ref_key not in series_map:              # 参考值不存在
                    continue
                ref = series_map[ref_key]

                for src in src_list:
                    if src == ref_key or src not in series_map:
                        continue
                    tgt = series_map[src]

                    # ─── 仅温度-MWR 需剔除指定时段 ───
                    if var == "Temp" and src == "MWR":
                        tgt = exclude_time_windows(tgt, EXCLUDE_WINDOWS)

                    # 对齐并求差
                    aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
                    diff    = aligned.iloc[:, 1] - aligned.iloc[:, 0]

                    recs.extend({
                        "resolution": res,
                        "variable"  : var,
                        "source"    : src,
                        "error"     : d,
                    } for d in diff.values)

    return pd.DataFrame.from_records(recs)

def summarize_errors(df: pd.DataFrame) -> pd.DataFrame:
    """
    计算每个 (resolution, variable, source) 组合的
    mean / median / q25 / q75 / IQR / whisker_low / whisker_high。
    与 matplotlib.boxplot 默认设置保持一致：
    whisker = q1 ± 1.5 × IQR（再截断到全局 min/max）。
    """
    grp = df.groupby(["resolution", "variable", "source"])["error"]

    summary = grp.agg(
        count  = "size",
        mean   = "mean",
        median = "median",
        q25    = lambda x: x.quantile(0.25),
        q75    = lambda x: x.quantile(0.75),
        min    = "min",
        max    = "max",
    )

    summary["iqr"] = summary.q75 - summary.q25
    summary["whisker_low"]  = (summary.q25 - 1.5 * summary.iqr).clip(lower=summary["min"])
    summary["whisker_high"] = (summary.q75 + 1.5 * summary.iqr).clip(upper=summary["max"])

    # 调整列顺序更直观
    summary = summary[[
        "count", "mean", "median", "q25", "q75",
        "iqr", "whisker_low", "whisker_high", "min", "max"
    ]]

    return summary
# -----------------------------------------------------------------------------
# Plotting
# -----------------------------------------------------------------------------
from matplotlib.lines import Line2D
def make_plot(df: pd.DataFrame, outfile: Path, top_caption: str, bottom_caption: str, xlims: dict, steps: dict):
    
    fig, axes = plt.subplots(nrows=len(VAR_GROUPS), ncols=len(COL_LABELS) , figsize=FIGSIZE, sharex=False, sharey=False)
    fig.subplots_adjust(wspace=0.4, hspace=0.4)
    box_offset = 0.5  # left of tick
    scatter_offset = -0.5  # right of tick (more separation than before)
    box_width = 0.5
    scatter_size = 18
    scatter_alpha = 1
    row_spacing_factor = 1.7
    
    for r, var in enumerate(VAR_GROUPS):
        # Row labels -------------------------------------------------------------
        fig.text(0.08, 0.82 - r * 0.215, ROW_LABELS[var], rotation=90, va="center", ha="center",fontsize = 22)
    
        for c, res in enumerate(COL_LABELS):
            if r == 0:
                fig.text(0.13 + (c + 0.5) * 0.92 / len(COL_LABELS), 0.90, res, ha="center", va="bottom",fontsize = 22)
    
            ax = axes[r, c]
            subset = df.query("variable == @var and resolution == @res")
            # if subset.empty:
            #     ax.text(0.5, 0.5, "No data", ha="center", va="center", alpha=0.6)
            #     ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
            #     continue
    
            # ★ CHG ★
            order = SRC_MAP_1MIN[var] if res == "1min" else SRC_MAP[var]
            display_order = [DISPLAY_NAMES.get(s, s) for s in order]

            y_pos = np.arange(len(order)) * row_spacing_factor   # ← greater spacing
            for i, src in enumerate(order):
                e = subset.loc[subset["source"] == src, "error"].values
                if e.size == 0:
                    continue
                ax.boxplot(e, zorder=2,
                           vert=False,
                           positions=[y_pos[i] + box_offset], 
                           widths=box_width,#showmeans=True,
                           patch_artist=True, manage_ticks=False,
                           boxprops=dict(facecolor=COLMAP[src], alpha=1, linewidth=2.5),
                           #meanprops=dict(marker="^", markerfacecolor="white", markeredgecolor='crimson', markersize=10, linewidth=0),
                           flierprops=dict(marker="x", markeredgecolor=COLMAP[src], markerfacecolor="none",
                                    markersize=7, linewidth=0.5, alpha=scatter_alpha),
                           medianprops=dict(color="lime", linewidth=2), 
                           whiskerprops=dict(color="black", linewidth=2.5), 
                           capprops=dict(color="black", linewidth=2.5)
                           )
                mean_val = e.mean()
                ax.scatter(mean_val, y_pos[i] + scatter_offset, marker="^",zorder=3,edgecolors="white",
                           color='crimson', s=180, linewidths=2, alpha=1)  # mean
                jitter_y = np.random.normal(y_pos[i] + scatter_offset, 0.08, size=len(e))
                ax.scatter(e, jitter_y, s=scatter_size, color=COLMAP[src], alpha=scatter_alpha, linewidths=0)
    
            ax.set_yticks(y_pos)
            ax.set_yticklabels(display_order)
            ax.set_ylim(y_pos[0] - 1, y_pos[-1] + 1)
            ax.grid(axis="x", linestyle=(0, (5, 10)), alpha=0.6,color = 'black',linewidth = 2.5,zorder=3)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_visible(True)
            ax.spines["left"].set_linewidth(2)
            ax.spines["bottom"].set_linewidth(2)
            # ─── x-limits & ticks ──────────────────────────────────────────────
            ax.set_xlim(*xlims[var])                                # ★ CHG ★
            major = steps[var]                                      # ★ CHG ★
            ax.xaxis.set_major_locator(mticker.MultipleLocator(major))
            ax.xaxis.set_minor_locator(mticker.MultipleLocator(major / 2))
            ax.tick_params(axis="x", which="major", length=9)       # ★ CHG ★
            ax.tick_params(axis="x", which="minor", length=4.5)     # ★ CHG ★
            ax.tick_params(axis="y", which="major", length=9)
        
# -----------------------------------------------------------------------------
# Legend
# -----------------------------------------------------------------------------

    handles = [mlines.Line2D([], [], color=COLMAP[s], marker="s", linestyle="", markersize=16,
                             markerfacecolor=COLMAP[s], label=DISPLAY_NAMES.get(s, s))
               for s in COLMAP]
    # 把 '\n' → ' '，这样 legend 里就是单行
    legend_labels = [h.get_label().replace('\n', ' ') for h in handles]
    fig.legend(handles,legend_labels, loc="upper center", ncol=6, frameon=False, bbox_to_anchor=(0.55, 0.99))
    
    handle_mean   = Line2D([], [], marker="^", markersize=14,
                           linestyle="", markeredgecolor="white",
                           markerfacecolor="crimson", color="crimson",
                           label="Mean (▲)")
    handle_median = Line2D([], [], linestyle="-", linewidth=4,
                           color="lime", label="Median (—)")
    
    fig.legend([handle_mean, handle_median],
               ["Mean", "Median"],
               loc="lower left",
               bbox_to_anchor=(0.04, 0.02),   # ← 左下角留一点边距
               frameon=False, fontsize=22)
    # fig.text(0.58, 0.045,
    #          "Observation Biases = Ground-based Remote Sensing Datasets/Reanalysis - Ref.",
    #          ha="center", va="bottom", fontsize=22)
    
    if top_caption:
        fig.text(0.55, 0.93, top_caption, ha="center", va="bottom",
                 fontsize=22, fontstyle="italic", fontweight="bold")
        
    fig.tight_layout(rect=[0.08, 0.05, 1, 0.92])
    fig.savefig(outfile, dpi=600, format="tif", transparent=True, bbox_inches="tight", pil_kwargs={"compression": "tiff_adobe_deflate"})
    print("Saved →", outfile)
    
# # ─────────────────────────────────────────────────────────────────────────────
# # 4.  Build two dataframes & plot
# # ─────────────────────────────────────────────────────────────────────────────

# print("Building DF (inside window)...")
# df_in = build_error_df(inside=True)
# make_plot(
#     df_in,
#     OUT_DIR / "Error_Eval_BoxJitter_WINDOW.tif",
#     top_caption=f"Rainstorm Events: {TIME_START.strftime('%Y-%m-%d %H:%M')} — {TIME_END.strftime('%Y-%m-%d %H:%M')}(BJT) Included",  ### CHG ###
#     bottom_caption=f"Bias inside {TIME_START} — {TIME_END}",
#     xlims=XLIMS_INSIDE, steps=STEP_INSIDE
# )

# print("Building DF (outside window)...")
# df_out = build_error_df(inside=False)
# make_plot(
#     df_out,
#     OUT_DIR / "Error_Eval_BoxJitter_EXCLUDE.tif",
#     top_caption=f"Rainstorm Events: {TIME_START.strftime('%Y-%m-%d %H:%M')} — {TIME_END.strftime('%Y-%m-%d %H:%M')}(BJT) Excluded",  ### CHG ###
#     bottom_caption=f"Bias excluding {TIME_START} — {TIME_END}",
#     xlims=XLIMS_OUTSIDE, steps=STEP_OUTSIDE
# )

# ─────────────────────────────────────────────────────────────
# 4.  Build dataframe & plot   ★ ONLY ONCE ★
# ─────────────────────────────────────────────────────────────
print("Building full-period dataframe …")
df_all = build_error_df()

make_plot(
    df_all,
    OUT_DIR / "Error_Eval_BoxJitter_window.tif",
    top_caption = (#f"MWR-Temp data between "
                   #f"{TIME_START:%Y-%m-%d %H:%M} — {TIME_END:%m-%d %H:%M} (BJT) "
                   "Rainfall Period analysis"),
    bottom_caption = "",
    xlims = XLIMS_OUTSIDE,   # ← 继续用你原先的一套范围
    steps = STEP_OUTSIDE
)

stats_df = summarize_errors(df_all)

# —— 写 Excel：每个 variable 单独一个 sheet，更方便查看 ——
# out_excel = OUT_DIR / "error_stats.xlsx"
# with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
#     for var, sub in stats_df.groupby(level="variable"):
#         # 取消 variable 这一层索引，保留其余两层可读性更好
#         sub.droplevel("variable").to_excel(writer, sheet_name=var)

# print(f"Statistics saved → {out_excel}")





#####这个代码一次能生成全部的，包含降水时间段的和不包含降水时间段的三张图####
#####包含特定时间窗的绘图####
OUT_DIR = Path(r"E:\Beijing2024\出图TIF")
OUT_DIR.mkdir(parents=True, exist_ok=True)
FIGSIZE = (16, 18)  # inches (width, height)
# ─ Time-window for filtering data ─
EXCLUDE_WINDOWS = [
    (pd.Timestamp("2024-08-09 12:46"), pd.Timestamp("2024-08-10 00:03")),
    (pd.Timestamp("2024-06-07 09:16"), pd.Timestamp("2024-06-07 15:00")),
    (pd.Timestamp("2024-08-10 16:20"), pd.Timestamp("2024-08-10 17:30"))
]
plt.rcParams.update({
    'font.family': 'sans-serif',                # 先指定家族
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     # 把 Arial 设为首选
    "font.size": 22,
    "axes.titlesize": 22,
    "axes.labelsize": 22,
    "xtick.labelsize": 22,
    "ytick.labelsize": 22,
    "legend.fontsize": 22,
    'xtick.major.width': 2,
    'xtick.minor.width': 2,
    'ytick.major.width': 2,
    'ytick.minor.width': 2,
})
COLMAP = {
    "MWR": "red",
    "TowerTurb": "slategray",
    "CRA": "magenta",
    "ERA5": "blue",
    "LiDAR": "darkorange",
}
DISPLAY_NAMES = {
    "TowerTurb": "ECM",   
    "LiDAR":"DWL"
}
VAR_GROUPS = ["Temp", "RH", "HWS", "VWS"]
ROW_LABELS = {
    "Temp": "Temperature (K)\nRef. : AWS",
    "RH": "Relative Humidity (%)\nRef. : AWS",
    "HWS": "Horizontal Wind Speed (m/s)\nRef. : AWS",
    "VWS": "Vertical Wind Speed (m/s)\nRef. : ECM",
}
COL_LABELS = ["1min",  "60min"]#"30min",
# -----------------------------------------------------------------------------
# Expected data in memory
# -----------------------------------------------------------------------------
series_dict_1min  = globals().get("series_dict_1min")
#series_dict_30min = globals().get("series_dict_30min")
series_dict_60min = globals().get("series_dict_60min")
RESOLUTION_DICT = {
    "1min": series_dict_1min,
    #"30min": series_dict_30min,
    "60min": series_dict_60min,
}
REF_MAP = {"Temp": "TowerAvg", "RH": "TowerAvg", "HWS": "TowerAvg", "VWS": "TowerTurb"}
SRC_MAP = {
    "Temp": ["TowerTurb", "MWR", "CRA", "ERA5"],
    "RH":   ["TowerTurb", "MWR", "CRA", "ERA5"],
    "HWS":  ["LiDAR", "TowerTurb", "ERA5", "CRA"],
    "VWS":  ["LiDAR", "ERA5", "CRA"],
}
# -----------------------------------------------------------------------------
# Helper: build long‑form dataframe of errors
# -----------------------------------------------------------------------------
def include_time_windows(series, windows, inside=True):
    """根据inside参数决定是包含还是排除指定时间窗口内的数据"""
    # 如果没有窗口，返回原数据
    if not windows:
        return series
    
    # 初始化一个全为False的布尔Series
    mask = pd.Series(False, index=series.index)
    
    # 对每个窗口，将窗口内的数据标记为True
    for start, end in windows:
        mask = mask | ((series.index >= start) & (series.index <= end))
    
    # 根据inside参数返回数据
    if inside:
        return series[mask]
    else:
        return series[~mask]

def align_diff(ref: pd.Series, tgt: pd.Series, *, inside: bool) -> pd.Series:
    """Return tgt−ref for either inside (True) or outside (False) time window."""
    # slice either inside or outside window
    if inside:
        ref = ref.loc[TIME_START:TIME_END]
        tgt = tgt.loc[TIME_START:TIME_END]
    else:
        ref = ref.loc[(ref.index < TIME_START) | (ref.index > TIME_END)]
        tgt = tgt.loc[(tgt.index < TIME_START) | (tgt.index > TIME_END)]
    aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
    return aligned.iloc[:, 1] - aligned.iloc[:, 0]
# ---------------------------------------------------------------------------
# 专门给 1 min 用的 source 映射   ★ CHG ★
# ---------------------------------------------------------------------------
SRC_MAP_1MIN = {
    "Temp": ["TowerTurb", "MWR"],
    "RH":   ["TowerTurb", "MWR"],
    "HWS":  ["LiDAR", "TowerTurb"],
    "VWS":  ["LiDAR"],
}

XLIMS_INSIDE = {"Temp": (-7.5, 17.5),   "RH": (-40, 40),
                "HWS": (-10, 15),    "VWS": (-7.5, 2.5)}
STEP_INSIDE  = {"Temp": 5, "RH": 10, "HWS": 2.5, "VWS": 2.5}
XLIMS_OUTSIDE = {"Temp": (-9, 13), "RH": (-60, 60),
                  "HWS": (-10, 7.5),     "VWS": (-7.5, 5)}
STEP_OUTSIDE  = {"Temp": 2, "RH": 20, "HWS": 2.5, "VWS": 2.5}
# ---------------------------------------------------------------------------
# Helper: build long-form dataframe of errors   ★ NEW ★
# ---------------------------------------------------------------------------
def build_error_df(inside=True) -> pd.DataFrame:
    """
    统计误差数据，根据inside参数决定是包含还是排除特定时间窗口内的数据。
    - inside=True: 只保留在EXCLUDE_WINDOWS定义的时间窗口内的数据
    - inside=False: 只保留在EXCLUDE_WINDOWS定义的时间窗口外的数据
    """
    recs = []
    for res, sdict in RESOLUTION_DICT.items():
        if sdict is None:
            continue
        for var in VAR_GROUPS:
            ref_key  = REF_MAP[var]
            src_list = SRC_MAP_1MIN[var] if res == "1min" else SRC_MAP[var]
            for h, series_map in sdict[var].items():
                if ref_key not in series_map:              # 参考值不存在
                    continue
                ref = series_map[ref_key]
                # ─── 对所有数据应用时间窗口过滤 ───
                ref = include_time_windows(ref, EXCLUDE_WINDOWS, inside=inside)
                for src in src_list:
                    if src == ref_key or src not in series_map:
                        continue
                    tgt = series_map[src]
                    # ─── 对所有数据应用时间窗口过滤 ───
                    tgt = include_time_windows(tgt, EXCLUDE_WINDOWS, inside=inside)
                    # 对齐并求差
                    aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
                    diff    = aligned.iloc[:, 1] - aligned.iloc[:, 0]
                    recs.extend({
                        "resolution": res,
                        "variable"  : var,
                        "source"    : src,
                        "error"     : d,
                    } for d in diff.values)
    return pd.DataFrame.from_records(recs)
def summarize_errors(df: pd.DataFrame) -> pd.DataFrame:
    """
    计算每个 (resolution, variable, source) 组合的
    mean / median / q25 / q75 / IQR / whisker_low / whisker_high。
    与 matplotlib.boxplot 默认设置保持一致：
    whisker = q1 ± 1.5 × IQR（再截断到全局 min/max）。
    """
    grp = df.groupby(["resolution", "variable", "source"])["error"]
    summary = grp.agg(
        count  = "size",
        mean   = "mean",
        median = "median",
        q25    = lambda x: x.quantile(0.25),
        q75    = lambda x: x.quantile(0.75),
        min    = "min",
        max    = "max",
    )
    summary["iqr"] = summary.q75 - summary.q25
    summary["whisker_low"]  = (summary.q25 - 1.5 * summary.iqr).clip(lower=summary["min"])
    summary["whisker_high"] = (summary.q75 + 1.5 * summary.iqr).clip(upper=summary["max"])
    # 调整列顺序更直观
    summary = summary[[
        "count", "mean", "median", "q25", "q75",
        "iqr", "whisker_low", "whisker_high", "min", "max"
    ]]
    return summary
# -----------------------------------------------------------------------------
# Plotting
# -----------------------------------------------------------------------------
from matplotlib.lines import Line2D
def make_plot(df: pd.DataFrame, outfile: Path, top_caption: str, bottom_caption: str, xlims: dict, steps: dict):
    
    fig, axes = plt.subplots(nrows=len(VAR_GROUPS), ncols=len(COL_LABELS) , figsize=FIGSIZE, sharex=False, sharey=False)
    fig.subplots_adjust(wspace=0.4, hspace=0.4)
    box_offset = 0.5  # left of tick
    scatter_offset = -0.5  # right of tick (more separation than before)
    box_width = 0.5
    scatter_size = 18
    scatter_alpha = 1
    row_spacing_factor = 1.7
    
    for r, var in enumerate(VAR_GROUPS):
        # Row labels -------------------------------------------------------------
        fig.text(0.08, 0.82 - r * 0.215, ROW_LABELS[var], rotation=90, va="center", ha="center",fontsize = 22)
    
        for c, res in enumerate(COL_LABELS):
            if r == 0:
                fig.text(0.13 + (c + 0.5) * 0.92 / len(COL_LABELS), 0.90, res, ha="center", va="bottom",fontsize = 22)
    
            ax = axes[r, c]
            subset = df.query("variable == @var and resolution == @res")
    
            # ★ CHG ★
            order = SRC_MAP_1MIN[var] if res == "1min" else SRC_MAP[var]
            display_order = [DISPLAY_NAMES.get(s, s) for s in order]
            y_pos = np.arange(len(order)) * row_spacing_factor   # ← greater spacing
            for i, src in enumerate(order):
                e = subset.loc[subset["source"] == src, "error"].values
                if e.size == 0:
                    continue
                ax.boxplot(e, zorder=2,
                           vert=False,
                           positions=[y_pos[i] + box_offset], 
                           widths=box_width,#showmeans=True,
                           patch_artist=True, manage_ticks=False,
                           boxprops=dict(facecolor=COLMAP[src], alpha=1, linewidth=2.5),
                           #meanprops=dict(marker="^", markerfacecolor="white", markeredgecolor='crimson', markersize=10, linewidth=0),
                           flierprops=dict(marker="x", markeredgecolor=COLMAP[src], markerfacecolor="none",
                                    markersize=7, linewidth=0.5, alpha=scatter_alpha),
                           medianprops=dict(color="lime", linewidth=2), 
                           whiskerprops=dict(color="black", linewidth=2.5), 
                           capprops=dict(color="black", linewidth=2.5)
                           )
                mean_val = e.mean()
                ax.scatter(mean_val, y_pos[i] + scatter_offset, marker="^",zorder=3,edgecolors="white",
                           color='crimson', s=180, linewidths=2, alpha=1)  # mean
                jitter_y = np.random.normal(y_pos[i] + scatter_offset, 0.08, size=len(e))
                ax.scatter(e, jitter_y, s=scatter_size, color=COLMAP[src], alpha=scatter_alpha, linewidths=0)
    
            ax.set_yticks(y_pos)
            ax.set_yticklabels(display_order)
            ax.set_ylim(y_pos[0] - 1, y_pos[-1] + 1)
            ax.grid(axis="x", linestyle=(0, (5, 10)), alpha=0.6,color = 'black',linewidth = 2.5,zorder=3)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_visible(True)
            ax.spines["left"].set_linewidth(2)
            ax.spines["bottom"].set_linewidth(2)
            # ─── x-limits & ticks ──────────────────────────────────────────────
            ax.set_xlim(*xlims[var])                                # ★ CHG ★
            major = steps[var]                                      # ★ CHG ★
            ax.xaxis.set_major_locator(mticker.MultipleLocator(major))
            ax.xaxis.set_minor_locator(mticker.MultipleLocator(major / 2))
            ax.tick_params(axis="x", which="major", length=9)       # ★ CHG ★
            ax.tick_params(axis="x", which="minor", length=4.5)     # ★ CHG ★
            ax.tick_params(axis="y", which="major", length=9)
        
# -----------------------------------------------------------------------------
# Legend
# -----------------------------------------------------------------------------
    handles = [mlines.Line2D([], [], color=COLMAP[s], marker="s", linestyle="", markersize=16,
                             markerfacecolor=COLMAP[s], label=DISPLAY_NAMES.get(s, s))
               for s in COLMAP]
    # 把 '\n' → ' '，这样 legend 里就是单行
    legend_labels = [h.get_label().replace('\n', ' ') for h in handles]
    fig.legend(handles,legend_labels, loc="upper center", ncol=6, frameon=False, bbox_to_anchor=(0.55, 0.99))
    
    handle_mean   = Line2D([], [], marker="^", markersize=14,
                           linestyle="", markeredgecolor="white",
                           markerfacecolor="crimson", color="crimson",
                           label="Mean (▲)")
    handle_median = Line2D([], [], linestyle="-", linewidth=4,
                           color="lime", label="Median (—)")
    
    fig.legend([handle_mean, handle_median],
               ["Mean", "Median"],
               loc="lower left",
               bbox_to_anchor=(0.04, 0.02),   # ← 左下角留一点边距
               frameon=False, fontsize=22)
    fig.text(0.58, 0.045,
             "Observation Biases = Ground-based Remote Sensing Datasets/Reanalysis - Ref.",
             ha="center", va="bottom", fontsize=22)
    
    if top_caption:
        fig.text(0.55, 0.93, top_caption, ha="center", va="bottom",
                 fontsize=22, fontstyle="italic", fontweight="bold")
        
    fig.tight_layout(rect=[0.08, 0.05, 1, 0.92])
    fig.savefig(outfile, dpi=600, format="tif", transparent=True, bbox_inches="tight", pil_kwargs={"compression": "tiff_adobe_deflate"})
    print("Saved →", outfile)
    
# # ─────────────────────────────────────────────────────────────────────────────
# # 4.  Build two dataframes & plot
# # ─────────────────────────────────────────────────────────────────────────────
print("Building DF (inside window)...")
df_in = build_error_df(inside=True)
make_plot(
    df_in,
    OUT_DIR / "Error_Eval_BoxJitter_WINDOW.tif",
    top_caption=f"",  #Rainstorm Events: {EXCLUDE_WINDOWS[0][0].strftime('%Y-%m-%d %H:%M')} — {EXCLUDE_WINDOWS[-1][1].strftime('%Y-%m-%d %H:%M')}(BJT) Included## CHG ###
    bottom_caption=f"Bias inside specified windows",
    xlims=XLIMS_INSIDE, steps=STEP_INSIDE
)
print("Building DF (outside window)...")
df_out = build_error_df(inside=False)
make_plot(
    df_out,
    OUT_DIR / "Error_Eval_BoxJitter_EXCLUDE.tif",
    top_caption=f"",  #Rainstorm Events: {EXCLUDE_WINDOWS[0][0].strftime('%Y-%m-%d %H:%M')} — {EXCLUDE_WINDOWS[-1][1].strftime('%Y-%m-%d %H:%M')}(BJT) Excluded## CHG ###
    bottom_caption=f"Bias excluding specified windows",
    xlims=XLIMS_OUTSIDE, steps=STEP_OUTSIDE
)
# ─────────────────────────────────────────────────────────────
# 4.  Build dataframe & plot   ★ ONLY ONCE ★
# ─────────────────────────────────────────────────────────────
print("Building full-period dataframe …")
# 对于全周期数据，我们不使用时间窗口过滤
def build_full_error_df() -> pd.DataFrame:
    """
    统计全周期误差数据，不进行时间窗口过滤
    """
    recs = []
    for res, sdict in RESOLUTION_DICT.items():
        if sdict is None:
            continue
        for var in VAR_GROUPS:
            ref_key  = REF_MAP[var]
            src_list = SRC_MAP_1MIN[var] if res == "1min" else SRC_MAP[var]
            for h, series_map in sdict[var].items():
                if ref_key not in series_map:              # 参考值不存在
                    continue
                ref = series_map[ref_key]
                for src in src_list:
                    if src == ref_key or src not in series_map:
                        continue
                    tgt = series_map[src]
                    # 对齐并求差
                    aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
                    diff    = aligned.iloc[:, 1] - aligned.iloc[:, 0]
                    recs.extend({
                        "resolution": res,
                        "variable"  : var,
                        "source"    : src,
                        "error"     : d,
                    } for d in diff.values)
    return pd.DataFrame.from_records(recs)

df_all = build_full_error_df()
make_plot(
    df_all,
    OUT_DIR / "Error_Eval_BoxJitter_ALL.tif",
    top_caption = "",
    bottom_caption = "",
    xlims = XLIMS_OUTSIDE,   # ← 继续用你原先的一套范围
    steps = STEP_OUTSIDE
)
stats_df = summarize_errors(df_all)

#—— 写 Excel：每个 variable 单独一个 sheet，更方便查看 ——
out_excel = OUT_DIR / "error_stats.xlsx"
with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
    for var, sub in stats_df.groupby(level="variable"):
        # 取消 variable 这一层索引，保留其余两层可读性更好
        sub.droplevel("variable").to_excel(writer, sheet_name=var)
print(f"Statistics saved → {out_excel}")

# 对于包含特定时间窗的数据统计及保存
stats_df_in = summarize_errors(df_in)

# 对于排除特定时间窗的数据统计及保存
stats_df_out = summarize_errors(df_out)

# 写入Excel：每个variable单独一个sheet，对于包含特定时间窗的数据
out_excel_in = OUT_DIR / "error_stats_inside_window.xlsx"
with pd.ExcelWriter(out_excel_in, engine="openpyxl") as writer:
    for var, sub in stats_df_in.groupby(level="variable"):
        # 取消variable这一层索引，保留其余两层可读性更好
        sub.droplevel("variable").to_excel(writer, sheet_name=var)
print(f"Statistics saved → {out_excel_in}")

# 写入Excel：每个variable单独一个sheet，对于排除特定时间窗的数据
out_excel_out = OUT_DIR / "error_stats_exclude_window.xlsx"
with pd.ExcelWriter(out_excel_out, engine="openpyxl") as writer:
    for var, sub in stats_df_out.groupby(level="variable"):
        # 取消variable这一层索引，保留其余两层可读性更好
        sub.droplevel("variable").to_excel(writer, sheet_name=var)
print(f"Statistics saved → {out_excel_out}")



import matplotlib.gridspec as gridspec  # 添加这一行
import matplotlib as mpl
import matplotlib.pyplot as plt
try:
    import seaborn as sns
    sns.reset_defaults()       # 还原 seaborn
except Exception:
    pass
mpl.rcParams.update(mpl.rcParamsDefault)  # 还原 Matplotlib 默认
plt.style.use('default')                  # 还原样式表

###############近地层温湿风误差评估############################
# =============================================================================
# 0. 依赖与全局常量（保持不变＋新增 RH 相关常量）
# =============================================================================

# ---- 你的已有常量 ----
HEIGHTS      = [80, 140, 200, 280]
# ── 数据源颜色 & 显示名（同前）
COLMAP = {"MWR":"red", "TowerTurb":"slategray",
          "CRA":"magenta", "ERA5":"blue", "LiDAR":"darkorange"}
DISPLAY_NAMES = {"TowerTurb":"ECM", "LiDAR":"DWL"}
# ── 每个分辨率允许出现的数据源
RES_SRC_RH = {
    "1min" : ["TowerTurb", "MWR"],
    "60min": ["TowerTurb", "MWR", "ERA5", "CRA"],
}
RES_SRC_TEMP = {
    "1min" : ["TowerTurb", "MWR"],
    "60min": ["TowerTurb", "MWR", "ERA5", "CRA"],
}
RES_SRC_HWS = {
    "1min": ["LiDAR", "TowerTurb"],
    "60min": ["LiDAR", "TowerTurb", "ERA5", "CRA"],
}
plt.rcParams.update({
    'font.size': 26,
    'font.family': 'sans-serif',                
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     
    'axes.titlepad': 8,
    'legend.fontsize': 26,
    'axes.labelsize': 26,
    'xtick.labelsize': 26,
    'ytick.labelsize': 26,
    'xtick.major.size': 10,
    'xtick.minor.size': 5,
    'ytick.major.size': 10,
    'ytick.minor.size': 5,
    'xtick.direction': 'out',
    'ytick.direction': 'out',
    'xtick.major.width': 2,
    'xtick.minor.width': 2,
    'ytick.major.width': 2,
    'ytick.minor.width': 2,
})

# ---- RH 档划分 & 颜色 ----
RH_BINS = {                                           # 左闭右开
    "Low RH (<50%)"      : (-np.inf, 50),
    "Middle RH (50–80%)"    : (50, 80),
    "High RH (>80%)"     : (80,  np.inf),
}
RH_ORDER  = list(RH_BINS.keys())                      # 为了固定顺序
RH_COLORS = {                                         # 箱/散点颜色
    "Low RH (<50%)"   : "#1f77b4",   # 蓝
    "Middle RH (50–80%)" : "#2ca02c",   # 绿
    "High RH (>80%)"  : "#d62728",   # 红
}

# ---- Temp 档划分 & 颜色 ----
TEMP_BINS = {                                # 左闭右开
    "Low T (<300 K)"  : (-np.inf, 27+273),
    "Middle T (300–303 K)": (27+273 , 30+273),
    "High T (≥303 K)" : (30+273 , np.inf),
}
TEMP_ORDER  = list(TEMP_BINS.keys())         # 固定画图顺序
TEMP_COLORS = {
    "Low T (<300 K)"  : "#1f77b4",
    "Middle T (300–303 K)": "#2ca02c",
    "High T (≥303 K)" : "#d62728",
}

# ---- HWS 档划分 & 颜色 ----
HWS_BINS = {
    "Low HWS (<3 m/s)": (-np.inf, 3),
    "Middle HWS (3–7 m/s)": (3, 7),
    "High HWS (≥7 m/s)": (7, np.inf),
}
HWS_ORDER = list(HWS_BINS.keys())
HWS_COLORS = {
    "Low HWS (<3 m/s)": "#1f77b4",
    "Middle HWS (3–7 m/s)": "#2ca02c",
    "High HWS (≥7 m/s)": "#d62728",
}

# ─────────────────────────────────────────────
# 时间窗口定义
# ─────────────────────────────────────────────
EXCLUDE_WINDOWS = [
    (pd.Timestamp("2024-08-09 12:46"), pd.Timestamp("2024-08-10 00:03")),
    (pd.Timestamp("2024-06-07 09:16"), pd.Timestamp("2024-06-07 15:00")),
    (pd.Timestamp("2024-08-10 16:20"), pd.Timestamp("2024-08-10 17:30"))
]

# =============================================================================
# 1. 通用函数：时间窗口过滤
# =============================================================================
def filter_by_time_windows(series, windows, inside=True):
    """
    根据时间窗口过滤数据
    :param series: 时间序列数据
    :param windows: 时间窗口列表
    :param inside: True-保留窗口内数据，False-保留窗口外数据
    :return: 过滤后的时间序列
    """
    if not windows:
        return series
    
    mask = pd.Series(False, index=series.index)
    
    for start, end in windows:
        window_mask = (series.index >= start) & (series.index <= end)
        if inside:
            mask = mask | window_mask
        else:
            # 对于outside，我们使用取反逻辑
            pass
    
    if inside:
        return series[mask]
    else:
        # 对于outside，我们需要标记所有窗口内的点，然后取反
        inside_mask = pd.Series(False, index=series.index)
        for start, end in windows:
            inside_mask = inside_mask | ((series.index >= start) & (series.index <= end))
        return series[~inside_mask]

# =============================================================================
# 2. RH 条件误差数据构建
# =============================================================================
def build_rh_error_df(inside=None):
    """
    针对 RH (TowerAvg 参考)：
        * 计算 src − TowerAvg
        * 按 TowerAvg RH 分入低/中/高档
        * 根据inside参数过滤时间窗口
    """
    recs = []
    for res_key, sdict in (("1min",  series_dict_1min),
                           ("60min", series_dict_60min)):
        if sdict is None:
            continue
        src_list = RES_SRC_RH[res_key]
        
        for h in HEIGHTS:
            if "TowerAvg" not in sdict["RH"][h]:
                continue
            ref = sdict["RH"][h]["TowerAvg"].dropna()
            
            # 应用时间窗口过滤（如果指定了inside参数）
            if inside is not None:
                ref = filter_by_time_windows(ref, EXCLUDE_WINDOWS, inside=inside)
            
            for src in src_list:
                if src not in sdict["RH"][h]:
                    continue
                tgt = sdict["RH"][h][src].dropna()
                
                # 应用时间窗口过滤（如果指定了inside参数）
                if inside is not None:
                    tgt = filter_by_time_windows(tgt, EXCLUDE_WINDOWS, inside=inside)
                
                aligned = pd.concat([ref, tgt], axis=1,
                                    join="inner").dropna()
                diff = aligned.iloc[:, 1] - aligned.iloc[:, 0]
                ref_rh = aligned.iloc[:, 0]
                
                for ts, e in diff.items():
                    rh_val = ref_rh.loc[ts]
                    for rh_bin, (lo, hi) in RH_BINS.items():
                        if lo <= rh_val < hi:
                            recs.append({
                                "resolution": res_key,
                                "height"    : h,
                                "rh_bin"    : rh_bin,
                                "source"    : src,
                                "error"     : e,
                            })
                            break
    return pd.DataFrame.from_records(recs)

# =============================================================================
# 3. Temp 条件误差数据构建
# =============================================================================
def build_temp_error_df(inside=None):
    """
    针对 Temp (TowerAvg 参考)：
        * 计算 src − TowerAvg
        * 按 TowerAvg Temp 分入低/中/高档
        * 根据inside参数过滤时间窗口
    """
    recs = []
    for res_key, sdict in (("1min",  series_dict_1min),
                           ("60min", series_dict_60min)):
        if sdict is None:
            continue
        src_list = RES_SRC_TEMP[res_key]
        
        for h in HEIGHTS:
            if "TowerAvg" not in sdict["Temp"][h]:
                continue
            ref = sdict["Temp"][h]["TowerAvg"].dropna()
            
            # 应用时间窗口过滤（如果指定了inside参数）
            if inside is not None:
                ref = filter_by_time_windows(ref, EXCLUDE_WINDOWS, inside=inside)
            
            for src in src_list:
                if src not in sdict["Temp"][h]:
                    continue
                tgt = sdict["Temp"][h][src].dropna()
                
                # 应用时间窗口过滤（如果指定了inside参数）
                if inside is not None:
                    tgt = filter_by_time_windows(tgt, EXCLUDE_WINDOWS, inside=inside)
                
                aligned = pd.concat([ref, tgt], axis=1,
                                    join="inner").dropna()
                diff = aligned.iloc[:, 1] - aligned.iloc[:, 0]
                ref_t = aligned.iloc[:, 0]
                
                for ts, e in diff.items():
                    t_val = ref_t.loc[ts]
                    for t_bin, (lo, hi) in TEMP_BINS.items():
                        if lo <= t_val < hi:
                            recs.append({
                                "resolution": res_key,
                                "height"    : h,
                                "temp_bin"  : t_bin,
                                "source"    : src,
                                "error"     : e,
                            })
                            break
    return pd.DataFrame.from_records(recs)

# =============================================================================
# 4. HWS 条件误差数据构建
# =============================================================================
def build_hws_error_df(inside=None):
    """
    针对 HWS (TowerAvg 参考)：
        * 计算 src − TowerAvg
        * 按 TowerAvg HWS 分入低/中/高档
        * 根据inside参数过滤时间窗口
    """
    recs = []
    for res_key, sdict in (("1min",  series_dict_1min),
                           ("60min", series_dict_60min)):
        if sdict is None:
            continue
        src_list = RES_SRC_HWS[res_key]
        
        for h in HEIGHTS:
            if "TowerAvg" not in sdict["HWS"][h]:
                continue
            ref = sdict["HWS"][h]["TowerAvg"].dropna()
            
            # 应用时间窗口过滤（如果指定了inside参数）
            if inside is not None:
                ref = filter_by_time_windows(ref, EXCLUDE_WINDOWS, inside=inside)
            
            for src in src_list:
                if src not in sdict["HWS"][h]:
                    continue
                tgt = sdict["HWS"][h][src].dropna()
                
                # 应用时间窗口过滤（如果指定了inside参数）
                if inside is not None:
                    tgt = filter_by_time_windows(tgt, EXCLUDE_WINDOWS, inside=inside)
                
                aligned = pd.concat([ref, tgt], axis=1,
                                    join="inner").dropna()
                diff = aligned.iloc[:, 1] - aligned.iloc[:, 0]
                ref_hws = aligned.iloc[:, 0]
                
                for ts, e in diff.items():
                    hws_val = ref_hws.loc[ts]
                    for hws_bin, (lo, hi) in HWS_BINS.items():
                        if lo <= hws_val < hi:
                            recs.append({
                                "resolution": res_key,
                                "height"    : h,
                                "hws_bin"   : hws_bin,
                                "source"    : src,
                                "error"     : e,
                            })
                            break
    return pd.DataFrame.from_records(recs)

# =============================================================================
# 5. 绘图函数（RH）
# =============================================================================
# def make_rh_plot(df: pd.DataFrame, outfile: Path, title_suffix=""):
#     """
#     横向 4 高度 × 纵向 2 分辨率 的子图网格。
#     每个子图里：
#         * y 轴 = 数据源（TowerTurb/MWR/ERA5/CRA）
#         * 同一源画 3 个并排箱图，对应低/中/高 RH 档
#     """
#     plt.rcParams.update({
#         "font.family":"sans-serif",
#         "font.sans-serif":["Arial","DejaVu Sans"],
#         "font.size":26,
#         "xtick.major.width":2,
#         "ytick.major.width":2,
#     })    
#     RES_LIST = ("1min",  "60min")          
#     ROWS, COLS = len(RES_LIST), len(HEIGHTS)
#     FIGSIZE = (30, 12)
#     fig, axes = plt.subplots(ROWS, COLS, figsize=FIGSIZE,
#                              sharex=False, sharey=False)
#     fig.subplots_adjust(wspace=0.6,hspace=0.8)  # 增大/减小子图行间距

#     # ───────────────────────────────────────────────────────────
#     # 布局参数
#     box_w   = 0.3                         # 单箱宽度
#     jit_s   = 10                           # jitter 直径
#     off     = np.array([-0.6, 0.0, 0.6]) # 低中高 RH 的 y 偏移
#     row_spacing = 1.9                      # 源之间的行距
    
#     # 为同一分辨率-高度子集遍历
#     for r, res_key in enumerate(RES_LIST):
#         srcs_this_row = RES_SRC_RH[res_key]
#         y_base = np.arange(len(srcs_this_row)) * row_spacing
#         for c, h in enumerate(HEIGHTS):
#             ax = axes[r, c]
#             ax.axvline(0, color='black', lw=2, linestyle=(0,(5,10)))

#             sub = df.query("resolution == @res_key and height == @h")
#             if sub.empty:
#                 ax.text(0.5, 0.5, "No data", ha="center", va="center",
#                         alpha=0.6)
#                 ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
#                 continue
                
#             # ―― 逐源 & RH 档画箱+散点 ―――――――――――――――――――――――
#             for i, src in enumerate(srcs_this_row):
#                 for j, rh_bin in enumerate(RH_ORDER):
#                     errs = sub.loc[(sub["source"] == src) &
#                                    (sub["rh_bin"] == rh_bin),
#                                    "error"].values
#                     if errs.size == 0:
#                         continue
                        
#                     # 箱图
#                     ax.boxplot(errs, vert=False,
#                                positions=[y_base[i] + off[j]],
#                                widths=box_w,
#                                patch_artist=True, manage_ticks=False,
#                                boxprops=dict(facecolor=RH_COLORS[rh_bin],
#                                              alpha=0.75, linewidth=2),
#                                flierprops=dict(marker="x", markersize=6,
#                                                markeredgecolor="black",
#                                                markerfacecolor="none",
#                                                linewidth=0.5),
#                                medianprops=dict(color="lime", linewidth=2),
#                                whiskerprops=dict(color="black", linewidth=1.8),
#                                capprops=dict(color="black", linewidth=1.8))
                               
#                     # jitter
#                     jitter_y = np.random.normal(y_base[i] + off[j], 0.06,
#                                                 size=len(errs))
#                     ax.scatter(errs, jitter_y, s=jit_s,
#                                color=RH_COLORS[rh_bin], alpha=0.8,
#                                linewidths=0)
                               
#             # ―― 轴刻度 & 标签 ―――――――――――――――――――――――――――――――――
#             ax.set_yticks(y_base)
#             disp_names = [DISPLAY_NAMES.get(s, s) for s in srcs_this_row]
#             ax.set_yticklabels(disp_names)
#             ax.set_ylim(y_base[0] - 1, y_base[-1] + 1)
#             ax.grid(axis="x", linestyle=(0, (5, 8)), alpha=0.6)
#             ax.spines["top"].set_visible(False)
#             ax.spines["right"].set_visible(False)
#             ax.spines["left"].set_linewidth(2)
#             ax.spines["bottom"].set_linewidth(2)
            
#             # X-lim & tick 间距根据经验给定，可自行调整
#             ax.set_xlim(-35, 50)
#             ax.xaxis.set_major_locator(mticker.MultipleLocator(20))
#             ax.xaxis.set_minor_locator(mticker.MultipleLocator(10))
            
#             if r == ROWS - 1:
#                 ax.set_xlabel("Bias (%)")          # 仅底排标注 x-label
#             if r == 0:
#                 ax.set_title(f"{h} m", pad=10, fontsize=26)
                
#     # # 行标签（分辨率）
#     # for r, res_key in enumerate(RES_LIST):
#     #     axes[r, 0].annotate(res_key, xy=(-0.80, 0.5),
#     #                         xycoords="axes fraction", rotation=90,
#     #                         va="center", ha="center", fontsize=26)
                            
#     # ───────────────────────────────────────────────────────────
#     # 图例：RH 档
#     handles_rh = [mlines.Line2D([], [], color=RH_COLORS[b],
#                                 marker="s", linestyle="", markersize=15,
#                                 label=b) for b in RH_ORDER]
#     # 图例：数据源
#     all_sources = set([src for src_list in RES_SRC_RH.values() for src in src_list])
#     handles_src = [mlines.Line2D([], [], color="black",
#                              marker="s", linestyle="", markersize=15,
#                              markerfacecolor=COLMAP[s],
#                              label=DISPLAY_NAMES.get(s, s))
#                for s in all_sources]
               
#     fig = plt.gcf()
#     fig.legend(handles_rh,
#                [h.get_label() for h in handles_rh + handles_src],
#                loc="upper center", ncol=7, frameon=False,
#                fontsize=26, bbox_to_anchor=(0.51, 0.98))
               
#     fig.text(0.52, -0.015,
#              "RH Bias = Ground-based remote sensing/Reanalysis − AWS",
#              ha="center", fontsize=26)
             
#     # 添加标题
#     if title_suffix:
#         fig.text(0.52, 0.98, f"RH Conditional Bias {title_suffix}",
#                 ha="center", fontsize=26, fontweight="bold")
                
#     fig.savefig(outfile, dpi=600, format="tif", transparent=True,
#                 bbox_inches="tight",
#                 pil_kwargs={"compression": "tiff_adobe_deflate"})
#     plt.close(fig)          # ← 立即关闭，释放内存
#     print("Saved →", outfile)


######new
def make_rh_plot(df: pd.DataFrame, outfile: Path, title_suffix=""):
    plt.rcParams.update({
        "font.family":"sans-serif",
        "font.sans-serif":["Arial","DejaVu Sans"],
        "font.size":26,
        "xtick.major.width":2,
        "ytick.major.width":2,
    })    
    RES_LIST = ("1min",  "60min")          
    ROWS, COLS = len(RES_LIST), len(HEIGHTS)
    
    # 计算每行所需的高度比例（基于数据源数量）
    row_heights = [len(RES_SRC_RH[res]) for res in RES_LIST]
    total_height = sum(row_heights)
    height_ratios = [h/total_height for h in row_heights]
    
    # 创建图形和GridSpec布局
    FIGSIZE = (30, 16)  # 增加图形高度
    fig = plt.figure(figsize=FIGSIZE)
    gs = gridspec.GridSpec(ROWS, COLS, height_ratios=height_ratios, 
                          hspace=0.15, wspace=0.25)
    
    # 创建子图数组
    axes = [[fig.add_subplot(gs[r, c]) for c in range(COLS)] for r in range(ROWS)]

    # ───────────────────────────────────────────────────────────
    # 布局参数
    box_w   = 0.35                         # 单箱宽度（固定）
    jit_s   = 12                           # jitter 直径
    off     = np.array([-0.6, 0.0, 0.6])   # 低中高 RH 的 y 偏移（固定）
    UNIT_HEIGHT = 1.8                      # 每个数据源占用的垂直空间（固定）
    
    # 为同一分辨率-高度子集遍历
    for r, res_key in enumerate(RES_LIST):
        srcs_this_row = RES_SRC_RH[res_key]
        # 计算该分辨率需要的总高度
        total_height = UNIT_HEIGHT * len(srcs_this_row)
        # 计算y_base位置
        y_base = np.arange(len(srcs_this_row)) * UNIT_HEIGHT
        for c, h in enumerate(HEIGHTS):
            ax = axes[r][c]
            ax.axvline(0, color='black', lw=2, linestyle=(0,(5,10)))

            sub = df.query("resolution == @res_key and height == @h")
            if sub.empty:
                ax.text(0.5, 0.5, "No data", ha="center", va="center",
                        alpha=0.6)
                ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
                continue
                
            # ―― 逐源 & RH 档画箱+散点 ―――――――――――――――――――――――
            for i, src in enumerate(srcs_this_row):
                for j, rh_bin in enumerate(RH_ORDER):
                    errs = sub.loc[(sub["source"] == src) &
                                   (sub["rh_bin"] == rh_bin),
                                   "error"].values
                    if errs.size == 0:
                        continue
                        
                    # 箱图
                    ax.boxplot(errs, vert=False,
                               positions=[y_base[i] + off[j]],
                               widths=box_w,
                               patch_artist=True, manage_ticks=False,
                               boxprops=dict(facecolor=RH_COLORS[rh_bin],
                                             alpha=0.75, linewidth=2),
                               flierprops=dict(marker="x", markersize=6,
                                               markeredgecolor="black",
                                               markerfacecolor="none",
                                               linewidth=0.5),
                               medianprops=dict(color="lime", linewidth=2),
                               whiskerprops=dict(color="black", linewidth=1.8),
                               capprops=dict(color="black", linewidth=1.8))
                               
                    # jitter
                    jitter_y = np.random.normal(y_base[i] + off[j], 0.06,
                                                size=len(errs))
                    ax.scatter(errs, jitter_y, s=jit_s,
                               color=RH_COLORS[rh_bin], alpha=0.8,
                               linewidths=0)
                               
            # ―― 轴刻度 & 标签 ―――――――――――――――――――――――――――――――――
            ax.set_yticks(y_base)
            disp_names = [DISPLAY_NAMES.get(s, s) for s in srcs_this_row]
            ax.set_yticklabels(disp_names)
            # 设置ylim，上下各留0.5的边距
            ax.set_ylim(-1, total_height - 1)
            ax.grid(axis="x", linestyle=(0, (5, 8)), alpha=0.6)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_linewidth(2)
            ax.spines["bottom"].set_linewidth(2)
            
            # X-lim & tick 间距根据经验给定，可自行调整
            ax.set_xlim(-35, 50)
            ax.xaxis.set_major_locator(mticker.MultipleLocator(20))
            ax.xaxis.set_minor_locator(mticker.MultipleLocator(10))
            
            if r == ROWS - 1:
                ax.set_xlabel("Bias (%)")          # 仅底排标注 x-label
            if r == 0:
                ax.set_title(f"{h} m", pad=10, fontsize=26)
                
    # # 行标签（分辨率）
    # for r, res_key in enumerate(RES_LIST):
    #     axes[r][0].annotate(res_key, xy=(-0.40, 0.5),
    #                         xycoords="axes fraction", rotation=90,
    #                         va="center", ha="center", fontsize=26)
                            
    # ───────────────────────────────────────────────────────────
    # 图例：RH 档
    handles_rh = [mlines.Line2D([], [], color=RH_COLORS[b],
                                marker="s", linestyle="", markersize=15,
                                label=b) for b in RH_ORDER]
    # 图例：数据源
    all_sources = set([src for src_list in RES_SRC_RH.values() for src in src_list])
    handles_src = [mlines.Line2D([], [], color="black",
                             marker="s", linestyle="", markersize=15,
                             markerfacecolor=COLMAP[s],
                             label=DISPLAY_NAMES.get(s, s))
               for s in all_sources]
               
    fig = plt.gcf()
    fig.legend(handles_rh,
               [h.get_label() for h in handles_rh + handles_src],
               loc="upper center", ncol=7, frameon=False,
               fontsize=26, bbox_to_anchor=(0.51, 0.98))
               
    fig.text(0.52, -0.015,
             "RH Bias = Ground-based remote sensing/Reanalysis − AWS",
             ha="center", fontsize=26)
             
    # 添加标题
    if title_suffix:
        fig.text(0.52, 0.98, f"RH Conditional Bias {title_suffix}",
                ha="center", fontsize=26, fontweight="bold")
                
    fig.savefig(outfile, dpi=600, format="tif", transparent=True,
                bbox_inches="tight",
                pil_kwargs={"compression": "tiff_adobe_deflate"})
    plt.close(fig)          # ← 立即关闭，释放内存
    print("Saved →", outfile)




# =============================================================================
# 6. 绘图函数（Temp）
# =============================================================================
# def make_temp_plot(df: pd.DataFrame, outfile: Path, title_suffix=""):
#     """
#     横向 4 高度 × 纵向 2 分辨率 的子图网格。
#     每个子图里：
#         * y 轴 = 数据源（TowerTurb/MWR/ERA5/CRA）
#         * 同一源画 3 个并排箱图，对应低/中/高 Temp 档
#     """
#     RES_LIST = ("1min",  "60min")          
#     ROWS, COLS = len(RES_LIST), len(HEIGHTS)
#     FIGSIZE = (30, 12)
#     fig, axes = plt.subplots(ROWS, COLS, figsize=FIGSIZE,
#                              sharex=False, sharey=False)
#     fig.subplots_adjust(wspace=0.6,hspace=0.8)  # 增大/减小子图行间距
#     plt.rcParams.update({
#         "font.family":"sans-serif",
#         "font.sans-serif":["Arial","DejaVu Sans"],
#         "font.size":26,
#         "xtick.major.width":2,
#         "ytick.major.width":2,
#     })
#     # ───────────────────────────────────────────────────────────
#     # 布局参数
#     box_w   = 0.3                         # 单箱宽度
#     jit_s   = 10                           # jitter 直径
#     off     = np.array([-0.6, 0.0, 0.6]) # 低中高 Temp 的 y 偏移
#     row_spacing = 1.9                      # 源之间的行距
    
#     # 为同一分辨率-高度子集遍历
#     for r, res_key in enumerate(RES_LIST):
#         srcs_this_row = RES_SRC_TEMP[res_key]
#         y_base = np.arange(len(srcs_this_row)) * row_spacing
#         for c, h in enumerate(HEIGHTS):
#             ax = axes[r, c]
#             ax.axvline(0, color='black', lw=2, linestyle=(0,(5,10)))

#             sub = df.query("resolution == @res_key and height == @h")
#             if sub.empty:
#                 ax.text(0.5, 0.5, "No data", ha="center", va="center",
#                         alpha=0.6)
#                 ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
#                 continue
                
#             # ―― 逐源 & Temp 档画箱+散点 ―――――――――――――――――――――――
#             for i, src in enumerate(srcs_this_row):
#                 for j, t_bin in enumerate(TEMP_ORDER):
#                     errs = sub.loc[(sub["source"] == src) &
#                                    (sub["temp_bin"] == t_bin),
#                                    "error"].values
#                     if errs.size == 0:
#                         continue
                        
#                     # 箱图
#                     ax.boxplot(errs, vert=False,
#                                positions=[y_base[i] + off[j]],
#                                widths=box_w,
#                                patch_artist=True, manage_ticks=False,
#                                boxprops=dict(facecolor=TEMP_COLORS[t_bin],
#                                              alpha=0.75, linewidth=2),
#                                flierprops=dict(marker="x", markersize=6,
#                                                markeredgecolor="black",
#                                                markerfacecolor="none",
#                                                linewidth=0.5),
#                                medianprops=dict(color="lime", linewidth=2),
#                                whiskerprops=dict(color="black", linewidth=1.8),
#                                capprops=dict(color="black", linewidth=1.8))
                               
#                     # jitter
#                     jitter_y = np.random.normal(y_base[i] + off[j], 0.06,
#                                                 size=len(errs))
#                     ax.scatter(errs, jitter_y, s=jit_s,
#                                color=TEMP_COLORS[t_bin], alpha=0.8,
#                                linewidths=0)
                               
#             # ―― 轴刻度 & 标签 ―――――――――――――――――――――――――――――――――
#             ax.set_yticks(y_base)
#             disp_names = [DISPLAY_NAMES.get(s, s) for s in srcs_this_row]
#             ax.set_yticklabels(disp_names)
#             ax.set_ylim(y_base[0] - 1, y_base[-1] + 1)
#             ax.grid(axis="x", linestyle=(0, (5, 8)), alpha=0.6)
#             ax.spines["top"].set_visible(False)
#             ax.spines["right"].set_visible(False)
#             ax.spines["left"].set_linewidth(2)
#             ax.spines["bottom"].set_linewidth(2)
            
#             # X-lim & tick 间距根据经验给定，可自行调整
#             ax.set_xlim(-8, 8)
#             ax.xaxis.set_major_locator(mticker.MultipleLocator(2))
#             ax.xaxis.set_minor_locator(mticker.MultipleLocator(1))
            
#             if r == ROWS - 1:
#                 ax.set_xlabel("Bias (K)")          # 仅底排标注 x-label
#             if r == 0:
#                 ax.set_title(f"{h} m", pad=10, fontsize=26)
                
#     # # 行标签（分辨率）
#     # for r, res_key in enumerate(RES_LIST):
#     #     axes[r, 0].annotate(res_key, xy=(-2.40, 0.5),
#     #                         xycoords="axes fraction", rotation=90,
#     #                         va="center", ha="center", fontsize=26)
                            
#     # ───────────────────────────────────────────────────────────
#     # 图例：Temp 档
#     handles_temp = [mlines.Line2D([], [], color=TEMP_COLORS[b],
#                                 marker="s", linestyle="", markersize=15,
#                                 label=b) for b in TEMP_ORDER]
#     # 图例：数据源
#     all_sources = set([src for src_list in RES_SRC_TEMP.values() for src in src_list])
#     handles_src = [mlines.Line2D([], [], color="black",
#                              marker="s", linestyle="", markersize=15,
#                              markerfacecolor=COLMAP[s],
#                              label=DISPLAY_NAMES.get(s, s))
#                for s in all_sources]
               
#     fig = plt.gcf()
#     fig.legend(handles_temp,
#                [h.get_label() for h in handles_temp + handles_src],
#                loc="upper center", ncol=7, frameon=False,
#                fontsize=26, bbox_to_anchor=(0.51, 0.98))
               
#     fig.text(0.52, -0.015,
#              "Temp Bias = Ground-based remote sensing/Reanalysis − AWS",
#              ha="center", fontsize=26)
             
#     # 添加标题
#     if title_suffix:
#         fig.text(0.52, 0.98, f"Temp Conditional Bias {title_suffix}",
#                 ha="center", fontsize=26, fontweight="bold")
                
#     fig.savefig(outfile, dpi=600, format="tif", transparent=True,
#                 bbox_inches="tight",
#                 pil_kwargs={"compression": "tiff_adobe_deflate"})
#     plt.close(fig)          # ← 立即关闭，释放内存
#     print("Saved →", outfile)
####new
def make_temp_plot(df: pd.DataFrame, outfile: Path, title_suffix=""):
    plt.rcParams.update({
        "font.family":"sans-serif",
        "font.sans-serif":["Arial","DejaVu Sans"],
        "font.size":26,
        "xtick.major.width":2,
        "ytick.major.width":2,
    })
    RES_LIST = ("1min",  "60min")          
    ROWS, COLS = len(RES_LIST), len(HEIGHTS)
    
    # 计算每行所需的高度比例（基于数据源数量）
    row_heights = [len(RES_SRC_TEMP[res]) for res in RES_LIST]
    total_height = sum(row_heights)
    height_ratios = [h/total_height for h in row_heights]
    
    # 创建图形和GridSpec布局
    FIGSIZE = (30, 16)  # 增加图形高度
    fig = plt.figure(figsize=FIGSIZE)
    gs = gridspec.GridSpec(ROWS, COLS, height_ratios=height_ratios, 
                          hspace=0.15, wspace=0.25)
    
    # 创建子图数组
    axes = [[fig.add_subplot(gs[r, c]) for c in range(COLS)] for r in range(ROWS)]
    
    # ───────────────────────────────────────────────────────────
    # 布局参数
    box_w   = 0.35                         # 单箱宽度（固定）
    jit_s   = 12                           # jitter 直径
    off     = np.array([-0.6, 0.0, 0.6])   # 低中高 Temp 的 y 偏移（固定）
    UNIT_HEIGHT = 1.8                      # 每个数据源占用的垂直空间（固定）
    
    # 为同一分辨率-高度子集遍历
    for r, res_key in enumerate(RES_LIST):
        srcs_this_row = RES_SRC_TEMP[res_key]
        # 计算该分辨率需要的总高度
        total_height = UNIT_HEIGHT * len(srcs_this_row)
        # 计算y_base位置
        y_base = np.arange(len(srcs_this_row)) * UNIT_HEIGHT
        for c, h in enumerate(HEIGHTS):
            ax = axes[r][c]
            ax.axvline(0, color='black', lw=2, linestyle=(0,(5,10)))

            sub = df.query("resolution == @res_key and height == @h")
            if sub.empty:
                ax.text(0.5, 0.5, "No data", ha="center", va="center",
                        alpha=0.6)
                ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
                continue
                
            # ―― 逐源 & Temp 档画箱+散点 ―――――――――――――――――――――――
            for i, src in enumerate(srcs_this_row):
                for j, t_bin in enumerate(TEMP_ORDER):
                    errs = sub.loc[(sub["source"] == src) &
                                   (sub["temp_bin"] == t_bin),
                                   "error"].values
                    if errs.size == 0:
                        continue
                        
                    # 箱图
                    ax.boxplot(errs, vert=False,
                               positions=[y_base[i] + off[j]],
                               widths=box_w,
                               patch_artist=True, manage_ticks=False,
                               boxprops=dict(facecolor=TEMP_COLORS[t_bin],
                                             alpha=0.75, linewidth=2),
                               flierprops=dict(marker="x", markersize=6,
                                               markeredgecolor="black",
                                               markerfacecolor="none",
                                               linewidth=0.5),
                               medianprops=dict(color="lime", linewidth=2),
                               whiskerprops=dict(color="black", linewidth=1.8),
                               capprops=dict(color="black", linewidth=1.8))
                               
                    # jitter
                    jitter_y = np.random.normal(y_base[i] + off[j], 0.06,
                                                size=len(errs))
                    ax.scatter(errs, jitter_y, s=jit_s,
                               color=TEMP_COLORS[t_bin], alpha=0.8,
                               linewidths=0)
                               
            # ―― 轴刻度 & 标签 ―――――――――――――――――――――――――――――――――
            ax.set_yticks(y_base)
            disp_names = [DISPLAY_NAMES.get(s, s) for s in srcs_this_row]
            ax.set_yticklabels(disp_names)
            # 设置ylim，上下各留0.5的边距
            ax.set_ylim(-1, total_height - 1)
            ax.grid(axis="x", linestyle=(0, (5, 8)), alpha=0.6)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_linewidth(2)
            ax.spines["bottom"].set_linewidth(2)
            
            # X-lim & tick 间距根据经验给定，可自行调整
            ax.set_xlim(-8, 8)
            ax.xaxis.set_major_locator(mticker.MultipleLocator(2))
            ax.xaxis.set_minor_locator(mticker.MultipleLocator(1))
            
            if r == ROWS - 1:
                ax.set_xlabel("Bias (K)")          # 仅底排标注 x-label
            if r == 0:
                ax.set_title(f"{h} m", pad=10, fontsize=26)
                
    # # 行标签（分辨率）
    # for r, res_key in enumerate(RES_LIST):
    #     axes[r][0].annotate(res_key, xy=(-0.40, 0.5),
    #                         xycoords="axes fraction", rotation=90,
    #                         va="center", ha="center", fontsize=26)
                            
    # ───────────────────────────────────────────────────────────
    # 图例：Temp 档
    handles_temp = [mlines.Line2D([], [], color=TEMP_COLORS[b],
                                marker="s", linestyle="", markersize=15,
                                label=b) for b in TEMP_ORDER]
    # 图例：数据源
    all_sources = set([src for src_list in RES_SRC_TEMP.values() for src in src_list])
    handles_src = [mlines.Line2D([], [], color="black",
                             marker="s", linestyle="", markersize=15,
                             markerfacecolor=COLMAP[s],
                             label=DISPLAY_NAMES.get(s, s))
               for s in all_sources]
               
    fig = plt.gcf()
    fig.legend(handles_temp,
               [h.get_label() for h in handles_temp + handles_src],
               loc="upper center", ncol=7, frameon=False,
               fontsize=26, bbox_to_anchor=(0.51, 0.98))
               
    fig.text(0.52, 0.015,
             "Temp Bias = Ground-based remote sensing/Reanalysis − AWS",
             ha="center", fontsize=26)
             
    # 添加标题
    if title_suffix:
        fig.text(0.52, 0.98, f"Temp Conditional Bias {title_suffix}",
                ha="center", fontsize=26, fontweight="bold")
                
    fig.savefig(outfile, dpi=600, format="tif", transparent=True,
                bbox_inches="tight",
                pil_kwargs={"compression": "tiff_adobe_deflate"})
    plt.close(fig)          # ← 立即关闭，释放内存
    print("Saved →", outfile)



# =============================================================================
# 7. 绘图函数（HWS）
# =============================================================================
# def make_hws_plot(df: pd.DataFrame, outfile: Path, title_suffix=""):
#     """
#     横向 4 高度 × 纵向 2 分辨率 的子图网格。
#     每个子图里：
#         * y 轴 = 数据源（LiDAR/TowerTurb/ERA5/CRA）
#         * 同一源画 3 个并排箱图，对应低/中/高 HWS 档
#     """
#     RES_LIST = ("1min",  "60min")          
#     ROWS, COLS = len(RES_LIST), len(HEIGHTS)
#     FIGSIZE = (30, 12)
#     fig, axes = plt.subplots(ROWS, COLS, figsize=FIGSIZE,
#                              sharex=False, sharey=False)
#     fig.subplots_adjust(wspace=0.6,hspace=0.8)  # 增大/减小子图行间距
#     plt.rcParams.update({
#         "font.family":"sans-serif",
#         "font.sans-serif":["Arial","DejaVu Sans"],
#         "font.size":26,
#         "xtick.major.width":2,
#         "ytick.major.width":2,
#     })
#     # ───────────────────────────────────────────────────────────
#     # 布局参数
#     box_w   = 0.3                         # 单箱宽度
#     jit_s   = 10                           # jitter 直径
#     off     = np.array([-0.6, 0.0, 0.6]) # 低中高 HWS 的 y 偏移
#     row_spacing = 1.9                      # 源之间的行距
    
#     # 为同一分辨率-高度子集遍历
#     for r, res_key in enumerate(RES_LIST):
#         srcs_this_row = RES_SRC_HWS[res_key]
#         y_base = np.arange(len(srcs_this_row)) * row_spacing
#         for c, h in enumerate(HEIGHTS):
#             ax = axes[r, c]
#             ax.axvline(0, color='black', lw=2, linestyle=(0,(5,10)))

#             sub = df.query("resolution == @res_key and height == @h")
#             if sub.empty:
#                 ax.text(0.5, 0.5, "No data", ha="center", va="center",
#                         alpha=0.6)
#                 ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
#                 continue
                
#             # ―― 逐源 & HWS 档画箱+散点 ―――――――――――――――――――――――
#             for i, src in enumerate(srcs_this_row):
#                 for j, hws_bin in enumerate(HWS_ORDER):
#                     errs = sub.loc[(sub["source"] == src) &
#                                    (sub["hws_bin"] == hws_bin),
#                                    "error"].values
#                     if errs.size == 0:
#                         continue
                        
#                     # 箱图
#                     ax.boxplot(errs, vert=False,
#                                positions=[y_base[i] + off[j]],
#                                widths=box_w,
#                                patch_artist=True, manage_ticks=False,
#                                boxprops=dict(facecolor=HWS_COLORS[hws_bin],
#                                              alpha=0.75, linewidth=2),
#                                flierprops=dict(marker="x", markersize=6,
#                                                markeredgecolor="black",
#                                                markerfacecolor="none",
#                                                linewidth=0.5),
#                                medianprops=dict(color="lime", linewidth=2),
#                                whiskerprops=dict(color="black", linewidth=1.8),
#                                capprops=dict(color="black", linewidth=1.8))
                               
#                     # jitter
#                     jitter_y = np.random.normal(y_base[i] + off[j], 0.06,
#                                                 size=len(errs))
#                     ax.scatter(errs, jitter_y, s=jit_s,
#                                color=HWS_COLORS[hws_bin], alpha=0.8,
#                                linewidths=0)
                               
#             # ―― 轴刻度 & 标签 ―――――――――――――――――――――――――――――――――
#             ax.set_yticks(y_base)
#             disp_names = [DISPLAY_NAMES.get(s, s) for s in srcs_this_row]
#             ax.set_yticklabels(disp_names)
#             ax.set_ylim(y_base[0] - 1, y_base[-1] + 1)
#             ax.grid(axis="x", linestyle=(0, (5, 8)), alpha=0.6)
#             ax.spines["top"].set_visible(False)
#             ax.spines["right"].set_visible(False)
#             ax.spines["left"].set_linewidth(2)
#             ax.spines["bottom"].set_linewidth(2)
            
#             # X-lim & tick 间距根据经验给定，可自行调整
#             ax.set_xlim(-8, 8)
#             ax.xaxis.set_major_locator(mticker.MultipleLocator(2))
#             ax.xaxis.set_minor_locator(mticker.MultipleLocator(1))
            
#             if r == ROWS - 1:
#                 ax.set_xlabel("Bias (m/s)")          # 仅底排标注 x-label
#             if r == 0:
#                 ax.set_title(f"{h} m", pad=10, fontsize=26)
                
#     # 行标签（分辨率）
#     # for r, res_key in enumerate(RES_LIST):
#     #     axes[r, 0].annotate(res_key, xy=(-0.40, 0.5),
#     #                         xycoords="axes fraction", rotation=90,
#     #                         va="center", ha="center", fontsize=26)
                            
#     # ───────────────────────────────────────────────────────────
#     # 图例：HWS 档
#     handles_hws = [mlines.Line2D([], [], color=HWS_COLORS[b],
#                                 marker="s", linestyle="", markersize=15,
#                                 label=b) for b in HWS_ORDER]
#     # 图例：数据源
#     all_sources = set([src for src_list in RES_SRC_HWS.values() for src in src_list])
#     handles_src = [mlines.Line2D([], [], color="black",
#                              marker="s", linestyle="", markersize=15,
#                              markerfacecolor=COLMAP[s],
#                              label=DISPLAY_NAMES.get(s, s))
#                for s in all_sources]
               
#     fig = plt.gcf()
#     fig.legend(handles_hws,
#                [h.get_label() for h in handles_hws + handles_src],
#                loc="upper center", ncol=7, frameon=False,
#                fontsize=26, bbox_to_anchor=(0.51, 0.98))
               
#     fig.text(0.52, -0.015,
#              "HWS Bias = Ground-based remote sensing/Reanalysis − AWS",
#              ha="center", fontsize=26)
             
#     # 添加标题
#     if title_suffix:
#         fig.text(0.52, 0.98, f"HWS Conditional Bias {title_suffix}",
#                 ha="center", fontsize=26, fontweight="bold")
                
#     fig.savefig(outfile, dpi=600, format="tif", transparent=True,
#                 bbox_inches="tight",
#                 pil_kwargs={"compression": "tiff_adobe_deflate"})
#     plt.close(fig)          # ← 立即关闭，释放内存
#     print("Saved →", outfile)
######new
def make_hws_plot(df: pd.DataFrame, outfile: Path, title_suffix=""):
    plt.rcParams.update({
        "font.family":"sans-serif",
        "font.sans-serif":["Arial","DejaVu Sans"],
        "font.size":26,
        "xtick.major.width":2,
        "ytick.major.width":2,
    })
    RES_LIST = ("1min",  "60min")          
    ROWS, COLS = len(RES_LIST), len(HEIGHTS)
    
    # 计算每行所需的高度比例（基于数据源数量）
    row_heights = [len(RES_SRC_HWS[res]) for res in RES_LIST]
    total_height = sum(row_heights)
    height_ratios = [h/total_height for h in row_heights]
    
    # 创建图形和GridSpec布局
    FIGSIZE = (30, 16)  # 增加图形高度
    fig = plt.figure(figsize=FIGSIZE)
    gs = gridspec.GridSpec(ROWS, COLS, height_ratios=height_ratios, 
                          hspace=0.15, wspace=0.25)
    
    # 创建子图数组
    axes = [[fig.add_subplot(gs[r, c]) for c in range(COLS)] for r in range(ROWS)]
    
    # ───────────────────────────────────────────────────────────
    # 布局参数
    box_w   = 0.35                         # 单箱宽度（固定）
    jit_s   = 12                           # jitter 直径
    off     = np.array([-0.6, 0.0, 0.6])   # 低中高 HWS 的 y 偏移（固定）
    UNIT_HEIGHT = 1.8                      # 每个数据源占用的垂直空间（固定）
    
    # 为同一分辨率-高度子集遍历
    for r, res_key in enumerate(RES_LIST):
        srcs_this_row = RES_SRC_HWS[res_key]
        # 计算该分辨率需要的总高度
        total_height = UNIT_HEIGHT * len(srcs_this_row)
        # 计算y_base位置
        y_base = np.arange(len(srcs_this_row)) * UNIT_HEIGHT
        for c, h in enumerate(HEIGHTS):
            ax = axes[r][c]
            ax.axvline(0, color='black', lw=2, linestyle=(0,(5,10)))

            sub = df.query("resolution == @res_key and height == @h")
            if sub.empty:
                ax.text(0.5, 0.5, "No data", ha="center", va="center",
                        alpha=0.6)
                ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
                continue
                
            # ―― 逐源 & HWS 档画箱+散点 ―――――――――――――――――――――――
            for i, src in enumerate(srcs_this_row):
                for j, hws_bin in enumerate(HWS_ORDER):
                    errs = sub.loc[(sub["source"] == src) &
                                   (sub["hws_bin"] == hws_bin),
                                   "error"].values
                    if errs.size == 0:
                        continue
                        
                    # 箱图
                    ax.boxplot(errs, vert=False,
                               positions=[y_base[i] + off[j]],
                               widths=box_w,
                               patch_artist=True, manage_ticks=False,
                               boxprops=dict(facecolor=HWS_COLORS[hws_bin],
                                             alpha=0.75, linewidth=2),
                               flierprops=dict(marker="x", markersize=6,
                                               markeredgecolor="black",
                                               markerfacecolor="none",
                                               linewidth=0.5),
                               medianprops=dict(color="lime", linewidth=2),
                               whiskerprops=dict(color="black", linewidth=1.8),
                               capprops=dict(color="black", linewidth=1.8))
                               
                    # jitter
                    jitter_y = np.random.normal(y_base[i] + off[j], 0.06,
                                                size=len(errs))
                    ax.scatter(errs, jitter_y, s=jit_s,
                               color=HWS_COLORS[hws_bin], alpha=0.8,
                               linewidths=0)
                               
            # ―― 轴刻度 & 标签 ―――――――――――――――――――――――――――――――――
            ax.set_yticks(y_base)
            disp_names = [DISPLAY_NAMES.get(s, s) for s in srcs_this_row]
            ax.set_yticklabels(disp_names)
            # 设置ylim，上下各留0.5的边距
            ax.set_ylim(-1, total_height - 1)
            ax.grid(axis="x", linestyle=(0, (5, 8)), alpha=0.6)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_linewidth(2)
            ax.spines["bottom"].set_linewidth(2)
            
            # X-lim & tick 间距根据经验给定，可自行调整
            ax.set_xlim(-8, 8)
            ax.xaxis.set_major_locator(mticker.MultipleLocator(2))
            ax.xaxis.set_minor_locator(mticker.MultipleLocator(1))
            
            if r == ROWS - 1:
                ax.set_xlabel("Bias (m/s)")          # 仅底排标注 x-label
            if r == 0:
                ax.set_title(f"{h} m", pad=10, fontsize=26)
                
    # # 行标签（分辨率）
    # for r, res_key in enumerate(RES_LIST):
    #     axes[r][0].annotate(res_key, xy=(-0.40, 0.5),
    #                         xycoords="axes fraction", rotation=90,
    #                         va="center", ha="center", fontsize=26)
                            
    # ───────────────────────────────────────────────────────────
    # 图例：HWS 档
    handles_hws = [mlines.Line2D([], [], color=HWS_COLORS[b],
                                marker="s", linestyle="", markersize=15,
                                label=b) for b in HWS_ORDER]
    # 图例：数据源
    all_sources = set([src for src_list in RES_SRC_HWS.values() for src in src_list])
    handles_src = [mlines.Line2D([], [], color="black",
                             marker="s", linestyle="", markersize=15,
                             markerfacecolor=COLMAP[s],
                             label=DISPLAY_NAMES.get(s, s))
               for s in all_sources]
               
    fig = plt.gcf()
    fig.legend(handles_hws,
               [h.get_label() for h in handles_hws + handles_src],
               loc="upper center", ncol=7, frameon=False,
               fontsize=26, bbox_to_anchor=(0.51, 0.98))
               
    fig.text(0.52, -0.015,
             "HWS Bias = Ground-based remote sensing/Reanalysis − AWS",
             ha="center", fontsize=26)
             
    # 添加标题
    if title_suffix:
        fig.text(0.52, 0.98, f"HWS Conditional Bias {title_suffix}",
                ha="center", fontsize=26, fontweight="bold")
                
    fig.savefig(outfile, dpi=600, format="tif", transparent=True,
                bbox_inches="tight",
                pil_kwargs={"compression": "tiff_adobe_deflate"})
    plt.close(fig)          # ← 立即关闭，释放内存
    print("Saved →", outfile)
# =============================================================================
# 8. 统计函数
# =============================================================================
def summarize_rh_errors(df: pd.DataFrame) -> pd.DataFrame:
    """
    按 (resolution, height, rh_bin, source) 分组统计：
    count / mean / median / q25 / q75 / IQR / whisker_low / whisker_high / min / max
    """
    grp = df.groupby(["resolution", "height", "rh_bin", "source"])["error"]
    summary = grp.agg(
        count  = "size",
        mean   = "mean",
        median = "median",
        q25    = lambda x: x.quantile(0.25),
        q75    = lambda x: x.quantile(0.75),
        min    = "min",
        max    = "max",
    )
    summary["iqr"] = summary.q75 - summary.q25
    summary["whisker_low"]  = (summary.q25 - 1.5 * summary.iqr).clip(lower=summary["min"])
    summary["whisker_high"] = (summary.q75 + 1.5 * summary.iqr).clip(upper=summary["max"])
    # 更易读的列顺序
    return summary[[
        "count", "mean", "median", "q25", "q75",
        "iqr", "whisker_low", "whisker_high", "min", "max"
    ]]

def summarize_temp_errors(df: pd.DataFrame) -> pd.DataFrame:
    """
    按 (resolution, height, temp_bin, source) 分组统计：
    count / mean / median / q25 / q75 / IQR / whisker_low / whisker_high / min / max
    """
    grp = df.groupby(["resolution", "height", "temp_bin", "source"])["error"]
    summary = grp.agg(
        count  = "size",
        mean   = "mean",
        median = "median",
        q25    = lambda x: x.quantile(0.25),
        q75    = lambda x: x.quantile(0.75),
        min    = "min",
        max    = "max",
    )
    summary["iqr"] = summary.q75 - summary.q25
    summary["whisker_low"]  = (summary.q25 - 1.5 * summary.iqr).clip(lower=summary["min"])
    summary["whisker_high"] = (summary.q75 + 1.5 * summary.iqr).clip(upper=summary["max"])
    return summary[[
        "count", "mean", "median", "q25", "q75",
        "iqr", "whisker_low", "whisker_high", "min", "max"
    ]]

def summarize_hws_errors(df: pd.DataFrame) -> pd.DataFrame:
    """
    按 (resolution, height, hws_bin, source) 分组统计：
    count / mean / median / q25 / q75 / IQR / whisker_low / whisker_high / min / max
    """
    grp = df.groupby(["resolution", "height", "hws_bin", "source"])["error"]
    summary = grp.agg(
        count  = "size",
        mean   = "mean",
        median = "median",
        q25    = lambda x: x.quantile(0.25),
        q75    = lambda x: x.quantile(0.75),
        min    = "min",
        max    = "max",
    )
    summary["iqr"] = summary.q75 - summary.q25
    summary["whisker_low"]  = (summary.q25 - 1.5 * summary.iqr).clip(lower=summary["min"])
    summary["whisker_high"] = (summary.q75 + 1.5 * summary.iqr).clip(upper=summary["max"])
    return summary[[
        "count", "mean", "median", "q25", "q75",
        "iqr", "whisker_low", "whisker_high", "min", "max"
    ]]

# =============================================================================
# 9. 主执行流程
# =============================================================================
# 创建输出目录
OUT_DIR = Path(r"E:\Beijing2024\出图TIF")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 时间窗口格式化字符串
time_window_str = f"{EXCLUDE_WINDOWS[0][0].strftime('%Y-%m-%d %H:%M')} to {EXCLUDE_WINDOWS[-1][1].strftime('%Y-%m-%d %H:%M')}"

# ==================== RH 分析 ====================
print("Building RH conditional error dataframes...")

# 1. 包含时间窗口的数据
print("  - Inside time windows...")
df_rh_inside = build_rh_error_df(inside=True)
make_rh_plot(
    df_rh_inside,
    OUT_DIR / "RH_Conditional_Bias_BoxJitter_Inside.tif",
    title_suffix=f"(Inside {time_window_str})"
)
rh_stats_inside = summarize_rh_errors(df_rh_inside)
with pd.ExcelWriter(OUT_DIR / "RH_error_stats_inside.xlsx", engine="openpyxl") as writer:
    for (rh_bin, h), sub in rh_stats_inside.groupby(level=["rh_bin", "height"]):
        sheet_name = f"{rh_bin.split(' ')[0]}_{h}m"           
        sub.droplevel(["rh_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - RH inside statistics saved to {OUT_DIR / 'RH_error_stats_inside.xlsx'}")

# 2. 不包含时间窗口的数据
print("  - Outside time windows...")
df_rh_outside = build_rh_error_df(inside=False)
make_rh_plot(
    df_rh_outside,
    OUT_DIR / "RH_Conditional_Bias_BoxJitter_Outside.tif",
    title_suffix="(Exclude Rainstorm)"
)
rh_stats_outside = summarize_rh_errors(df_rh_outside)
with pd.ExcelWriter(OUT_DIR / "RH_error_stats_outside.xlsx", engine="openpyxl") as writer:
    for (rh_bin, h), sub in rh_stats_outside.groupby(level=["rh_bin", "height"]):
        sheet_name = f"{rh_bin.split(' ')[0]}_{h}m"           
        sub.droplevel(["rh_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - RH outside statistics saved to {OUT_DIR / 'RH_error_stats_outside.xlsx'}")

# 3. 全部数据
print("  - All data...")
df_rh_all = build_rh_error_df()  # 不传递inside参数
make_rh_plot(
    df_rh_all,
    OUT_DIR / "RH_Conditional_Bias_BoxJitter_All.tif",
    title_suffix="(All Data)"
)
rh_stats_all = summarize_rh_errors(df_rh_all)
with pd.ExcelWriter(OUT_DIR / "RH_error_stats_all.xlsx", engine="openpyxl") as writer:
    for (rh_bin, h), sub in rh_stats_all.groupby(level=["rh_bin", "height"]):
        sheet_name = f"{rh_bin.split(' ')[0]}_{h}m"           
        sub.droplevel(["rh_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - RH all statistics saved to {OUT_DIR / 'RH_error_stats_all.xlsx'}")

# ==================== Temp 分析 ====================
print("\nBuilding Temp conditional error dataframes...")

# 1. 包含时间窗口的数据
print("  - Inside time windows...")
df_temp_inside = build_temp_error_df(inside=True)
make_temp_plot(
    df_temp_inside,
    OUT_DIR / "Temp_Conditional_Bias_BoxJitter_Inside.tif",
    title_suffix=f"(Include {time_window_str})"
)
temp_stats_inside = summarize_temp_errors(df_temp_inside)
with pd.ExcelWriter(OUT_DIR / "Temp_error_stats_inside.xlsx", engine="openpyxl") as writer:
    for (t_bin, h), sub in temp_stats_inside.groupby(level=["temp_bin", "height"]):
        sheet_name = f"{t_bin.split(' ')[0]}_{h}m"    
        sub.droplevel(["temp_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - Temp inside statistics saved to {OUT_DIR / 'Temp_error_stats_inside.xlsx'}")

# 2. 不包含时间窗口的数据
print("  - Outside time windows...")
df_temp_outside = build_temp_error_df(inside=False)
make_temp_plot(
    df_temp_outside,
    OUT_DIR / "Temp_Conditional_Bias_BoxJitter_Outside.tif",
    title_suffix="(Exlude Rainstorm)"
)
temp_stats_outside = summarize_temp_errors(df_temp_outside)
with pd.ExcelWriter(OUT_DIR / "Temp_error_stats_outside.xlsx", engine="openpyxl") as writer:
    for (t_bin, h), sub in temp_stats_outside.groupby(level=["temp_bin", "height"]):
        sheet_name = f"{t_bin.split(' ')[0]}_{h}m"    
        sub.droplevel(["temp_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - Temp outside statistics saved to {OUT_DIR / 'Temp_error_stats_outside.xlsx'}")

# 3. 全部数据
print("  - All data...")
df_temp_all = build_temp_error_df()  # 不传递inside参数
make_temp_plot(
    df_temp_all,
    OUT_DIR / "Temp_Conditional_Bias_BoxJitter_All.tif",
    title_suffix="(All Data)"
)
temp_stats_all = summarize_temp_errors(df_temp_all)
with pd.ExcelWriter(OUT_DIR / "Temp_error_stats_all.xlsx", engine="openpyxl") as writer:
    for (t_bin, h), sub in temp_stats_all.groupby(level=["temp_bin", "height"]):
        sheet_name = f"{t_bin.split(' ')[0]}_{h}m"    
        sub.droplevel(["temp_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - Temp all statistics saved to {OUT_DIR / 'Temp_error_stats_all.xlsx'}")

# ==================== HWS 分析 ====================
print("\nBuilding HWS conditional error dataframes...")

# 1. 包含时间窗口的数据
print("  - Inside time windows...")
df_hws_inside = build_hws_error_df(inside=True)
make_hws_plot(
    df_hws_inside,
    OUT_DIR / "HWS_Conditional_Bias_BoxJitter_Inside.tif",
    title_suffix=f"(Include {time_window_str})"
)
hws_stats_inside = summarize_hws_errors(df_hws_inside)
with pd.ExcelWriter(OUT_DIR / "HWS_error_stats_inside.xlsx", engine="openpyxl") as writer:
    for (hws_bin, h), sub in hws_stats_inside.groupby(level=["hws_bin", "height"]):
        sheet_name = f"{hws_bin.split(' ')[0]}_{h}m"    
        sub.droplevel(["hws_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - HWS inside statistics saved to {OUT_DIR / 'HWS_error_stats_inside.xlsx'}")

# 2. 不包含时间窗口的数据
print("  - Outside time windows...")
df_hws_outside = build_hws_error_df(inside=False)
make_hws_plot(
    df_hws_outside,
    OUT_DIR / "HWS_Conditional_Bias_BoxJitter_Outside.tif",
    title_suffix="(Exclude Rainstorm)"
)
hws_stats_outside = summarize_hws_errors(df_hws_outside)
with pd.ExcelWriter(OUT_DIR / "HWS_error_stats_outside.xlsx", engine="openpyxl") as writer:
    for (hws_bin, h), sub in hws_stats_outside.groupby(level=["hws_bin", "height"]):
        sheet_name = f"{hws_bin.split(' ')[0]}_{h}m"    
        sub.droplevel(["hws_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - HWS outside statistics saved to {OUT_DIR / 'HWS_error_stats_outside.xlsx'}")

# 3. 全部数据
print("  - All data...")
df_hws_all = build_hws_error_df()  # 不传递inside参数
make_hws_plot(
    df_hws_all,
    OUT_DIR / "HWS_Conditional_Bias_BoxJitter_All.tif",
    title_suffix="(All Data)"
)
hws_stats_all = summarize_hws_errors(df_hws_all)
with pd.ExcelWriter(OUT_DIR / "HWS_error_stats_all.xlsx", engine="openpyxl") as writer:
    for (hws_bin, h), sub in hws_stats_all.groupby(level=["hws_bin", "height"]):
        sheet_name = f"{hws_bin.split(' ')[0]}_{h}m"    
        sub.droplevel(["hws_bin", "height"]).to_excel(writer, sheet_name=sheet_name)
print(f"  - HWS all statistics saved to {OUT_DIR / 'HWS_error_stats_all.xlsx'}")

print("\nAll analysis completed successfully!")








import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.ticker as mticker
import matplotlib.lines as mlines
from pathlib import Path

# =============================================================================
# 0. 全局常量（修改DISPLAY_NAMES）
# =============================================================================
HEIGHTS = [80, 140, 200, 280]
COLMAP = {"MWR":"red", "TowerTurb":"slategray",
          "CRA":"magenta", "ERA5":"blue", "LiDAR":"darkorange"}
# 修改：将Turb.改为ECFMS
DISPLAY_NAMES = {"TowerTurb":"ECM", "LiDAR":"DWL"}

# ---- RH 档划分 & 颜色 ----
RH_BINS = {
    "Low RH (<50%)"      : (-np.inf, 50),
    "Middle RH (50–80%)"    : (50, 80),
    "High RH (>80%)"     : (80,  np.inf),
}
RH_ORDER  = list(RH_BINS.keys())
RH_COLORS = {
    "Low RH (<50%)"   : "#1f77b4",
    "Middle RH (50–80%)" : "#2ca02c",
    "High RH (>80%)"  : "#d62728",
}

# ---- Temp 档划分 & 颜色 ----
TEMP_BINS = {
    "Low T (<300 K)"  : (-np.inf, 27+273),
    "Middle T (300–303 K)": (27+273 , 30+273),
    "High T (≥303 K)" : (30+273 , np.inf),
}
TEMP_ORDER  = list(TEMP_BINS.keys())
TEMP_COLORS = {
    "Low T (<300 K)"  : "#1f77b4",
    "Middle T (300–303 K)": "#2ca02c",
    "High T (≥303 K)" : "#d62728",
}

# ---- HWS 档划分 & 颜色 ----
HWS_BINS = {
    "Low HWS (<3 m/s)": (-np.inf, 3),
    "Middle HWS (3–7 m/s)": (3, 7),
    "High HWS (≥7 m/s)": (7, np.inf),
}
HWS_ORDER = list(HWS_BINS.keys())
HWS_COLORS = {
    "Low HWS (<3 m/s)": "#1f77b4",
    "Middle HWS (3–7 m/s)": "#2ca02c",
    "High HWS (≥7 m/s)": "#d62728",
}

# ---- 时间窗口定义 ----
EXCLUDE_WINDOWS = [
    (pd.Timestamp("2024-08-09 12:46"), pd.Timestamp("2024-08-10 00:03")),
    (pd.Timestamp("2024-06-07 09:15"), pd.Timestamp("2024-06-07 15:00")),
    (pd.Timestamp("2024-08-10 16:20"), pd.Timestamp("2024-08-10 17:30"))
]

# ---- 60分钟分辨率的数据源 ----
RES_SRC_RH_60min = RES_SRC_TEMP_60min = ["TowerTurb", "MWR", "ERA5", "CRA"]
RES_SRC_HWS_60min = ["LiDAR", "TowerTurb", "ERA5", "CRA"]

# =============================================================================
# 1. 通用函数：时间窗口过滤（与原代码相同）
# =============================================================================
def filter_by_time_windows(series, windows, inside=True):
    if not windows:
        return series
    
    mask = pd.Series(False, index=series.index)
    
    for start, end in windows:
        window_mask = (series.index >= start) & (series.index <= end)
        if inside:
            mask = mask | window_mask
        else:
            pass
    
    if inside:
        return series[mask]
    else:
        inside_mask = pd.Series(False, index=series.index)
        for start, end in windows:
            inside_mask = inside_mask | ((series.index >= start) & (series.index <= end))
        return series[~inside_mask]

# =============================================================================
# 2. 数据构建函数（只处理60分钟数据）
# =============================================================================
def build_rh_error_df_60min(inside=None):
    """构建60分钟RH误差数据"""
    recs = []
    sdict = series_dict_60min  # 假设这个变量已定义
    if sdict is None:
        return pd.DataFrame()
        
    for h in HEIGHTS:
        if "TowerAvg" not in sdict["RH"][h]:
            continue
        ref = sdict["RH"][h]["TowerAvg"].dropna()
        
        if inside is not None:
            ref = filter_by_time_windows(ref, EXCLUDE_WINDOWS, inside=inside)
        
        for src in RES_SRC_RH_60min:
            if src not in sdict["RH"][h]:
                continue
            tgt = sdict["RH"][h][src].dropna()
            
            if inside is not None:
                tgt = filter_by_time_windows(tgt, EXCLUDE_WINDOWS, inside=inside)
            
            aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
            diff = aligned.iloc[:, 1] - aligned.iloc[:, 0]
            ref_rh = aligned.iloc[:, 0]
            
            for ts, e in diff.items():
                rh_val = ref_rh.loc[ts]
                for rh_bin, (lo, hi) in RH_BINS.items():
                    if lo <= rh_val < hi:
                        recs.append({
                            "height"    : h,
                            "rh_bin"    : rh_bin,
                            "source"    : src,
                            "error"     : e,
                        })
                        break
    return pd.DataFrame.from_records(recs)

def build_temp_error_df_60min(inside=None):
    """构建60分钟温度误差数据"""
    recs = []
    sdict = series_dict_60min  # 假设这个变量已定义
    if sdict is None:
        return pd.DataFrame()
        
    for h in HEIGHTS:
        if "TowerAvg" not in sdict["Temp"][h]:
            continue
        ref = sdict["Temp"][h]["TowerAvg"].dropna()
        
        if inside is not None:
            ref = filter_by_time_windows(ref, EXCLUDE_WINDOWS, inside=inside)
        
        for src in RES_SRC_TEMP_60min:
            if src not in sdict["Temp"][h]:
                continue
            tgt = sdict["Temp"][h][src].dropna()
            
            if inside is not None:
                tgt = filter_by_time_windows(tgt, EXCLUDE_WINDOWS, inside=inside)
            
            aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
            diff = aligned.iloc[:, 1] - aligned.iloc[:, 0]
            ref_t = aligned.iloc[:, 0]
            
            for ts, e in diff.items():
                t_val = ref_t.loc[ts]
                for t_bin, (lo, hi) in TEMP_BINS.items():
                    if lo <= t_val < hi:
                        recs.append({
                            "height"    : h,
                            "temp_bin"  : t_bin,
                            "source"    : src,
                            "error"     : e,
                        })
                        break
    return pd.DataFrame.from_records(recs)

def build_hws_error_df_60min(inside=None):
    """构建60分钟风速误差数据"""
    recs = []
    sdict = series_dict_60min  # 假设这个变量已定义
    if sdict is None:
        return pd.DataFrame()
        
    for h in HEIGHTS:
        if "TowerAvg" not in sdict["HWS"][h]:
            continue
        ref = sdict["HWS"][h]["TowerAvg"].dropna()
        
        if inside is not None:
            ref = filter_by_time_windows(ref, EXCLUDE_WINDOWS, inside=inside)
        
        for src in RES_SRC_HWS_60min:
            if src not in sdict["HWS"][h]:
                continue
            tgt = sdict["HWS"][h][src].dropna()
            
            if inside is not None:
                tgt = filter_by_time_windows(tgt, EXCLUDE_WINDOWS, inside=inside)
            
            aligned = pd.concat([ref, tgt], axis=1, join="inner").dropna()
            diff = aligned.iloc[:, 1] - aligned.iloc[:, 0]
            ref_hws = aligned.iloc[:, 0]
            
            for ts, e in diff.items():
                hws_val = ref_hws.loc[ts]
                for hws_bin, (lo, hi) in HWS_BINS.items():
                    if lo <= hws_val < hi:
                        recs.append({
                            "height"    : h,
                            "hws_bin"   : hws_bin,
                            "source"    : src,
                            "error"     : e,
                        })
                        break
    return pd.DataFrame.from_records(recs)

# =============================================================================
# 3. 绘图函数：组合温度、湿度和风速误差
# =============================================================================
plt.rcParams.update(plt.rcParamsDefault)
panel_labels = ["(a)", "(b)", "(c)", "(d)", "(e)", "(f)"]
def make_combined_plot(outfile: Path):
    """创建包含温度、湿度和风速误差的组合图"""
    # 设置全局字体为Arial，字号为22
    plt.rcParams.update({
        'font.size': 22,
        'font.family': 'sans-serif',                
        'font.sans-serif': ['Arial', 'DejaVu Sans'],     
        'axes.titlepad': 10,
        'legend.fontsize': 22,
        'axes.labelsize': 22,
        'xtick.labelsize': 22,
        'ytick.labelsize': 22,
        'xtick.major.size': 12,
        'xtick.minor.size': 7,
        'ytick.major.size': 12,
        'ytick.minor.size': 7,
        'xtick.direction': 'out',
        'ytick.direction': 'out',
        'xtick.major.width': 2,
        'xtick.minor.width': 2,
        'ytick.major.width': 2,
        'ytick.minor.width': 2,
    })
    
    # 创建6行4列的图形，使用GridSpec以便更精确控制间距
    ROWS, COLS = 6, len(HEIGHTS)
    FIGSIZE = (20, 40)  # 增加高度以容纳图例
    fig = plt.figure(figsize=FIGSIZE)
    
    # 使用GridSpec并设置高度比例，以便控制变量内部分组间距
    # 每个变量组占2行，变量组之间间距为0.5，变量内部分组间距为0.2
    height_ratios = [1, 1, 0.1, 1, 1, 0.1, 1, 1, 0.1, 1, 1]  # 6个子图+5个间距
    gs = gridspec.GridSpec(11, COLS, height_ratios=height_ratios, 
                          hspace=0.35, wspace=0.5)
    
    # 创建子图数组，跳过间距行
    axes = []
    row_indices = [0, 1, 3, 4, 6, 7]  # 实际子图所在的行索引
    for r in row_indices:
        row_axes = []
        for c in range(COLS):
            row_axes.append(fig.add_subplot(gs[r, c]))
        axes.append(row_axes)
    
    # 布局参数
    box_w = 0.4
    jit_s = 12
    off = np.array([-0.8, 0.0, 0.8])
    row_spacing = 2.2
    
    # 行标签和变量设置 - 修改顺序：include在上，exclude在下
    row_settings = [
        # (变量类型, 分组, 数据源列表, 颜色字典, 档顺序, x轴标签, x轴范围, x轴刻度)
        ("temp", "exclude", RES_SRC_TEMP_60min, TEMP_COLORS, TEMP_ORDER, "Bias (K)", (-9, 9.5), 3),
        ("temp", "include", RES_SRC_TEMP_60min, TEMP_COLORS, TEMP_ORDER, "Bias (K)", (-9, 9.5), 3),
        
        ("rh", "exclude", RES_SRC_RH_60min, RH_COLORS, RH_ORDER, "Bias (%)", (-54, 50), 20),
        ("rh", "include", RES_SRC_RH_60min, RH_COLORS, RH_ORDER, "Bias (%)", (-54, 50), 20),

        ("hws", "exclude", RES_SRC_HWS_60min, HWS_COLORS, HWS_ORDER, "Bias (m/s)", (-9.5, 9), 3),
        ("hws", "include", RES_SRC_HWS_60min, HWS_COLORS, HWS_ORDER, "Bias (m/s)", (-9.5, 9), 3),
    ]
    
    # 构建数据
    df_temp_exclude = build_temp_error_df_60min(inside=False)
    df_temp_include = build_temp_error_df_60min(inside=True)
    df_rh_exclude = build_rh_error_df_60min(inside=False)
    df_rh_include = build_rh_error_df_60min(inside=True)
    df_hws_exclude = build_hws_error_df_60min(inside=False)
    df_hws_include = build_hws_error_df_60min(inside=True)
    
    # 数据映射
    data_map = {
        ("temp", "exclude"): df_temp_exclude,
        ("temp", "include"): df_temp_include,
        ("rh", "exclude"): df_rh_exclude,
        ("rh", "include"): df_rh_include,
        ("hws", "exclude"): df_hws_exclude,
        ("hws", "include"): df_hws_include,
    }
    
    # 变量名缩写映射
    VAR_ABBR = {"temp": "Temp", "rh": "RH", "hws": "HWS"}
    #VAR_ABBR1 = {"temp": "Temp", "rh": "RH", "hws": "HWS"}
    # 绘制每个子图
    for r, (var, group, src_list, colors, bin_order, xlabel, xlim, xtick) in enumerate(row_settings):
        y_base = np.arange(len(src_list)) * row_spacing
        df = data_map[(var, group)]
        
        for c, h in enumerate(HEIGHTS):
            ax = axes[r][c]
            ax.axvline(0, color='black', lw=1.5, linestyle=(0,(5,10)))
            
            sub = df.query("height == @h")
            if sub.empty:
                ax.text(0.5, 0.5, "No data", ha="center", va="center", alpha=0.6)
                ax.set_xticks([]); ax.set_yticks([]); ax.set_frame_on(False)
                continue
            
            # 绘制箱线图和散点
            for i, src in enumerate(src_list):
                for j, bin_name in enumerate(bin_order):
                    if var == "temp":
                        errs = sub.loc[(sub["source"] == src) & (sub["temp_bin"] == bin_name), "error"].values
                    elif var == "rh":
                        errs = sub.loc[(sub["source"] == src) & (sub["rh_bin"] == bin_name), "error"].values
                    else:  # hws
                        errs = sub.loc[(sub["source"] == src) & (sub["hws_bin"] == bin_name), "error"].values
                    
                    if errs.size == 0:
                        continue
                    
                    # 箱线图
                    ax.boxplot(errs, vert=False,
                               positions=[y_base[i] + off[j]],
                               widths=box_w,
                               patch_artist=True, manage_ticks=False,
                               boxprops=dict(facecolor=colors[bin_name], alpha=0.75, linewidth=1.5),
                               flierprops=dict(marker="x", markersize=5, markeredgecolor="black", 
                                               markerfacecolor="none", linewidth=0.5),
                               medianprops=dict(color="lime", linewidth=1.5),
                               whiskerprops=dict(color="black", linewidth=1.5),
                               capprops=dict(color="black", linewidth=1.5))
                    
                    # 散点
                    jitter_y = np.random.normal(y_base[i] + off[j], 0.06, size=len(errs))
                    ax.scatter(errs, jitter_y, s=jit_s, color=colors[bin_name], alpha=0.8, linewidths=0)
            
            # 设置轴标签和刻度
            ax.set_yticks(y_base)
            disp_names = [DISPLAY_NAMES.get(s, s) for s in src_list]
            ax.set_yticklabels(disp_names)
            ax.set_ylim(y_base[0] - 1.5, y_base[-1] + 1.5)
            ax.grid(axis="x", linestyle=(0, (5, 8)), alpha=0.6)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_linewidth(1.5)
            ax.spines["bottom"].set_linewidth(1.5)
            
            ax.set_xlim(xlim)
            ax.xaxis.set_major_locator(mticker.MultipleLocator(xtick))
            ax.xaxis.set_minor_locator(mticker.MultipleLocator(xtick/2))
            
            # 只在最左侧添加变量标签，在最底部添加x轴标签
            if c == 0:
               var_abbr = VAR_ABBR[var]
               if group == "exclude":
                   label_text = f"{var_abbr}\nExclude Rainstorm"
               else:
                   label_text = f"{var_abbr}\nInclude Rainstorm"
               ax.set_ylabel(label_text, fontsize=22)
            else:
               ax.set_ylabel("")
            
            # 在顶部添加高度标签
            if r == 0:
                ax.set_title(f"{h} m", fontsize=22, pad=10,fontweight="bold")
            if c == 0:
                ax.text(-0.25, 1.02, panel_labels[r],transform=ax.transAxes,fontsize=26,fontweight="bold", va="bottom", ha="right")
    # 添加变量组图例 - 调整位置以适应新的布局
    # 温度组图例（放在第0行上方）
    handles_temp = [mlines.Line2D([], [], color=TEMP_COLORS[b], marker="s", linestyle="", 
                                 markersize=12, label=b) for b in TEMP_ORDER]
    temp_legend = fig.legend(handles_temp, [h.get_label() for h in handles_temp],
                            loc='center', bbox_to_anchor=(0.5, 0.90), ncol=3, frameon=False,
                           title_fontsize=22)
    
    # 湿度组图例（放在第2行上方）
    handles_rh = [mlines.Line2D([], [], color=RH_COLORS[b], marker="s", linestyle="", 
                              markersize=12, label=b) for b in RH_ORDER]
    rh_legend = fig.legend(handles_rh, [h.get_label() for h in handles_rh],
                          loc='center', bbox_to_anchor=(0.5, 0.695), ncol=3, frameon=False,
                          title="Temperature bias(K)\n", title_fontsize=22)
    
    # 风速组图例（放在第4行上方）
    handles_hws = [mlines.Line2D([], [], color=HWS_COLORS[b], marker="s", linestyle="", 
                               markersize=12, label=b) for b in HWS_ORDER]
    hws_legend = fig.legend(handles_hws, [h.get_label() for h in handles_hws],
                           loc='center', bbox_to_anchor=(0.5, 0.49), ncol=3, frameon=False,
                           title="Relative Humidity bias(%)\n", title_fontsize=22)
    fig.text(
        0.5, 0.295,                  # 0.04 可以根据效果微调高低
        "Horizontal Wind Speed bias(m/s)",
        ha="center", va="center",
        fontsize=22
    )
    # 保存图形
    fig.savefig(outfile, dpi=600, format="tif", transparent=True, bbox_inches="tight",
                pil_kwargs={"compression": "tiff_adobe_deflate"})
    plt.close(fig)
    print(f"Saved → {outfile}")

# =============================================================================
# 4. 主执行流程
# =============================================================================
# 假设 series_dict_60min 已定义（包含60分钟分辨率的数据）

# 创建输出目录
OUT_DIR = Path(r"E:\Beijing2024\出图TIF")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# 生成组合图
make_combined_plot(OUT_DIR / "Combined_Conditional_Bias_60min.tif")

print("Combined plot generation completed successfully!")

gc.collect()















################### 绘制温度(探空时刻±30min平均版) – 带 Inset ###################
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import metpy.calc as mpcalc
from metpy.units import units
from matplotlib.ticker import MultipleLocator
from mpl_toolkits.axes_grid1.inset_locator import inset_axes  # ← 新增：inset 工具

# —————————————————— 用户配置 ——————————————————
excel_path   = r"E:\Beijing2024\探空数据-54511\original combined2.xlsx"
sheet_names  = pd.ExcelFile(excel_path).sheet_names           # 全部时次
out_dir      = r"E:\Beijing2024\出图TIF"
out_fname    = "Temperature_Profile_1h平均_with_inset.tif"  # ← 文件名更新
os.makedirs(out_dir, exist_ok=True)
site_elev = 49

# —————————————————— 统一的绘图风格 ——————————————————
plt.rcParams.update({
    'font.family': 'sans-serif',                # 先指定家族
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     # 把 Arial 设为首选
    'font.size'      : 36,
    'axes.linewidth' : 3,
    'axes.labelsize' : 36,
    'xtick.labelsize': 36,
    'ytick.labelsize': 36
})

# 标准高度/气压刻度
alt_ticks = np.array([0, 500, 1000, 1500, 2000, 2500, 3000, 4000, 6000, 8000, 10000])   # m
p_ticks   = np.array([1000, 925, 850, 700, 500, 275])                         # hPa
p2alt = lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude

# —————————————————— 建立 4×3 画布 ——————————————————
n_rows, n_cols = 4, 3
fig, axes = plt.subplots(n_rows, n_cols, figsize=(30, 40), sharex=False, sharey=False)
axes = axes.ravel()             # 方便用 1-D 索引
### ← 新增：先准备图例用到的句柄与标签 ###
legend_handles = [
    plt.Line2D([], [], color='blue',    lw=6, label='ERA5'),
    plt.Line2D([], [], color='magenta', lw=6, label='CRA'),
    plt.Line2D([], [], color='red',     lw=6, label='MWR'),
    plt.Line2D([], [], color='black',   lw=6, label='Radiosonde')
]

# —————————————————— 循环所有时次 ——————————————————
for idx, sheet in enumerate(sheet_names):
    if idx >= n_rows * n_cols:            # 超出 12 张子图就跳过
        break

    ax = axes[idx]
    ax.spines['bottom'].set_color('red')
    ax.spines['bottom'].set_linewidth(ax.spines['left'].get_linewidth())
    ax.spines['left'].set_color('red')
    # 1) 读取探空表
    snd = pd.read_excel(excel_path, sheet_name=sheet)

    # 2) 当前 target_time（由工作表名解析）
    target_time = pd.to_datetime(sheet, format='%Y%m%d%H')

    # 3) 按原逻辑过滤/处理 ERA5、CRA、MWR
    era5_profile = ERA5_interp_BJ[ERA5_interp_BJ['valid_time'] == target_time].copy()
    cra_regular['timestamps'] = pd.to_datetime(cra_regular['timestamps'])
    cra_profile = cra_regular[cra_regular['timestamps'] == target_time].copy()
    # 处理MWR数据：取探空时刻前后半小时平均
    start_time = target_time - pd.Timedelta(minutes=30)
    end_time = target_time + pd.Timedelta(minutes=30)
    mwr_filtered = t_mwr_long.reset_index()  # 将双重索引转换为列
    mask = (mwr_filtered['timestamps'] >= start_time) & (mwr_filtered['timestamps'] <= end_time)
    mwr_filtered = mwr_filtered.loc[mask]
    if not mwr_filtered.empty:
        mwr_avg = mwr_filtered.groupby('height', as_index=False)['T(K)'].mean()
        mwr_profile = mwr_avg.sort_values('height')
    else:
        print(f"⚠️ 当前时次 {target_time.strftime('%Y-%m-%d %H:%M')} 的 MWR 数据不存在")
        mwr_profile = pd.DataFrame(columns=['height', 'T(K)'])  # 空DataFrame避免报错
    
    # — 单位转换 —
    era5_profile['height'] = mpcalc.pressure_to_height_std(
        era5_profile['pressure_level'].to_list() * units.hPa
    ).to('m').magnitude - site_elev
    cra_profile['height'] = mpcalc.pressure_to_height_std(
        cra_profile['Level (hPa)'].to_list() * units.hPa
    ).to('m').magnitude - site_elev
    snd['temperature_K'] = snd['temperature_C'] + 273.15
    snd['height'] = mpcalc.pressure_to_height_std(
        snd['pressure_hPa'].to_list() * units.hPa
    ).to('m').magnitude - site_elev

    # — 排序 —
    era5_profile = era5_profile.sort_values('height')
    cra_profile  = cra_profile.sort_values('height').iloc[:37]
    snd          = snd.sort_values('height')

    # ———————————————— 绘图（主图：0–3000 m） ————————————————
    ax.plot(era5_profile['t'],       era5_profile['height'],color='blue',    lw=5)
    ax.plot(cra_profile['Temp (K)'], cra_profile['height'], color='magenta', lw=5)
    ax.plot(mwr_profile['T(K)'],     mwr_profile['height'], color='red',     lw=5)
    ax.plot(snd['temperature_K'],    snd['height'],         color='black',   lw=5)
    
    # 主图 y 范围：0–3000 m
    ax.set_ylim(0, 3500)  # ← 修改：限定主图高度
    ax.set_xlim(275, 310)
    # — 虚线参考高度/气压 —
    for y in alt_ticks:
        ax.axhline(y, color='red',   lw=2, ls=(0, (5, 10)), zorder=0)
    for p in p_ticks:
        ax.axhline(p2alt(p), color='blue', lw=2, ls=(0, (5, 10)), zorder=0)
    # — 左右双轴 —
    secax = ax.secondary_yaxis(
        'right',
        functions=(
            lambda h: mpcalc.height_to_pressure_std(h * units.m).to('hPa').magnitude,
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude
        )
    )
    secax.set_ylim(1000, 700)  # 对应 0–3000 m 的压强范围近似

    # y-ticks（主）
    ax.set_yticks([0, 500, 1000, 1500, 2000, 2500, 3000, 3500])
    secax.set_yticks([1000, 925, 900, 850, 700])

    # x 轴细节
    ax.xaxis.set_major_locator(MultipleLocator(10))
    ax.xaxis.set_minor_locator(MultipleLocator(5))    
    ax.tick_params(axis='x', colors='red', length=12, width=2)
    ax.tick_params(axis='x', which='minor', length=6, width=2, colors='red')
    if row == 3 and col == 0:
        ax.set_xlabel('Temperature (K)', color='red', fontsize=50, weight='bold')

    # y 轴刻度颜色
    ax.tick_params(axis='y', colors='red', length=12, width=2)
    secax.tick_params(axis='y', colors='blue', length=12, width=2)

    # ———————————————— Inset：3000–10000 m ————————————————
    axins = inset_axes(ax, width="31%", height="48%", loc='upper right', borderpad=0.8)  # ← 新增
       # ★★★ 关键：给 inset 轴加一块不透明背景并抬高 z-order ★★★
    axins.set_facecolor('white')         # 用白色把主图线条遮掉
    axins.patch.set_alpha(1)             # 保证不透明
    axins.patch.set_zorder(2)            # z-order 要高于主图里 axhline 的 0
    # 画同一条曲线
    axins.plot(era5_profile['t'],      era5_profile['height'], color='blue',    lw=3)
    axins.plot(cra_profile['Temp (K)'], cra_profile['height'], color='magenta', lw=3)
    axins.plot(mwr_profile['T(K)'],     mwr_profile['height'], color='red',     lw=3)
    axins.plot(snd['temperature_K'],    snd['height'],         color='black',   lw=3)

    axins.set_xlim(230, 290)
    axins.set_ylim(3500, 10000)  # 只看高空
    axins.set_facecolor('white')
    # Inset 轴刻度更小即可（关闭 minor）
    axins.xaxis.set_major_locator(MultipleLocator(20))
    axins.yaxis.set_major_locator(MultipleLocator(2000))
    # —— 统一设置 x 轴元素为红色 ——
    axins.tick_params(axis='x', which='both', colors='red', labelsize=30,
                      width=ax.spines['left'].get_linewidth(), length=8)
    axins.tick_params(axis='y', which='both', labelsize=30,
                      width=ax.spines['left'].get_linewidth(), length=8)
    # —— Inset spines 设置 ——
    axins.spines['top'  ].set_visible(False)   # ← 顶部不可见
    axins.spines['right'].set_visible(False)   # ← 右侧不可见
    axins.spines['bottom'].set_color('red')
    # 底 / 左保持同等线宽
    for spine in ['left', 'bottom']:
        axins.spines[spine].set_linewidth(ax.spines['left'].get_linewidth())

    # 不需要次轴、标签
    axins.set_xlabel('', color='red')
    axins.set_ylabel('')

    # ———————— 轴刻度、标签管理（主图） ————————
    col = idx % n_cols
    row = idx // n_cols

    # 仅保留左/右边 & 最底行的轴标签
    if col == 0:                # 最左列 → 高度标签
        ax.set_ylabel(' ', color='black')
    else:
        ax.set_ylabel('')
        ax.tick_params(axis='y', labelleft=False)

    if col == n_cols - 1:       # 最右列 → Pressure 标签
        secax.set_ylabel(' ', color='black', rotation=270, labelpad=20)
    else:
        secax.set_ylabel('')
        secax.tick_params(axis='y', labelright=False)

    if row == n_rows - 1:       # 最底行 → Temperature 标签
        ax.xaxis.set_major_locator(MultipleLocator(10))
        ax.xaxis.set_minor_locator(MultipleLocator(5))
        ax.tick_params(axis='x', which='major', length=12, width=2, colors='red')
        ax.tick_params(axis='x', which='minor', length=6, width=2, colors='red')
    else:
        ax.set_xlabel('', color='red')
        ax.tick_params(axis='x', labelbottom=False, colors='red')
    if row == 3 and col == 1:
        ax.set_xlabel('Temperature (K)', color='red', weight='bold')
    
    
    # 仅在 Inset 上绘制高层虚线高度
    for y in alt_ticks[(alt_ticks >= 3500) & (alt_ticks <= 10000)]:
        axins.axhline(y, color='black', lw=2, ls=(0, (5, 5)), zorder=0)
    # — 子图标题 —
    ax.set_title(target_time.strftime('%Y-%m-%d %H:%M (BJT)'), pad=12, fontsize=36, weight='bold')
    axins.set_ylim(3500, 10100)
    axins.set_xlim(230,  290)
    axins.spines['left'].set_color('red')
    axins.tick_params(axis='y', colors='red')
# —————————————————— 隐藏空白子图（如果有） ——————————————————
for j in range(idx + 1, n_rows * n_cols):
    fig.delaxes(axes[j])

fig.text(0.01, 0.5, 'Height (m a.g.l.)', rotation=90, va='center', ha='left',
         color='black', weight='bold', fontsize=50)
fig.text(0.995, 0.5, 'Pressure (hPa)', rotation=270, va='center', ha='right',
         color='black', weight='bold', fontsize=50)

# —————————————————— 保存高分辨率 TIF ——————————————————
fig.legend(handles=legend_handles,
           loc='upper center',
           bbox_to_anchor=(0.5, 0.995),
           ncol=4,
           frameon=False,
           fontsize=32)

fig.text(0.5, 0.993,
         'Temperature Radiosonde Comparison – 1h Temporal Resolution',
         ha='center', va='bottom',
         fontsize=40, weight='bold')

fig.tight_layout(rect=[0, 0, 1, 0.98])
save_path = os.path.join(out_dir, out_fname)
fig.savefig(save_path, dpi=300, #transparent=True,
            pil_kwargs={"compression": "tiff_adobe_deflate"},
            format='tif', bbox_inches='tight')
print(f"✅ 已保存到：{save_path}")






################### 绘制温度(探空匹配版) — 带 Inset ###################
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import metpy.calc as mpcalc
from metpy.units import units
from matplotlib.ticker import MultipleLocator
from scipy.interpolate import interpn
from mpl_toolkits.axes_grid1.inset_locator import inset_axes   # ← 新增：inset 工具

# —————————————————— 用户配置 ——————————————————
excel_path   = r"E:\\Beijing2024\\探空数据-54511\\original combined2.xlsx"
sheet_names  = pd.ExcelFile(excel_path).sheet_names           # 全部时次
out_dir      = r"E:\Beijing2024\出图TIF"
out_fname    = "Temperature_Profile_MWR匹配探空-1min_with_inset2.tif"
os.makedirs(out_dir, exist_ok=True)
site_elev = Altitude   # ← 请替换为实际海拔高度 (m)

# 提取 t_mwr_long 的网格点（时间戳和高度）
time_points   = t_mwr_long.index.get_level_values('timestamps').unique().sort_values()
height_points = t_mwr_long.index.get_level_values('height').unique().sort_values()

# 将时间转换为 UNIX 秒
time_numeric = (time_points - pd.Timestamp("1970-01-01")) // pd.Timedelta('1s')

# 提取 t_mwr_long 的温度数据并重塑为二维网格
values_grid = t_mwr_long.unstack().values  # 形状 = (len(time_points), len(height_points))

# —————————————————— 统一的绘图风格 ——————————————————
plt.rcParams.update({
    'font.family'      : 'sans-serif',
    'font.sans-serif'  : ['Arial', 'DejaVu Sans'],
    'font.size'        : 36,
    'axes.linewidth'   : 3,
    'axes.labelsize'   : 36,
    'xtick.labelsize'  : 36,
    'ytick.labelsize'  : 36
})


alt_ticks = np.array([0, 500, 1000, 1500, 2000, 2500, 3000, 4000, 6000, 8000, 10000])   # m
p_ticks   = np.array([1000, 925, 850, 700, 500, 275])     
p2alt = lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude

# —————————————————— 建立 4×3 画布 ——————————————————
n_rows, n_cols = 4, 3
fig, axes = plt.subplots(n_rows, n_cols, figsize=(30, 40), sharex=False, sharey=False)
axes = axes.ravel()  # 方便用 1‑D 索引

# 图例用到的句柄与标签 (主图 & inset 共用)
legend_handles = [
    #plt.Line2D([], [], color='blue',    lw=6, label='ERA5'),
    #plt.Line2D([], [], color='magenta', lw=6, label='CRA'),
    plt.Line2D([], [], color='red',     lw=6, label='MWR'),
    plt.Line2D([], [], color='black',   lw=6, label='Radiosonde')
]

# —————————————————— 循环所有时次 ——————————————————
for idx, sheet in enumerate(sheet_names):
    if idx >= n_rows * n_cols:   # 超出 12 张子图就跳过
        break

    ax = axes[idx]

    # 1) 读取探空表
    snd = pd.read_excel(excel_path, sheet_name=sheet)

    # 2) 当前 target_time（由工作表名解析）
    target_time = pd.to_datetime(sheet, format='%Y%m%d%H')

    # 3) 按原逻辑过滤 / 处理 ERA5、CRA、MWR
    era5_profile = ERA5_interp_BJ[ERA5_interp_BJ['valid_time'] == target_time].copy()
    cra_regular['timestamps'] = pd.to_datetime(cra_regular['timestamps'])
    cra_profile = cra_regular[cra_regular['timestamps'] == target_time].copy()

    # 若目标时刻不在 MWR 网格时序中 → 填充 NaN
    if target_time not in time_points:
        interpolated_T = np.full(snd.shape[0], np.nan)
    else:
        snd['time'] = pd.to_datetime(snd['time'])
        snd_time_numeric = (snd['time'] - pd.Timestamp("1970-01-01")) // pd.Timedelta('1s')
        snd_height = snd['pressure_hPa'].apply(
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude - site_elev
        )
        xi = np.array([snd_time_numeric, snd_height]).T
        interpolated_T = interpn(
            points=(time_numeric, height_points),
            values=values_grid,
            xi=xi,
            method='linear',
            bounds_error=False,
            fill_value=np.nan
        )
    snd['T_mwr_interpolated'] = interpolated_T

    # — 单位转换 —
    era5_profile['height'] = mpcalc.pressure_to_height_std(
        era5_profile['pressure_level'].to_list() * units.hPa).to('m').magnitude - site_elev
    cra_profile['height'] = mpcalc.pressure_to_height_std(
        cra_profile['Level (hPa)'].to_list() * units.hPa).to('m').magnitude - site_elev
    snd['temperature_K'] = snd['temperature_C'] + 273.15
    snd['height'] = mpcalc.pressure_to_height_std(
        snd['pressure_hPa'].to_list() * units.hPa).to('m').magnitude - site_elev

    # — 排序 —
    era5_profile = era5_profile.sort_values('height')
    cra_profile  = cra_profile.sort_values('height').iloc[:37]
    snd          = snd.sort_values('height')

    # =========================== 主图 ===========================
    # 只显示低层 (0–3500 m)
    ax.set_ylim(0, 3500)

    #ax.plot(era5_profile['t'],         era5_profile['height'],         color='blue',    lw=3)
    #ax.plot(cra_profile['Temp (K)'],   cra_profile['height'],         color='magenta', lw=3)
    ax.plot(snd['T_mwr_interpolated'], snd['height'],                 color='red',     lw=3)
    ax.plot(snd['temperature_K'],      snd['height'],                 color='black',   lw=3)

    # — 左右双轴 —
    secax = ax.secondary_yaxis(
        'right',
        functions=(
            lambda h: mpcalc.height_to_pressure_std(h * units.m).to('hPa').magnitude,
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude)
    )
    secax.set_ylim(1000, 275)   # 气压范围对应 0–3500 m (近似值即可)

    # — 彩色细节
    ax.set_xlim(275, 310)
    ax.spines['bottom'].set_color('red')
    ax.spines['left'  ].set_color('black')
    ax.spines['left'].set_color('red')
    
    # 主刻度 / 次刻度
    ax.xaxis.set_major_locator(MultipleLocator(10))
    ax.xaxis.set_minor_locator(MultipleLocator(5))

    # y‑tick / x‑tick 配色 & 样式
    ax.tick_params(axis='y', colors='red',   length=12, width=2)
    ax.tick_params(axis='x', colors='red',   length=12, width=2)
    ax.tick_params(axis='x', which='minor',  colors='red', length=6, width=2)
    secax.tick_params(axis='y', colors='blue', length=12, width=2)

    # — 虚线参考高度 / 气压 (仅低层)
    for y in alt_ticks:
        ax.axhline(y, color='red',   lw=2, ls=(0, (5, 10)), zorder=0)
    for p in p_ticks:
        ax.axhline(p2alt(p), color='blue', lw=2, ls=(0, (5, 10)), zorder=0)

    # ———————— 轴刻度、标签管理 ————————
    col = idx % n_cols
    row = idx // n_cols

    # 仅最左列显示高度刻度
    if col == 0:
        ax.set_ylabel(' ', color='black')
    else:
        ax.set_ylabel('')
        ax.tick_params(axis='y', labelleft=False)

    # 仅最右列显示气压刻度
    if col == n_cols - 1:
        secax.set_ylabel(' ', color='black', rotation=270, labelpad=20)
    else:
        secax.set_ylabel('')
        secax.tick_params(axis='y', labelright=False)

    # 仅最底行显示温度标签
    if row == n_rows - 1:
        ax.xaxis.set_major_locator(MultipleLocator(10))
        ax.xaxis.set_minor_locator(MultipleLocator(5))
    else:
        ax.set_xlabel('')
        ax.tick_params(axis='x', labelbottom=False)
    if row == 3 and col == 1:
        ax.set_xlabel('Temperature (K)', color='red', fontsize=50, weight='bold')

    # — 子图标题 —
    ax.set_title(target_time.strftime('%Y-%m-%d %H:%M (BJT)'), pad=12, fontsize=36, weight='bold')

    # =========================== Inset Plot (3500–10000 m) ===========================
    ax_inset = inset_axes(
        ax,
        width="31%", height="48%",   # 占主图面积的比例
        loc='upper right',
        borderpad=0.6
    )

       # ★★★ 关键：给 inset 轴加一块不透明背景并抬高 z-order ★★★
    ax_inset.set_facecolor('white')         # 用白色把主图线条遮掉
    ax_inset.patch.set_alpha(1)             # 保证不透明
    ax_inset.patch.set_zorder(2)            # z-order 要高于主图里 axhline 的 0

    # （如果想让 inset 的外框也压在最上面，可以再抬高一下脊线）
    for spine in ax_inset.spines.values():
        spine.set_zorder(3)
    
    # 选取 3500–10000 m 的数据
    era5_high = era5_profile[(era5_profile['height'] >= 3500) & (era5_profile['height'] <= 10000)]
    cra_high  = cra_profile [(cra_profile ['height'] >= 3500) & (cra_profile ['height'] <= 10000)]
    snd_high  = snd         [(snd['height']          >= 3500) & (snd['height']          <= 10000)]

    #ax_inset.plot(era5_high['t'],         era5_high['height'],         color='blue',    lw=3)
    #ax_inset.plot(cra_high['Temp (K)'],   cra_high['height'],          color='magenta', lw=3)
    ax_inset.plot(snd_high['T_mwr_interpolated'], snd_high['height'],  color='red',     lw=3)
    ax_inset.plot(snd_high['temperature_K'],      snd_high['height'],  color='black',   lw=3)
    
    # Inset 坐标轴设置
    ax_inset.set_xlim(ax.get_xlim())
    ax_inset.set_ylim(3500, 10100)
    ax_inset.set_xlim(230,  290)
    # Inset 轴刻度更小即可（关闭 minor）
    ax_inset.xaxis.set_major_locator(MultipleLocator(20))
    #ax_inset.yaxis.set_major_locator(MultipleLocator(2000))
    # 隐藏顶部和右侧边框
    ax_inset.spines['top' ].set_visible(False)
    ax_inset.spines['right'].set_visible(False)

    # 左 / 底边框颜色 & tick 颜色与主图保持一致
    ax_inset.spines['left' ].set_color(ax.spines['left' ].get_edgecolor())
    ax_inset.spines['bottom'].set_color(ax.spines['bottom'].get_edgecolor())

    base_color = 'red'            # 与主图一致
    ax_inset.tick_params(axis='x', which='major',
                         length=8, width=2, labelsize=30,
                         colors=base_color)
    ax_inset.tick_params(axis='y', which='major',
                         length=8, width=2, labelsize=30,
                         colors='black')
    
    # 仅在 Inset 上绘制高层虚线高度
    for y in alt_ticks[(alt_ticks >= 3500) & (alt_ticks <= 10000)]:
        ax_inset.axhline(y, color='black', lw=2, ls=(0, (5, 5)), zorder=0)

    # Inset 不需要轴标签
    ax_inset.set_xlabel('')
    ax_inset.set_ylabel('')
    ax_inset.spines['left'].set_color('red')
    ax_inset.tick_params(axis='y', colors='red')
# —————————————————— 隐藏空白子图（如果有） ——————————————————
for j in range(idx + 1, n_rows * n_cols):
    fig.delaxes(axes[j])

# —————————————————— 统一的全局标签 ——————————————————
fig.text(0.01, 0.5, 'Height (m a.g.l.)', rotation=90, va='center', ha='left',
         color='black', weight='bold', fontsize=50)
fig.text(0.995, 0.5, 'Pressure (hPa)',   rotation=270, va='center', ha='right',
         color='black', weight='bold', fontsize=50)

# —————————————————— 全局图例 & 标题 ——————————————————
fig.legend(handles=legend_handles,
           loc='upper center', bbox_to_anchor=(0.5, 0.995), ncol=4,
           frameon=False, fontsize=32)
fig.text(0.5, 0.993, 'Temperature Radiosonde Comparison – 1 min Temporal Resolution',
         ha='center', va='bottom', fontsize=40, weight='bold')

# —————————————————— 保存高分辨率 TIF ——————————————————
fig.tight_layout(rect=[0, 0, 1, 0.98])   # 留出顶部空间
save_path = os.path.join(out_dir, out_fname)
fig.savefig(save_path, dpi=300, format='tif', #transparent=True,
            bbox_inches='tight', pil_kwargs={"compression": "tiff_adobe_deflate"})
print(f"✅ 已保存到：{save_path}")



###################绘制湿度(探空时刻±30min平均版-1h分辨率)###################
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import metpy.calc as mpcalc
from metpy.units import units
from matplotlib.ticker import MultipleLocator

# —————————————————— 用户配置 ——————————————————
excel_path   = r"E:\\Beijing2024\\探空数据-54511\\original combined2.xlsx"
sheet_names  = pd.ExcelFile(excel_path).sheet_names           # 全部时次
out_dir      = r"E:\Beijing2024\出图TIF"
out_fname    = "RH_Profile_平均.tif"
os.makedirs(out_dir, exist_ok=True)
site_elev = 49

# —————————————————— 统一的绘图风格 ——————————————————
plt.rcParams.update({
    'font.family': 'sans-serif',                # 先指定家族
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     # 把 Arial 设为首选
    'font.size'      : 36,
    'axes.linewidth' : 3,
    'axes.labelsize' : 36,
    'xtick.labelsize': 36,
    'ytick.labelsize': 36
})

# 标准高度/气压刻度
alt_ticks = np.array([0, 500, 1500, 3000, 5000, 8000, 10000, 12500, 15000])   # m
p_ticks   = np.array([1000, 925, 850, 700, 500, 275])                         # hPa
p2alt = lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude

# —————————————————— 建立 4×3 画布 ——————————————————
n_rows, n_cols = 4, 3
fig, axes = plt.subplots(n_rows, n_cols, figsize=(30, 40), sharex=False, sharey=False)
axes = axes.ravel()             # 方便用 1-D 索引
# 图例用到的句柄与标签 (主图 & inset 共用)
legend_handles = [
    plt.Line2D([], [], color='blue',    lw=6, label='ERA5'),
    plt.Line2D([], [], color='magenta', lw=6, label='CRA'),
    plt.Line2D([], [], color='green',   lw=6, label='MWR'),
    plt.Line2D([], [], color='black',   lw=6, label='Radiosonde')
]

# —————————————————— 循环所有时次 ——————————————————
for idx, sheet in enumerate(sheet_names):
    if idx >= n_rows * n_cols:            # 超出 12 张子图就跳过
        break

    ax = axes[idx]

    # 1) 读取探空表
    snd = pd.read_excel(excel_path, sheet_name=sheet)

    # 2) 当前 target_time（由工作表名解析）
    target_time = pd.to_datetime(sheet, format='%Y%m%d%H')

    # 3) 按原逻辑过滤/处理 ERA5、CRA、MWR
    era5_profile = ERA5_interp_BJ[ERA5_interp_BJ['valid_time'] == target_time].copy()
    cra_regular['timestamps'] = pd.to_datetime(cra_regular['timestamps'])
    cra_profile = cra_regular[cra_regular['timestamps'] == target_time].copy()
    # 处理MWR数据：取探空时刻前后半小时平均
    start_time = target_time - pd.Timedelta(minutes=30)
    end_time = target_time + pd.Timedelta(minutes=30)
    mwr_filtered = rh_mwr_long.reset_index()  # 将双重索引转换为列
    mask = (mwr_filtered['timestamps'] >= start_time) & (mwr_filtered['timestamps'] <= end_time)
    mwr_filtered = mwr_filtered.loc[mask]
    if not mwr_filtered.empty:
        mwr_avg = mwr_filtered.groupby('height', as_index=False)['RH(%)'].mean()
        mwr_profile = mwr_avg.sort_values('height')
    else:
        print(f"⚠️ 当前时次 {target_time.strftime('%Y-%m-%d %H:%M')} 的 MWR 数据不存在")
        mwr_profile = pd.DataFrame(columns=['height', 'RH(%)'])  # 空DataFrame避免报错
    
    # — 单位转换 —
    era5_profile['height'] = mpcalc.pressure_to_height_std(
        era5_profile['pressure_level'].to_list() * units.hPa
    ).to('m').magnitude - site_elev
    cra_profile['height'] = mpcalc.pressure_to_height_std(
        cra_profile['Level (hPa)'].to_list() * units.hPa
    ).to('m').magnitude - site_elev
    snd['temperature_K'] = snd['temperature_C'] + 273.15
    snd['height'] = mpcalc.pressure_to_height_std(
        snd['pressure_hPa'].to_list() * units.hPa
    ).to('m').magnitude - site_elev

    # — 排序 —
    era5_profile = era5_profile.sort_values('height')
    cra_profile  = cra_profile.sort_values('height').iloc[:37]
    snd          = snd.sort_values('height')

    # ———————————————— 绘图 ————————————————
    ax.plot(era5_profile['r'],             era5_profile['height'], color='blue',    lw=3)
    ax.plot(cra_profile['rh (%)'],         cra_profile['height'],  color='magenta', lw=3)
    ax.plot(mwr_profile['RH(%)'],          mwr_profile['height'],  color='green',   lw=3)
    ax.plot(snd['relative humidity_%'],    snd['height'],          color='black',   lw=3)
    
    # — 左右双轴 —
    secax = ax.secondary_yaxis(
        'right',
        functions=(
            lambda h: mpcalc.height_to_pressure_std(h * units.m).to('hPa').magnitude,
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude
        )
    )
    secax.set_ylim(1000, 275)
    ax.set_ylim(0, 10000)
    ax.set_xlim(-2, 103)
    ax.xaxis.set_major_locator(MultipleLocator(10))
    ax.xaxis.set_minor_locator(MultipleLocator(5))
    ax.tick_params(axis='x', colors='green', length=12,width=2) 
    ax.tick_params(axis='x', which='minor', length=6, width=2,color='green')
    ax.set_yticks(alt_ticks)  # 设置左侧主刻度
    secax.tick_params(axis='y', colors='blue', length=12, width=2,labelcolor='blue')
    secax.set_yticks(p_ticks)   # 设置右侧主刻度
    ax.set_ylim(0, 10000) 
    # ax.xaxis.set_major_locator(MultipleLocator(10))
    # ax.xaxis.set_minor_locator(MultipleLocator(5))
    # ax.tick_params(axis='x',which='minor',length=6,width=2,color='black')
    #plt.tick_params(axis='x', length=12, width=2)
    #ax.set_xlim(240, 310)
    

    # — 虚线参考高度/气压 —
    for y in alt_ticks:
        ax.axhline(y, color='red',   lw=2, ls=(0, (5, 10)), zorder=0)
    for p in p_ticks:
        ax.axhline(p2alt(p), color='blue', lw=2, ls=(0, (5, 10)), zorder=0)

    # ———————— 轴刻度、标签管理 ————————
    col = idx % n_cols
    row = idx // n_cols

    # y-轴 & sec-y 颜色
    ax.spines['left' ].set_color('red')
    ax.spines['right'].set_color('black')
    ax.spines['bottom'].set_color('green')
    ax.tick_params(axis='y', colors='red', length=12, width=2)
    #secax.tick_params(axis='y', colors='black', length=12, width=2)

    # 仅保留左/右边 & 最底行的轴标签
    if col == 0:                # 最左列 → 高度标签
        ax.set_ylabel(' ', color='red')
    else:
        ax.set_ylabel('')
        ax.tick_params(axis='y', labelleft=False)

    if col == n_cols - 1:       # 最右列 → Pressure 标签
        secax.set_ylabel(' ', color='blue', rotation=270, labelpad=20)
    else:
        secax.set_ylabel('')
        secax.tick_params(axis='y', labelright=False)

    if row == n_rows - 1:       # 最底行 → Temperature 标签
        #ax.set_xlabel('Temperature (K)')
        ax.xaxis.set_major_locator(MultipleLocator(10))
        ax.xaxis.set_minor_locator(MultipleLocator(5))
        ax.tick_params(axis='x', which='major', length=12, width=2)
        ax.tick_params(axis='x', which='minor', length=6, width=2,color='green')
    else:
        ax.set_xlabel('')
        ax.tick_params(axis='x', labelbottom=False)
    if row == 3 and col == 1:
        ax.set_xlabel('Relative Humidity(%)',fontsize=50,weight='bold',color='green')
    # — 子图标题 —
    ax.set_title(target_time.strftime('%Y-%m-%d %H:%M (BJT)'), pad=12, fontsize=36,weight='bold')

# —————————————————— 隐藏空白子图（如果有） ——————————————————

for j in range(idx + 1, n_rows * n_cols):
    fig.delaxes(axes[j])
fig.text(0.01, 0.5, 'Height (m a.g.l.)', 
        rotation=90, va='center', ha='left',
        color='black', weight='bold',fontsize=50)
fig.text(0.995, 0.5, 'Pressure (hPa)', 
        rotation=270, va='center', ha='right',
        color='black', weight='bold',fontsize=50)
# —————————————————— 保存高分辨率 TIF ——————————————————
### ← 新增：全局图例（放在顶部） ###
fig.legend(handles=legend_handles,
           loc='upper center', bbox_to_anchor=(0.5, 0.995), ncol=4,
           frameon=False, fontsize=32)
fig.text(0.5, 0.993, 'Relative Humidity Radiosonde Comparison – 1h Temporal Resolution',
         ha='center', va='bottom', fontsize=40, weight='bold')

fig.tight_layout(rect=[0, 0, 1, 0.98])   # 留出顶部空间
save_path = os.path.join(out_dir, out_fname)
fig.savefig(save_path, dpi=300, transparent=True,pil_kwargs={"compression": "tiff_adobe_deflate"},
            format='tif', bbox_inches='tight')
print(f"✅ 已保存到：{save_path}")


###################绘制湿度(探空匹配版)-1min###################
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import metpy.calc as mpcalc
from metpy.units import units
from matplotlib.ticker import MultipleLocator
from scipy.interpolate import interpn

# —————————————————— 用户配置 ——————————————————
excel_path   = r"E:\\Beijing2024\\探空数据-54511\\original combined2.xlsx"
sheet_names  = pd.ExcelFile(excel_path).sheet_names           # 全部时次
out_dir      = r"E:\Beijing2024\出图TIF"
out_fname    = "RH_Profile_MWR匹配探空.tif"
os.makedirs(out_dir, exist_ok=True)
site_elev = Altitude

# 提取 rh_mwr_long 的网格点（时间戳和高度）
time_points = rh_mwr_long.index.get_level_values('timestamps').unique().sort_values()
height_points = rh_mwr_long.index.get_level_values('height').unique().sort_values()

# 将时间转换为数值（例如相对于某个基准时间的秒数）
time_numeric = (time_points - pd.Timestamp("1970-01-01")) // pd.Timedelta('1s')

# 提取 rh_mwr_long 的温度数据并重塑为二维网格
values_grid = rh_mwr_long.unstack().values  # 形状为 (len(time_points), len(height_points))

# —————————————————— 统一的绘图风格 ——————————————————
plt.rcParams.update({
    'font.family': 'sans-serif',                # 先指定家族
    'font.sans-serif': ['Arial', 'DejaVu Sans'],     # 把 Arial 设为首选
    'font.size'      : 36,
    'axes.linewidth' : 3,
    'axes.labelsize' : 36,
    'xtick.labelsize': 36,
    'ytick.labelsize': 36
})

# 标准高度/气压刻度
alt_ticks = np.array([0, 500, 2000, 3000, 5000, 8000, 10000, 12500, 15000])   # m
p_ticks   = np.array([1000, 925, 850, 700, 500, 275])                         # hPa
p2alt = lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude
legend_handles = [
    #plt.Line2D([], [], color='blue',    lw=6, label='ERA5'),
    #plt.Line2D([], [], color='magenta', lw=6, label='CRA'),
    plt.Line2D([], [], color='green',     lw=6, label='MWR'),
    plt.Line2D([], [], color='black',   lw=6, label='Radiosonde')
]

# —————————————————— 建立 4×3 画布 ——————————————————
n_rows, n_cols = 4, 3
fig, axes = plt.subplots(n_rows, n_cols, figsize=(30, 40), sharex=False, sharey=False)
axes = axes.ravel()             # 方便用 1-D 索引

# —————————————————— 循环所有时次 ——————————————————
for idx, sheet in enumerate(sheet_names):
    if idx >= n_rows * n_cols:            # 超出 12 张子图就跳过
        break

    ax = axes[idx]

    # 1) 读取探空表
    snd = pd.read_excel(excel_path, sheet_name=sheet)

    # 2) 当前 target_time（由工作表名解析）
    target_time = pd.to_datetime(sheet, format='%Y%m%d%H')

    # 3) 按原逻辑过滤/处理 ERA5、CRA、MWR
    era5_profile = ERA5_interp_BJ[ERA5_interp_BJ['valid_time'] == target_time].copy()
    cra_regular['timestamps'] = pd.to_datetime(cra_regular['timestamps'])
    cra_profile = cra_regular[cra_regular['timestamps'] == target_time].copy()
    # 判断当前时次是否存在MWR数据
    if target_time not in time_points:
        # 若不存在，填充NaN并跳过插值
        interpolated_rh = np.full(snd.shape[0], np.nan)
    else:
    # 从 snd 中提取目标插值点的时间和高度
        snd['time'] = pd.to_datetime(snd['time'])
    
        snd_time_numeric = (snd['time'] - pd.Timestamp("1970-01-01")) // pd.Timedelta('1s')
        snd_height = snd['pressure_hPa'].apply(
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude - site_elev
        )
    
        # 构建插值目标点坐标 (时间, 高度)
        xi = np.array([snd_time_numeric, snd_height]).T
    
        # 执行线性插值（时间和高度维度）
        interpolated_rh = interpn(
            points=(time_numeric, height_points),
            values=values_grid,
            xi=xi,
            method='linear',
            bounds_error=False,  # 允许外推
            fill_value=np.nan     # 超出范围填充为NaN
        )
    
        # 将插值结果添加到 snd 数据中
    snd['rh_mwr_interpolated'] = interpolated_rh

    # 生成插值后数据框
    mwr_profile = snd['rh_mwr_interpolated']

    # — 单位转换 —
    era5_profile['height'] = mpcalc.pressure_to_height_std(
        era5_profile['pressure_level'].to_list() * units.hPa
    ).to('m').magnitude - site_elev
    cra_profile['height'] = mpcalc.pressure_to_height_std(
        cra_profile['Level (hPa)'].to_list() * units.hPa
    ).to('m').magnitude - site_elev
    snd['temperature_K'] = snd['temperature_C'] + 273.15
    snd['height'] = mpcalc.pressure_to_height_std(
        snd['pressure_hPa'].to_list() * units.hPa
    ).to('m').magnitude - site_elev

    # — 排序 —
    era5_profile = era5_profile.sort_values('height')
    cra_profile  = cra_profile.sort_values('height').iloc[:37]
    snd          = snd.sort_values('height')

    # ———————————————— 绘图 ————————————————
    #ax.plot(era5_profile['r'],         era5_profile['height'],color='blue',    lw=3)
    #ax.plot(cra_profile['rh (%)'],   cra_profile['height'], color='magenta',     lw=3)
    ax.plot(snd['rh_mwr_interpolated'], snd['height'],         color='green', lw=3)
    ax.plot(snd['relative humidity_%'],      snd['height'],         color='black',   lw=3)
    
    # — 左右双轴 —
    secax = ax.secondary_yaxis(
        'right',
        functions=(
            lambda h: mpcalc.height_to_pressure_std(h * units.m).to('hPa').magnitude,
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude
        )
    )
    secax.set_ylim(1000, 275)
    ax.set_ylim(0, 10000)
    ax.set_xlim(-2, 103)
    ax.xaxis.set_major_locator(MultipleLocator(10))
    ax.xaxis.set_minor_locator(MultipleLocator(5))
    ax.tick_params(axis='x', colors='green', length=12,width=2) 
    ax.tick_params(axis='x', which='minor', length=6, width=2,color='green')
    ax.set_yticks(alt_ticks)  # 设置左侧主刻度
    secax.tick_params(axis='y', colors='blue', length=12, width=2,labelcolor='blue')
    secax.set_yticks(p_ticks)   # 设置右侧主刻度
    ax.set_ylim(0, 10000) 

    # — 虚线参考高度/气压 —
    for y in alt_ticks:
        ax.axhline(y, color='red',   lw=2, ls=(0, (5, 10)), zorder=0)
    for p in p_ticks:
        ax.axhline(p2alt(p), color='blue', lw=2, ls=(0, (5, 10)), zorder=0)

    # ———————— 轴刻度、标签管理 ————————
    col = idx % n_cols
    row = idx // n_cols

    # y-轴 & sec-y 颜色
    ax.spines['left' ].set_color('red')
    ax.spines['right'].set_color('black')
    ax.spines['bottom'].set_color('green')
    ax.tick_params(axis='y', colors='red', length=12, width=2)

    # 仅保留左/右边 & 最底行的轴标签
    if col == 0:                # 最左列 → 高度标签
        ax.set_ylabel(' ', color='red')
    else:
        ax.set_ylabel('')
        ax.tick_params(axis='y', labelleft=False)

    if col == n_cols - 1:       # 最右列 → Pressure 标签
        secax.set_ylabel(' ', color='blue', rotation=270, labelpad=20)
    else:
        secax.set_ylabel('')
        secax.tick_params(axis='y', labelright=False)

    if row == n_rows - 1:       # 最底行 → Temperature 标签
        ax.xaxis.set_major_locator(MultipleLocator(10))
        ax.xaxis.set_minor_locator(MultipleLocator(5))
        ax.tick_params(axis='x', which='major', length=12, width=2)
        ax.tick_params(axis='x', which='minor', length=6, width=2,color='green')
    else:
        ax.set_xlabel('')
        ax.tick_params(axis='x', labelbottom=False)
    if row == 3 and col == 1:
        ax.set_xlabel('Relative Humidity(%)',fontsize=50,weight='bold',color='green')
    # — 子图标题 —
    ax.set_title(target_time.strftime('%Y-%m-%d %H:%M (BJT)'), pad=12, fontsize=36,weight='bold')

# —————————————————— 隐藏空白子图（如果有） ——————————————————

for j in range(idx + 1, n_rows * n_cols):
    fig.delaxes(axes[j])
fig.text(0.01, 0.5, 'Height (m a.g.l.)', 
        rotation=90, va='center', ha='left',
        color='black', weight='bold',fontsize=50)
fig.text(0.995, 0.5, 'Pressure (hPa)', 
        rotation=270, va='center', ha='right',
        color='black', weight='bold',fontsize=50)
# —————————————————— 保存高分辨率 TIF ——————————————————
### ← 新增：全局图例（放在顶部） ###
fig.legend(handles=legend_handles,
           loc='upper center', bbox_to_anchor=(0.5, 0.995), ncol=4,
           frameon=False, fontsize=32)
fig.text(0.5, 0.993, 'Relative Humidity Radiosonde Comparison – 1min Temporal Resolution',
         ha='center', va='bottom', fontsize=40, weight='bold')

fig.tight_layout(rect=[0, 0, 1, 0.98])   # 留出顶部空间
save_path = os.path.join(out_dir, out_fname)
fig.savefig(save_path, dpi=300, transparent=True,pil_kwargs={"compression": "tiff_adobe_deflate"},
            format='tif', bbox_inches='tight')
print(f"✅ 已保存到：{save_path}")



############################# Horizontal Wind Speed Profiles – 1 h #############################
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import metpy.calc as mpcalc
from metpy.units import units
from matplotlib.ticker import MultipleLocator
from mpl_toolkits.axes_grid1.inset_locator import inset_axes

# ─────────────────────────── USER CONFIG ──────────────────────────────────────
excel_path   = r"E:\\Beijing2024\\探空数据-54511\\original combined2.xlsx"
out_dir      = r"E:\\Beijing2024\\出图TIF"
out_fname    = "wind_Profile-1h_with_inset.tif"
os.makedirs(out_dir, exist_ok=True)

# ─────────────────────────── PLOT STYLE ───────────────────────────────────────
plt.rcParams.update({
    'font.family'   : 'sans-serif',
    'font.sans-serif': ['Arial', 'DejaVu Sans'],
    'font.size'      : 36,
    'axes.linewidth' : 3,
    'axes.labelsize' : 36,
    'xtick.labelsize': 36,
    'ytick.labelsize': 36
})

# ─────────────── Standard altitude / pressure ticks ───────────────────────────
alt_ticks = np.array([0, 500, 1000, 1500, 2000, 2500, 3000, 4000, 6000, 8000, 10000])  # m
p_ticks   = np.array([1000, 925, 850, 800, 750, 700, 500, 275])                                  # hPa
p2alt = lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude

# ───────────────────── Helper: extract DWL profile ────────────────────────────

def _get_profile_from_series_dict(dic, timestamp):
    """Return a DataFrame of Doppler‑Lidar HWS for the given *timestamp*.
    The user must ensure the dictionary structure matches the assumed format.
    Expected layout: dic['HWS'][height]['Doppler'] (or similar) is a Series
    indexed by timestamps. Adjust if necessary.
    """
    # Build list of (height, value) pairs
    data = []
    for h, sources in dic['HWS'].items():
        # ---- adapt key name 'Doppler' to match your actual inner key ----
        s = sources.get('Doppler', None)  # replace 'Doppler' if needed
        if s is None or timestamp not in s.index:
            continue
        data.append((h, s.loc[timestamp]))
    prof = pd.DataFrame(data, columns=['height', 'hws']).sort_values('height')
    return prof

# ──────────────────────────── Figure grid ─────────────────────────────────────
from pandas import ExcelFile
sheet_names = ExcelFile(excel_path).sheet_names         # 12 time‑stamps max
n_rows, n_cols = 4, 3
fig, axes = plt.subplots(n_rows, n_cols, figsize=(30, 40), sharex=False, sharey=False)
axes = axes.ravel()

# ───────── 时间列先标准化 ─────────
cleaned_h['BJT'] = pd.to_datetime(cleaned_h['BJT'])

# ───────── 60 min 重采样（按 height 分组）─────────
cleaned_h_60min = (
    cleaned_h
    .groupby(
        ['height',                     # ← 分高度
         pd.Grouper(key='BJT', freq='60min')]   # ← 对时间列做 60 min 分箱
    , as_index=False)                 # 保持列形式，避免成为索引
    .mean(numeric_only=True)          # 你只关心数值列，这样更快
    
)

Subplot_yaxis=2500

# ──────────────────────────── Main loop ───────────────────────────────────────
for idx, sheet in enumerate(sheet_names):
    if idx >= n_rows * n_cols:  # safety guard
        break
    ax = axes[idx]

    # ---- 1) Read radiosonde table (sheet‑wise) --------------------------------
    snd = pd.read_excel(excel_path, sheet_name=sheet)

    # ---- 2) current timestamp -------------------------------------------------
    target_time = pd.to_datetime(sheet, format='%Y%m%d%H')

    # ---- 3) ERA5, CRA profiles (prepared upstream) ---------------------------
    era5_profile = ERA5_interp_BJ.loc[ERA5_interp_BJ['valid_time'] == target_time].copy()
    cra_profile  = cra_regular.loc[pd.to_datetime(cra_regular['timestamps']) == target_time].copy()

    # ---- 4) Doppler Wind‑Lidar profile from series_dict_60min -----------------
    dwl_profile = (
            cleaned_h_60min[cleaned_h_60min['BJT'] == target_time]   # 取当前整点
            .loc[:, ['height', 'hws']]
            .sort_values('height')
        )

    # ---- 5) Unit conversion ---------------------------------------------------
    era5_profile['height'] = mpcalc.pressure_to_height_std(
        era5_profile['pressure_level'].to_numpy() * units.hPa).to('m').magnitude
    cra_profile['height'] = mpcalc.pressure_to_height_std(
        cra_profile['Level (hPa)'].to_numpy() * units.hPa).to('m').magnitude
    snd['height'] = mpcalc.pressure_to_height_std(
        snd['pressure_hPa'].to_numpy() * units.hPa).to('m').magnitude

    # rename radiosonde wind‑speed column if needed
    if 'wind speed_m/s' in snd.columns:
        snd.rename(columns={'wind speed_m/s': 'hws'}, inplace=True)
    else:
        snd.rename(columns={snd.filter(like='wind').columns[0]: 'hws'}, inplace=True)

    # ERA5 / CRA wind‑speed columns may differ; harmonise to 'hws'
    era5_profile.rename(columns={era5_profile.filter(like='hws').columns[0]: 'hws'}, inplace=True)
    cra_profile.rename(columns={cra_profile.filter(like='hws').columns[0]: 'hws'}, inplace=True)

    # Keep only needed columns; sort by height
    era5_profile = era5_profile[['height', 'hws']].sort_values('height')
    cra_profile  = cra_profile[['height', 'hws']].sort_values('height').iloc[:37]
    snd          = snd[['height', 'hws']].sort_values('height')
    dwl_profile  = dwl_profile.sort_values('height')

    # ---------------------------- PLOT MAIN AXIS -----------------------------
    ax.plot(era5_profile['hws'], era5_profile['height'], color='blue',    lw=3)
    ax.plot(cra_profile['hws'],  cra_profile['height'],  color='magenta', lw=3)
    ax.plot(dwl_profile['hws'], dwl_profile['height'], color='darkorange',      lw=3)
    ax.plot(snd['hws'],         snd['height'],         color='black',    lw=3)

    # ---- dual y‑axis (height ↔ pressure) -------------------------------------
    secax = ax.secondary_yaxis(
        'right',
        functions=(
            lambda h: mpcalc.height_to_pressure_std(h * units.m).to('hPa').magnitude,
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude
        )
    )

    ax.set_xlim(0, 40)
    alt_ticks_plot = alt_ticks[alt_ticks >= Subplot_yaxis]
    ax.set_yticks(alt_ticks)
    p_ticks_plot = [p for p in p_ticks if p2alt(p) >= Subplot_yaxis]
    secax.set_yticks(p_ticks)
    # main & secondary limits
    ax.set_ylim(0, Subplot_yaxis)
    # x‑ticks formatting (orange)
    ax.xaxis.set_major_locator(MultipleLocator(5))
    ax.tick_params(axis='x', which='major', length=12, width=2, colors='darkorange')
    ax.tick_params(axis='x', which='minor', length=6, width=2, colors='darkorange')

    # spine colours
    ax.spines['left' ].set_color('red')
    ax.spines['bottom'].set_color('darkorange')
    ax.spines['right'].set_color('black')
    ax.tick_params(axis='y', colors='red', length=12, width=2)
    secax.tick_params(axis='y', colors='blue', length=12, width=2)

    # ---- dashed reference lines (unchanged) ----------------------------------
    for y in alt_ticks:
        ax.axhline(y, color='red',   lw=2, ls=(0, (5, 10)), zorder=0)
    for p in p_ticks:
        ax.axhline(p2alt(p), color='blue',  lw=2, ls=(0, (5, 10)), zorder=0)

    # --------------------------- INSET AXIS ----------------------------------
    ax_in = inset_axes(ax, width="31%", height="48%", loc='upper right',
                       borderpad=1)
    ax_inset.set_facecolor('white')         # 用白色把主图线条遮掉
    ax_inset.patch.set_alpha(1)             # 保证不透明
    ax_inset.patch.set_zorder(2) 

    # plot same datasets, restricted to < 500 m
    ax_in.plot(era5_profile.loc[era5_profile['height'] >= Subplot_yaxis, 'hws'],
               era5_profile.loc[era5_profile['height'] >= Subplot_yaxis, 'height'],
               color='blue',   lw=3)
    ax_in.plot(cra_profile.loc[cra_profile['height'] >= Subplot_yaxis, 'hws'],
               cra_profile.loc[cra_profile['height'] >= Subplot_yaxis, 'height'],
               color='magenta', lw=3)
    ax_in.plot(dwl_profile.loc[dwl_profile['height'] >= Subplot_yaxis, 'hws'],
               dwl_profile.loc[dwl_profile['height'] >= Subplot_yaxis, 'height'],
               color='red',    lw=3)
    ax_in.plot(snd.loc[snd['height'] >= Subplot_yaxis, 'hws'],
               snd.loc[snd['height'] >= Subplot_yaxis, 'height'],
               color='black',  lw=3)

    # inset axis limits & ticks
    ax_in.set_ylim(Subplot_yaxis, 10010)
    ax_in.set_xlim(0, 45)
    # >>> 强制在 y 轴最底部显示 Subplot_yaxis 这一个刻度 <<<
    y_ticks_inset = np.concatenate([[Subplot_yaxis],
                                    alt_ticks[(alt_ticks > Subplot_yaxis) &
                                              (alt_ticks >= 4000)]])
    ax_in.set_yticks(y_ticks_inset)
    ax_in.tick_params(axis='both', which='major', length=8, width=2, labelsize=30)
    ax_in.tick_params(axis='both', which='minor', length=4, width=2)

    # inset spines styling
    ax_in.spines['top'   ].set_visible(False)
    ax_in.spines['right' ].set_visible(False)
    ax_in.spines['left'  ].set_color('red')
    ax_in.spines['bottom'].set_color('darkorange')
    ax_in.tick_params(axis='y', colors='red')
    ax_in.tick_params(axis='x', colors='darkorange')
    ax_in.xaxis.set_major_locator(MultipleLocator(10))

    # -------------------- subplot‑specific labels / title --------------------
    col = idx % n_cols
    row = idx // n_cols
    center_col = n_cols // 2 
    if col == 0:
        ax.set_ylabel('')  # we move global y‑label outside fig
    else:
        ax.tick_params(axis='y', labelleft=False)
    if col == n_cols - 1:
        secax.set_ylabel('')  # right‑side label moved to global text
    else:
        secax.tick_params(axis='y', labelright=False)

    if row == n_rows - 1:                       # 底行
        ax.tick_params(axis='x', labelbottom=True)   # 刻度数字保留
        
        if col == center_col:                   # 只有中间那格要 xlabel
            ax.set_xlabel('Horizontal Wind Speed (m/s)',
                      fontsize=50, weight='bold', color='darkorange',
                      labelpad=20)
    else:                                       # 其余行
        ax.tick_params(axis='x', labelbottom=False)  # 刻度数字隐藏

    # panel title
    ax.set_title(target_time.strftime('%Y-%m-%d %H:%M (BJT)'), pad=12, fontsize=36,weight='bold')

# ───────────────────── Hide any empty sub‑axes ────────────────────────────────
for j in range(idx + 1, n_rows * n_cols):
    fig.delaxes(axes[j])

# ───────────────────── Global axes‑level annotations ──────────────────────────
fig.text(-0.01, 0.5, 'Height (m a.g.l.)', rotation=90, va='center', ha='left',
         color='black', weight='bold', fontsize=50)
fig.text(1.01, 0.5, 'Pressure (hPa)', rotation=270, va='center', ha='right',
         color='black', weight='bold', fontsize=50)

# Legend & title (top)
legend_handles = [
    plt.Line2D([], [], color='blue',      lw=6, label='ERA5'),
    plt.Line2D([], [], color='magenta',   lw=6, label='CRA'),
    plt.Line2D([], [], color='darkorange',lw=6, label='Doppler Wind Lidar'),
    plt.Line2D([], [], color='black',     lw=6, label='Radiosonde')
]
fig.legend(handles=legend_handles, loc='upper center', bbox_to_anchor=(0.5, 0.995),
           ncol=4, frameon=False, fontsize=32)
fig.text(0.5, 0.993, 'Horizontal Wind Speed Radiosonde Comparison – 1h Temporal Resolution',
         ha='center', va='bottom', fontsize=40, weight='bold')

# ───────────────────── Save figure (no transparency) ──────────────────────────
fig.tight_layout(rect=[0, 0, 1, 0.985])
save_path = os.path.join(out_dir, out_fname)
fig.savefig(save_path, dpi=300, pil_kwargs={"compression": "tiff_adobe_deflate"},
            format='tif', bbox_inches='tight')
print(f"✅ Figure saved to: {save_path}")




import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import metpy.calc as mpcalc
from metpy.units import units
from matplotlib.ticker import MultipleLocator
from mpl_toolkits.axes_grid1.inset_locator import inset_axes

# ─────────────────────────── USER CONFIG ──────────────────────────────────────
excel_path   = r"E:\\Beijing2024\\探空数据-54511\\original combined2.xlsx"
out_dir      = r"E:\\Beijing2024\\出图TIF"
out_fname    = "wind_Profile-5min_with_inset.tif"
os.makedirs(out_dir, exist_ok=True)

# ─────────────────────────── PLOT STYLE ───────────────────────────────────────
plt.rcParams.update({
    'font.family'   : 'sans-serif',
    'font.sans-serif': ['Arial', 'DejaVu Sans'],
    'font.size'      : 36,
    'axes.linewidth' : 3,
    'axes.labelsize' : 36,
    'xtick.labelsize': 36,
    'ytick.labelsize': 36
})

# ─────────────── Standard altitude / pressure ticks ───────────────────────────
alt_ticks = np.array([0, 500, 1000, 1500, 2000, 2500, 3000, 4000, 6000, 8000, 10000])  # m
p_ticks   = np.array([1000, 925, 850, 800, 750, 700, 500, 275])                                    # hPa
p2alt = lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude




# ──────────────────────────── Figure grid ─────────────────────────────────────
from pandas import ExcelFile
sheet_names = ExcelFile(excel_path).sheet_names         # 12 time‑stamps max
n_rows, n_cols = 4, 3
fig, axes = plt.subplots(n_rows, n_cols, figsize=(30, 40), sharex=False, sharey=False)
axes = axes.ravel()

# ───────── 时间列先标准化 ─────────
cleaned_h['BJT'] = pd.to_datetime(cleaned_h['BJT'])

# ───────── 60 min 重采样（按 height 分组）─────────
cleaned_h_5min = (
    cleaned_h
    .groupby(
        ['height',                     # ← 分高度
         pd.Grouper(key='BJT', freq='5min')]   # ← 对时间列做 60 min 分箱
    , as_index=False)                 # 保持列形式，避免成为索引
    .mean(numeric_only=True)          # 你只关心数值列，这样更快
    #.dropna(subset=['hws'])           # 可选：去掉 hws 缺测行
)

Subplot_yaxis=2500

# ──────────────────────────── Main loop ───────────────────────────────────────
for idx, sheet in enumerate(sheet_names):
    if idx >= n_rows * n_cols:  # safety guard
        break
    ax = axes[idx]

    # ---- 1) Read radiosonde table (sheet‑wise) --------------------------------
    snd = pd.read_excel(excel_path, sheet_name=sheet)

    # ---- 2) current timestamp -------------------------------------------------
    target_time = pd.to_datetime(sheet, format='%Y%m%d%H')

    # ---- 3) ERA5, CRA profiles (peared upstream) ---------------------------
    era5_profile = ERA5_interp_BJ.loc[ERA5_interp_BJ['valid_time'] == target_time].copy()
    cra_profile  = cra_regular.loc[pd.to_datetime(cra_regular['timestamps']) == target_time].copy()

    # ---- 4) Doppler Wind‑Lidar profile from series_dict_60min -----------------
    dwl_profile = (
            cleaned_h_5min[cleaned_h_5min['BJT'] == target_time]   # 取当前整点
            .loc[:, ['height', 'hws']]
            .sort_values('height')
        )

    # ---- 5) Unit conversion ---------------------------------------------------
    era5_profile['height'] = mpcalc.pressure_to_height_std(
        era5_profile['pressure_level'].to_numpy() * units.hPa).to('m').magnitude
    cra_profile['height'] = mpcalc.pressure_to_height_std(
        cra_profile['Level (hPa)'].to_numpy() * units.hPa).to('m').magnitude
    snd['height'] = mpcalc.pressure_to_height_std(
        snd['pressure_hPa'].to_numpy() * units.hPa).to('m').magnitude

    # rename radiosonde wind‑speed column if needed
    if 'wind speed_m/s' in snd.columns:
        snd.rename(columns={'wind speed_m/s': 'hws'}, inplace=True)
    else:
        snd.rename(columns={snd.filter(like='wind').columns[0]: 'hws'}, inplace=True)

    # ERA5 / CRA wind‑speed columns may differ; harmonise to 'hws'
    era5_profile.rename(columns={era5_profile.filter(like='hws').columns[0]: 'hws'}, inplace=True)
    cra_profile.rename(columns={cra_profile.filter(like='hws').columns[0]: 'hws'}, inplace=True)

    # Keep only needed columns; sort by height
    era5_profile = era5_profile[['height', 'hws']].sort_values('height')
    cra_profile  = cra_profile[['height', 'hws']].sort_values('height').iloc[:37]
    snd          = snd[['height', 'hws']].sort_values('height')
    dwl_profile  = dwl_profile.sort_values('height')

    # ---------------------------- PLOT MAIN AXIS -----------------------------
    #ax.plot(era5_profile['hws'], era5_profile['height'], color='blue',    lw=3)
    #ax.plot(cra_profile['hws'],  cra_profile['height'],  color='magenta', lw=3)
    ax.plot(dwl_profile['hws'], dwl_profile['height'], color='darkorange',      lw=3)
    ax.plot(snd['hws'],         snd['height'],         color='black',    lw=3)

    # ---- dual y‑axis (height ↔ pressure) -------------------------------------
    secax = ax.secondary_yaxis(
        'right',
        functions=(
            lambda h: mpcalc.height_to_pressure_std(h * units.m).to('hPa').magnitude,
            lambda p: mpcalc.pressure_to_height_std(p * units.hPa).to('m').magnitude
        )
    )

    ax.set_xlim(0, 40)
    
    alt_ticks_plot = alt_ticks[alt_ticks >= Subplot_yaxis]
    ax.set_yticks(alt_ticks)
    
    p_ticks_plot = [p for p in p_ticks if p2alt(p) >= Subplot_yaxis]
    secax.set_yticks(p_ticks)
    
    ax.set_ylim(0, Subplot_yaxis)
    ax.xaxis.set_major_locator(MultipleLocator(5))
    ax.tick_params(axis='x', which='major', length=12, width=2, colors='darkorange')
    ax.tick_params(axis='x', which='minor', length=6, width=2, colors='darkorange')

    # spine colours
    ax.spines['left' ].set_color('red')
    ax.spines['bottom'].set_color('darkorange')
    ax.spines['right'].set_color('black')
    ax.tick_params(axis='y', colors='red', length=12, width=2)
    secax.tick_params(axis='y', colors='blue', length=12, width=2)

    # ---- dashed reference lines (unchanged) ----------------------------------
    for y in alt_ticks:
        ax.axhline(y, color='red',   lw=2, ls=(0, (5, 10)), zorder=0)
    for p in p_ticks:
        ax.axhline(p2alt(p), color='blue',  lw=2, ls=(0, (5, 10)), zorder=0)

    # # --------------------------- INSET AXIS ----------------------------------
    # ax_in = inset_axes(ax, width="31%", height="48%", loc='upper right',
    #                    borderpad=1)
    # ax_inset.set_facecolor('white')         # 用白色把主图线条遮掉
    # ax_inset.patch.set_alpha(1)             # 保证不透明
    # ax_inset.patch.set_zorder(2) 

    # # plot same datasets, restricted to < 500 m
    # # ax_in.plot(era5_profile.loc[era5_profile['height'] >= Subplot_yaxis, 'hws'],
    # #            era5_profile.loc[era5_profile['height'] >= Subplot_yaxis, 'height'],
    # #            color='blue',   lw=3)
    # # ax_in.plot(cra_profile.loc[cra_profile['height'] >= Subplot_yaxis, 'hws'],
    # #            cra_profile.loc[cra_profile['height'] >= Subplot_yaxis, 'height'],
    # #            color='magenta', lw=3)
    # ax_in.plot(dwl_profile.loc[dwl_profile['height'] >= Subplot_yaxis, 'hws'],
    #            dwl_profile.loc[dwl_profile['height'] >= Subplot_yaxis, 'height'],
    #            color='red',    lw=3)
    # ax_in.plot(snd.loc[snd['height'] >= Subplot_yaxis, 'hws'],
    #            snd.loc[snd['height'] >= Subplot_yaxis, 'height'],
    #            color='black',  lw=3)

    # # inset axis limits & ticks
    # ax_in.set_ylim(Subplot_yaxis, 10010)
    # ax_in.set_xlim(0, 45)
    
    # # >>> 强制在 y 轴最底部显示 Subplot_yaxis 这一个刻度 <<<
    # y_ticks_inset = np.concatenate([[Subplot_yaxis],
    #                                 alt_ticks[(alt_ticks > Subplot_yaxis) &
    #                                           (alt_ticks >= 4000)]])
    # ax_in.set_yticks(y_ticks_inset)
    
    
    # ax_in.tick_params(axis='both', which='major', length=8, width=2, labelsize=30)
    # ax_in.tick_params(axis='both', which='minor', length=4, width=2)

    # # inset spines styling
    # ax_in.spines['top'   ].set_visible(False)
    # ax_in.spines['right' ].set_visible(False)
    # ax_in.spines['left'  ].set_color('red')
    # ax_in.spines['bottom'].set_color('darkorange')
    # ax_in.tick_params(axis='y', colors='red')
    # ax_in.tick_params(axis='x', colors='darkorange')
    # ax_in.xaxis.set_major_locator(MultipleLocator(10))

    # -------------------- subplot‑specific labels / title --------------------
    col = idx % n_cols
    row = idx // n_cols
    center_col = n_cols // 2 
    if col == 0:
        ax.set_ylabel('')  # we move global y‑label outside fig
    else:
        ax.tick_params(axis='y', labelleft=False)
    if col == n_cols - 1:
        secax.set_ylabel('')  # right‑side label moved to global text
    else:
        secax.tick_params(axis='y', labelright=False)

    if row == n_rows - 1:                       # 底行
        ax.tick_params(axis='x', labelbottom=True)   # 刻度数字保留
        if col == center_col:                   # 只有中间那格要 xlabel
            ax.set_xlabel('Horizontal Wind Speed (m/s)',
                      fontsize=50, weight='bold', color='darkorange',
                      labelpad=20)
    else:                                       # 其余行
        ax.tick_params(axis='x', labelbottom=False)  # 刻度数字隐藏

    # panel title
    ax.set_title(target_time.strftime('%Y-%m-%d %H:%M (BJT)'), pad=12, fontsize=36,weight='bold')

# ───────────────────── Hide any empty sub‑axes ────────────────────────────────
for j in range(idx + 1, n_rows * n_cols):
    fig.delaxes(axes[j])

# ───────────────────── Global axes‑level annotations ──────────────────────────
fig.text(-0.01, 0.5, 'Height (m a.g.l.)', rotation=90, va='center', ha='left',
         color='black', weight='bold', fontsize=50)
fig.text(1.01, 0.5, 'Pressure (hPa)', rotation=270, va='center', ha='right',
         color='black', weight='bold', fontsize=50)

# Legend & title (top)
legend_handles = [
    #plt.Line2D([], [], color='blue',    lw=6, label='ERA5'),
    #plt.Line2D([], [], color='magenta', lw=6, label='CRA'),
    plt.Line2D([], [], color='darkorange',     lw=6, label='Doppler Wind Lidar'),
    plt.Line2D([], [], color='black',   lw=6, label='Radiosonde')
]
fig.legend(handles=legend_handles, loc='upper center', bbox_to_anchor=(0.5, 0.995),
           ncol=4, frameon=False, fontsize=32)
fig.text(0.5, 0.993, 'Horizontal Wind Speed Radiosonde Comparison – 5min Temporal Resolution',
         ha='center', va='bottom', fontsize=40, weight='bold')

# ───────────────────── Save figure (no transparency) ──────────────────────────
fig.tight_layout(rect=[0, 0, 1, 0.985])
save_path = os.path.join(out_dir, out_fname)
fig.savefig(save_path, dpi=300, pil_kwargs={"compression": "tiff_adobe_deflate"},
            format='tif', bbox_inches='tight')
print(f"✅ Figure saved to: {save_path}")

##########################三维轨迹图+地面地图##############

##########################需要在Jupyter中运行来开启三维交互##############
# import os
# import requests
# import numpy as np
# import pandas as pd
# import plotly.graph_objects as go
# from PIL import Image
# from metpy.calc import pressure_to_height_std
# from metpy.units import units      # ← 注意是 from … import units
# from io import BytesIO


# # 设置文件路径和输出目录
# excel_path = r"E:\Beijing2024\探空数据-54511\original combined2.xlsx"
# out_dir = r"E:\Beijing2024\出图TIF"
# os.makedirs(out_dir, exist_ok=True)

# # 1. 读取所有工作表并清洗数据
# xls = pd.ExcelFile(excel_path)
# data_sheets = {}
# for sheet in xls.sheet_names:
#     df = pd.read_excel(xls, sheet_name=sheet, header=0)  # 默认无表头
#     # 提取第2列（index1）、第3列（index2）、第4列（index3）
#     df = df.iloc[:, [1, 2, 3]]
#     df.columns = ['lon', 'lat', 'pressure_hPa']
#     df = df.astype(float)                # 强制转换为浮点
#     df = df.dropna(how='any')           # 删除含 NaN 的行:contentReference[oaicite:6]{index=6}
#     if df.empty:
#         continue
#     data_sheets[sheet] = df

# # 2. 压力转换为高度（标准大气）:contentReference[oaicite:7]{index=7}
# for sheet, df in data_sheets.items():
#     pressure = df['pressure_hPa'].values * units('hPa')
#     height = pressure_to_height_std(pressure)  # 返回带单位的高度
#     df['height_m'] = height.to('m').magnitude

# # 3. 获取高德静态地图图像（经纬范围 116–117E, 39.5–40.5N）
# api_key = "f53bda1e2e8685bdbb4529a24d03f063"
# center = "116.5,40.0"   # 地图中心
# zoom = 9
# img_size = "6400*1600"
# static_map_url = (
#     f"https://restapi.amap.com/v3/staticmap?"
#     f"location={center}&zoom={zoom}&size={img_size}&key={api_key}"
# )
# resp = requests.get(static_map_url)
# img = Image.open(BytesIO(resp.content))
# # 转换为8位调色板图像以获取颜色映射
# eight_bit_img = img.convert('P', palette=Image.WEB, dither=None)
# palette = np.array(eight_bit_img.getpalette()).reshape((-1, 3))
# colorscale = [[i/255.0, f"rgb({r},{g},{b})"] for i,(r,g,b) in enumerate(palette)]

# # 构造地图面 x,y,z
# im = np.array(img)
# H, W = im.shape[:2]
# lon_min, lon_max = 116.0, 117.0
# lat_min, lat_max = 39.5, 40.5
# x = np.linspace(lon_min, lon_max, W)
# y = np.linspace(lat_max, lat_min, H)  # 注意纬度方向取反，使图像不倒置
# z = np.zeros((W, H))  # z 为零平面:contentReference[oaicite:8]{index=8}

# # 4. 创建 3D 图形并绘制轨迹和地图底面
# fig = go.Figure()

# # 各时次轨迹，每条线使用不同颜色和名称
# color_list = ["red", "green", "blue", "orange", "purple", "cyan", "magenta"]
# for i, (sheet, df) in enumerate(data_sheets.items()):
#     fig.add_trace(go.Scatter3d(
#         x=df['lon'], y=df['lat'], z=df['height_m'],
#         mode='lines',
        
#         marker=dict(size=4),
#         line=dict(color=color_list[i % len(color_list)], width=2),
#         name=sheet  # 图例显示时次
#     ))
# line=dict(
#     #color='rgba(0, 0, 255, 0.8)',  # 半透明蓝色
#     width=8
# )
# # 添加高德地图平面（Surface）:contentReference[oaicite:9]{index=9}
# fig.add_trace(go.Surface(
#     x=x, y=y, z=z,
#     surfacecolor=np.array(eight_bit_img),  # 调色板索引图
#     colorscale=colorscale,
#     cmin=0, cmax=255,
#     showscale=False,
#     lighting_diffuse=1, lighting_ambient=1,
#     lighting_fresnel=1, lighting_roughness=1, lighting_specular=0.5
# ))

# # 5. 设置布局：透明背景、轴标签、视角等:contentReference[oaicite:10]{index=10}
# fig.update_layout(
#     title="北京地区探空轨迹（3D）",
#     scene=dict(
#         xaxis_title='Longitude (°E)',
#         yaxis_title='Latitude (°N)',
#         zaxis_title='Height (m a.g.l.)',
#         xaxis=dict(range=[lon_min, lon_max],gridcolor='gray',title_font=dict(size=20),
#             tickfont=dict(size=13)),     # X轴刻度字号
#         yaxis=dict(range=[lat_min, lat_max],title_font=dict(size=20),  # X轴标题字号
#             tickfont=dict(size=13),     # X轴刻度字号
#                    gridcolor='gray'),
#         zaxis=dict(range=[0, 10000], title_font=dict(size=20),  # X轴标题字号
#             tickfont=dict(size=13),     # X轴刻度字号
#                    tickformat=".0f", 
#                    gridcolor='gray'  # 设置z轴网格线颜色为黑色
#         )
#     ),
#     margin=dict(l=0, r=0, b=0, t=50),
#     legend=dict(title="时次"),
#     paper_bgcolor='rgba(0,0,0,0)',  # 背景透明:contentReference[oaicite:11]{index=11}
#     plot_bgcolor='rgba(0,0,0,0)'
# )

# # 设置初始视角，方便观察
# # 在原有代码的布局设置部分增加以下参数
# fig.update_layout(
#     height=1500, 
#     width=1100, 
#     # 原有布局参数保持不变...
#     scene=dict(
#         # 原有轴设置保持不变...
#         # 新增参数调整显示区域
#         aspectmode="manual",  # 手动控制比例
#         aspectratio=dict(x=1, y=1, z=0.8),  # 调整三维空间比例
#         camera=dict(  # 调整相机视角参数
#             up=dict(x=0, y=0, z=1),
#             center=dict(x=0, y=0, z=0),
#             eye=dict(x=1.5, y=1.5, z=1.5)  # 调整观察点位置
#         )
#     )
    
# )


# ───────────────────────────  export_plot_profiles.py  ──────────────────────────
# ─── precheck_plot_export.py ─────────────────────────────────────────

import pandas as pd
import numpy as np
import sys

# ❶ 定义“必须存在”的变量及其关键列
REQUIRED = {
    "ERA5_interp_BJ" : ["pressure_level", "t"],
    "cra_regular"    : ["Level (hPa)", "rh (%)", "hws(m/s)"],
    "t_mwr_long"     : [],                # 已是多层索引，列检查可省
    "rh_mwr_long"    : [],
    "cleaned_h"      : ["BJT", "hws"],
    "cleaned_h_60min": ["BJT", "hws"],
    "cleaned_h_5min" : ["BJT", "hws"],
}

SCALARS = ["Altitude", "site_elev"]       # 二者至少要有一个

errors = []

# ❷ 变量是否存在、是否 DataFrame、关键列是否在
for var, cols in REQUIRED.items():
    if var not in globals():
        errors.append(f"❌ 缺少变量  {var}")
        continue
    obj = globals()[var]
    if not isinstance(obj, pd.DataFrame):
        errors.append(f"❌ {var} 不是 pandas.DataFrame (而是 {type(obj)})")
        continue
    for c in cols:
        if c not in obj.columns:
            errors.append(f"❌ {var} 缺少列 '{c}'")
    if obj.empty:
        errors.append(f"⚠️  {var} 是空 DataFrame")

# ❸ 标高 / 站点海拔是否提供
if not any(s in globals() for s in SCALARS):
    errors.append("❌ 未找到 Altitude 或 site_elev")

# ❹ 打印结果并视需要退出
if errors:
    print("\n".join(errors))
    sys.exit("‼️  数据检查未通过，请先准备好以上对象再执行导出。")
else:
    print("✅ 所有必要数据已就绪，可安全执行导出。")
# ───────────────────────────────────────────────────────────────────

# =============================================================================
# 项目：多源廓线数据融合与导出（Radiosonde / MWR / DWL / ERA5 / CRA）
#
# 功能概述：
# 1) 读取探空 Excel（多 sheet，sheet 名形如 YYYYMMDDHH），将“气压→高度”并作为全流程的高度基准；
# 2) 1 分钟层（time-height 逐点配准）：
#    - 锚点：直接使用“探空的原始时间与高度”作为 (time, height) 取样点（不做整分重采样）；
#    - MWR（温度/湿度）：从 (timestamps × height) 的长表构建规则网格，使用 scipy.interpn
#      在 (time, height) 上双线性插值到探空锚点；随后按每次发射（launch）
#      “从首个非 NaN 前的连续 NaN 跳过、从第一个有效值开始，遇到首个 NaN 即停止”的策略裁剪，
#      确保 MWR 与探空在 1min 层一一配对；
#    - DWL（风速 1min）：由 cleaned_h 聚合得到 cleaned_h_1min（按高度×1min 分箱取均值），
#      构建 (time × height) 网格并插值到探空锚点；**不截断**，仅保留 DWL 插值后非空点，
#      与探空在 (launch, time, height) 上内连接配对，保证时空匹配；
#    - 将上述配对后的 Radiosonde/MWR/DWL 的 1min 数据分别写入 temp_1min / rh_1min / hws_1min。
#
# 3) 1 小时层（高度对齐）：
#    - 对每个时次（sheet）：以该时次探空的高度集合（去重升序）为“目标高度网格”；
#    - 将 ERA5（压力层→高度）、CRA（压力层→高度）、MWR（±30 min 平均剖面）、
#      DWL（整点 60min 剖面）沿“高度维”线性插值到该目标高度；
#    - 探空自身的温/湿/风在目标高度上做按高聚合（无需再插值）；
#    - 将上述结果写入 temp_1h / rh_1h / hws_1h。
#
# 4) 统一限制高度：全流程仅保留 height ≤ MAX_H（含），默认 MAX_H=10000 m。
#
# 5) 导出到 Excel：
#    - 统一列：time, launch, source, height, value；
#    - 统一排序：time → height → source（source 固定顺序：Radiosonde → MWR → DWL → ERA5 → CRA）；
#    - 对 hws_1h 额外丢弃 value 为 NaN 的行；
#    - 所有工作表列宽统一设置为 25；
#    - 输出文件路径由 OUT_XLSX 指定（覆盖式写出）。
#
# 主要输入（外部依赖的 DataFrame，需在运行前准备好）：
# - ERA5_interp_BJ：列含 ['valid_time','pressure_level','t','r','hws(m/s)']；
# - cra_regular   ：列含 ['timestamps','Level (hPa)','Temp (K)','rh (%)','hws(m/s)']；
# - t_mwr_long    ：MultiIndex (timestamps, height) 的长表/Series（列名/值名 'T(K)'）；
# - rh_mwr_long   ：同上（'RH(%)'）；
# - cleaned_h     ：DWL 原始逐条数据，列含 ['BJT','height','hws']（用于生成 1min）；
# - cleaned_h_60min：DWL 整点 1 小时剖面，列含 ['BJT','height','hws']；
# - 探空 Excel（EXCEL_IN）：每个 sheet 至少含 ['time','pressure_hPa','temperature_C','relative humidity_%',
#   'geopotential height_m','wind speed_m/s'] 中的若干列；气压将被转换为高度作为主轴。
#
# 关键算法/实现细节：
# - p2h：使用 MetPy 标准大气（pressure_to_height_std）将“裸 hPa”转换为几何高度并减去站点海拔；
# - 2D 插值：scipy.interpolate.interpn，在 (time, height) 规则网格上对探空锚点做双线性插值；
# - 1D 插值：scipy.interpolate.interp1d，仅沿高度维线性插值，越界填 NaN（不外推）；
# - 1min 温/湿裁剪（仅 MWR）：按每个 launch，从首个有效值开始保留，遇到首个 NaN 即停止；
# - DWL 1min 不截断：仅保留非空并与探空配对；
# - 统一类型清洗：对用于计算/插值的列强制数值化（errors='coerce'），并在必要处去 NaN；
# - 安全性：构建网格时保证时间与高度轴严格单调（去重、合并重复高度、近似重复列按 3 位小数再合并）。
#
# 主要可调参数：
# - EXCEL_IN  ：探空多 sheet Excel 路径；
# - OUT_XLSX  ：输出 Excel 路径（覆盖写）；
# - SITE_ELEV ：站点海拔（m），用于从标准大气高度中减去；
# - MAX_H     ：统一的高度上限（默认 10000 m，含）。
#
# 已移除/不包含：
# - 早先的 DWL 5 分钟整点对齐与插值逻辑（prepare_dwl_5min_block 及相关入库）已全面删除。
# =============================================================================

import re
import pandas as pd
import numpy as np
from pathlib import Path
import metpy.calc as mpcalc
from metpy.units import units
from scipy.interpolate import interpn, interp1d
from pandas.api.types import CategoricalDtype

# ================================ 配置区 ================================ #
EXCEL_IN   = r"E:\Beijing2024\探空数据-54511\original combined2.xlsx"   # 探空原始表（多 sheet）
OUT_XLSX   = Path(r"E:\Beijing2024\出图TIF\plot_data for representativeness.xlsx")                            # 输出 Excel
SITE_ELEV  = 49.0        # 站点海拔，m
MAX_H      = 10000.0     # 高度上限（含）
SHEET_LIST = pd.ExcelFile(EXCEL_IN).sheet_names
# ======================================================================= #

# ────────────────────────── 基础工具 ────────────────────────── #
def p2h(p_hpa_series, site_elev=SITE_ELEV):
    """'裸 hPa' → m a.g.l.（减去站点海拔）"""
    p_vals = np.asarray(p_hpa_series, dtype='float64')
    p_q = p_vals * units.hPa
    return mpcalc.pressure_to_height_std(p_q).to('m').magnitude - site_elev

def _to_epoch_seconds(ts_series):
    return (pd.to_datetime(ts_series) - pd.Timestamp("1970-01-01")) // pd.Timedelta('1s')

def _launch_from_sheet(sheet_name: str) -> pd.Timestamp:
    """工作表名 'YYYYMMDDHH'（例如 2024080608） → Timestamp"""
    if not re.fullmatch(r"\d{10}", str(sheet_name)):
        raise ValueError(f"工作表名不符合 YYYYMMDDHH 格式: {sheet_name}")
    return pd.to_datetime(sheet_name, format="%Y%m%d%H", errors="raise")

def _stash(container, df, variable, source, ts, val_col, time_col=None, launch=None, launch_col='launch'):
    """
    收集曲线 df → 统一列：time, launch, source, height, value, variable
    - 若 df 里有 launch 列则保留；否则用 launch 参数（通常是该 sheet 的 launch_ts）。
    - 若提供 time_col，则使用逐行时间戳；否则统一使用 ts。
    """
    if df is None or df.empty:
        return
    rec = df.rename(columns={val_col: 'value'}).copy()
    # time
    if time_col and (time_col in rec.columns):
        rec['time'] = pd.to_datetime(rec[time_col])
    else:
        rec['time'] = ts
    # launch
    if launch_col in rec.columns:
        rec['launch'] = pd.to_datetime(rec[launch_col], errors='coerce')
    else:
        rec['launch'] = pd.to_datetime(launch) if launch is not None else pd.NaT
    # 统一列
    rec = (rec[['time', 'launch', 'height', 'value']]
           .assign(variable=variable, source=source))
    container.append(rec)

# ────────────────────────── 网格构建（1min/长表） ────────────────────────── #
def _grid_from_long(long_df, value_col=None):
    """
    接受：
      - Series（MultiIndex=(timestamps, height)）
      - DataFrame + value_col 指定列名；或单列 DataFrame
    返回: (t_axis_num, h_axis, grid_vals)
    """
    if long_df is None or len(long_df) == 0:
        return None, None, None

    # 取到 Series
    if isinstance(long_df, pd.Series):
        series = long_df
    elif isinstance(long_df, pd.DataFrame):
        if value_col is not None and value_col in long_df.columns:
            series = long_df[value_col]
        elif long_df.shape[1] == 1:
            series = long_df.iloc[:, 0]
        else:
            raise ValueError("_grid_from_long: 传入 DataFrame 请提供 value_col。")
    else:
        raise TypeError("_grid_from_long: 仅支持 Series 或 DataFrame。")

    # 展宽：index=timestamps, columns=height
    wide = series.unstack()
    wide.index = pd.to_datetime(wide.index, errors='coerce')
    wide = wide.sort_index().groupby(wide.index).mean()

    # 高度列为数值并限高
    wide.columns = pd.to_numeric(wide.columns, errors='coerce')
    wide = wide.loc[:, ~np.isnan(wide.columns)]
    if 'MAX_H' in globals():
        wide = wide.loc[:, [c for c in wide.columns if c <= MAX_H]]
    if wide.shape[1] == 0:
        return None, None, None

    # 合并重复高度并升序
    wide = wide.T.groupby(level=0).mean().T
    wide = wide.reindex(sorted(wide.columns), axis=1)

    # 轴
    t_axis = wide.index
    h_axis = wide.columns.to_numpy(dtype='float64')
    t_num  = ((t_axis - pd.Timestamp("1970-01-01")) // pd.Timedelta('1s')).to_numpy(dtype='int64')

    # 去重时间
    if np.any(np.diff(np.sort(t_num)) == 0):
        _, first_idx = np.unique(t_num, return_index=True)
        first_idx.sort()
        wide = wide.iloc[first_idx]
        t_axis = wide.index
        t_num  = ((t_axis - pd.Timestamp("1970-01-01")) // pd.Timedelta('1s')).to_numpy(dtype='int64')

    # 近似重复高度再合并一次
    if not np.all(np.diff(h_axis) > 0):
        wide_round = wide.copy()
        wide_round.columns = np.round(wide_round.columns.astype(float), 3)
        wide_round = wide_round.T.groupby(level=0).mean().T
        wide_round = wide_round.reindex(sorted(wide_round.columns), axis=1)
        h_axis = wide_round.columns.to_numpy(dtype='float64')
        wide   = wide_round

    if len(t_num) < 2 or len(h_axis) < 2:
        return None, None, None
    if not np.all(np.diff(h_axis) > 0) or not np.all(np.diff(np.sort(t_num)) > 0):
        return None, None, None

    return t_num, h_axis, wide.to_numpy()

def _grid_from_1min_df(df, time_col='BJT', height_col='height', value_col='hws'):
    """
    从 'cleaned_h_1min' 构建 (time×height) 网格以供 interpn：
    返回 (t_axis_num, h_axis, grid_vals)
    """
    if df is None or df.empty:
        return None, None, None
    tmp = df[[time_col, height_col, value_col]].copy()
    tmp[time_col]   = pd.to_datetime(tmp[time_col], errors='coerce')
    tmp[height_col] = pd.to_numeric(tmp[height_col], errors='coerce')
    tmp[value_col]  = pd.to_numeric(tmp[value_col], errors='coerce')
    tmp = tmp.dropna(subset=[time_col, height_col, value_col])
    if 'MAX_H' in globals():
        tmp = tmp.loc[tmp[height_col] <= MAX_H]
    if tmp.empty:
        return None, None, None

    wide = tmp.pivot_table(index=time_col, columns=height_col, values=value_col, aggfunc='mean')
    wide = wide.sort_index()
    wide = wide.groupby(wide.index).mean()

    cols = pd.to_numeric(pd.Index(wide.columns), errors='coerce')
    wide.columns = cols
    wide = wide.loc[:, ~np.isnan(wide.columns)]
    wide = wide.T.groupby(level=0).mean().T
    wide = wide.reindex(sorted(wide.columns), axis=1)
    if wide.shape[1] < 2 or wide.shape[0] < 2:
        return None, None, None

    t_axis = wide.index
    h_axis = wide.columns.to_numpy(dtype='float64')
    t_num  = _to_epoch_seconds(t_axis).to_numpy(dtype='int64')

    if np.any(np.diff(np.sort(t_num)) == 0):
        _, first_idx = np.unique(t_num, return_index=True)
        first_idx.sort()
        wide = wide.iloc[first_idx]
        t_axis = wide.index
        t_num  = _to_epoch_seconds(t_axis).to_numpy(dtype='int64')

    if not (np.all(np.diff(h_axis) > 0) and np.all(np.diff(np.sort(t_num)) > 0)):
        return None, None, None

    return t_num, h_axis, wide.to_numpy()

# ────────────────────────── 1D 高度插值（1 小时用） ────────────────────────── #
def _unique_sorted_nanfloat(arr_like):
    a = pd.to_numeric(pd.Series(arr_like), errors='coerce').astype('float64').to_numpy()
    a = a[np.isfinite(a)]
    return np.unique(np.sort(a))

def _interp_profile_to_targets(src_h, src_v, tgt_h):
    src_h = pd.to_numeric(pd.Series(src_h), errors='coerce').to_numpy(dtype='float64')
    src_v = pd.to_numeric(pd.Series(src_v), errors='coerce').to_numpy(dtype='float64')
    msk = np.isfinite(src_h) & np.isfinite(src_v)
    if msk.sum() < 2 or len(tgt_h) == 0:
        return np.full(len(tgt_h), np.nan, dtype='float64')
    df = (pd.DataFrame({'h': src_h[msk], 'v': src_v[msk]})
            .groupby('h', as_index=False)['v'].mean()
            .sort_values('h'))
    if df.shape[0] < 2:
        return np.full(len(tgt_h), np.nan, dtype='float64')
    f = interp1d(df['h'].to_numpy(), df['v'].to_numpy(),
                 kind='linear', bounds_error=False, fill_value=np.nan, assume_sorted=True)
    return f(np.asarray(tgt_h, dtype='float64'))

# ────────────────────────── MWR 1min 的“从首个有效值到首个 NaN”裁剪 ────────────────────────── #
def _trim_from_first_finite_then_stop_at_first_nan(df: pd.DataFrame, value_col: str, order_col: str = 'time') -> pd.DataFrame:
    """
    每个 launch 内，按 order_col 排序：
      - 跳过开头连续的 NaN
      - 从第一个非 NaN 开始，直到遇到第一个 NaN 为止（不含该 NaN 行）
    返回时显式带回 'launch' 列，兼容新旧 pandas。
    """
    if df.empty:
        return df

    def _one(g: pd.DataFrame) -> pd.DataFrame:
        g = g.sort_values(order_col)
        vals = g[value_col].to_numpy()
        isn  = np.isnan(vals)
        finite_idx = np.flatnonzero(~isn)
        if finite_idx.size == 0:
            return g.iloc[0:0]
        start = int(finite_idx[0])
        tail_isn = isn[start:]
        nnz = np.flatnonzero(tail_isn)
        end = start + int(nnz[0]) if nnz.size > 0 else len(vals)
        return g.iloc[start:end]

    try:
        # 新版：把分组列从子表里排除，但保留为外层索引，之后还原成一列
        out = df.groupby('launch', group_keys=True).apply(_one, include_groups=False)
        # launch 作为外层索引层名通常就是 'launch'；若无名则回退到 'level_0'
        lvl_name = out.index.names[0] if isinstance(out.index, pd.MultiIndex) else None
        out = out.reset_index(level=0).rename(columns={lvl_name or 'level_0': 'launch'})
        return out
    except TypeError:
        # 旧版 pandas：不支持 include_groups；仍然用 group_keys=True，之后从索引恢复
        out = df.groupby('launch', group_keys=True).apply(_one)
        lvl_name = out.index.names[0] if isinstance(out.index, pd.MultiIndex) else None
        out = out.reset_index(level=0).rename(columns={lvl_name or 'level_0': 'launch'})
        return out


# ────────────────────────── 1 分钟：探空锚点 + MWR 温/湿 + DWL 风速 ────────────────────────── #
records = []

print("▶ 汇总 Radiosonde 原始时间与高度锚点…")
_rs_dict = pd.read_excel(EXCEL_IN, sheet_name=None)

def _prep_rs_one(sheet_name, df):
    want = ['time', 'pressure_hPa', 'temperature_C', 'relative humidity_%',
            'geopotential height_m', 'wind speed_m/s']
    present = [c for c in want if c in df.columns]
    out = df[present].copy()
    # 类型
    out['launch'] = _launch_from_sheet(sheet_name)  # ← 来自工作表名
    out['time']   = pd.to_datetime(out['time'])
    for col in ['pressure_hPa','temperature_C','relative humidity_%','geopotential height_m','wind speed_m/s']:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce')
    # 高度
    out = out[out['pressure_hPa'].between(50, 1100)]
    out['height'] = p2h(out['pressure_hPa'])
    out = out.loc[out['height'] <= MAX_H]
    # 物理量
    out['temperature_K'] = out['temperature_C'] + 273.15
    out['value_pc']      = out['relative humidity_%']
    out['hws']           = out.get('wind speed_m/s')
    return out[['launch','time','height','temperature_K','value_pc','hws']]

rs_native = pd.concat([_prep_rs_one(k, v) for k, v in _rs_dict.items()], ignore_index=True)

RS_TIME   = pd.to_datetime(rs_native['time'])
RS_HGT    = rs_native['height'].to_numpy(dtype='float64')
RS_LAUNCH = rs_native['launch'].to_numpy()
RS_PTS_2D = np.column_stack([_to_epoch_seconds(RS_TIME), RS_HGT])

# —— MWR：在 (time,height) 上直接插值到探空锚点；按 launch 裁剪；与 RS 成对 —— #
print("▶ 1min 温/湿（MWR → RS 锚点 & 裁剪配对）…")

# 温度
tn_T = hp_T = grid_T = None
if 't_mwr_long' in globals() and t_mwr_long is not None and len(t_mwr_long) > 0:
    tn_T, hp_T, grid_T = _grid_from_long(t_mwr_long, 'T(K)')

# 湿度
tn_R = hp_R = grid_R = None
if 'rh_mwr_long' in globals() and rh_mwr_long is not None and len(rh_mwr_long) > 0:
    tn_R, hp_R, grid_R = _grid_from_long(rh_mwr_long, 'RH(%)')

# 插值
mwr_T_df = pd.DataFrame(columns=['launch','time','height','T(K)'])
if tn_T is not None:
    mwr_T_on_rs = interpn((tn_T, hp_T), grid_T, RS_PTS_2D,
                          method='linear', bounds_error=False, fill_value=np.nan)
    mwr_T_df = pd.DataFrame({'launch': RS_LAUNCH, 'time': RS_TIME.to_numpy(),
                             'height': RS_HGT, 'T(K)': mwr_T_on_rs})

mwr_R_df = pd.DataFrame(columns=['launch','time','height','RH(%)'])
if tn_R is not None:
    mwr_R_on_rs = interpn((tn_R, hp_R), grid_R, RS_PTS_2D,
                          method='linear', bounds_error=False, fill_value=np.nan)
    mwr_R_df = pd.DataFrame({'launch': RS_LAUNCH, 'time': RS_TIME.to_numpy(),
                             'height': RS_HGT, 'RH(%)': mwr_R_on_rs})

# 裁剪（仅 MWR）
mwr_T_cut = _trim_from_first_finite_then_stop_at_first_nan(mwr_T_df, value_col='T(K)',  order_col='time')
mwr_R_cut = _trim_from_first_finite_then_stop_at_first_nan(mwr_R_df, value_col='RH(%)', order_col='time')

# 用裁剪后的 (launch,time,height) 作为键，与 RS 匹配，保证一一对应
keys_T = mwr_T_cut[['launch','time','height']].drop_duplicates()
keys_R = mwr_R_cut[['launch','time','height']].drop_duplicates()

rs_T_cut = (rs_native[['launch','time','height','temperature_K']]
            .merge(keys_T, on=['launch','time','height'], how='inner'))
rs_R_cut = (rs_native[['launch','time','height','value_pc']]
            .merge(keys_R, on=['launch','time','height'], how='inner'))

# 入库（temp_1min / rh_1min：RS 与 MWR 成对，包含 launch）
_stash(records, rs_T_cut[['launch','time','height','temperature_K']],
       'temp_1min', 'Radiosonde', pd.NaT, 'temperature_K', time_col='time')
_stash(records, mwr_T_cut[['launch','time','height','T(K)']],
       'temp_1min', 'MWR',        pd.NaT, 'T(K)',          time_col='time')

_stash(records, rs_R_cut[['launch','time','height','value_pc']],
       'rh_1min',   'Radiosonde', pd.NaT, 'value_pc',      time_col='time')
_stash(records, mwr_R_cut[['launch','time','height','RH(%)']],
       'rh_1min',   'MWR',        pd.NaT, 'RH(%)',         time_col='time')

print(f"   → 成对样本（温度）：{len(rs_T_cut)} | 成对样本（湿度）：{len(rs_R_cut)}")

# —— DWL 1min 风速：不截断，仅保留非空并与 RS 配对 —— #
print("▶ 1min 风速（DWL → RS 锚点 & 配对，不截断）…")

# 1) 从 cleaned_h 生成 cleaned_h_1min（高度×1min 聚合）
cleaned_h_1min = pd.DataFrame(columns=['height','BJT','hws'])
if 'cleaned_h' in globals() and cleaned_h is not None and len(cleaned_h) > 0:
    tmp = cleaned_h.copy()
    tmp['BJT']    = pd.to_datetime(tmp['BJT'], errors='coerce')
    tmp['height'] = pd.to_numeric(tmp['height'], errors='coerce')
    tmp['hws']    = pd.to_numeric(tmp['hws'], errors='coerce')
    cleaned_h_1min = (
        tmp.groupby(['height', pd.Grouper(key='BJT', freq='1min')], as_index=False)
           .mean(numeric_only=True)
           .dropna(subset=['hws'])
    )
    cleaned_h_1min = cleaned_h_1min.loc[cleaned_h_1min['height'] <= MAX_H]

# 2) 构建 (time×height) 网格并 2D 插值到 RS 锚点
dwl_tn = dwl_haxis = dwl_grid = None
if not cleaned_h_1min.empty:
    dwl_tn, dwl_haxis, dwl_grid = _grid_from_1min_df(
        cleaned_h_1min, time_col='BJT', height_col='height', value_col='hws'
    )

dwl_1min_df = pd.DataFrame(columns=['launch','time','height','hws'])
if dwl_tn is not None:
    dwl_on_rs = interpn((dwl_tn, dwl_haxis), dwl_grid, RS_PTS_2D,
                        method='linear', bounds_error=False, fill_value=np.nan)
    dwl_1min_df = pd.DataFrame({
        'launch': RS_LAUNCH,
        'time':   RS_TIME.to_numpy(),
        'height': RS_HGT,
        'hws':    dwl_on_rs
    })

# 3) 仅保留 DWL 非空点，并与 RS 在 (launch,time,height) 上配对（时空匹配）
dwl_nonnull = dwl_1min_df.dropna(subset=['hws'])
if not dwl_nonnull.empty:
    keys_H = dwl_nonnull[['launch','time','height']].drop_duplicates()
    rs_H_match = (rs_native[['launch','time','height','hws']]
                  .merge(keys_H, on=['launch','time','height'], how='inner'))
    # 入库（包含 launch）
    _stash(records, rs_H_match[['launch','time','height','hws']],
           'hws_1min', 'Radiosonde', pd.NaT, 'hws', time_col='time')
    _stash(records, dwl_nonnull[['launch','time','height','hws']],
           'hws_1min', 'DWL',        pd.NaT, 'hws', time_col='time')
    print(f"   → 成对样本（风速）：{len(rs_H_match)}")
else:
    print("⚠️ DWL→RS 1min 风速：插值后全为 NaN 或无匹配点；本轮不写入。")

# ────────────────────────── 1 小时：高度对齐到 Radiosonde（≤10000 m） ────────────────────────── #
print("▶ 1 小时数据入库（高度对齐到 Radiosonde，≤10000 m）…")

for sheet in SHEET_LIST:
    launch_ts = _launch_from_sheet(sheet)  # 来自表名
    t0 = launch_ts

    # 该时次探空 → 目标高度
    snd = pd.read_excel(EXCEL_IN, sheet_name=sheet)
    snd['pressure_hPa'] = pd.to_numeric(snd['pressure_hPa'], errors='coerce')
    snd['height'] = p2h(snd['pressure_hPa'])
    snd = snd.loc[snd['height'] <= MAX_H].copy()
    tgt_h = _unique_sorted_nanfloat(snd['height'])
    if tgt_h.size == 0:
        continue

    # Radiosonde 自身
    if 'temperature_C' in snd.columns:
        snd['temperature_C'] = pd.to_numeric(snd['temperature_C'], errors='coerce')
        rs_T = (snd.assign(temperature_K=snd['temperature_C'] + 273.15)
                   .dropna(subset=['height', 'temperature_K'])
                   .groupby('height', as_index=False)['temperature_K'].mean())
        rs_T = rs_T.set_index('height').reindex(tgt_h).reset_index()
        _stash(records, rs_T[['height','temperature_K']],
               'temp_1h', 'Radiosonde', t0, 'temperature_K', launch=launch_ts)

    if 'relative humidity_%' in snd.columns:
        snd['relative humidity_%'] = pd.to_numeric(snd['relative humidity_%'], errors='coerce')
        rs_RH = (snd.rename(columns={'relative humidity_%':'value_pc'})
                    .dropna(subset=['height', 'value_pc'])
                    .groupby('height', as_index=False)['value_pc'].mean())
        rs_RH = rs_RH.set_index('height').reindex(tgt_h).reset_index()
        _stash(records, rs_RH[['height','value_pc']],
               'rh_1h', 'Radiosonde', t0, 'value_pc', launch=launch_ts)

    if 'wind speed_m/s' in snd.columns:
        snd['wind speed_m/s'] = pd.to_numeric(snd['wind speed_m/s'], errors='coerce')
        rs_WS = (snd[['height','wind speed_m/s']]
                 .rename(columns={'wind speed_m/s':'hws'})
                 .dropna(subset=['height', 'hws'])
                 .groupby('height', as_index=False)['hws'].mean())
        rs_WS = rs_WS.set_index('height').reindex(tgt_h).reset_index()
        _stash(records, rs_WS[['height','hws']],
               'hws_1h', 'Radiosonde', t0, 'hws', launch=launch_ts)

    # ERA5
    if 'ERA5_interp_BJ' in globals():
        era5 = (ERA5_interp_BJ[ERA5_interp_BJ['valid_time'] == t0]
                .assign(height=lambda d: p2h(d['pressure_level'])))
        if not era5.empty:
            h_src = era5['height'].to_numpy()
            if 't' in era5.columns:
                v = _interp_profile_to_targets(h_src, era5['t'], tgt_h)
                _stash(records, pd.DataFrame({'height': tgt_h, 't': v}),
                       'temp_1h', 'ERA5', t0, 't', launch=launch_ts)
            if 'r' in era5.columns:
                v = _interp_profile_to_targets(h_src, era5['r'], tgt_h)
                _stash(records, pd.DataFrame({'height': tgt_h, 'r': v}),
                       'rh_1h', 'ERA5', t0, 'r', launch=launch_ts)
            if 'hws(m/s)' in era5.columns:
                v = _interp_profile_to_targets(h_src, era5['hws(m/s)'], tgt_h)
                _stash(records, pd.DataFrame({'height': tgt_h, 'hws(m/s)': v}),
                       'hws_1h', 'ERA5', t0, 'hws(m/s)', launch=launch_ts)

    # CRA
    if 'cra_regular' in globals():
        cra = (cra_regular
               .assign(timestamps=pd.to_datetime(cra_regular['timestamps'], errors='coerce'))
               .loc[lambda d: d['timestamps'] == t0]
               .assign(height=lambda d: p2h(d['Level (hPa)'])))
        if not cra.empty:
            h_src = cra['height'].to_numpy()
            if 'Temp (K)' in cra.columns:
                v = _interp_profile_to_targets(h_src, cra['Temp (K)'], tgt_h)
                _stash(records, pd.DataFrame({'height': tgt_h, 'Temp (K)': v}),
                       'temp_1h', 'CRA', t0, 'Temp (K)', launch=launch_ts)
            if 'rh (%)' in cra.columns:
                v = _interp_profile_to_targets(h_src, cra['rh (%)'], tgt_h)
                _stash(records, pd.DataFrame({'height': tgt_h, 'rh (%)': v}),
                       'rh_1h', 'CRA', t0, 'rh (%)', launch=launch_ts)
            if 'hws(m/s)' in cra.columns:
                v = _interp_profile_to_targets(h_src, cra['hws(m/s)'], tgt_h)
                _stash(records, pd.DataFrame({'height': tgt_h, 'hws(m/s)': v}),
                       'hws_1h', 'CRA', t0, 'hws(m/s)', launch=launch_ts)

    # MWR（±30min 平均）→ 高度插值
    win30, wout30 = t0 - pd.Timedelta(minutes=30), t0 + pd.Timedelta(minutes=30)
    if 't_mwr_long' in globals():
        mwr_T_1h = (t_mwr_long.reset_index()
                    .query("timestamps >= @win30 and timestamps <= @wout30")
                    .groupby('height', as_index=False)['T(K)'].mean())
        if not mwr_T_1h.empty:
            v = _interp_profile_to_targets(mwr_T_1h['height'], mwr_T_1h['T(K)'], tgt_h)
            _stash(records, pd.DataFrame({'height': tgt_h, 'T(K)': v}),
                   'temp_1h', 'MWR', t0, 'T(K)', launch=launch_ts)
    if 'rh_mwr_long' in globals():
        mwr_RH_1h = (rh_mwr_long.reset_index()
                     .query("timestamps >= @win30 and timestamps <= @wout30")
                     .groupby('height', as_index=False)['RH(%)'].mean())
        if not mwr_RH_1h.empty:
            v = _interp_profile_to_targets(mwr_RH_1h['height'], mwr_RH_1h['RH(%)'], tgt_h)
            _stash(records, pd.DataFrame({'height': tgt_h, 'RH(%)': v}),
                   'rh_1h', 'MWR', t0, 'RH(%)', launch=launch_ts)

    # DWL 1 小时（整点）
    if 'cleaned_h_60min' in globals():
        dwl_1h = (cleaned_h_60min
                  .loc[lambda d: pd.to_datetime(d['BJT']) == t0, ['height', 'hws']]
                  .dropna())
        if not dwl_1h.empty:
            v = _interp_profile_to_targets(dwl_1h['height'], dwl_1h['hws'], tgt_h)
            _stash(records, pd.DataFrame({'height': tgt_h, 'hws': v}),
                   'hws_1h', 'DWL', t0, 'hws', launch=launch_ts)

# ────────────────────────── 写出 Excel（带 launch，过滤/排序/列宽=25） ────────────────────────── #
from openpyxl.utils import get_column_letter

if not records:
    raise RuntimeError("records 为空，可能前面的 _stash 未触发。")

# 汇总
df_all = pd.concat(records, ignore_index=True)
df_all['time']   = pd.to_datetime(df_all['time'])
df_all['launch'] = pd.to_datetime(df_all['launch'])

# 固定来源顺序
source_order = ['Radiosonde', 'MWR', 'DWL', 'ERA5', 'CRA']
df_all['source'] = df_all['source'].astype(CategoricalDtype(categories=source_order, ordered=True))

# 统一排序键（与 rh_1min 一致）
SORT_KEYS = ['time', 'height', 'source']

# 导出前检查 launch
missing = df_all['launch'].isna().mean()
print(f"Launch 缺失比例：{missing:.1%}")

# 输出前先清理旧文件
if OUT_XLSX.exists():
    OUT_XLSX.unlink()

with pd.ExcelWriter(OUT_XLSX, engine='openpyxl') as xls:
    for name, grp in df_all.groupby('variable', sort=False):
        g = grp.copy()

        # 仅对 hws_1h 丢弃空值
        if name == 'hws_1h':
            g = g.dropna(subset=['value'])

        # 统一排序（稳定排序避免同值乱序）
        g = g.sort_values(SORT_KEYS, kind='mergesort')

        # 统一列顺序输出（包含 launch）
        cols = ['time', 'launch', 'source', 'height', 'value']
        sheet_name = name[:31]
        g[cols].to_excel(xls, sheet_name=sheet_name, index=False)

        # 设置列宽：所有表列宽统一为 25
        ws = xls.sheets[sheet_name]
        for col_idx in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

print(f"✅ 已写入 {OUT_XLSX}（所有表已包含 launch；hws_1h 已去空；统一排序；列宽=25）")
# =============================================================================

# =============================================================================
# 按高度分箱计算 epsilon 的核心实现（含详细中文注释）
# 目标：以 Radiosonde 为参考，在“严格相同时空点（launch, time, height）”上
#       计算各来源与 RS 的差值 diff，并按高度 0~10000 m（步长 100 m）分箱，
#       在每个高度箱内计算：
#         - B        = mean(diff)                       # 偏差（系统误差）
#         - mean_sq  = mean(diff**2)                    # 平方均值
#         - B2       = B**2
#         - sigma    = sqrt(mean_sq)                    # 注意：按你的新定义，不减 B2
#         - epsilon  = mean(|diff| <= sigma)            # σ 内比例（箱内样本为分母）
# 输出：每个来源、每个变量在每个高度箱的一行统计（不含 bin_center）
# 说明：
#   1) “相同时空点”是通过对 (launch, time, height) 做内连接保证的；
#   2) 分箱为右闭区间 (a, b]，并包含最小端（0）；使用 cat.codes 回填边界，
#      避免 pandas 默认显示的 -0.001 伪影；
#   3) 丢弃 height/diff 中的 NaN，防止统计被污染；
#   4) N 为每个高度箱的有效样本数，用于质量检查；
#   5) 若某来源/变量在某些箱没有数据，则该箱不会出现在结果中（自然缺省）。
# =============================================================================
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from pathlib import Path

# 1) 配置：变量清单与分箱
VAR_GROUPS = {
    "1h":   ["temp_1h", "rh_1h", "hws_1h"],
    "1min": ["temp_1min", "rh_1min", "hws_1min"],
}
BIN_EDGES = np.arange(0, 10000 + 100, 100)   # [0,100,200,...,10000]（右闭，含10000）

# 2) 工具：解析变量名
def _var_parts(var_name: str):
    """'temp_1h' -> ('temp','1h') ; 'hws_1min' -> ('hws','1min')"""
    base, res = var_name.split("_", 1)
    return base, res

# 3) 工具：该变量下有哪些非RS来源
def _sources_for_var(df_all, variable):
    s = (df_all[df_all["variable"] == variable]["source"]
         .dropna().unique().tolist())
    return [x for x in s if x != "Radiosonde"]

# 4) 在相同时空点配对（严格内连接）
def _pair_same_spacetime(df_all, variable, src, exclude_launches=None):
    """返回一个 DataFrame，含列 ['time','launch','height','diff']，diff=src-ref"""
    ref = df_all[(df_all["variable"] == variable) & (df_all["source"] == "Radiosonde")].copy()
    dat = df_all[(df_all["variable"] == variable) & (df_all["source"] == src)].copy()

    if exclude_launches:
        ref = ref[~ref["launch"].isin(exclude_launches)]
        dat = dat[~dat["launch"].isin(exclude_launches)]

    if ref.empty or dat.empty:
        return pd.DataFrame(columns=["time","launch","height","diff"])

    # 为保证严格同一发射次/时刻/高度，对 (launch,time,height) 做内连接
    key_cols = ["launch", "time", "height"]
    merged = (dat[key_cols + ["value"]]
              .merge(ref[key_cols + ["value"]],
                     on=key_cols, how="inner",
                     suffixes=("_src","_ref")))
    if merged.empty:
        return pd.DataFrame(columns=["time","launch","height","diff"])

    merged["diff"] = merged["value_src"] - merged["value_ref"]
    return merged[["time","launch","height","diff"]]

# 5) 按高度分箱计算指标
def _binned_stats(df_pairs, bin_edges):
    """
    输入：同一来源与RS配对得到的差值表 df_pairs（列含 height, diff）
    输出：各高度分箱统计结果 DataFrame
          列：['bin_low','bin_high','bin_center','N','B','mean_sq','B2','sigma','epsilon']
    """
    if df_pairs is None or df_pairs.empty:
        return pd.DataFrame(columns=["bin_low","bin_high","bin_center","N","B","mean_sq","B2","sigma","epsilon"])

    tmp = df_pairs.copy()
    tmp["height"] = pd.to_numeric(tmp["height"], errors="coerce")
    tmp["diff"]   = pd.to_numeric(tmp["diff"],   errors="coerce")
    tmp = tmp.dropna(subset=["height","diff"])
    if tmp.empty:
        return pd.DataFrame(columns=["bin_low","bin_high","bin_center","N","B","mean_sq","B2","sigma","epsilon"])

    # 分箱（右闭，含最低端），observed=True 仅保留出现过的箱
    tmp["bin"] = pd.cut(tmp["height"], bins=bin_edges, right=True, include_lowest=True, ordered=True)
    g = tmp.groupby("bin", observed=True)["diff"]

    # 逐箱聚合
    stat = pd.DataFrame({
        "N": g.size(),
        "B": g.mean(),
        "mean_sq": g.apply(lambda x: np.mean(np.square(x)) if len(x) else np.nan),
    })
    stat["B2"]   = stat["B"]**2
    stat["sigma"] = np.sqrt(stat["mean_sq"])

    # epsilon：箱内 |diff| <= sigma 的比例
    def _eps_one(x):
        if len(x) == 0:
            return np.nan
        sig = np.sqrt(np.mean(np.square(x)))
        return float((np.abs(x) <= sig).mean())
    stat["epsilon"] = g.apply(_eps_one)

    # 还原区间左右边 & 中心 —— 先转成对象再取 left/right，并强制为 float
    stat = stat.reset_index()
    bins_obj = stat["bin"].astype("object")
    stat["bin_low"]    = bins_obj.map(lambda iv: float(iv.left)  if pd.notna(iv) else np.nan)
    stat["bin_high"]   = bins_obj.map(lambda iv: float(iv.right) if pd.notna(iv) else np.nan)
    stat["bin_center"] = (stat["bin_low"].astype(float) + stat["bin_high"].astype(float)) / 2.0
    
    first_edge = float(bin_edges[0])
    last_edge  = float(bin_edges[-1])
    stat["bin_low"]  = stat["bin_low"].astype(float).clip(lower=first_edge)
    stat["bin_high"] = stat["bin_high"].astype(float).clip(upper=last_edge)
   
    # 收尾
    stat = stat.drop(columns=["bin"])
    stat = stat[["bin_low","bin_high","N","B","mean_sq","B2","sigma","epsilon"]]
    stat = stat.sort_values("bin_low").reset_index(drop=True)
    return stat

# 6) 计算某一分辨率（多个变量）的两套结果
def _compute_binned_for_vars(df_all, var_list, bin_edges, exclude_launches=None):
    """
    返回：一个长表，列：
      ['resolution','variable','source','bin_low','bin_high','bin_center','N','B','mean_sq','B2','sigma','epsilon']
    """
    rows = []
    for var in var_list:
        base, res = _var_parts(var)
        for src in _sources_for_var(df_all, var):
            pairs = _pair_same_spacetime(df_all, var, src, exclude_launches=exclude_launches)
            stat  = _binned_stats(pairs, bin_edges)
            if stat.empty:
                continue
            stat.insert(0, "source", src)
            stat.insert(0, "variable", base)   # 只保留物理量名（temp/rh/hws）
            stat.insert(0, "resolution", res)  # 1h / 1min
            rows.append(stat)
    if not rows:
        return pd.DataFrame(columns=["resolution","variable","source","bin_low","bin_high","bin_center","N","B","mean_sq","B2","sigma","epsilon"])
    return pd.concat(rows, ignore_index=True)

# 7) 需要排除的 launch（用你之前的解析函数 _launch_from_sheet）
EXCLUDE_LAUNCH_STRS = ["2024080908", "2024080920", "2024081008"]
EXCLUDE_LAUNCHES = [_launch_from_sheet(s) for s in EXCLUDE_LAUNCH_STRS]

# —— 全局过滤版本：先得到“已排除”的总表 —— #
df_all_excl = df_all[~df_all["launch"].isin(EXCLUDE_LAUNCHES)].copy()

# —— 计算四个场景：1h/all, 1h/excl, 1min/all, 1min/excl —— #
# “all” 用完整 df_all；“excl” 用全局过滤后的 df_all_excl
# 注意：这里把 exclude_launches 参数设为 None（或去掉），避免重复过滤
res1h_all   = _compute_binned_for_vars(df_all,       VAR_GROUPS["1h"],   BIN_EDGES, exclude_launches=None)
res1h_excl  = _compute_binned_for_vars(df_all_excl,  VAR_GROUPS["1h"],   BIN_EDGES, exclude_launches=None)
res1m_all   = _compute_binned_for_vars(df_all,       VAR_GROUPS["1min"], BIN_EDGES, exclude_launches=None)
res1m_excl  = _compute_binned_for_vars(df_all_excl,  VAR_GROUPS["1min"], BIN_EDGES, exclude_launches=None)


# 8) 写入 Excel（每个变量/场景一个工作表；列宽=25）
EPS_BIN_XLSX = OUT_XLSX.parent / "eps_binned_by_height_1h_1min.xlsx"

def _write_binned_to_xlsx(filepath: Path, df_all_res, tag: str):
    """
    df_all_res: 单一场景（如 1h_all）的结果长表
    tag: 场景后缀，如 '1h_all', '1h_excl', '1min_all', '1min_excl'
    - 每个变量（temp/rh/hws）各一张表，表名：<tag>_<variable>
    """
    # 按变量拆表
    for var in ["temp","rh","hws"]:
        sub = df_all_res[df_all_res["variable"] == var].copy()
        if sub.empty:
            continue
        # 列顺序与排序
        cols = ["resolution","variable","source","bin_low","bin_high","N","B","mean_sq","B2","sigma","epsilon"]
        sub = sub[cols].sort_values(["resolution","source","bin_low"], kind="mergesort")

        sheet = f"{tag}_{var}"[:31]
        sub.to_excel(_xls_writer, sheet_name=sheet, index=False)
        # 列宽统一 25
        ws = _xls_writer.sheets[sheet]
        for i in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

# 覆盖写
if EPS_BIN_XLSX.exists():
    EPS_BIN_XLSX.unlink()

with pd.ExcelWriter(EPS_BIN_XLSX, engine="openpyxl") as _xls_writer:
    _write_binned_to_xlsx(EPS_BIN_XLSX, res1h_all,  "1h_all")
    _write_binned_to_xlsx(EPS_BIN_XLSX, res1h_excl, "1h_excl")
    _write_binned_to_xlsx(EPS_BIN_XLSX, res1m_all,  "1min_all")
    _write_binned_to_xlsx(EPS_BIN_XLSX, res1m_excl, "1min_excl")

print(f"✅ 高度分箱 epsilon 结果已写入：{EPS_BIN_XLSX}")



# =============================================================================
# ε(launch) — 仅随发射时间变化的 epsilon（无高度分箱）
# 以 Radiosonde 为参考：在每个 launch 内跨 time×height 的所有同位点差值一起计算
# 结果覆盖 1h 与 1min；场景：all / exclude 指定 launch；输出到 Excel（列宽=25）
# =============================================================================
from openpyxl.utils import get_column_letter
from pathlib import Path

# 分辨率变量清单
VAR_GROUPS = {
    "1h":   ["temp_1h", "rh_1h", "hws_1h"],
    "1min": ["temp_1min", "rh_1min", "hws_1min"],
}

def _var_parts(var_name: str):
    """'temp_1h' -> ('temp','1h') ; 'hws_1min' -> ('hws','1min')"""
    base, res = var_name.split("_", 1)
    return base, res

def _sources_for_var(df_all, variable):
    s = (df_all[df_all["variable"] == variable]["source"]
         .dropna().unique().tolist())
    return [x for x in s if x != "Radiosonde"]

def _pair_same_spacetime(df_all, variable, src, exclude_launches=None):
    """返回列 ['time','launch','height','diff']；diff=src-ref，仅包含同一 (launch,time,height)。"""
    ref = df_all[(df_all["variable"] == variable) & (df_all["source"] == "Radiosonde")].copy()
    dat = df_all[(df_all["variable"] == variable) & (df_all["source"] == src)].copy()
    if exclude_launches is not None and len(exclude_launches):
        ref = ref[~ref["launch"].isin(exclude_launches)]
        dat = dat[~dat["launch"].isin(exclude_launches)]
    if ref.empty or dat.empty:
        return pd.DataFrame(columns=["time","launch","height","diff"])
    key_cols = ["launch", "time", "height"]
    merged = (dat[key_cols + ["value"]]
              .merge(ref[key_cols + ["value"]],
                     on=key_cols, how="inner",
                     suffixes=("_src","_ref")))
    if merged.empty:
        return pd.DataFrame(columns=["time","launch","height","diff"])
    merged["diff"] = merged["value_src"] - merged["value_ref"]
    return merged[["time","launch","height","diff"]]

def _launch_stats(df_pairs: pd.DataFrame) -> pd.DataFrame:
    """
    在每个 launch 内聚合：
      N, B=mean(diff), mean_sq=mean(diff**2), B2=B**2, sigma=sqrt(mean_sq), 
      epsilon=mean(|diff|<=sigma)
    返回列：['launch','N','B','mean_sq','B2','sigma','epsilon']
    """
    if df_pairs is None or df_pairs.empty:
        return pd.DataFrame(columns=['launch','N','B','mean_sq','B2','sigma','epsilon'])
    tmp = df_pairs.copy()
    tmp['diff'] = pd.to_numeric(tmp['diff'], errors='coerce')
    tmp = tmp.dropna(subset=['diff', 'launch'])
    if tmp.empty:
        return pd.DataFrame(columns=['launch','N','B','mean_sq','B2','sigma','epsilon'])

    g = tmp.groupby('launch')['diff']
    stat = pd.DataFrame({
        'N': g.size(),
        'B': g.mean(),
        'mean_sq': g.apply(lambda x: np.mean(np.square(x)) if len(x) else np.nan),
    })
    stat['B2']    = stat['B']**2
    stat['sigma'] = np.sqrt(stat['mean_sq'])

    def _eps_one(x):
        if len(x) == 0:
            return np.nan
        sig = float(np.sqrt(np.mean(np.square(x))))
        return float((np.abs(x) <= sig).mean())
    stat['epsilon'] = g.apply(_eps_one)

    stat = stat.reset_index().sort_values('launch').reset_index(drop=True)
    return stat[['launch','N','B','mean_sq','B2','sigma','epsilon']]

def _compute_launch_for_vars(df_use: pd.DataFrame, var_list):
    """
    在给定 df_use（全量或已全局排除）上，计算每个 variable×source 的 launch 级 ε。
    返回长表：['resolution','variable','source','launch','N','B','mean_sq','B2','sigma','epsilon']
    """
    rows = []
    for var in var_list:
        base, res = _var_parts(var)
        for src in _sources_for_var(df_use, var):
            pairs = _pair_same_spacetime(df_use, var, src, exclude_launches=None)  # 全局过滤由 df_use 决定
            stat  = _launch_stats(pairs)
            if stat.empty:
                continue
            stat.insert(0, 'source', src)
            stat.insert(0, 'variable', base)
            stat.insert(0, 'resolution', res)
            rows.append(stat)
    if not rows:
        return pd.DataFrame(columns=['resolution','variable','source','launch','N','B','mean_sq','B2','sigma','epsilon'])
    out = pd.concat(rows, ignore_index=True)
    out['launch'] = pd.to_datetime(out['launch'])
    return out

# —— 全局排除指定 launch —— #
EXCLUDE_LAUNCH_STRS = ["2024080908", "2024080920", "2024081008"]
EXCLUDE_LAUNCHES = [_launch_from_sheet(s) for s in EXCLUDE_LAUNCH_STRS]
df_all_excl = df_all[~df_all["launch"].isin(EXCLUDE_LAUNCHES)].copy()

# —— 计算四个场景 —— #
res_launch_1h_all   = _compute_launch_for_vars(df_all,      VAR_GROUPS["1h"])
res_launch_1h_excl  = _compute_launch_for_vars(df_all_excl, VAR_GROUPS["1h"])
res_launch_1m_all   = _compute_launch_for_vars(df_all,      VAR_GROUPS["1min"])
res_launch_1m_excl  = _compute_launch_for_vars(df_all_excl, VAR_GROUPS["1min"])

# —— 写 Excel（每个 <分辨率×变量×场景> 一个工作表；列宽=25） —— #
EPS_LAUNCH_XLSX = OUT_XLSX.parent / "eps_launch_only_1h_1min.xlsx"

def _write_launch_to_xlsx(writer, df_res: pd.DataFrame, res_tag: str, scenario_tag: str):
    """
    将某分辨率(res_tag: '1h'/'1min')、某场景(scenario_tag: 'all'/'excl')的
    launch 级结果写入多个工作表（每个变量一个表）。
    sheet名：<res_tag>_launch_<variable>_<scenario_tag> 例：1h_launch_temp_all
    """
    for base in ['temp','rh','hws']:
        sub = df_res[(df_res['variable']==base) & (df_res['resolution']==res_tag)].copy()
        if sub.empty:
            continue
        cols = ['resolution','variable','source','launch','N','B','mean_sq','B2','sigma','epsilon']
        sub = sub[cols].sort_values(['source','launch'], kind='mergesort')
        sheet = f"{res_tag}_launch_{base}_{scenario_tag}"[:31]
        sub.to_excel(writer, sheet_name=sheet, index=False)
        ws = writer.sheets[sheet]
        for i in range(1, len(cols)+1):
            ws.column_dimensions[get_column_letter(i)].width = 25

# 覆盖写
if EPS_LAUNCH_XLSX.exists():
    EPS_LAUNCH_XLSX.unlink()

with pd.ExcelWriter(EPS_LAUNCH_XLSX, engine="openpyxl") as xls:
    _write_launch_to_xlsx(xls, res_launch_1h_all,  "1h",   "all")
    _write_launch_to_xlsx(xls, res_launch_1h_excl, "1h",   "excl")
    _write_launch_to_xlsx(xls, res_launch_1m_all,  "1min", "all")
    _write_launch_to_xlsx(xls, res_launch_1m_excl, "1min", "excl")

print(f"✅ 仅随发射时间变化的 epsilon 已写入：{EPS_LAUNCH_XLSX}（无高度分箱）")
